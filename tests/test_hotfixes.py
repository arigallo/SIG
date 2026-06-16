import os
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

os.environ["INIT_DB"] = "false"

from flask import render_template
from openpyxl import Workbook, load_workbook

import app
from services import calendario as calendario_service
from services import notificaciones as notificaciones_service
from services import portal as portal_service


class HotfixTests(unittest.TestCase):
    def test_timezone_aware_datetime_is_serialized_before_saving(self):
        wb = Workbook()
        ws = wb.active

        app.append_fila_reporte(
            ws,
            [datetime(2026, 5, 20, 14, 0, 51, tzinfo=timezone.utc)],
        )

        self.assertEqual(ws["A1"].value, "2026-05-20 11:00:51")

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb.save(path)
            reloaded = load_workbook(path)
            self.assertEqual(reloaded.active["A1"].value, "2026-05-20 11:00:51")
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_usuarios_template_accepts_string_ultimo_login(self):
        usuario = {
            "id": 1,
            "username": "admin",
            "email": "admin@example.com",
            "rol": "admin",
            "debe_cambiar_password": 0,
            "onboarding_visto": 1,
            "ultimo_login": "2026-05-20 12:34:56+00",
        }

        with app.app.test_request_context("/usuarios"):
            html = render_template("usuarios.html", usuarios=[usuario])

        self.assertIn("2026-05-20 09:34", html)

    def test_no_known_unsafe_fecha_vencimiento_casts(self):
        source = Path("app.py").read_text(encoding="utf-8")
        unsafe_patterns = [
            "fecha_vencimiento <> ''\n                     AND fecha_vencimiento::date",
            "c.fecha_vencimiento <> ''\n                     AND c.fecha_vencimiento::date",
            "f.fecha_vencimiento <> ''\n          AND f.fecha_vencimiento::date",
            "c.fecha_vencimiento <> ''\n             AND c.fecha_vencimiento::date",
        ]

        for pattern in unsafe_patterns:
            with self.subTest(pattern=pattern):
                self.assertNotIn(pattern, source)

    def test_asistencia_nuevo_redirects_to_calendario_form(self):
        with patch.object(app, "obtener_config_mantenimiento", return_value={"activo": False}):
            client = app.app.test_client()
            with client.session_transaction() as session:
                session["user_id"] = 1
                session["rol"] = "entrenador"
                session["permisos"] = ["asistencia_gestionar"]

            response = client.get("/asistencia/nuevo")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/calendario/nuevo?origen=asistencia", response.headers["Location"])

    def test_asistencia_listing_orders_upcoming_events_first(self):
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("WHEN e.fecha >= CURRENT_DATE::text THEN 0", source)
        self.assertIn("WHEN e.fecha >= CURRENT_DATE::text THEN e.fecha", source)
        self.assertIn("WHEN e.fecha < CURRENT_DATE::text THEN e.fecha", source)

    def test_calendario_generates_monthly_training_dates(self):
        fechas = app.generar_fechas_recurrentes_mes("2026-06", ["1", "3"])

        self.assertEqual(
            fechas,
            [
                "2026-06-02",
                "2026-06-04",
                "2026-06-09",
                "2026-06-11",
                "2026-06-16",
                "2026-06-18",
                "2026-06-23",
                "2026-06-25",
                "2026-06-30",
            ],
        )
        self.assertIs(app.generar_fechas_recurrentes_mes, calendario_service.generar_fechas_recurrentes_mes)

    def test_calendario_form_exposes_monthly_training_batch(self):
        base = Path("templates/base.html").read_text(encoding="utf-8")
        dashboard = Path("templates/dashboard.html").read_text(encoding="utf-8")
        template = Path("templates/calendario_evento_form.html").read_text(encoding="utf-8")
        calendario = Path("templates/calendario.html").read_text(encoding="utf-8")
        asistencia = Path("templates/asistencia_eventos.html").read_text(encoding="utf-8")
        styles = Path("static/styles.css").read_text(encoding="utf-8")
        source = Path("app.py").read_text(encoding="utf-8")
        service = Path("services/calendario.py").read_text(encoding="utf-8")
        repository = Path("repositories/calendario.py").read_text(encoding="utf-8")

        self.assertIn("Crear varios entrenamientos del mes", template)
        self.assertIn("batch-event-panel", template)
        self.assertEqual(base.count('class="main-nav"'), 1)
        self.assertNotIn("workspace-bar", base)
        self.assertIn("quick-module-grid", dashboard)
        self.assertIn("dashboard-hero", dashboard)
        self.assertIn("repeticion_mes", template)
        self.assertIn("repeticion_dias", template)
        self.assertIn("Entrenamientos del mes", calendario)
        self.assertIn("Entrenamientos del mes", asistencia)
        self.assertIn("module-switcher", calendario)
        self.assertIn("module-switcher", asistencia)
        self.assertIn("module-callout", calendario)
        self.assertNotIn(".workspace-bar", styles)
        self.assertIn(".quick-module-card", styles)
        self.assertIn(".module-switcher", styles)
        self.assertIn(".module-callout", styles)
        self.assertIn("modo='entrenamientos_mes'", calendario)
        self.assertIn("modo='entrenamientos_mes'", asistencia)
        self.assertIn("generar_fechas_recurrentes_mes", source)
        self.assertIn("from services.calendario import", source)
        self.assertIn("crear_eventos_calendario", source)
        self.assertIn("from repositories.calendario import", service)
        self.assertIn("def crear_eventos_calendario", service)
        self.assertIn("def existe_evento_calendario", repository)
        self.assertIn("def crear_evento_calendario_desde_data", repository)

    def test_portal_asistencia_labels_change_for_partidos(self):
        partido = {"tipo": "Partido"}
        entrenamiento = {"tipo": "Entrenamiento"}
        source = Path("app.py").read_text(encoding="utf-8")
        service = Path("services/portal.py").read_text(encoding="utf-8")

        self.assertEqual(
            [opcion["label"] for opcion in app.asistencia_portal_opciones(partido)],
            ["Voy y Juego", "Voy y no juego", "No voy"],
        )
        self.assertEqual(app.asistencia_portal_label(partido, "dudoso"), "Voy y no juego")
        self.assertTrue(app.es_evento_partido({"tipo": "Partidos"}))
        self.assertIs(app.es_evento_partido, portal_service.es_evento_partido)
        self.assertIn("from services.portal import", source)
        self.assertIn("def resumen_bienestar_confirmacion", service)
        self.assertEqual(
            [opcion["label"] for opcion in app.asistencia_portal_opciones(entrenamiento)],
            ["Voy", "Dudoso", "No voy"],
        )

    def test_asistencia_event_export_is_excel_and_includes_unsaved_players(self):
        source = Path("app.py").read_text(encoding="utf-8")
        template = Path("templates/tomar_asistencia.html").read_text(encoding="utf-8")

        self.assertIn("download_name=f\"asistencia_{evento_id}.xlsx\"", source)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", source)
        self.assertIn("FROM jugadores j", source)
        self.assertIn("LEFT JOIN asistencias a", source)
        self.assertIn("estado_asistencia or \"Sin registrar\"", source)
        self.assertIn("Exportar Excel", template)
        self.assertNotIn("Exportar CSV", template)

    def test_whatsapp_unsupported_message_shows_meta_reason(self):
        resumen = app.resumir_contenido_whatsapp(
            "unsupported",
            {
                "type": "unsupported",
                "errors": [
                    {
                        "code": 131051,
                        "title": "Unsupported message type",
                        "details": "Message type is not currently supported",
                    }
                ],
            },
        )

        self.assertEqual(resumen, "[Mensaje no compatible] Unsupported message type")

    def test_whatsapp_interactive_and_contact_messages_are_readable(self):
        interactive = app.resumir_contenido_whatsapp(
            "interactive",
            {"interactive": {"button_reply": {"id": "ok", "title": "Confirmar"}}},
        )
        contacto = app.resumir_contenido_whatsapp(
            "contacts",
            {"contacts": [{"name": {"formatted_name": "Ariel Gallo"}}]},
        )

        self.assertEqual(interactive, "Confirmar")
        self.assertEqual(contacto, "[Contacto] Ariel Gallo")

    def test_whatsapp_inbox_email_notification_uses_configured_recipients(self):
        enviados = []

        def fake_enviar_email(destinatario, asunto, cuerpo):
            enviados.append((destinatario, asunto, cuerpo))
            return True, None

        with app.app.test_request_context("/webhooks/whatsapp", base_url="https://sig.example.test"):
            with patch.object(app, "WHATSAPP_INBOX_NOTIFY_EMAILS", ["avisos@example.com"]):
                with patch.object(app, "suprimir_email_whatsapp_por_presencia", return_value=False):
                    with patch.object(app, "enviar_email", side_effect=fake_enviar_email):
                        enviado = app.enviar_notificacion_whatsapp_inbox_email(
                            mensaje={"tipo": "text", "texto": "Hola"},
                            telefono="5491112345678",
                            jugador={"apellido": "Gallo", "nombre": "Ariel"},
                        )

        self.assertTrue(enviado)
        self.assertEqual(enviados[0][0], "avisos@example.com")
        self.assertIn("Nueva respuesta WhatsApp - Gallo, Ariel", enviados[0][1])
        self.assertIn("Mensaje: Hola", enviados[0][2])
        self.assertIn("/comunicacion/whatsapp?telefono=5491112345678", enviados[0][2])

    def test_whatsapp_inbox_email_notification_falls_back_to_smtp_from(self):
        with patch.object(app, "WHATSAPP_INBOX_NOTIFY_EMAILS", []):
            with patch.object(app, "SMTP_FROM", "tesoreria@example.com"):
                self.assertEqual(
                    app.destinatarios_notificacion_whatsapp_inbox(),
                    ["tesoreria@example.com"],
                )

    def test_login_template_exposes_three_entry_actions(self):
        template = Path("templates/login.html").read_text(encoding="utf-8")

        self.assertIn("Acceso Administrativo", template)
        self.assertIn("Portal del Jugador", template)
        self.assertIn("Sugerencias / Recomendaciones", template)
        self.assertIn("url_for('sugerencias_recomendaciones')", template)
        self.assertIn("avisos_login", template)
        self.assertIn("entry-login-notice", template)
        self.assertIn("entry-login-notice-urgent", template)
        self.assertIn("configurar_avisos_login", Path("app.py").read_text(encoding="utf-8-sig"))
        admin_template = Path("templates/avisos_login_admin.html").read_text(encoding="utf-8")
        self.assertIn("Avisos del login", admin_template)
        self.assertIn("urgente_", admin_template)

    def test_error_500_email_alerts_are_configurable(self):
        source = Path("app.py").read_text(encoding="utf-8-sig")
        sistema = Path("templates/sistema_admin.html").read_text(encoding="utf-8")

        self.assertIn("@app.errorhandler(500)", source)
        self.assertIn("def notificar_error_500", source)
        self.assertIn("ERROR_500_ALERT_EMAILS_KEY", source)
        self.assertIn("def configurar_alertas_500", source)
        self.assertIn("Alertas error 500", sistema)
        self.assertIn("configurar_alertas_500", sistema)

    def test_sugerencias_use_configured_recipients(self):
        config = {
            "directiva_emails": ["directiva@example.com"],
            "actualizado_en": None,
            "actualizado_por": None,
        }

        with patch.object(app, "obtener_sugerencias_config", return_value=config):
            destinatarios = app.obtener_destinatarios_sugerencias(object())

        self.assertEqual(destinatarios, ["directiva@example.com"])

    def test_sistema_admin_links_sugerencias_config(self):
        template = Path("templates/sistema_admin.html").read_text(encoding="utf-8")
        nav = Path("templates/base.html").read_text(encoding="utf-8")
        admin_template = Path("templates/sugerencias_recomendaciones_admin.html").read_text(encoding="utf-8")

        self.assertIn("url_for('listar_sugerencias_recomendaciones')", template)
        self.assertIn("listar_sugerencias_recomendaciones", nav)
        self.assertIn("url_for('configurar_sugerencias_recomendaciones')", admin_template)
        self.assertIn("configurar_sugerencias_recomendaciones", nav)

    def test_sugerencias_admin_template_explains_pending_email_states(self):
        template = Path("templates/sugerencias_recomendaciones_admin.html").read_text(encoding="utf-8")
        public_template = Path("templates/sugerencias_recomendaciones.html").read_text(encoding="utf-8")
        source = Path("app.py").read_text(encoding="utf-8")

        self.assertIn("Bandeja interna", template)
        self.assertIn("Email pendiente", template)
        self.assertIn("registro.email_info.descripcion", template)
        self.assertIn("Notas internas", template)
        self.assertIn("Reenviar email", template)
        self.assertIn("actualizar_sugerencia_recomendacion", source)
        self.assertIn("reenviar_sugerencia_recomendacion", source)
        self.assertIn("No pudimos enviar el aviso por email", source)
        self.assertIn("data-anonymous-select", public_template)
        self.assertNotIn("Denuncia", public_template)

    def test_sugerencias_permissions_are_registered(self):
        for permiso in [
            "sugerencias_ver",
            "sugerencias_gestionar",
            "sugerencias_configurar",
        ]:
            self.assertIn(permiso, app.PERMISOS)

        self.assertNotIn("denuncias_ver", app.PERMISOS)
        self.assertIn("sugerencias_gestionar", app.ROLE_PRESETS["admin"])

    def test_closed_shared_expenses_remain_visible_as_player_debt(self):
        source = Path("app.py").read_text(encoding="utf-8")
        player_template = Path("templates/jugador_detalle.html").read_text(encoding="utf-8")
        portal_template = Path("templates/portal_jugador.html").read_text(encoding="utf-8")

        self.assertIn("OR i.estado = 'pendiente'", source)
        self.assertIn("deuda_gastos_compartidos", source)
        self.assertIn("gastos_compartidos_pendientes", player_template)
        self.assertIn("Cerrado con deuda", player_template)
        self.assertIn("Cerrado con saldo pendiente", portal_template)
        self.assertIn("Cuotas: {{ deuda_cuotas | moneda }}", portal_template)

    def test_drive_runtime_error_reports_missing_secretaria_config(self):
        mensaje = app.mensaje_error_drive(
            RuntimeError("Falta configurar GOOGLE_DRIVE_SECRETARIA_FOLDER_ID, GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID o GOOGLE_DRIVE_SHARED_DRIVE_ID."),
            carpeta="Secretaria",
            accion="guardar el documento",
        )

        self.assertIn("Falta configurar GOOGLE_DRIVE_SECRETARIA_FOLDER_ID", mensaje)
        self.assertIn("GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID", mensaje)
        self.assertIn("No se pudo guardar el documento en Google Drive.", mensaje)

    def test_drive_runtime_error_reports_missing_secretaria_folder(self):
        mensaje = app.mensaje_error_drive(
            RuntimeError("No se encontro la carpeta 'Secretaria' en la unidad compartida configurada."),
            carpeta="Secretaria",
            accion="guardar el documento",
        )

        self.assertIn("No se encontro la carpeta 'Secretaria'", mensaje)
        self.assertIn("GOOGLE_DRIVE_SECRETARIA_FOLDER_ID", mensaje)

    def test_secretaria_drive_creates_base_year_and_month_folders(self):
        class FakeExecute:
            def __init__(self, response):
                self.response = response

            def execute(self):
                return self.response

        class FakeFiles:
            def __init__(self):
                self.created = []

            def list(self, **kwargs):
                return FakeExecute({"files": []})

            def create(self, **kwargs):
                body = kwargs["body"]
                folder_id = f"folder-{len(self.created) + 1}"
                self.created.append({"id": folder_id, "body": body, "kwargs": kwargs})
                return FakeExecute({"id": folder_id})

        class FakeService:
            def __init__(self):
                self._files = FakeFiles()

            def files(self):
                return self._files

        service = FakeService()
        with patch.object(app, "DRIVE_SECRETARIA_FOLDER_ID", ""):
            with patch.object(app, "DRIVE_SHARED_DRIVE_ID", "shared-drive-id"):
                with patch.object(app, "DRIVE_SECRETARIA_SUBFOLDER", "Secretaria"):
                    folder_id = app.get_drive_secretaria_folder(service, "Actas", fecha_base="2026-06-01")

        self.assertEqual(folder_id, "folder-4")
        nombres = [item["body"]["name"] for item in service._files.created]
        padres = [item["body"]["parents"][0] for item in service._files.created]
        self.assertEqual(nombres, ["Secretaria", "2026", "Junio", "Actas"])
        self.assertEqual(padres, ["shared-drive-id", "folder-1", "folder-2", "folder-3"])

    def test_secretaria_drive_uses_comprobantes_folder_as_fallback(self):
        service = object()
        with patch.object(app, "DRIVE_SECRETARIA_FOLDER_ID", "comprobantes-root"):
            with patch.object(app, "DRIVE_SECRETARIA_SUBFOLDER", "Secretaria"):
                with patch.object(app, "get_or_create_drive_subfolder", return_value="secretaria-root") as crear:
                    folder_id = app.get_drive_secretaria_base_folder(service)

        self.assertEqual(folder_id, "secretaria-root")
        crear.assert_called_once_with(service, "comprobantes-root", "Secretaria")

    def test_whatsapp_email_notification_is_suppressed_by_presence(self):
        with patch.object(app, "suprimir_email_whatsapp_por_presencia", return_value=True):
            with patch.object(app, "enviar_email") as enviar_email:
                enviado = app.enviar_notificacion_whatsapp_inbox_email(
                    mensaje={"tipo": "text", "texto": "Hola"},
                    telefono="5491112345678",
                    jugador=None,
                )

        self.assertFalse(enviado)
        enviar_email.assert_not_called()

    def test_presence_heartbeat_requires_session_and_records_user(self):
        with patch.object(app, "obtener_config_mantenimiento", return_value={"activo": False}):
            with patch.object(app, "registrar_presencia_usuario", return_value=True) as registrar:
                client = app.app.test_client()
                with client.session_transaction() as session:
                    session["user_id"] = 1
                    session["username"] = "arielgallo"
                    session["rol"] = "admin"
                    session["permisos"] = []
                    session["_csrf_token"] = "token"

                response = client.post(
                    "/presencia/heartbeat",
                    headers={"X-CSRF-Token": "token"},
                )

        self.assertEqual(response.status_code, 200)
        registrar.assert_called_once_with("arielgallo")

    def test_whatsapp_status_endpoint_returns_json(self):
        with patch.object(app, "obtener_config_mantenimiento", return_value={"activo": False}):
            with patch.object(app, "permiso_requerido", return_value=None):
                with patch.object(app, "obtener_estado_whatsapp_inbox", return_value={"sin_leer": 2, "ultimo_id": 10}):
                    client = app.app.test_client()
                    with client.session_transaction() as session:
                        session["user_id"] = 1
                        session["username"] = "arielgallo"
                        session["rol"] = "admin"
                        session["permisos"] = ["comunicaciones_ver"]

                    response = client.get("/comunicacion/whatsapp/estado")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"sin_leer": 2, "ultimo_id": 10})

    def test_notificaciones_template_highlights_sig_user_events(self):
        with app.app.test_request_context("/notificaciones"):
            with patch.object(app, "obtener_contador_notificaciones", return_value=3):
                with patch.object(app, "obtener_contador_whatsapp_inbox", return_value=1):
                    cambios = [{
                        "apellido": "Perez",
                        "nombre": "Juan",
                        "detalle_resumen": "Cambio telefono",
                        "fecha": f"2026-05-23 09:0{i}",
                        "jugador_id": 2 + i,
                        "_notificacion_tipo": "cambio_portal",
                        "_notificacion_id": str(20 + i),
                        "_notificacion_key": f"cambio_portal:{20 + i}",
                    } for i in range(6)]
                    comprobantes = [{
                        "apellido": "Gallo",
                        "nombre": "Ariel",
                        "periodo": f"2026-0{i + 1}",
                        "importe": 1000,
                        "comprobante_fecha": "2026-05-23",
                        "comprobante_usuario": "portal",
                        "comprobante_nombre": "ticket.pdf",
                        "jugador_id": 1,
                        "_notificacion_tipo": "comprobante",
                        "_notificacion_id": f"{i + 1}:2026-05-23",
                        "_notificacion_key": f"comprobante:{i + 1}:2026-05-23",
                    } for i in range(6)]
                    html = render_template(
                        "notificaciones.html",
                        cuotas_vencidas=[],
                        cuotas_por_vencer=[],
                        fichas=[],
                        asistencia_baja=[],
                        comprobantes_pendientes=comprobantes,
                        whatsapp_conversaciones=[{
                            "apellido": None,
                            "nombre": None,
                            "jugador_id": None,
                            "telefono": "5491111111111",
                            "categoria": "",
                            "texto": "Hola",
                            "tipo": "text",
                            "sin_leer": 1,
                            "creado_en": "2026-05-23 10:00",
                            "_notificacion_tipo": "whatsapp",
                            "_notificacion_id": "5491111111111:10",
                            "_notificacion_key": "whatsapp:5491111111111:10",
                        }],
                        secretaria_documentos=[],
                        ahijadxs_objetivo=[],
                        cambios_portal=cambios,
                        whatsapp_api_activa=False,
                    )

        self.assertIn("Hay un nuevo mensaje de WhatsApp", html)
        self.assertIn("Sin vincular", html)
        self.assertIn("Perez, Juan modific", html)
        self.assertIn("Hay comprobantes por verificar", html)
        self.assertIn("Ignorar", html)
        self.assertIn("+2 m", html)
        self.assertIn("Seleccionar visibles", html)
        self.assertIn("Ignorar seleccionadas", html)
        self.assertIn("Ignorar todas", html)

    def test_dismiss_notification_records_user_setting(self):
        with patch.object(app, "obtener_config_mantenimiento", return_value={"activo": False}):
            with patch.object(app, "descartar_notificacion_usuario", return_value=True) as descartar:
                with patch.object(app, "registrar_auditoria") as auditar:
                    client = app.app.test_client()
                    with client.session_transaction() as session:
                        session["user_id"] = 1
                        session["username"] = "arielgallo"
                        session["rol"] = "admin"
                        session["permisos"] = ["comunicaciones_ver"]
                        session["_csrf_token"] = "token"

                    response = client.post(
                        "/notificaciones/descartar",
                        data={
                            "_csrf_token": "token",
                            "tipo": "comprobante",
                            "entidad_id": "1:2026-05-23",
                        },
                    )

        self.assertEqual(response.status_code, 302)
        self.assertIn("/notificaciones", response.headers["Location"])
        descartar.assert_called_once_with("comprobante", "1:2026-05-23")
        auditar.assert_any_call(
            "descartar",
            "notificacion",
            "comprobante:1:2026-05-23",
            {"tipo": "comprobante", "entidad_id": "1:2026-05-23"},
        )

    def test_bulk_dismiss_notifications_uses_selected_or_all_items(self):
        with patch.object(app, "obtener_config_mantenimiento", return_value={"activo": False}):
            with patch.object(app, "descartar_notificaciones_usuario", return_value=2) as descartar:
                client = app.app.test_client()
                with client.session_transaction() as session:
                    session["user_id"] = 1
                    session["username"] = "arielgallo"
                    session["rol"] = "admin"
                    session["permisos"] = ["comunicaciones_ver"]
                    session["_csrf_token"] = "token"

                response = client.post(
                    "/notificaciones/descartar-lote",
                    data={
                        "_csrf_token": "token",
                        "modo": "todas",
                        "items": ["whatsapp|5491111111111:10"],
                        "all_items": ["whatsapp|5491111111111:10", "comprobante|1:2026-05-23"],
                    },
                )

        self.assertEqual(response.status_code, 302)
        self.assertIn("/notificaciones", response.headers["Location"])
        descartar.assert_called_once_with([
            ("whatsapp", "5491111111111:10"),
            ("comprobante", "1:2026-05-23"),
        ])

    def test_operacion_template_renders_daily_review_and_tasks(self):
        with app.app.test_request_context("/operacion"):
            with patch.object(app, "obtener_contador_notificaciones", return_value=0):
                with patch.object(app, "obtener_contador_whatsapp_inbox", return_value=0):
                    html = render_template(
                        "operacion.html",
                        revision={
                            "whatsapp": 1,
                            "comprobantes": 2,
                            "cambios_portal": 3,
                            "cuotas": 4,
                            "fichas": 5,
                            "asistencia_baja": 0,
                            "secretaria": 1,
                            "ahijadxs": 0,
                            "proximos_eventos": [],
                            "tareas_vencidas": 1,
                        },
                        tareas=[{
                            "id": 1,
                            "titulo": "Revisar comprobante",
                            "descripcion": "Pago informado",
                            "modulo": "finanzas",
                            "prioridad": "alta",
                            "responsable": "arielgallo",
                            "fecha_vencimiento": "2026-05-24",
                            "estado": "pendiente",
                            "jugador_id": None,
                            "apellido": None,
                            "nombre": None,
                        }],
                        estado="pendiente",
                        puede_gestionar_tareas=True,
                    )

        self.assertIn("Operaci", html)
        self.assertIn("Centro de tareas", html)
        self.assertIn("Revisar comprobante", html)

    def test_cobranzas_template_renders_pipeline(self):
        with app.app.test_request_context("/finanzas/cobranzas"):
            with patch.object(app, "obtener_contador_notificaciones", return_value=0):
                with patch.object(app, "obtener_contador_whatsapp_inbox", return_value=0):
                    html = render_template(
                        "cobranzas.html",
                        panel={
                            "resumen": {
                                "emitidas": 10,
                                "pagadas": 6,
                                "pendientes": 4,
                                "vencidas": 2,
                                "comprobantes": 1,
                                "pendiente_importe": 4000,
                            },
                            "avance": 60,
                            "por_categoria": [{"categoria": "Plantel Superior", "pendientes": 2, "deuda": 2000}],
                            "comprobantes_recientes": [],
                        },
                    )

        self.assertIn("Panel de cobranzas", html)
        self.assertIn("Plantel Superior", html)
        self.assertIn("60%", html)

    def test_whatsapp_inbox_template_has_quick_replies(self):
        with app.app.test_request_context("/comunicacion/whatsapp"):
            with patch.object(app, "obtener_contador_notificaciones", return_value=0):
                with patch.object(app, "obtener_contador_whatsapp_inbox", return_value=0):
                    html = render_template(
                        "whatsapp_inbox.html",
                        conversaciones=[],
                        conversacion_actual={"nombre": "Sin vincular", "telefono": "5491111111111", "jugador_id": None},
                        mensajes=[],
                        telefono_actual="5491111111111",
                        jugadores_disponibles=[],
                        webhook_eventos=[],
                        respuestas_rapidas=["Gracias, lo revisamos."],
                    )

        self.assertIn("data-whatsapp-reply", html)
        self.assertIn("Gracias, lo revisamos.", html)

    def test_urba_sync_timeout_is_user_facing_error(self):
        with patch.object(app, "urlopen", side_effect=TimeoutError()):
            with self.assertRaises(RuntimeError) as contexto:
                app.sincronizar_circulares_urba(None, 2026, "admin")

        self.assertIn("URBA no respondio", str(contexto.exception))

    def test_management_package_keeps_closed_debt_and_later_payment(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        template = (Path(app.__file__).parent / "templates" / "gasto_compartido_detalle.html").read_text(encoding="utf-8-sig")

        self.assertIn("def registrar_pago_posterior_gasto_compartido", source)
        self.assertIn("Pago posterior gasto compartido", source)
        self.assertIn("i.estado = 'pendiente'", source)
        self.assertIn("Registrar pago", template)

    def test_management_package_exposes_ledger_automation_and_rate_limit(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        portal = (Path(app.__file__).parent / "templates" / "portal_jugador.html").read_text(encoding="utf-8-sig")
        sistema = (Path(app.__file__).parent / "templates" / "sistema_admin.html").read_text(encoding="utf-8-sig")
        service = (Path(app.__file__).parent / "services" / "finanzas.py").read_text(encoding="utf-8-sig")
        repository = (Path(app.__file__).parent / "repositories" / "finanzas.py").read_text(encoding="utf-8-sig")

        self.assertIn("from services.finanzas import", source)
        self.assertIn("def periodo_inicio_plan", service)
        self.assertIn("def calcular_importe_cuota_mensual", service)
        self.assertIn("def recalcular_cuotas_planes_pago", service)
        self.assertIn("def obtener_cuenta_corriente_jugador", repository)
        self.assertIn("def obtener_cuotas_impagas_para_plan", repository)
        self.assertIn("def ejecutar_automatizaciones", source)
        self.assertIn("def consumir_limite_publico", source)
        self.assertIn("/tasks/automatizaciones", source)
        self.assertIn("Cuenta corriente", portal)
        self.assertIn("Automatizaciones", sistema)

    def test_pwa_install_and_push_infrastructure_is_present(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        base = (Path(app.__file__).parent / "templates" / "base.html").read_text(encoding="utf-8-sig")
        login = (Path(app.__file__).parent / "templates" / "login.html").read_text(encoding="utf-8-sig")
        js = (Path(app.__file__).parent / "static" / "app.js").read_text(encoding="utf-8-sig")
        sw = (Path(app.__file__).parent / "static" / "service-worker.js").read_text(encoding="utf-8-sig")

        self.assertIn("/manifest.webmanifest", source)
        self.assertIn("pwa_push_subscribe", source)
        self.assertIn("pwa_push_subscriptions", source)
        self.assertIn("rel=\"manifest\"", base)
        self.assertIn("Instalar app", login)
        self.assertIn("Android", login)
        self.assertIn("iPhone", login)
        self.assertIn("serviceWorker.register", js)
        self.assertIn("data-pwa-install-inline", js)
        self.assertIn("pushManager.subscribe", js)
        self.assertIn("sig:pwa:test-ok", js)
        self.assertIn("sig:pwa:saved", js)
        self.assertIn("savePwaSubscription", js)
        self.assertIn("showNotification", sw)

    def test_pwa_notifications_are_split_into_service_and_repository(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        service = (Path(app.__file__).parent / "services" / "notificaciones.py").read_text(encoding="utf-8-sig")
        repository = (Path(app.__file__).parent / "repositories" / "notificaciones.py").read_text(encoding="utf-8-sig")

        self.assertIn("from services.notificaciones import", source)
        self.assertIn("usuario_id=session.get(\"user_id\")", source)
        self.assertIn("def actor_push_actual", service)
        self.assertIn("def normalizar_url_push", service)
        self.assertIn("def guardar_suscripcion_push", repository)
        self.assertIn("def obtener_destinatarios_push_manual", repository)

    def test_asistencia_and_portal_have_initial_services_and_repositories(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        asistencia_service = (Path(app.__file__).parent / "services" / "asistencia.py").read_text(encoding="utf-8-sig")
        asistencia_repo = (Path(app.__file__).parent / "repositories" / "asistencia.py").read_text(encoding="utf-8-sig")
        portal_service = (Path(app.__file__).parent / "services" / "portal.py").read_text(encoding="utf-8-sig")
        portal_repo = (Path(app.__file__).parent / "repositories" / "portal.py").read_text(encoding="utf-8-sig")

        self.assertIn("from services.asistencia import", source)
        self.assertIn("obtener_evento_asistencia", source)
        self.assertIn("def listar_eventos_asistencia", asistencia_repo)
        self.assertIn("def obtener_evento_asistencia", asistencia_repo)
        self.assertIn("obtener_evento_asistencia", asistencia_service)
        self.assertIn("def obtener_eventos_deportivos_portal", portal_service)
        self.assertIn("def listar_eventos_deportivos_portal", portal_repo)

    def test_push_actor_prefers_portal_token_over_admin_session(self):
        class Result:
            def fetchone(self):
                return {"id": 42}

        class Conn:
            def execute(self, *_args, **_kwargs):
                return Result()

        actor = notificaciones_service.actor_push_actual(
            Conn(),
            portal_token="token-demo",
            usuario_id=7,
        )

        self.assertEqual(actor["tipo"], "portal")
        self.assertEqual(actor["jugador_id"], 42)
        self.assertIsNone(actor["usuario_id"])

    def test_manual_app_notifications_admin_screen_is_present(self):
        source = Path(app.__file__).read_text(encoding="utf-8-sig")
        base = (Path(app.__file__).parent / "templates" / "base.html").read_text(encoding="utf-8-sig")
        template = (Path(app.__file__).parent / "templates" / "notificaciones_app.html").read_text(encoding="utf-8-sig")
        portal = (Path(app.__file__).parent / "templates" / "portal_jugador.html").read_text(encoding="utf-8-sig")
        jugador_detalle = (Path(app.__file__).parent / "templates" / "jugador_detalle.html").read_text(encoding="utf-8-sig")

        self.assertIn("def enviar_notificacion_app_manual", source)
        self.assertIn("pwa_push_envios", source)
        self.assertIn("obtener_destinatarios_push_manual", source)
        self.assertIn("notificaciones_portal", source)
        self.assertIn("s.actualizado_en", source)
        self.assertIn("obtener_comunicaciones_portal_dia", source)
        self.assertIn("obtener_avisos_login_publicos", source)
        self.assertIn("LOGIN_AVISOS_KEY", source)
        self.assertIn("urgente", source)
        self.assertIn("Notificaciones app", base)
        self.assertIn("Todos los portales suscriptos", template)
        self.assertIn("mostrar_portal", template)
        self.assertNotIn("mostrar_login", template)
        self.assertIn("Historial de env", template)
        self.assertIn("Comunicaciones del d", portal)
        self.assertIn("data-pwa-enable-push", portal)
        self.assertIn("data-portal-token", portal)
        self.assertIn("Notificaciones app", jugador_detalle)
        self.assertIn("notificaciones_portal", jugador_detalle)


if __name__ == "__main__":
    unittest.main()
