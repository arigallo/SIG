from flask import Flask, render_template, request, redirect, url_for, flash, Response, send_file, session, has_request_context, abort, g
import psycopg
from flask import jsonify
from psycopg.rows import dict_row
from pathlib import Path
import csv
import base64
import html
import io
import json
import os
import posixpath
import re
import mimetypes
import imaplib
import smtplib
import secrets
import unicodedata
import hashlib
import hmac
from datetime import date, datetime, timedelta
from email import policy
from email.header import decode_header, make_header
from email.message import EmailMessage
from email.parser import BytesParser
from email.utils import formataddr, make_msgid, parseaddr, parsedate_to_datetime
from urllib.parse import quote
from urllib.parse import urljoin
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import HTTPError, URLError
from html.parser import HTMLParser
from zoneinfo import ZoneInfo

from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from google.auth import default as google_auth_default
    from googleapiclient.discovery import build as google_build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
except ImportError:
    google_auth_default = None
    google_build = None
    HttpError = None
    MediaIoBaseDownload = None
    MediaIoBaseUpload = None

try:
    from pywebpush import WebPushException, webpush
except ImportError:
    WebPushException = None
    webpush = None

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-esto-por-una-clave-segura")

app.config["SESSION_PERMANENT"] = False
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = timedelta(hours=6)
app.config["SESSION_COOKIE_NAME"] = "sig_session"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = os.environ.get(
    "SESSION_COOKIE_SECURE",
    "true",
).lower() in {"1", "true", "yes", "on"}
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")
MAX_REQUEST_BYTES = max(
    int(os.environ.get("MAX_UPLOAD_BYTES", str(16 * 1024 * 1024))),
    int(os.environ.get("MAX_BATCH_UPLOAD_BYTES", str(64 * 1024 * 1024))),
)
app.config["MAX_CONTENT_LENGTH"] = MAX_REQUEST_BYTES

app.permanent_session_lifetime = timedelta(minutes=30)

APP_TIMEZONE = os.environ.get("APP_TIMEZONE", "America/Argentina/Buenos_Aires")
APP_TZ = ZoneInfo(APP_TIMEZONE)
PWA_VAPID_PUBLIC_KEY = os.environ.get("PWA_VAPID_PUBLIC_KEY", "").strip()
PWA_VAPID_PRIVATE_KEY = os.environ.get("PWA_VAPID_PRIVATE_KEY", "").strip()
PWA_VAPID_CLAIMS_SUB = os.environ.get("PWA_VAPID_CLAIMS_SUB", "mailto:admin@sig.local").strip()


def ahora_sig():
    return datetime.now(APP_TZ)


def fecha_hora_sig(valor):
    if isinstance(valor, datetime):
        if valor.tzinfo is not None and valor.utcoffset() is not None:
            return valor.astimezone(APP_TZ)
        return valor

    texto = str(valor or "").strip()
    if not texto:
        return None
    if len(texto) < 16:
        return None
    try:
        normalizado = texto.replace("Z", "+00:00")
        fecha = datetime.fromisoformat(normalizado)
    except ValueError:
        return None
    if fecha.tzinfo is not None and fecha.utcoffset() is not None:
        return fecha.astimezone(APP_TZ)
    return fecha


def formato_moneda(valor):
    try:
        return "${:,.0f}".format(float(valor)).replace(",", ".")
    except (TypeError, ValueError):
        return "$0"

app.jinja_env.filters["moneda"] = formato_moneda


def formato_fecha_hora(valor):
    if not valor:
        return "-"
    fecha = fecha_hora_sig(valor)
    if fecha:
        return fecha.strftime("%Y-%m-%d %H:%M")
    texto = str(valor).strip()
    if len(texto) >= 16:
        return texto[:16].replace("T", " ")
    return texto


app.jinja_env.filters["fecha_hora"] = formato_fecha_hora

@app.context_processor
def inject_now():
    return {
        "now": ahora_sig,
        "csrf_token": csrf_token,
        "puede": tiene_permiso,
        "mantenimiento": getattr(g, "mantenimiento", None) if has_request_context() else None,
        "notificaciones_count": obtener_contador_notificaciones() if has_request_context() else 0,
        "whatsapp_inbox_count": obtener_contador_whatsapp_inbox() if has_request_context() else 0,
        "whatsapp_api_activa": whatsapp_api_disponible(),
        "static_asset": static_asset,
        "pwa_public_key": PWA_VAPID_PUBLIC_KEY,
        "pwa_icon_url": pwa_icon_url,
        "asistencia_portal_opciones": asistencia_portal_opciones,
        "asistencia_portal_label": asistencia_portal_label,
        "asistencia_portal_badge_class": asistencia_portal_badge_class,
        "es_evento_partido": es_evento_partido,
        "current_month": lambda: ahora_sig().strftime("%Y-%m"),
    }

BASE_DIR = Path(__file__).resolve().parent
WHATSAPP_ENABLED = str(os.environ.get("WHATSAPP_ENABLED", "false")).strip().lower() in {"1", "true", "yes", "on"}
WHATSAPP_ACCESS_TOKEN = os.environ.get("WHATSAPP_ACCESS_TOKEN", "").strip()
WHATSAPP_PHONE_NUMBER_ID = os.environ.get("WHATSAPP_PHONE_NUMBER_ID", "").strip()
WHATSAPP_WABA_ID = os.environ.get("WHATSAPP_WABA_ID", "").strip()
WHATSAPP_VERIFY_TOKEN = os.environ.get("WHATSAPP_VERIFY_TOKEN", "").strip()
WHATSAPP_APP_SECRET = os.environ.get("WHATSAPP_APP_SECRET", "").strip()
WHATSAPP_API_VERSION = os.environ.get("WHATSAPP_API_VERSION", "v25.0").strip() or "v25.0"
WHATSAPP_GRAPH_BASE = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}"
WHATSAPP_TEMPLATE_CUOTA = os.environ.get("WHATSAPP_TEMPLATE_CUOTA", "recordatorio_cuota").strip() or "recordatorio_cuota"
WHATSAPP_TEMPLATE_CUOTA_LANG = os.environ.get("WHATSAPP_TEMPLATE_CUOTA_LANG", "es_AR").strip() or "es_AR"
WHATSAPP_EMAIL_SUPPRESS_USERNAMES = {
    item.strip().lower()
    for item in re.split(r"[;,]", os.environ.get("WHATSAPP_EMAIL_SUPPRESS_USERNAMES", "arielgallo"))
    if item.strip()
}
WHATSAPP_EMAIL_SUPPRESS_SECONDS = int(os.environ.get("WHATSAPP_EMAIL_SUPPRESS_SECONDS", "120"))
WHATSAPP_RESPUESTAS_RAPIDAS = [
    "Gracias, lo revisamos y te avisamos.",
    "Recibido. En breve verificamos el comprobante.",
    "Tu mensaje quedo registrado en el SIG.",
    "Cuando puedas, envianos el comprobante o el detalle del pago.",
    "Perfecto, lo derivamos al area correspondiente.",
]


def static_asset(filename):
    try:
        version = int((BASE_DIR / "static" / filename).stat().st_mtime)
    except OSError:
        version = 1
    return url_for("static", filename=filename, v=version)


def pwa_icon_url(size="192"):
    normalized = "512" if str(size) == "512" else "192"
    return url_for("static", filename=f"img/pwa-icon-{normalized}.png")


def base64url_decode(value):
    texto = (value or "").strip()
    if not texto:
        return b""
    padding = "=" * (-len(texto) % 4)
    return base64.urlsafe_b64decode(texto + padding)


def csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def csrf_valido():
    token_session = session.get("_csrf_token")
    token_enviado = request.form.get("_csrf_token") or request.headers.get("X-CSRF-Token")
    return bool(
        token_session
        and token_enviado
        and secrets.compare_digest(token_session, token_enviado)
    )


def destino_interno(destino, fallback="index"):
    destino = (destino or "").strip()
    if destino.startswith("/") and not destino.startswith("//"):
        return destino
    return url_for(fallback)


def permisos_default_rol(rol):
    return list(ROLE_PRESETS.get(rol, []))


def normalizar_permisos(permisos):
    return sorted({permiso for permiso in permisos if permiso in PERMISOS})


def serializar_permisos(permisos):
    return json.dumps(normalizar_permisos(permisos), ensure_ascii=False)


def deserializar_permisos(valor, rol=None):
    if not valor:
        return permisos_default_rol(rol)
    try:
        permisos = json.loads(valor)
    except (TypeError, ValueError):
        return permisos_default_rol(rol)
    if not isinstance(permisos, list):
        return permisos_default_rol(rol)
    return normalizar_permisos(permisos)


def resumen_auditoria_portal(detalle):
    try:
        datos = json.loads(detalle or "{}")
    except (TypeError, ValueError):
        return detalle or "Actualizacion de datos personales"

    cambios = datos.get("cambios") or {}
    if not cambios:
        campos = datos.get("campos") or []
        if campos:
            return "Campos actualizados: " + ", ".join(str(campo) for campo in campos)
        return "Actualizacion de datos personales"

    partes = []
    for cambio in cambios.values():
        label = cambio.get("label") or "Campo"
        antes = cambio.get("antes") or "-"
        despues = cambio.get("despues") or "-"
        partes.append(f"{label}: {antes} -> {despues}")
    return "; ".join(partes)


def nombre_jugador_auditoria(jugador):
    apellido = (jugador.get("apellido") or "").strip()
    nombre = (jugador.get("nombre") or "").strip()
    jugador_id = jugador.get("jugador_id") or jugador.get("id")
    nombre_completo = ", ".join(parte for parte in (apellido, nombre) if parte)
    if not nombre_completo:
        nombre_completo = "Jugador"
    if jugador_id:
        return f"{nombre_completo} #{jugador_id}"
    return nombre_completo


def username_portal_jugador(jugador):
    return f"portal - {nombre_jugador_auditoria(jugador)}"


def detalle_actor_portal(jugador):
    return {
        "jugador_id": jugador.get("jugador_id") or jugador.get("id"),
        "jugador": nombre_jugador_auditoria(jugador),
        "categoria": jugador.get("categoria") or "",
        "dni": jugador.get("dni") or "",
    }


def enriquecer_actores_auditoria(conn, registros):
    registros = [dict(registro) for registro in registros]
    jugador_ids = sorted({
        int(registro["entidad_id"])
        for registro in registros
        if registro.get("entidad") == "portal_jugador"
        and str(registro.get("entidad_id") or "").isdigit()
    })
    jugadores_por_id = {}
    if jugador_ids:
        jugadores = conn.execute("""
            SELECT id, nombre, apellido
            FROM jugadores
            WHERE id = ANY(%s)
        """, (jugador_ids,)).fetchall()
        jugadores_por_id = {jugador["id"]: jugador for jugador in jugadores}

    for registro in registros:
        actor = registro.get("username")
        if registro.get("entidad") == "portal_jugador":
            entidad_id = str(registro.get("entidad_id") or "")
            jugador = jugadores_por_id.get(int(entidad_id)) if entidad_id.isdigit() else None
            if jugador:
                if actor and actor.startswith("portal - "):
                    registro["actor_display"] = actor
                else:
                    registro["actor_display"] = (
                        f"{actor or 'portal'} - {jugador['apellido']}, {jugador['nombre']} #{jugador['id']}"
                    )
                continue
        registro["actor_display"] = actor
    return registros


def grupos_permisos():
    grupos = {}
    for clave, permiso in PERMISOS.items():
        grupos.setdefault(permiso["grupo"], []).append({
            "clave": clave,
            **permiso,
        })
    return grupos


def cargar_permisos_rol(conn, rol):
    fila = conn.execute("""
        SELECT permisos
        FROM roles
        WHERE nombre = %s
    """, (rol,)).fetchone()
    if not fila:
        return permisos_default_rol(rol)
    return deserializar_permisos(fila["permisos"], rol)

DATABASE_URL = os.environ.get("DATABASE_URL")
DB_NAME = os.environ.get("DB_NAME") or os.environ.get("POSTGRES_DB", "sig")
DB_USER = os.environ.get("DB_USER") or os.environ.get("POSTGRES_USER", "sig_user")
DB_PASS = os.environ.get("DB_PASS") or os.environ.get("DB_PASSWORD", "")
DB_HOST = os.environ.get("DB_HOST")
DB_PORT = os.environ.get("DB_PORT", "5432")
DB_SSLMODE = os.environ.get("DB_SSLMODE")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
FORCE_ADMIN_PASSWORD_UPDATE = os.environ.get(
    "FORCE_ADMIN_PASSWORD_UPDATE", ""
).lower() in {"1", "true", "yes", "on"}
CLOUD_SQL_CONNECTION_NAME = (
    os.environ.get("INSTANCE_CONNECTION_NAME")
    or os.environ.get("CLOUD_SQL_CONNECTION_NAME")
)
DB_SOCKET_DIR = os.environ.get("DB_SOCKET_DIR", "/cloudsql")
DB_CONNECT_TIMEOUT = int(os.environ.get("DB_CONNECT_TIMEOUT", "10"))
DRIVE_COMPROBANTES_FOLDER_ID = (
    os.environ.get("GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID")
    or os.environ.get("DRIVE_COMPROBANTES_FOLDER_ID")
)
DRIVE_SHARED_DRIVE_ID = (
    os.environ.get("GOOGLE_DRIVE_SHARED_DRIVE_ID")
    or os.environ.get("DRIVE_SHARED_DRIVE_ID")
)
DRIVE_COMPROBANTES_SUBFOLDER = (
    os.environ.get("GOOGLE_DRIVE_COMPROBANTES_SUBFOLDER")
    or os.environ.get("DRIVE_COMPROBANTES_SUBFOLDER")
)
DRIVE_FICHAS_MEDICAS_FOLDER_ID = (
    os.environ.get("GOOGLE_DRIVE_FICHAS_MEDICAS_FOLDER_ID")
    or os.environ.get("DRIVE_FICHAS_MEDICAS_FOLDER_ID")
    or DRIVE_COMPROBANTES_FOLDER_ID
)
DRIVE_FICHAS_MEDICAS_SUBFOLDER = (
    os.environ.get("GOOGLE_DRIVE_FICHAS_MEDICAS_SUBFOLDER")
    or os.environ.get("DRIVE_FICHAS_MEDICAS_SUBFOLDER")
    or "Fichas m?dicas"
)
DRIVE_SECRETARIA_FOLDER_ID = (
    os.environ.get("GOOGLE_DRIVE_SECRETARIA_FOLDER_ID")
    or os.environ.get("DRIVE_SECRETARIA_FOLDER_ID")
    or DRIVE_COMPROBANTES_FOLDER_ID
)
DRIVE_SECRETARIA_SUBFOLDER = (
    os.environ.get("GOOGLE_DRIVE_SECRETARIA_SUBFOLDER")
    or os.environ.get("DRIVE_SECRETARIA_SUBFOLDER")
    or "Secretaria"
)
COMPROBANTE_MAX_BYTES = int(os.environ.get("COMPROBANTE_MAX_BYTES", str(10 * 1024 * 1024)))
COMPROBANTE_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
}
FACTURA_EMAIL_EXTENSIONS = {
    **COMPROBANTE_EXTENSIONS,
    ".html": "text/html",
    ".htm": "text/html",
}
FACTURA_EMAIL_MAX_BYTES = int(os.environ.get("FACTURA_EMAIL_MAX_BYTES", str(12 * 1024 * 1024)))
FACTURA_EMAIL_IMAP_HOST = os.environ.get("FACTURA_EMAIL_IMAP_HOST", "").strip()
FACTURA_EMAIL_IMAP_PORT = int(os.environ.get("FACTURA_EMAIL_IMAP_PORT", "993"))
FACTURA_EMAIL_IMAP_USER = os.environ.get("FACTURA_EMAIL_IMAP_USER", "").strip()
FACTURA_EMAIL_IMAP_PASSWORD = os.environ.get("FACTURA_EMAIL_IMAP_PASSWORD", "")
FACTURA_EMAIL_IMAP_FOLDER = os.environ.get("FACTURA_EMAIL_IMAP_FOLDER", "INBOX").strip() or "INBOX"
FACTURA_EMAIL_IMAP_USE_SSL = os.environ.get("FACTURA_EMAIL_IMAP_USE_SSL", "true").lower() in {"1", "true", "yes", "on"}
FACTURA_EMAIL_SEARCH_DAYS = int(os.environ.get("FACTURA_EMAIL_SEARCH_DAYS", "45"))
FACTURA_EMAIL_MAX_MESSAGES = int(os.environ.get("FACTURA_EMAIL_MAX_MESSAGES", "80"))
FACTURA_EMAIL_SECRET_NAME = os.environ.get("FACTURA_EMAIL_SECRET_NAME", "sig-factura-email-imap-password").strip()
FACTURA_EMAIL2_IMAP_HOST = os.environ.get("FACTURA_EMAIL2_IMAP_HOST", "").strip()
FACTURA_EMAIL2_IMAP_PORT = int(os.environ.get("FACTURA_EMAIL2_IMAP_PORT", "993"))
FACTURA_EMAIL2_IMAP_USER = os.environ.get("FACTURA_EMAIL2_IMAP_USER", "").strip()
FACTURA_EMAIL2_IMAP_PASSWORD = os.environ.get("FACTURA_EMAIL2_IMAP_PASSWORD", "")
FACTURA_EMAIL2_IMAP_FOLDER = os.environ.get("FACTURA_EMAIL2_IMAP_FOLDER", "INBOX").strip() or "INBOX"
FACTURA_EMAIL2_IMAP_USE_SSL = os.environ.get("FACTURA_EMAIL2_IMAP_USE_SSL", "true").lower() in {"1", "true", "yes", "on"}
FACTURA_EMAIL2_SEARCH_DAYS = int(os.environ.get("FACTURA_EMAIL2_SEARCH_DAYS", str(FACTURA_EMAIL_SEARCH_DAYS)))
FACTURA_EMAIL2_MAX_MESSAGES = int(os.environ.get("FACTURA_EMAIL2_MAX_MESSAGES", str(FACTURA_EMAIL_MAX_MESSAGES)))
FACTURA_EMAIL2_SECRET_NAME = os.environ.get("FACTURA_EMAIL2_SECRET_NAME", "sig-factura-email-imap-password-2").strip()
FACTURA_EMAIL_DEFAULT_FILTERS = [
    {"proveedor": "Meta", "remitente_patron": "meta", "asunto_patron": "invoice"},
    {"proveedor": "Meta", "remitente_patron": "meta", "asunto_patron": "factura"},
    {"proveedor": "Meta", "remitente_patron": "meta", "asunto_patron": "recibo"},
    {"proveedor": "Meta", "remitente_patron": "facebook", "asunto_patron": "invoice"},
    {"proveedor": "Meta", "remitente_patron": "facebook", "asunto_patron": "factura"},
    {"proveedor": "Meta", "remitente_patron": "facebook", "asunto_patron": "recibo"},
    {"proveedor": "Meta", "remitente_patron": "instagram", "asunto_patron": "invoice"},
    {"proveedor": "Meta", "remitente_patron": "instagram", "asunto_patron": "factura"},
    {"proveedor": "Meta", "remitente_patron": "instagram", "asunto_patron": "recibo"},
    {"proveedor": "Canva", "remitente_patron": "canva", "asunto_patron": "invoice"},
    {"proveedor": "Canva", "remitente_patron": "canva", "asunto_patron": "receipt"},
    {"proveedor": "Canva", "remitente_patron": "canva", "asunto_patron": "factura"},
    {"proveedor": "General", "remitente_patron": "", "asunto_patron": "factura"},
    {"proveedor": "General", "remitente_patron": "", "asunto_patron": "comprobante"},
    {"proveedor": "General", "remitente_patron": "", "asunto_patron": "recibo"},
]
FICHA_MEDICA_MAX_BYTES = int(os.environ.get("FICHA_MEDICA_MAX_BYTES", str(16 * 1024 * 1024)))
FICHA_MEDICA_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
}
SECRETARIA_MAX_BYTES = int(os.environ.get("SECRETARIA_MAX_BYTES", str(25 * 1024 * 1024)))
SECRETARIA_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".doc": "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xls": "application/vnd.ms-excel",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".ppt": "application/vnd.ms-powerpoint",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".csv": "text/csv",
    ".txt": "text/plain",
}
SECRETARIA_CATEGORIAS = [
    "Actas",
    "Asambleas",
    "Notas",
    "Legales",
    "Tesoreria",
    "Convenios",
    "Proveedores",
    "Institucional",
    "Otro",
]
FICHA_MEDICA_OCR_LANGUAGE = os.environ.get("FICHA_MEDICA_OCR_LANGUAGE", "es")
MAX_LOGIN_ATTEMPTS = int(os.environ.get("MAX_LOGIN_ATTEMPTS", "5"))
LOGIN_ATTEMPT_WINDOW_MINUTES = int(os.environ.get("LOGIN_ATTEMPT_WINDOW_MINUTES", "15"))
PASSWORD_RESET_TOKEN_MINUTES = int(os.environ.get("PASSWORD_RESET_TOKEN_MINUTES", "45"))
SMTP_HOST = os.environ.get("SMTP_HOST", "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "").strip()
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USER).strip()
SMTP_USE_TLS = os.environ.get("SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"}
SMTP_FROM_NAME = os.environ.get("SMTP_FROM_NAME", "Tesoreria - RMR").strip() or "Tesoreria - RMR"
SUGERENCIAS_DIRECTIVA_EMAILS_KEY = "sugerencias_directiva_emails"
SUGERENCIAS_CONFIG_KEYS = [
    SUGERENCIAS_DIRECTIVA_EMAILS_KEY,
]
WHATSAPP_INBOX_NOTIFY_EMAILS = [
    email.strip()
    for email in re.split(r"[;,]", os.environ.get("WHATSAPP_INBOX_NOTIFY_EMAILS", ""))
    if email.strip()
]
TESORERO_FIRMA_NOMBRE = os.environ.get("TESORERO_FIRMA_NOMBRE", "Ariel Gallo").strip() or "Ariel Gallo"
TESORERO_FIRMA_CARGO = os.environ.get("TESORERO_FIRMA_CARGO", "Tesorero").strip() or "Tesorero"
ASPIRANTE_ENTRENAMIENTOS_OBJETIVO = int(os.environ.get("ASPIRANTE_ENTRENAMIENTOS_OBJETIVO", "8"))
ASPIRANTE_ESTADOS = {"Aspirante", "Ingresado", "Baja"}
APP_VERSION = os.environ.get("APP_VERSION", "local")
CLOUD_SQL_BACKUP_WINDOW = os.environ.get("CLOUD_SQL_BACKUP_WINDOW", "12:00 a.m. - 4:00 a.m.")
CLOUD_SQL_BACKUP_RETENTION_DAYS = os.environ.get("CLOUD_SQL_BACKUP_RETENTION_DAYS", "7")
CLOUD_SQL_PITR_DAYS = os.environ.get("CLOUD_SQL_PITR_DAYS", "7")
CLOUD_SQL_PROJECT = os.environ.get("CLOUD_SQL_PROJECT") or os.environ.get("GOOGLE_CLOUD_PROJECT", "")
CLOUD_SQL_INSTANCE = os.environ.get("CLOUD_SQL_INSTANCE", "")
COMPROBANTE_ESTADOS = {"sin_comprobante", "pendiente", "aceptado", "rechazado"}
CALENDARIO_DEPORTIVO_TIPOS = {"Entrenamiento", "Partido", "Evento", "Otro"}
CALENDARIO_ASISTENCIA_TIPOS = {"Entrenamiento", "Partido"}
CALENDARIO_TZ = "America/Argentina/Buenos_Aires"
PORTAL_ASISTENCIA_ESTADOS = {"confirmado", "dudoso", "no_asiste"}
PORTAL_ASISTENCIA_LABELS = {
    "default": {
        "confirmado": "Voy",
        "dudoso": "Dudoso",
        "no_asiste": "No voy",
    },
    "partido": {
        "confirmado": "Voy y Juego",
        "dudoso": "Voy y no juego",
        "no_asiste": "No voy",
    },
}
PORTAL_ASISTENCIA_BADGES = {
    "confirmado": "badge-success",
    "dudoso": "badge-warning",
    "no_asiste": "badge-danger",
}
TIPOS_MIEMBRO = {"Jugador", "Socio activo", "Colaborador"}
ESTADOS_JUGADOR = ["Activo", "Inactivo", "Suspendido", "Baja"]

BIENESTAR_HORAS_OPCIONES = ["<5 h", "5-6 h", "6-7 h", "7-8 h", ">8 h"]
BIENESTAR_HORAS_SCORE = {"<5 h": 1, "5-6 h": 2, "6-7 h": 3, "7-8 h": 4, ">8 h": 5}
BIENESTAR_DOLOR_ZONAS = ["No", "Cuello", "Hombro", "Brazo", "Zona lumbar", "Cadera", "Muslo", "Rodilla", "Pantorrilla", "Tobillo", "Pie", "Otro"]


def es_evento_partido(evento):
    tipo = (evento.get("tipo") if evento else "") or ""
    tipo = unicodedata.normalize("NFKD", str(tipo).strip().lower())
    tipo = "".join(ch for ch in tipo if not unicodedata.combining(ch))
    return tipo in {"partido", "partidos"}


def asistencia_portal_labels(evento):
    return PORTAL_ASISTENCIA_LABELS["partido" if es_evento_partido(evento) else "default"]


def asistencia_portal_opciones(evento):
    labels = asistencia_portal_labels(evento)
    return [
        {"valor": estado, "label": labels[estado]}
        for estado in ("confirmado", "dudoso", "no_asiste")
    ]


def asistencia_portal_label(evento, estado):
    return asistencia_portal_labels(evento).get(estado, PORTAL_ASISTENCIA_LABELS["default"]["confirmado"])


def asistencia_portal_badge_class(estado):
    return PORTAL_ASISTENCIA_BADGES.get(estado, "badge-muted")


BITACORA_TIPOS = {
    "general": "General",
    "finanzas": "Finanzas",
    "salud": "Salud",
    "deportivo": "Deportivo",
}
MAINTENANCE_DEFAULT_MESSAGE = os.environ.get(
    "MAINTENANCE_DEFAULT_MESSAGE",
    "El sistema esta en mantenimiento. Volve a intentar en unos minutos.",
)
MESES_ES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]
PERMISOS = {
    "jugadores_ver": {
        "grupo": "Jugadores",
        "nombre": "Ver jugadores",
        "descripcion": "Listado y detalle general de jugadores.",
    },
    "jugadores_gestionar": {
        "grupo": "Jugadores",
        "nombre": "Crear y editar jugadores",
        "descripcion": "Alta, edición, importación y acciones masivas.",
    },
    "jugadores_eliminar": {
        "grupo": "Jugadores",
        "nombre": "Eliminar jugadores",
        "descripcion": "Baja definitiva de jugadores.",
    },
    "aspirantes_ver": {
        "grupo": "Jugadores",
        "nombre": "Ver ahijadxs",
        "descripcion": "Listado, detalle y seguimiento de ahijadxs.",
    },
    "aspirantes_gestionar": {
        "grupo": "Jugadores",
        "nombre": "Gestionar ahijadxs",
        "descripcion": "Alta, edicion, asistencia y conversion a jugador activo.",
    },
    "cuotas_ver": {
        "grupo": "Cuotas y caja",
        "nombre": "Ver cuotas",
        "descripcion": "Consulta de cuotas, recibos y comprobantes.",
    },
    "cuotas_gestionar": {
        "grupo": "Cuotas y caja",
        "nombre": "Gestionar cuotas",
        "descripcion": "Crear, generar, cobrar, eliminar y adjuntar comprobantes.",
    },
    "planes_pago_ver": {
        "grupo": "Cuotas y caja",
        "nombre": "Ver planes de pago",
        "descripcion": "Consulta de acuerdos y planes de regularizacion.",
    },
    "planes_pago_gestionar": {
        "grupo": "Cuotas y caja",
        "nombre": "Gestionar planes de pago",
        "descripcion": "Alta, seguimiento y cierre de planes de pago.",
    },
    "caja_ver": {
        "grupo": "Cuotas y caja",
        "nombre": "Ver caja",
        "descripcion": "Consulta y exportación de movimientos de caja.",
    },
    "caja_gestionar": {
        "grupo": "Cuotas y caja",
        "nombre": "Gestionar caja",
        "descripcion": "Crear, editar, anular movimientos y cerrar meses.",
    },
    "facturas_recibidas_ver": {
        "grupo": "Cuotas y caja",
        "nombre": "Ver facturas recibidas",
        "descripcion": "Consultar facturas detectadas desde emails de proveedores.",
    },
    "facturas_recibidas_gestionar": {
        "grupo": "Cuotas y caja",
        "nombre": "Gestionar facturas recibidas",
        "descripcion": "Sincronizar emails, administrar filtros y convertir facturas en egresos.",
    },
    "presupuesto_ver": {
        "grupo": "Cuotas y caja",
        "nombre": "Ver presupuesto",
        "descripcion": "Consultar presupuestos mensuales y proyecciones financieras.",
    },
    "presupuesto_gestionar": {
        "grupo": "Cuotas y caja",
        "nombre": "Gestionar presupuesto",
        "descripcion": "Crear y administrar gastos e ingresos presupuestados.",
    },
    "reportes_ver": {
        "grupo": "Reportes",
        "nombre": "Ver reportes",
        "descripcion": "Reportes y exportaciones generales.",
    },
    "comunicaciones_ver": {
        "grupo": "Reportes",
        "nombre": "Ver comunicaciones",
        "descripcion": "Morosos, plantillas y exportaciones de comunicación.",
    },
    "salud_ver": {
        "grupo": "Salud",
        "nombre": "Ver salud",
        "descripcion": "Fichas médicas, lesiones y alertas médicas.",
    },
    "salud_gestionar": {
        "grupo": "Salud",
        "nombre": "Gestionar salud",
        "descripcion": "Editar fichas m?dicas, crear, editar y eliminar lesiones.",
    },
    "documentos_ver": {
        "grupo": "Salud",
        "nombre": "Ver documentos",
        "descripcion": "Consulta de documentos, vencimientos y faltantes.",
    },
    "documentos_gestionar": {
        "grupo": "Salud",
        "nombre": "Gestionar documentos",
        "descripcion": "Carga, edicion y baja de documentos manuales.",
    },
    "secretaria_ver": {
        "grupo": "Administracion",
        "nombre": "Ver secretaria",
        "descripcion": "Consultar documentos administrativos internos y su archivo historico.",
    },
    "secretaria_gestionar": {
        "grupo": "Administracion",
        "nombre": "Gestionar secretaria",
        "descripcion": "Subir, categorizar y eliminar documentos administrativos internos.",
    },
    "sugerencias_ver": {
        "grupo": "Administracion",
        "nombre": "Ver sugerencias",
        "descripcion": "Consultar sugerencias y recomendaciones enviadas desde el formulario publico.",
    },
    "sugerencias_gestionar": {
        "grupo": "Administracion",
        "nombre": "Gestionar sugerencias y recomendaciones",
        "descripcion": "Actualizar estados, notas internas y reenviar notificaciones.",
    },
    "sugerencias_configurar": {
        "grupo": "Administracion",
        "nombre": "Configurar sugerencias y recomendaciones",
        "descripcion": "Definir los destinatarios de las notificaciones.",
    },
    "calendario_ver": {
        "grupo": "Deportivo",
        "nombre": "Ver calendario",
        "descripcion": "Consulta del calendario.",
    },
    "calendario_gestionar": {
        "grupo": "Deportivo",
        "nombre": "Gestionar calendario",
        "descripcion": "Crear eventos de calendario.",
    },
    "asistencia_ver": {
        "grupo": "Deportivo",
        "nombre": "Ver asistencia",
        "descripcion": "Consulta de eventos e historial de asistencia.",
    },
    "asistencia_gestionar": {
        "grupo": "Deportivo",
        "nombre": "Tomar asistencia",
        "descripcion": "Crear eventos y registrar asistencia.",
    },
    "tests_ver": {
        "grupo": "Deportivo",
        "nombre": "Ver tests deportivos",
        "descripcion": "Consulta de mediciones, rankings y graficos deportivos.",
    },
    "tests_gestionar": {
        "grupo": "Deportivo",
        "nombre": "Gestionar tests deportivos",
        "descripcion": "Crear tests, cargar puntajes e importar mediciones.",
    },
    "alertas_finanzas": {
        "grupo": "Alertas",
        "nombre": "Alertas financieras",
        "descripcion": "Cuotas vencidas, caja y movimientos inusuales.",
    },
    "alertas_salud": {
        "grupo": "Alertas",
        "nombre": "Alertas médicas",
        "descripcion": "Fichas vencidas y lesiones activas.",
    },
    "alertas_portal": {
        "grupo": "Alertas",
        "nombre": "Cambios desde portal",
        "descripcion": "Notificaciones por datos personales actualizados desde el portal del jugador.",
    },
    "auditoria_ver": {
        "grupo": "Administración",
        "nombre": "Ver auditoría",
        "descripcion": "Consulta de actividad del sistema.",
    },
    "roles_gestionar": {
        "grupo": "Administracion",
        "nombre": "Gestionar roles",
        "descripcion": "Crear roles y editar permisos, incluidos los roles base.",
    },
    "backup_ver": {
        "grupo": "Administración",
        "nombre": "Ver backup",
        "descripcion": "Acceso a información de backup.",
    },
    "backup_gestionar": {
        "grupo": "Administracion",
        "nombre": "Ejecutar backup",
        "descripcion": "Solicitar backups manuales de Cloud SQL desde el panel de sistema.",
    },
    "portal_jugador_gestionar": {
        "grupo": "Administracion",
        "nombre": "Gestionar portal externo",
        "descripcion": "Generar y desactivar enlaces externos para jugadores y familias.",
    },
    "seguridad_ver": {
        "grupo": "Administracion",
        "nombre": "Ver seguridad",
        "descripcion": "Consulta de logins, bloqueos y actividad sensible.",
    },
}
TODOS_LOS_PERMISOS = list(PERMISOS.keys())
ROLE_PRESETS = {
    "admin": TODOS_LOS_PERMISOS,
    "tesorero": [
        "jugadores_ver",
        "cuotas_ver",
        "cuotas_gestionar",
        "planes_pago_ver",
        "planes_pago_gestionar",
        "caja_ver",
        "caja_gestionar",
        "facturas_recibidas_ver",
        "facturas_recibidas_gestionar",
        "presupuesto_ver",
        "presupuesto_gestionar",
        "reportes_ver",
        "comunicaciones_ver",
        "calendario_ver",
        "calendario_gestionar",
        "asistencia_ver",
        "asistencia_gestionar",
        "alertas_finanzas",
        "alertas_portal",
    ],
    "medico": [
        "jugadores_ver",
        "salud_ver",
        "salud_gestionar",
        "documentos_ver",
        "documentos_gestionar",
        "alertas_salud",
    ],
    "entrenador": [
        "jugadores_ver",
        "jugadores_gestionar",
        "aspirantes_ver",
        "aspirantes_gestionar",
        "calendario_ver",
        "calendario_gestionar",
        "asistencia_ver",
        "asistencia_gestionar",
        "tests_ver",
        "tests_gestionar",
        "alertas_portal",
    ],
}


def get_db_connect_args():
    kwargs = {
        "row_factory": dict_row,
        "connect_timeout": DB_CONNECT_TIMEOUT,
    }

    if DB_SSLMODE:
        kwargs["sslmode"] = DB_SSLMODE

    if DATABASE_URL:
        return (DATABASE_URL,), kwargs

    host = DB_HOST
    if not host and CLOUD_SQL_CONNECTION_NAME:
        host = posixpath.join(DB_SOCKET_DIR, CLOUD_SQL_CONNECTION_NAME)

    kwargs.update({
        "dbname": DB_NAME,
        "user": DB_USER,
        "password": DB_PASS,
        "host": host or "localhost",
        "port": DB_PORT,
    })

    return (), kwargs


class DBConnection:
    def __init__(self):
        args, kwargs = get_db_connect_args()
        self.conn = psycopg.connect(*args, **kwargs)
        self.conn.execute("SELECT set_config('TimeZone', %s, false)", (APP_TIMEZONE,))

    def execute(self, query, params=None):
        cur = self.conn.cursor()
        cur.execute(query, params or ())
        return cur

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        if self.conn and not self.conn.closed:
            self.conn.close()


def get_connection():
    return DBConnection()


def get_columns(conn, table_name):
    columnas = conn.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
    """, (table_name,)).fetchall()
    return [col["column_name"] for col in columnas]


def parse_bool_setting(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def obtener_config_mantenimiento(conn=None):
    own_conn = conn is None
    conn = conn or get_connection()
    rows = conn.execute("""
        SELECT clave, valor, actualizado_en, actualizado_por
        FROM app_settings
        WHERE clave IN ('maintenance_mode', 'maintenance_message')
    """).fetchall()
    if own_conn:
        conn.close()

    settings = {row["clave"]: row for row in rows}
    mode_row = settings.get("maintenance_mode") or {}
    message_row = settings.get("maintenance_message") or {}
    return {
        "activo": parse_bool_setting(mode_row.get("valor")),
        "mensaje": message_row.get("valor") or MAINTENANCE_DEFAULT_MESSAGE,
        "actualizado_en": mode_row.get("actualizado_en") or message_row.get("actualizado_en"),
        "actualizado_por": mode_row.get("actualizado_por") or message_row.get("actualizado_por"),
    }


def guardar_app_setting(conn, clave, valor, usuario=None):
    conn.execute("""
        INSERT INTO app_settings (clave, valor, actualizado_en, actualizado_por)
        VALUES (%s, %s, CURRENT_TIMESTAMP, %s)
        ON CONFLICT(clave) DO UPDATE SET
            valor = EXCLUDED.valor,
            actualizado_en = EXCLUDED.actualizado_en,
            actualizado_por = EXCLUDED.actualizado_por
    """, (clave, valor, usuario))


def obtener_app_settings(conn, claves):
    if not claves:
        return {}
    placeholders = ", ".join(["%s"] * len(claves))
    rows = conn.execute(f"""
        SELECT clave, valor, actualizado_en, actualizado_por
        FROM app_settings
        WHERE clave IN ({placeholders})
    """, list(claves)).fetchall()
    return {row["clave"]: row for row in rows}


def obtener_cuenta_corriente_jugador(conn, jugador_id, limite=30):
    return conn.execute("""
        SELECT *
        FROM (
            SELECT
                c.fecha_vencimiento AS fecha,
                'cuota' AS origen,
                c.id AS origen_id,
                'Cuota ' || c.periodo AS concepto,
                c.importe AS importe,
                CASE
                    WHEN COALESCE(c.anulada, 0) = 1 THEN 'anulado'
                    WHEN c.pagado = 1 THEN 'pagado'
                    ELSE 'pendiente'
                END AS estado,
                c.fecha_pago
            FROM cuotas c
            WHERE c.jugador_id = %s

            UNION ALL

            SELECT
                COALESCE(g.fecha_vencimiento, g.fecha_evento, g.creado_en::text) AS fecha,
                'gasto_compartido' AS origen,
                i.id AS origen_id,
                'Gasto compartido: ' || g.titulo AS concepto,
                i.importe AS importe,
                i.estado,
                i.fecha_pago
            FROM gasto_compartido_items i
            JOIN gastos_compartidos g ON g.id = i.gasto_id
            WHERE i.jugador_id = %s
        ) cuenta
        ORDER BY COALESCE(fecha_pago, fecha) DESC NULLS LAST, origen_id DESC
        LIMIT %s
    """, (jugador_id, jugador_id, limite)).fetchall()


AUTOMATION_SETTING_KEYS = (
    "automation_reminders_enabled",
    "automation_reminders_days_before",
    "automation_invoices_enabled",
    "automation_last_run",
    "automation_last_status",
    "automation_last_detail",
)


def obtener_config_automatizaciones(conn=None):
    own_conn = conn is None
    conn = conn or get_connection()
    rows = obtener_app_settings(conn, AUTOMATION_SETTING_KEYS)
    if own_conn:
        conn.close()
    return {
        "recordatorios_activos": parse_bool_setting((rows.get("automation_reminders_enabled") or {}).get("valor")),
        "dias_antes": int_setting((rows.get("automation_reminders_days_before") or {}).get("valor"), 3, 0, 30),
        "facturas_activas": parse_bool_setting((rows.get("automation_invoices_enabled") or {}).get("valor")),
        "ultima_ejecucion": (rows.get("automation_last_run") or {}).get("valor"),
        "ultimo_estado": (rows.get("automation_last_status") or {}).get("valor"),
        "ultimo_detalle": (rows.get("automation_last_detail") or {}).get("valor"),
    }


def consumir_limite_publico(endpoint, max_intentos=5, minutos=60):
    ip = audit_request_ip() or "desconocida"
    conn = get_connection()
    conn.execute("""
        DELETE FROM public_rate_limits
        WHERE creado_en < CURRENT_TIMESTAMP - INTERVAL '2 days'
    """)
    cantidad = conn.execute("""
        SELECT COUNT(*) AS total
        FROM public_rate_limits
        WHERE endpoint = %s
          AND ip = %s
          AND creado_en >= CURRENT_TIMESTAMP - (%s * INTERVAL '1 minute')
    """, (endpoint, ip, minutos)).fetchone()["total"]
    if cantidad >= max_intentos:
        conn.commit()
        conn.close()
        return False
    conn.execute("""
        INSERT INTO public_rate_limits (endpoint, ip)
        VALUES (%s, %s)
    """, (endpoint, ip))
    conn.commit()
    conn.close()
    return True


def ejecutar_automatizaciones(usuario="sistema"):
    conn = get_connection()
    config = obtener_config_automatizaciones(conn)
    resultado = {"recordatorios_enviados": 0, "recordatorios_omitidos": 0, "facturas": None, "errores": []}
    hoy = ahora_sig().strftime("%Y-%m-%d")

    if config["recordatorios_activos"]:
        cuotas = conn.execute("""
            SELECT
                c.id AS cuota_id,
                c.jugador_id,
                c.periodo,
                c.importe,
                c.fecha_vencimiento,
                j.nombre,
                j.apellido,
                j.email,
                j.email_tutor,
                j.portal_token,
                j.portal_activo
            FROM cuotas c
            JOIN jugadores j ON j.id = c.jugador_id
            WHERE c.pagado = 0
              AND COALESCE(c.anulada, 0) = 0
              AND COALESCE(c.importe, 0) > 0
              AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
              AND c.fecha_vencimiento::date <= CURRENT_DATE + (%s * INTERVAL '1 day')
            ORDER BY c.fecha_vencimiento, c.id
        """, (config["dias_antes"],)).fetchall()

        for cuota in cuotas:
            clave = f"{cuota['cuota_id']}:{hoy}"
            reserva = conn.execute("""
                INSERT INTO automatizacion_ejecuciones (tipo, clave, estado, detalle)
                VALUES ('recordatorio_cuota', %s, 'procesando', NULL)
                ON CONFLICT(tipo, clave) DO NOTHING
                RETURNING id
            """, (clave,)).fetchone()
            if not reserva:
                resultado["recordatorios_omitidos"] += 1
                continue

            cuerpo = construir_texto_recordatorio_cuota(cuota)
            enviado, destinatario, motivo = enviar_email_jugador(
                cuota,
                f"Recordatorio de cuota {cuota['periodo']}",
                cuerpo,
            )
            conn.execute("""
                UPDATE automatizacion_ejecuciones
                SET estado = %s, detalle = %s
                WHERE id = %s
            """, ("enviado" if enviado else "error", destinatario or motivo, reserva["id"]))
            if enviado:
                resultado["recordatorios_enviados"] += 1
            else:
                resultado["errores"].append(f"Cuota {cuota['cuota_id']}: {motivo}")
            enviar_push_por_actor("portal", {
                "title": "Cuota pendiente",
                "body": f"Tenes pendiente la cuota {cuota['periodo']} por {formato_moneda(cuota['importe'])}.",
                "url": url_for("portal_jugador", token=cuota["portal_token"]) if cuota.get("portal_token") else "/portal",
                "icon": pwa_icon_url("192"),
            }, jugador_id=cuota["jugador_id"])

        gastos = conn.execute("""
            SELECT
                i.id,
                i.jugador_id,
                i.importe,
                g.titulo,
                g.concepto,
                g.fecha_vencimiento,
                g.estado AS gasto_estado,
                j.nombre,
                j.apellido,
                j.email,
                j.email_tutor,
                j.portal_token
            FROM gasto_compartido_items i
            JOIN gastos_compartidos g ON g.id = i.gasto_id
            JOIN jugadores j ON j.id = i.jugador_id
            WHERE i.estado = 'pendiente'
              AND COALESCE(i.importe, 0) > 0
              AND g.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
              AND g.fecha_vencimiento::date <= CURRENT_DATE + (%s * INTERVAL '1 day')
            ORDER BY g.fecha_vencimiento, i.id
        """, (config["dias_antes"],)).fetchall()

        for item in gastos:
            clave = f"{item['id']}:{hoy}"
            reserva = conn.execute("""
                INSERT INTO automatizacion_ejecuciones (tipo, clave, estado, detalle)
                VALUES ('recordatorio_gasto_compartido', %s, 'procesando', NULL)
                ON CONFLICT(tipo, clave) DO NOTHING
                RETURNING id
            """, (clave,)).fetchone()
            if not reserva:
                resultado["recordatorios_omitidos"] += 1
                continue

            enviado, destinatario, motivo = enviar_email_jugador(
                item,
                f"Gasto compartido pendiente - {item.get('titulo') or 'Ruda Macho Rugby Club'}",
                construir_texto_gasto_compartido(item),
            )
            conn.execute("""
                UPDATE automatizacion_ejecuciones
                SET estado = %s, detalle = %s
                WHERE id = %s
            """, ("enviado" if enviado else "error", destinatario or motivo, reserva["id"]))
            if enviado:
                resultado["recordatorios_enviados"] += 1
            else:
                resultado["errores"].append(f"Gasto compartido {item['id']}: {motivo}")
            enviar_push_por_actor("portal", {
                "title": "Gasto compartido pendiente",
                "body": f"{item.get('titulo') or 'Gasto compartido'}: {formato_moneda(item.get('importe') or 0)} pendiente.",
                "url": url_for("portal_jugador", token=item["portal_token"]) if item.get("portal_token") else "/portal",
                "icon": pwa_icon_url("192"),
            }, jugador_id=item["jugador_id"])

    if config["facturas_activas"]:
        try:
            resultado["facturas"] = sincronizar_facturas_email(conn, usuario=usuario)
            resultado["errores"].extend(
                f"Facturas: {error}"
                for error in (resultado["facturas"].get("errores") or [])
            )
        except Exception as error:
            app.logger.exception("Fallo la sincronizacion automatica de facturas.")
            resultado["errores"].append(f"Facturas: {error}")

    estado = "error" if resultado["errores"] else "ok"
    detalle = json.dumps(resultado, ensure_ascii=False, default=str)
    guardar_app_setting(conn, "automation_last_run", ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), usuario)
    guardar_app_setting(conn, "automation_last_status", estado, usuario)
    guardar_app_setting(conn, "automation_last_detail", detalle, usuario)
    conn.commit()
    conn.close()
    return resultado


def presence_key(username):
    username = normalizar_username(username)
    return f"presence:{username}" if username else ""


def registrar_presencia_usuario(username):
    clave = presence_key(username)
    if not clave:
        return False
    conn = get_connection()
    guardar_app_setting(conn, clave, "active", usuario=username)
    conn.commit()
    conn.close()
    return True


def hash_portal_token(token):
    token = (token or "").strip()
    return hashlib.sha256(token.encode("utf-8")).hexdigest() if token else None


def actor_push_actual(conn, portal_token=None):
    if session.get("user_id"):
        return {
            "tipo": "usuario",
            "usuario_id": session.get("user_id"),
            "jugador_id": None,
            "portal_token_hash": None,
        }

    portal_token = (portal_token or "").strip()
    if portal_token:
        jugador = conn.execute("""
            SELECT id
            FROM jugadores
            WHERE portal_token = %s
              AND COALESCE(portal_activo, 0) = 1
        """, (portal_token,)).fetchone()
        if jugador:
            return {
                "tipo": "portal",
                "usuario_id": None,
                "jugador_id": jugador["id"],
                "portal_token_hash": hash_portal_token(portal_token),
            }

    return None


def guardar_suscripcion_push(conn, subscription, actor, user_agent=None):
    endpoint = (subscription or {}).get("endpoint")
    if not endpoint:
        raise ValueError("Suscripcion push sin endpoint.")
    conn.execute("""
        INSERT INTO pwa_push_subscriptions (
            endpoint, actor_tipo, usuario_id, jugador_id, portal_token_hash,
            subscription_json, user_agent, enabled, actualizado_en
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, 1, CURRENT_TIMESTAMP)
        ON CONFLICT(endpoint) DO UPDATE SET
            actor_tipo = EXCLUDED.actor_tipo,
            usuario_id = EXCLUDED.usuario_id,
            jugador_id = EXCLUDED.jugador_id,
            portal_token_hash = EXCLUDED.portal_token_hash,
            subscription_json = EXCLUDED.subscription_json,
            user_agent = EXCLUDED.user_agent,
            enabled = 1,
            actualizado_en = CURRENT_TIMESTAMP
    """, (
        endpoint,
        actor["tipo"],
        actor.get("usuario_id"),
        actor.get("jugador_id"),
        actor.get("portal_token_hash"),
        json.dumps(subscription),
        truncate_audit_value(user_agent or "", 500),
    ))


def desactivar_suscripcion_push(conn, endpoint):
    if not endpoint:
        return
    conn.execute("""
        UPDATE pwa_push_subscriptions
        SET enabled = 0, actualizado_en = CURRENT_TIMESTAMP
        WHERE endpoint = %s
    """, (endpoint,))


def enviar_push_subscription(subscription, payload):
    if not webpush:
        return False, "Falta instalar pywebpush."
    if not (PWA_VAPID_PUBLIC_KEY and PWA_VAPID_PRIVATE_KEY):
        return False, "Faltan PWA_VAPID_PUBLIC_KEY y PWA_VAPID_PRIVATE_KEY."
    try:
        webpush(
            subscription_info=subscription,
            data=json.dumps(payload, ensure_ascii=False),
            vapid_private_key=PWA_VAPID_PRIVATE_KEY,
            vapid_claims={"sub": PWA_VAPID_CLAIMS_SUB},
        )
        return True, None
    except WebPushException as error:
        return False, str(error)


def enviar_push_por_actor(actor_tipo, payload, usuario_id=None, jugador_id=None):
    conn = get_connection()
    filtros = ["enabled = 1", "actor_tipo = %s"]
    params = [actor_tipo]
    if usuario_id is not None:
        filtros.append("usuario_id = %s")
        params.append(usuario_id)
    if jugador_id is not None:
        filtros.append("jugador_id = %s")
        params.append(jugador_id)
    where = " AND ".join(filtros)
    rows = conn.execute(f"""
        SELECT endpoint, subscription_json
        FROM pwa_push_subscriptions
        WHERE {where}
    """, params).fetchall()

    enviados = 0
    errores = []
    for row in rows:
        try:
            subscription = json.loads(row["subscription_json"] or "{}")
        except (TypeError, ValueError):
            subscription = {}
        ok, error = enviar_push_subscription(subscription, payload)
        if ok:
            enviados += 1
        else:
            errores.append(error)
            if error and ("410" in error or "404" in error):
                desactivar_suscripcion_push(conn, row["endpoint"])
    conn.commit()
    conn.close()
    return {"enviados": enviados, "errores": errores}


def obtener_destinatarios_push_manual(conn, destino, categoria=None, jugador_id=None):
    destino = (destino or "").strip()
    params = []
    filtros = ["s.enabled = 1"]
    joins = []

    if destino == "admins":
        joins.append("JOIN usuarios u ON u.id = s.usuario_id")
        filtros.append("s.actor_tipo = 'usuario'")
        filtros.append("u.rol = 'admin'")
    elif destino == "usuarios":
        filtros.append("s.actor_tipo = 'usuario'")
    elif destino == "portal_todos":
        filtros.append("s.actor_tipo = 'portal'")
    elif destino == "categoria":
        joins.append("JOIN jugadores j ON j.id = s.jugador_id")
        filtros.append("s.actor_tipo = 'portal'")
        if categoria == "Sin categoria":
            filtros.append("COALESCE(NULLIF(j.categoria, ''), 'Sin categoria') = 'Sin categoria'")
        else:
            filtros.append("COALESCE(j.categoria, '') = %s")
            params.append(categoria or "")
    elif destino == "jugador":
        filtros.append("s.actor_tipo = 'portal'")
        filtros.append("s.jugador_id = %s")
        params.append(jugador_id)
    else:
        return []

    join_sql = "\n".join(joins)
    where_sql = " AND ".join(filtros)
    return conn.execute(f"""
        SELECT DISTINCT
            s.endpoint,
            s.subscription_json,
            s.actor_tipo,
            s.usuario_id,
            s.jugador_id,
            s.actualizado_en
        FROM pwa_push_subscriptions s
        {join_sql}
        WHERE {where_sql}
        ORDER BY s.actualizado_en DESC
    """, params).fetchall()


def resumen_suscripciones_push(conn):
    return conn.execute("""
        SELECT
            COUNT(*) FILTER (WHERE enabled = 1) AS activas,
            COUNT(*) FILTER (WHERE enabled = 1 AND actor_tipo = 'portal') AS portal,
            COUNT(*) FILTER (WHERE enabled = 1 AND actor_tipo = 'usuario') AS usuarios,
            COUNT(*) FILTER (WHERE enabled = 1 AND actor_tipo = 'usuario' AND usuario_id IN (
                SELECT id FROM usuarios WHERE rol = 'admin'
            )) AS admins
        FROM pwa_push_subscriptions
    """).fetchone()


def normalizar_url_push(valor, fallback="/"):
    valor = (valor or "").strip()
    if not valor:
        return fallback
    if valor.startswith("/") and not valor.startswith("//"):
        return valor
    if valor.startswith("https://") or valor.startswith("http://"):
        return valor
    return fallback


def obtener_comunicaciones_portal_dia(conn, jugador, limite=5):
    hoy = ahora_sig().strftime("%Y-%m-%d")
    categoria = jugador.get("categoria") or ""
    return conn.execute("""
        SELECT id, titulo, mensaje, destino, categoria, jugador_id, url, creado_en, visible_hasta
        FROM pwa_push_envios
        WHERE COALESCE(mostrar_portal, 0) = 1
          AND (
              visible_hasta IS NULL
              OR visible_hasta = ''
              OR visible_hasta >= %s
          )
          AND (
              destino = 'portal_todos'
              OR (destino = 'categoria' AND COALESCE(categoria, '') = %s)
              OR (destino = 'jugador' AND jugador_id = %s)
          )
        ORDER BY creado_en DESC, id DESC
        LIMIT %s
    """, (hoy, categoria, jugador["id"], limite)).fetchall()


def usuario_activo_reciente(username, segundos=120):
    clave = presence_key(username)
    if not clave:
        return False
    try:
        conn = get_connection()
        fila = conn.execute("""
            SELECT 1 AS activo
            FROM app_settings
            WHERE clave = %s
              AND actualizado_en >= CURRENT_TIMESTAMP - (%s * INTERVAL '1 second')
            LIMIT 1
        """, (clave, max(1, int(segundos)))).fetchone()
        conn.close()
        return bool(fila)
    except Exception:
        app.logger.exception("No se pudo consultar presencia reciente de %s.", username)
        return False


def suprimir_email_whatsapp_por_presencia():
    for username in WHATSAPP_EMAIL_SUPPRESS_USERNAMES:
        if usuario_activo_reciente(username, WHATSAPP_EMAIL_SUPPRESS_SECONDS):
            return True
    return False


def normalizar_clave_notificacion(valor):
    texto = str(valor or "").strip().lower()
    return re.sub(r"[^a-z0-9_.:-]+", "_", texto).strip("_")


def clave_notificaciones_descartadas(username=None):
    username = normalizar_username(username or session.get("username", ""))
    return f"notifications:dismissed:{username}" if username else ""


def clave_notificacion(tipo, entidad_id):
    tipo = normalizar_clave_notificacion(tipo)
    entidad_id = normalizar_clave_notificacion(entidad_id)
    return f"{tipo}:{entidad_id}" if tipo and entidad_id else ""


def obtener_notificaciones_descartadas(username=None):
    clave = clave_notificaciones_descartadas(username)
    if not clave:
        return set()
    try:
        conn = get_connection()
        fila = conn.execute("""
            SELECT valor
            FROM app_settings
            WHERE clave = %s
        """, (clave,)).fetchone()
        conn.close()
        valores = json.loads((fila or {}).get("valor") or "[]")
        return {normalizar_clave_notificacion(item) for item in valores if item}
    except Exception:
        app.logger.exception("No se pudieron leer notificaciones descartadas.")
        return set()


def descartar_notificacion_usuario(tipo, entidad_id, username=None):
    username = normalizar_username(username or session.get("username", ""))
    clave_setting = clave_notificaciones_descartadas(username)
    clave_item = clave_notificacion(tipo, entidad_id)
    if not clave_setting or not clave_item:
        return False

    descartadas = obtener_notificaciones_descartadas(username)
    descartadas.add(clave_item)
    conn = get_connection()
    guardar_app_setting(
        conn,
        clave_setting,
        json.dumps(sorted(descartadas), ensure_ascii=False),
        username,
    )
    conn.commit()
    conn.close()
    return True


def descartar_notificaciones_usuario(items, username=None):
    username = normalizar_username(username or session.get("username", ""))
    clave_setting = clave_notificaciones_descartadas(username)
    if not clave_setting:
        return 0

    claves = {
        clave_notificacion(tipo, entidad_id)
        for tipo, entidad_id in items
        if clave_notificacion(tipo, entidad_id)
    }
    if not claves:
        return 0

    descartadas = obtener_notificaciones_descartadas(username)
    nuevas = claves - descartadas
    if not nuevas:
        return 0

    descartadas.update(nuevas)
    conn = get_connection()
    guardar_app_setting(
        conn,
        clave_setting,
        json.dumps(sorted(descartadas), ensure_ascii=False),
        username,
    )
    conn.commit()
    conn.close()
    return len(nuevas)


def parsear_notificacion_form_value(valor):
    tipo, separador, entidad_id = str(valor or "").partition("|")
    if not separador:
        return "", ""
    return tipo, entidad_id


def preparar_notificaciones_para_usuario(items, tipo, id_func, descartadas):
    preparadas = []
    for item in items:
        fila = dict(item)
        fila["_notificacion_tipo"] = tipo
        fila["_notificacion_id"] = str(id_func(fila))
        fila["_notificacion_key"] = clave_notificacion(tipo, fila["_notificacion_id"])
        if fila["_notificacion_key"] not in descartadas:
            preparadas.append(fila)
    return preparadas


def drive_service():
    if (
        google_auth_default is None
        or google_build is None
        or MediaIoBaseDownload is None
        or MediaIoBaseUpload is None
    ):
        raise RuntimeError(
            "Faltan dependencias de Google Drive. Instalá google-api-python-client y google-auth."
        )

    credentials, _ = google_auth_default(
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return google_build("drive", "v3", credentials=credentials, cache_discovery=False)


def cloud_sql_admin_service():
    if google_auth_default is None or google_build is None:
        raise RuntimeError(
            "Faltan dependencias de Google Cloud. Instalá google-api-python-client y google-auth."
        )

    credentials, _ = google_auth_default(
        scopes=["https://www.googleapis.com/auth/sqlservice.admin"]
    )
    return google_build("sqladmin", "v1beta4", credentials=credentials, cache_discovery=False)


def secret_manager_service():
    if google_auth_default is None or google_build is None:
        raise RuntimeError(
            "Faltan dependencias de Google Secret Manager. Instalá google-api-python-client y google-auth."
        )

    credentials, _ = google_auth_default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    return google_build("secretmanager", "v1", credentials=credentials, cache_discovery=False)


def cloud_sql_instance_context():
    project = CLOUD_SQL_PROJECT
    instance = CLOUD_SQL_INSTANCE

    if CLOUD_SQL_CONNECTION_NAME:
        partes = CLOUD_SQL_CONNECTION_NAME.split(":")
        if len(partes) == 3:
            project = project or partes[0]
            instance = instance or partes[2]

    return {
        "project": project,
        "instance": instance,
        "connection_name": CLOUD_SQL_CONNECTION_NAME or "",
    }


def google_cloud_project_id():
    for clave in ("CLOUD_SQL_PROJECT", "GOOGLE_CLOUD_PROJECT", "GCP_PROJECT", "GCLOUD_PROJECT"):
        valor = os.environ.get(clave, "").strip()
        if valor:
            return valor

    if google_auth_default is not None:
        try:
            _, project_id = google_auth_default()
            if project_id:
                return project_id
        except Exception:
            pass

    try:
        req = UrlRequest(
            "http://metadata.google.internal/computeMetadata/v1/project/project-id",
            headers={"Metadata-Flavor": "Google"},
        )
        with urlopen(req, timeout=2) as resp:
            return resp.read().decode("utf-8").strip()
    except Exception:
        return ""


def secret_resource_name(secret_name, version="latest"):
    secret_name = (secret_name or "").strip()
    if not secret_name:
        return ""
    if secret_name.startswith("projects/"):
        if "/versions/" in secret_name:
            return secret_name
        return f"{secret_name}/versions/{version}"
    project = google_cloud_project_id()
    if not project:
        return ""
    return f"projects/{project}/secrets/{secret_name}/versions/{version}"


def secret_parent_name():
    project = google_cloud_project_id()
    return f"projects/{project}" if project else ""


def guardar_secret_manager(secret_name, valor):
    parent = secret_parent_name()
    if not parent:
        raise RuntimeError("Falta configurar GOOGLE_CLOUD_PROJECT o CLOUD_SQL_PROJECT.")
    secret_name = (secret_name or FACTURA_EMAIL_SECRET_NAME or "").strip()
    if not secret_name:
        raise RuntimeError("Falta indicar el nombre del secreto.")
    if secret_name.startswith("projects/"):
        secret_path = secret_name.split("/versions/", 1)[0]
        secret_id = secret_path.rsplit("/", 1)[-1]
    else:
        secret_id = secret_name
        secret_path = f"{parent}/secrets/{secret_id}"

    service = secret_manager_service()
    try:
        service.projects().secrets().get(name=secret_path).execute()
    except Exception:
        service.projects().secrets().create(
            parent=parent,
            secretId=secret_id,
            body={"replication": {"automatic": {}}},
        ).execute()

    payload = base64.b64encode((valor or "").encode("utf-8")).decode("ascii")
    service.projects().secrets().addVersion(
        parent=secret_path,
        body={"payload": {"data": payload}},
    ).execute()
    return secret_path


def leer_secret_manager(secret_name):
    name = secret_resource_name(secret_name)
    if not name:
        return ""
    service = secret_manager_service()
    response = service.projects().secrets().versions().access(name=name).execute()
    data = ((response.get("payload") or {}).get("data") or "").encode("ascii")
    return base64.b64decode(data).decode("utf-8")


def normalizar_backup_cloud_sql(backup):
    if not backup:
        return None
    return {
        "id": backup.get("id") or backup.get("selfLink") or "",
        "status": backup.get("status") or "-",
        "type": backup.get("type") or "-",
        "start_time": backup.get("startTime") or "",
        "end_time": backup.get("endTime") or "",
        "window_start_time": backup.get("windowStartTime") or "",
        "description": backup.get("description") or "",
    }


def obtener_info_backups_cloud_sql():
    contexto = cloud_sql_instance_context()
    info = {
        "api_disponible": False,
        "error": None,
        "project": contexto["project"],
        "instance": contexto["instance"],
        "connection_name": contexto["connection_name"],
        "ultimo": None,
        "en_ejecucion": [],
    }

    if not contexto["project"] or not contexto["instance"]:
        info["error"] = "Falta configurar CLOUD_SQL_PROJECT/CLOUD_SQL_INSTANCE o CLOUD_SQL_CONNECTION_NAME."
        return info

    try:
        service = cloud_sql_admin_service()
        response = service.backupRuns().list(
            project=contexto["project"],
            instance=contexto["instance"],
            maxResults=10,
        ).execute()
        backups = response.get("items") or []
        backups.sort(
            key=lambda item: item.get("endTime") or item.get("startTime") or item.get("windowStartTime") or "",
            reverse=True,
        )
        info["api_disponible"] = True
        info["ultimo"] = normalizar_backup_cloud_sql(backups[0]) if backups else None
        info["en_ejecucion"] = [
            normalizar_backup_cloud_sql(item)
            for item in backups
            if item.get("status") in {"RUNNING", "PENDING"}
        ]
    except Exception as error:
        app.logger.exception("No se pudo consultar Cloud SQL Admin API.")
        info["error"] = str(error)

    return info


def solicitar_backup_cloud_sql():
    contexto = cloud_sql_instance_context()
    if not contexto["project"] or not contexto["instance"]:
        raise RuntimeError("Falta configurar CLOUD_SQL_PROJECT/CLOUD_SQL_INSTANCE o CLOUD_SQL_CONNECTION_NAME.")

    service = cloud_sql_admin_service()
    return service.backupRuns().insert(
        project=contexto["project"],
        instance=contexto["instance"],
        body={
            "description": f"Backup manual solicitado desde SIG por {session.get('username') or 'sistema'}"
        },
    ).execute()


def require_drive_comprobantes_folder():
    if not DRIVE_COMPROBANTES_FOLDER_ID and not DRIVE_SHARED_DRIVE_ID:
        raise RuntimeError(
            "Falta configurar GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID o GOOGLE_DRIVE_SHARED_DRIVE_ID."
        )
    return DRIVE_COMPROBANTES_FOLDER_ID


def require_drive_fichas_medicas_folder():
    if not DRIVE_FICHAS_MEDICAS_FOLDER_ID and not DRIVE_SHARED_DRIVE_ID:
        raise RuntimeError(
            "Falta configurar GOOGLE_DRIVE_FICHAS_MEDICAS_FOLDER_ID, GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID o GOOGLE_DRIVE_SHARED_DRIVE_ID."
        )
    return DRIVE_FICHAS_MEDICAS_FOLDER_ID


def require_drive_secretaria_folder():
    if not DRIVE_SECRETARIA_FOLDER_ID and not DRIVE_SHARED_DRIVE_ID:
        raise RuntimeError(
            "Falta configurar GOOGLE_DRIVE_SECRETARIA_FOLDER_ID, GOOGLE_DRIVE_COMPROBANTES_FOLDER_ID o GOOGLE_DRIVE_SHARED_DRIVE_ID."
        )
    return DRIVE_SECRETARIA_FOLDER_ID


def drive_query_escape(value):
    return str(value or "").replace("\\", "\\\\").replace("'", "\\'")


def find_drive_folder(service, name, parent_id=None, drive_id=None):
    safe_name = drive_query_escape(name)
    query_parts = [
        "mimeType = 'application/vnd.google-apps.folder'",
        f"name = '{safe_name}'",
        "trashed = false",
    ]
    if parent_id:
        safe_parent = drive_query_escape(parent_id)
        query_parts.append(f"'{safe_parent}' in parents")

    list_kwargs = {
        "q": " and ".join(query_parts),
        "fields": "files(id, name)",
        "pageSize": 1,
        "supportsAllDrives": True,
        "includeItemsFromAllDrives": True,
    }
    if drive_id:
        list_kwargs["corpora"] = "drive"
        list_kwargs["driveId"] = drive_id

    result = service.files().list(**list_kwargs).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"]
    return None


def get_drive_root_subfolder(service, drive_id, name):
    folder_id = find_drive_folder(service, name, parent_id="root", drive_id=drive_id)
    if folder_id:
        return folder_id
    raise RuntimeError(
        f"No se encontro la carpeta '{name}' en la unidad compartida configurada."
    )


def create_drive_subfolder(service, parent_id, name):
    folder = service.files().create(
        body={
            "name": name,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_id],
        },
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return folder["id"]


def get_or_create_drive_root_subfolder(service, drive_id, name):
    folder_id = find_drive_folder(service, name, parent_id="root", drive_id=drive_id)
    if folder_id:
        return folder_id
    return create_drive_subfolder(service, drive_id, name)


def get_or_create_drive_subfolder(service, parent_id, name):
    folder_id = find_drive_folder(service, name, parent_id=parent_id)
    if folder_id:
        return folder_id

    return create_drive_subfolder(service, parent_id, name)


def get_drive_periodo_folder(service, root_folder, periodo):
    try:
        fecha_periodo = datetime.strptime(periodo, "%Y-%m")
    except (TypeError, ValueError):
        fecha_periodo = ahora_sig()

    year_folder = get_or_create_drive_subfolder(
        service,
        root_folder,
        str(fecha_periodo.year),
    )
    return get_or_create_drive_subfolder(
        service,
        year_folder,
        MESES_ES[fecha_periodo.month - 1],
    )


def get_drive_comprobantes_base_folder(service):
    root_folder = require_drive_comprobantes_folder()
    if root_folder:
        subfolder = (DRIVE_COMPROBANTES_SUBFOLDER or "").strip()
        if not subfolder:
            return root_folder
        return get_or_create_drive_subfolder(service, root_folder, subfolder)

    subfolder = (DRIVE_COMPROBANTES_SUBFOLDER or "Comprobantes").strip()
    return get_drive_root_subfolder(service, DRIVE_SHARED_DRIVE_ID, subfolder)


def get_drive_fichas_medicas_base_folder(service):
    root_folder = require_drive_fichas_medicas_folder()
    if root_folder:
        subfolder = (DRIVE_FICHAS_MEDICAS_SUBFOLDER or "").strip()
        if not subfolder:
            return root_folder
        return get_or_create_drive_subfolder(service, root_folder, subfolder)

    subfolder = (DRIVE_FICHAS_MEDICAS_SUBFOLDER or "Fichas m?dicas").strip()
    return get_drive_root_subfolder(service, DRIVE_SHARED_DRIVE_ID, subfolder)


def get_drive_secretaria_base_folder(service):
    root_folder = require_drive_secretaria_folder()
    if root_folder:
        subfolder = (DRIVE_SECRETARIA_SUBFOLDER or "").strip()
        if not subfolder:
            return root_folder
        return get_or_create_drive_subfolder(service, root_folder, subfolder)

    subfolder = (DRIVE_SECRETARIA_SUBFOLDER or "Secretaria").strip()
    return get_or_create_drive_root_subfolder(service, DRIVE_SHARED_DRIVE_ID, subfolder)


def get_drive_jugador_folder(service, root_folder, jugador):
    jugador_slug = secure_filename(
        f"{jugador['apellido']}_{jugador['nombre']}_{jugador['id']}"
    ) or f"jugador_{jugador['id']}"
    return get_or_create_drive_subfolder(service, root_folder, jugador_slug)


def subir_ficha_medica_batch_pendiente(validado, batch_id):
    if not validado:
        return None

    service = drive_service()
    root_folder = get_drive_fichas_medicas_base_folder(service)
    pendientes_folder = get_or_create_drive_subfolder(service, root_folder, "_batch_pendientes")
    batch_folder = get_or_create_drive_subfolder(service, pendientes_folder, batch_id)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    filename = validado["filename"] or f"ficha_medica{validado['ext']}"
    drive_name = f"pendiente_{timestamp}_{filename}"
    media = MediaIoBaseUpload(
        io.BytesIO(validado["content"]),
        mimetype=validado["mime_type"],
        resumable=False,
    )

    uploaded = service.files().create(
        body={"name": drive_name, "parents": [batch_folder]},
        media_body=media,
        fields="id, name, mimeType, size, webViewLink",
        supportsAllDrives=True,
    ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or filename,
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": batch_folder,
    }


def mover_ficha_medica_batch_a_jugador(file_id, jugador, ext, source_folder_id=None):
    service = drive_service()
    root_folder = get_drive_fichas_medicas_base_folder(service)
    target_folder = get_drive_jugador_folder(service, root_folder, jugador)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    jugador_slug = secure_filename(f"{jugador['apellido']}_{jugador['nombre']}") or f"jugador_{jugador['id']}"
    drive_name = f"ficha_medica_{jugador['id']}_{jugador_slug}_{timestamp}{ext or ''}"

    update_kwargs = {
        "fileId": file_id,
        "body": {"name": drive_name},
        "addParents": target_folder,
        "fields": "id, name, mimeType, size, webViewLink",
        "supportsAllDrives": True,
    }
    if source_folder_id:
        update_kwargs["removeParents"] = source_folder_id

    updated = service.files().update(**update_kwargs).execute()
    return {
        "file_id": updated["id"],
        "nombre": updated.get("name") or drive_name,
        "mime_type": updated.get("mimeType"),
        "tamano": int(updated.get("size") or 0),
        "web_url": updated.get("webViewLink"),
        "folder_id": target_folder,
    }


def validar_comprobante_upload(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    filename = secure_filename(file_storage.filename)
    ext = Path(filename).suffix.lower()
    if ext not in COMPROBANTE_EXTENSIONS:
        raise ValueError("El comprobante debe ser PDF, JPG o PNG.")

    content = file_storage.read()
    if not content:
        raise ValueError("El comprobante está vacío.")
    if len(content) > COMPROBANTE_MAX_BYTES:
        max_mb = max(1, COMPROBANTE_MAX_BYTES // (1024 * 1024))
        raise ValueError(f"El comprobante supera el tamaño máximo permitido ({max_mb} MB).")

    mime_type = COMPROBANTE_EXTENSIONS[ext]
    guessed_mime = mimetypes.guess_type(filename)[0]
    if guessed_mime in COMPROBANTE_EXTENSIONS.values():
        mime_type = guessed_mime

    return filename, ext, content, mime_type


def validar_comprobante_upload_dict(file_storage):
    validado = validar_comprobante_upload(file_storage)
    if not validado:
        return None

    filename, ext, content, mime_type = validado
    return {
        "filename": filename,
        "ext": ext,
        "content": content,
        "mime_type": mime_type,
    }


def validar_documento_secretaria_upload(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    filename = secure_filename(file_storage.filename)
    ext = Path(filename).suffix.lower()
    if ext not in SECRETARIA_EXTENSIONS:
        raise ValueError("El archivo debe ser PDF, imagen, Word, Excel, PowerPoint, CSV o TXT.")

    content = file_storage.read()
    if not content:
        raise ValueError("El archivo está vacío.")
    if len(content) > SECRETARIA_MAX_BYTES:
        max_mb = max(1, SECRETARIA_MAX_BYTES // (1024 * 1024))
        raise ValueError(f"El archivo supera el tamaño máximo permitido ({max_mb} MB).")

    mime_type = SECRETARIA_EXTENSIONS[ext]
    guessed_mime = mimetypes.guess_type(filename)[0]
    if guessed_mime:
        mime_type = guessed_mime

    return {
        "filename": filename,
        "ext": ext,
        "content": content,
        "mime_type": mime_type,
    }


def get_drive_secretaria_folder(service, categoria, fecha_base=None):
    root_folder = get_drive_secretaria_base_folder(service)
    periodo = (fecha_base or ahora_sig().strftime("%Y-%m-%d"))[:7]
    periodo_folder = get_drive_periodo_folder(service, root_folder, periodo)
    categoria_slug = secure_filename(categoria or "General") or "General"
    return get_or_create_drive_subfolder(service, periodo_folder, categoria_slug)


def subir_documento_secretaria_a_drive(validado, categoria, titulo, fecha_base=None):
    if not validado:
        return None

    service = drive_service()
    folder_id = get_drive_secretaria_folder(service, categoria, fecha_base=fecha_base)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    titulo_slug = secure_filename(titulo or Path(validado["filename"]).stem or "documento") or "documento"
    drive_name = f"secretaria_{titulo_slug}_{timestamp}{validado['ext']}"
    media = MediaIoBaseUpload(io.BytesIO(validado["content"]), mimetype=validado["mime_type"], resumable=False)

    uploaded = service.files().create(
        body={"name": drive_name, "parents": [folder_id]},
        media_body=media,
        fields="id, name, mimeType, size, webViewLink",
        supportsAllDrives=True,
    ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or validado["filename"],
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": folder_id,
    }


def subir_comprobante_a_drive(file_storage, cuota):
    validado = validar_comprobante_upload(file_storage)
    if not validado:
        return None

    filename, ext, content, mime_type = validado
    service = drive_service()
    root_folder = get_drive_comprobantes_base_folder(service)
    periodo = cuota["periodo"] or ahora_sig().strftime("%Y-%m")
    folder_id = get_drive_periodo_folder(service, root_folder, periodo)
    jugador_slug = secure_filename(f"{cuota['apellido']}_{cuota['nombre']}") or f"jugador_{cuota['jugador_id']}"
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    drive_name = f"cuota_{cuota['id']}_{periodo}_{jugador_slug}_{timestamp}{ext}"
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=mime_type, resumable=False)
    body = {"name": drive_name, "parents": [folder_id]}

    existing_file_id = cuota.get("comprobante_drive_file_id")
    if existing_file_id:
        uploaded = service.files().update(
            fileId=existing_file_id,
            body={"name": drive_name},
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()
    else:
        uploaded = service.files().create(
            body=body,
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or filename,
        "mime_type": uploaded.get("mimeType") or mime_type,
        "tamano": int(uploaded.get("size") or len(content)),
        "web_url": uploaded.get("webViewLink"),
    }


def get_drive_gastos_compartidos_folder(service, fecha_base=None):
    root_folder = get_drive_comprobantes_base_folder(service)
    gastos_folder = get_or_create_drive_subfolder(service, root_folder, "Gastos compartidos")
    periodo = (fecha_base or ahora_sig().strftime("%Y-%m-%d"))[:7]
    return get_drive_periodo_folder(service, gastos_folder, periodo)


def subir_comprobante_gasto_compartido_a_drive(file_storage, item):
    validado = validar_comprobante_upload(file_storage)
    if not validado:
        return None

    filename, ext, content, mime_type = validado
    service = drive_service()
    folder_id = get_drive_gastos_compartidos_folder(service, item.get("fecha_vencimiento"))
    jugador_slug = secure_filename(f"{item['apellido']}_{item['nombre']}") or f"jugador_{item['jugador_id']}"
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    drive_name = f"gasto_{item['gasto_id']}_{item['id']}_{jugador_slug}_{timestamp}{ext}"
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=mime_type, resumable=False)
    body = {"name": drive_name, "parents": [folder_id]}

    existing_file_id = item.get("comprobante_drive_file_id")
    if existing_file_id:
        try:
            uploaded = service.files().update(
                fileId=existing_file_id,
                body={"name": drive_name},
                media_body=media,
                fields="id, name, mimeType, size, webViewLink",
                supportsAllDrives=True,
            ).execute()
        except Exception:
            app.logger.exception(
                "No se pudo actualizar comprobante existente de gasto compartido %s; se intentara crear uno nuevo.",
                item.get("id"),
            )
            uploaded = service.files().create(
                body=body,
                media_body=media,
                fields="id, name, mimeType, size, webViewLink",
                supportsAllDrives=True,
            ).execute()
    else:
        uploaded = service.files().create(
            body=body,
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or filename,
        "mime_type": uploaded.get("mimeType") or mime_type,
        "tamano": int(uploaded.get("size") or len(content)),
        "web_url": uploaded.get("webViewLink"),
    }


def get_drive_caja_folder(service, fecha):
    root_folder = get_drive_comprobantes_base_folder(service)
    caja_folder = get_or_create_drive_subfolder(service, root_folder, "Caja")
    periodo = (fecha or ahora_sig().strftime("%Y-%m-%d"))[:7]
    return get_drive_periodo_folder(service, caja_folder, periodo)


def subir_comprobante_movimiento_a_drive(validado, movimiento, existing_file_id=None):
    if not validado:
        return None

    service = drive_service()
    folder_id = get_drive_caja_folder(service, movimiento.get("fecha"))
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    concepto_slug = secure_filename(movimiento.get("concepto") or "movimiento") or "movimiento"
    tipo = secure_filename(movimiento.get("tipo") or "movimiento") or "movimiento"
    movimiento_id = movimiento.get("id") or "nuevo"
    drive_name = (
        f"caja_{movimiento_id}_{movimiento.get('fecha') or 'sin_fecha'}_"
        f"{tipo}_{concepto_slug}_{timestamp}{validado['ext']}"
    )
    media = MediaIoBaseUpload(io.BytesIO(validado["content"]), mimetype=validado["mime_type"], resumable=False)

    if existing_file_id:
        uploaded = service.files().update(
            fileId=existing_file_id,
            body={"name": drive_name},
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()
    else:
        uploaded = service.files().create(
            body={"name": drive_name, "parents": [folder_id]},
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or validado["filename"],
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": folder_id,
    }


def int_setting(valor, default, minimo=None, maximo=None):
    try:
        numero = int(valor)
    except (TypeError, ValueError):
        numero = default
    if minimo is not None:
        numero = max(minimo, numero)
    if maximo is not None:
        numero = min(maximo, numero)
    return numero


def factura_email_defaults(indice=1):
    if indice == 2:
        return {
            "host": FACTURA_EMAIL2_IMAP_HOST,
            "port": FACTURA_EMAIL2_IMAP_PORT,
            "user": FACTURA_EMAIL2_IMAP_USER,
            "password": FACTURA_EMAIL2_IMAP_PASSWORD,
            "folder": FACTURA_EMAIL2_IMAP_FOLDER,
            "use_ssl": FACTURA_EMAIL2_IMAP_USE_SSL,
            "search_days": FACTURA_EMAIL2_SEARCH_DAYS,
            "max_messages": FACTURA_EMAIL2_MAX_MESSAGES,
            "secret_name": FACTURA_EMAIL2_SECRET_NAME,
        }
    return {
        "host": FACTURA_EMAIL_IMAP_HOST,
        "port": FACTURA_EMAIL_IMAP_PORT,
        "user": FACTURA_EMAIL_IMAP_USER,
        "password": FACTURA_EMAIL_IMAP_PASSWORD,
        "folder": FACTURA_EMAIL_IMAP_FOLDER,
        "use_ssl": FACTURA_EMAIL_IMAP_USE_SSL,
        "search_days": FACTURA_EMAIL_SEARCH_DAYS,
        "max_messages": FACTURA_EMAIL_MAX_MESSAGES,
        "secret_name": FACTURA_EMAIL_SECRET_NAME,
    }


def factura_email_setting_key(nombre, indice=1):
    return f"factura_email{indice}_{nombre}" if indice != 1 else f"factura_email_{nombre}"


def obtener_factura_email_config(conn=None, incluir_password=False, indice=1):
    own_conn = conn is None
    if own_conn:
        conn = get_connection()
    keys = [
        factura_email_setting_key("imap_host", indice),
        factura_email_setting_key("imap_port", indice),
        factura_email_setting_key("imap_user", indice),
        factura_email_setting_key("imap_folder", indice),
        factura_email_setting_key("imap_use_ssl", indice),
        factura_email_setting_key("search_days", indice),
        factura_email_setting_key("max_messages", indice),
        factura_email_setting_key("secret_name", indice),
        factura_email_setting_key("secret_actualizado_en", indice),
        factura_email_setting_key("secret_actualizado_por", indice),
    ]
    settings = obtener_app_settings(conn, keys)
    if own_conn:
        conn.close()

    defaults = factura_email_defaults(indice)
    secret_name_key = factura_email_setting_key("secret_name", indice)
    secret_actualizado_key = factura_email_setting_key("secret_actualizado_en", indice)
    secret_actualizado_por_key = factura_email_setting_key("secret_actualizado_por", indice)
    secret_setting = (settings.get(secret_name_key) or {}).get("valor")
    secret_name = (
        os.environ.get("FACTURA_EMAIL_SECRET_NAME" if indice == 1 else "FACTURA_EMAIL2_SECRET_NAME", "").strip()
        or secret_setting
        or defaults["secret_name"]
    )
    secret_actualizado_en = (settings.get(secret_actualizado_key) or {}).get("valor")
    config = {
        "indice": indice,
        "label": f"Cuenta {indice}",
        "host": (settings.get(factura_email_setting_key("imap_host", indice)) or {}).get("valor") or defaults["host"],
        "port": int_setting((settings.get(factura_email_setting_key("imap_port", indice)) or {}).get("valor"), defaults["port"], 1, 65535),
        "user": (settings.get(factura_email_setting_key("imap_user", indice)) or {}).get("valor") or defaults["user"],
        "folder": (settings.get(factura_email_setting_key("imap_folder", indice)) or {}).get("valor") or defaults["folder"],
        "use_ssl": parse_bool_setting((settings.get(factura_email_setting_key("imap_use_ssl", indice)) or {}).get("valor")) if settings.get(factura_email_setting_key("imap_use_ssl", indice)) else defaults["use_ssl"],
        "search_days": int_setting((settings.get(factura_email_setting_key("search_days", indice)) or {}).get("valor"), defaults["search_days"], 1, 365),
        "max_messages": int_setting((settings.get(factura_email_setting_key("max_messages", indice)) or {}).get("valor"), defaults["max_messages"], 1, 500),
        "secret_name": secret_name,
        "secret_actualizado_en": secret_actualizado_en,
        "secret_actualizado_por": (settings.get(secret_actualizado_por_key) or {}).get("valor"),
        "password_env": bool(defaults["password"]),
        "password_configurado": bool(defaults["password"] or secret_actualizado_en or secret_setting),
    }
    if incluir_password:
        password = defaults["password"]
        if not password and secret_name:
            password = leer_secret_manager(secret_name)
        config["password"] = password
        config["password_configurado"] = bool(password)
    return config


def obtener_factura_email_configs(conn=None, incluir_password=False, solo_configuradas=False):
    configs_base = [
        obtener_factura_email_config(conn, incluir_password=False, indice=1),
        obtener_factura_email_config(conn, incluir_password=False, indice=2),
    ]
    if solo_configuradas:
        configs_base = [
            config for config in configs_base
            if config["host"] and config["user"] and config["password_configurado"]
        ]
    if incluir_password:
        return [
            obtener_factura_email_config(conn, incluir_password=True, indice=config["indice"])
            for config in configs_base
        ]
    return configs_base


def factura_email_configurado():
    return bool(obtener_factura_email_configs(solo_configuradas=True))


def get_drive_facturas_recibidas_folder(service, fecha_base=None):
    root_folder = get_drive_comprobantes_base_folder(service)
    facturas_folder = get_or_create_drive_subfolder(service, root_folder, "Facturas recibidas")
    periodo = (fecha_base or ahora_sig().strftime("%Y-%m-%d"))[:7]
    return get_drive_periodo_folder(service, facturas_folder, periodo)


def validar_factura_email_adjunto(filename, content, content_type=""):
    filename = secure_filename(filename or "factura.pdf")
    ext = Path(filename).suffix.lower()
    if ext not in FACTURA_EMAIL_EXTENSIONS:
        return None
    if not content:
        return None
    if len(content) > FACTURA_EMAIL_MAX_BYTES:
        raise ValueError(f"{filename} supera el tamaño maximo permitido para facturas.")

    mime_type = FACTURA_EMAIL_EXTENSIONS[ext]
    guessed_mime = mimetypes.guess_type(filename)[0]
    if guessed_mime in FACTURA_EMAIL_EXTENSIONS.values():
        mime_type = guessed_mime
    elif content_type in FACTURA_EMAIL_EXTENSIONS.values():
        mime_type = content_type

    return {
        "filename": filename,
        "ext": ext,
        "content": content,
        "mime_type": mime_type,
    }


def subir_factura_recibida_a_drive(validado, proveedor, fecha_base=None):
    if not validado:
        return None

    service = drive_service()
    folder_id = get_drive_facturas_recibidas_folder(service, fecha_base=fecha_base)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    proveedor_slug = secure_filename(proveedor or "proveedor") or "proveedor"
    archivo_slug = secure_filename(Path(validado["filename"]).stem or "factura") or "factura"
    drive_name = f"factura_{proveedor_slug}_{archivo_slug}_{timestamp}{validado['ext']}"
    media = MediaIoBaseUpload(io.BytesIO(validado["content"]), mimetype=validado["mime_type"], resumable=False)
    uploaded = service.files().create(
        body={"name": drive_name, "parents": [folder_id]},
        media_body=media,
        fields="id, name, mimeType, size, webViewLink",
        supportsAllDrives=True,
    ).execute()
    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or validado["filename"],
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": folder_id,
    }


def decodificar_header_email(valor):
    if not valor:
        return ""
    try:
        return str(make_header(decode_header(valor))).strip()
    except Exception:
        return str(valor or "").strip()


def fecha_email_iso(valor):
    if not valor:
        return None
    try:
        fecha = parsedate_to_datetime(valor)
        return fecha.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def normalizar_patron_factura(valor):
    return re.sub(r"\s+", " ", (valor or "").strip().lower())


def factura_email_coincide(filtro, remitente, asunto):
    remitente_patron = normalizar_patron_factura(filtro.get("remitente_patron"))
    asunto_patron = normalizar_patron_factura(filtro.get("asunto_patron"))
    remitente = normalizar_patron_factura(remitente)
    asunto = normalizar_patron_factura(asunto)
    if remitente_patron and remitente_patron not in remitente:
        return False
    if asunto_patron and asunto_patron not in asunto:
        return False
    return bool(remitente_patron or asunto_patron)


def buscar_filtro_factura_email(filtros, remitente, asunto):
    for filtro in filtros:
        if filtro.get("activo") and factura_email_coincide(filtro, remitente, asunto):
            return filtro
    return None


def contenido_texto_email(parte):
    try:
        return parte.get_content()
    except Exception:
        payload = parte.get_payload(decode=True) or b""
        charset = parte.get_content_charset() or "utf-8"
        return payload.decode(charset, errors="replace")


def factura_email_desde_cuerpo(mensaje, asunto):
    cuerpo_html = None
    cuerpo_texto = None
    partes = mensaje.walk() if mensaje.is_multipart() else [mensaje]
    for parte in partes:
        if parte.is_multipart():
            continue
        if (parte.get_content_disposition() or "").lower() == "attachment":
            continue
        content_type = (parte.get_content_type() or "").lower()
        if content_type == "text/html" and not cuerpo_html:
            cuerpo_html = contenido_texto_email(parte)
        elif content_type == "text/plain" and not cuerpo_texto:
            cuerpo_texto = contenido_texto_email(parte)

    if cuerpo_html:
        contenido = cuerpo_html
    elif cuerpo_texto:
        contenido = (
            "<!doctype html><html><head><meta charset=\"utf-8\"></head>"
            f"<body><pre>{html.escape(cuerpo_texto)}</pre></body></html>"
        )
    else:
        return None

    titulo = secure_filename(asunto or "factura_email") or "factura_email"
    filename = f"{titulo[:80]}.html"
    content = contenido.encode("utf-8", errors="replace")
    return validar_factura_email_adjunto(filename, content, "text/html")


def extraer_adjuntos_factura_email(mensaje, asunto=None):
    adjuntos = []
    for parte in mensaje.walk():
        if parte.is_multipart():
            continue
        filename = decodificar_header_email(parte.get_filename() or "")
        disposition = (parte.get_content_disposition() or "").lower()
        if not filename and disposition != "attachment":
            continue
        if not filename:
            filename = "factura.pdf"
        content = parte.get_payload(decode=True) or b""
        validado = validar_factura_email_adjunto(filename, content, parte.get_content_type())
        if validado:
            adjuntos.append(validado)
    if not adjuntos:
        factura_cuerpo = factura_email_desde_cuerpo(mensaje, asunto)
        if factura_cuerpo:
            adjuntos.append(factura_cuerpo)
    return adjuntos


def facturas_email_filtros_activos(conn):
    return conn.execute("""
        SELECT *
        FROM facturas_email_filtros
        WHERE activo = 1
        ORDER BY
            CASE WHEN COALESCE(remitente_patron, '') <> '' THEN 0 ELSE 1 END,
            proveedor,
            id
    """).fetchall()


def sincronizar_facturas_email_cuenta(conn, config, filtros, usuario=None):
    cliente = None
    procesados = 0
    nuevas = 0
    omitidas = 0
    errores = []
    try:
        if config["use_ssl"]:
            cliente = imaplib.IMAP4_SSL(config["host"], config["port"])
        else:
            cliente = imaplib.IMAP4(config["host"], config["port"])
        cliente.login(config["user"], config["password"])
        estado, _ = cliente.select(config["folder"])
        if estado != "OK":
            raise RuntimeError(f"No se pudo abrir la carpeta IMAP {config['folder']}.")

        desde = (date.today() - timedelta(days=max(1, config["search_days"]))).strftime("%d-%b-%Y")
        estado, data = cliente.search(None, "SINCE", desde)
        if estado != "OK":
            raise RuntimeError("No se pudo buscar emails en la casilla configurada.")

        ids = (data[0] or b"").split()
        ids = ids[-max(1, config["max_messages"]):]
        for email_id in reversed(ids):
            estado, partes = cliente.fetch(email_id, "(RFC822)")
            if estado != "OK" or not partes:
                omitidas += 1
                continue
            raw = None
            for parte in partes:
                if isinstance(parte, tuple):
                    raw = parte[1]
                    break
            if not raw:
                omitidas += 1
                continue

            mensaje = BytesParser(policy=policy.default).parsebytes(raw)
            asunto = decodificar_header_email(mensaje.get("Subject"))
            remitente_header = decodificar_header_email(mensaje.get("From"))
            _, remitente_email = parseaddr(remitente_header)
            remitente_match = f"{remitente_header} {remitente_email}".strip()
            filtro = buscar_filtro_factura_email(filtros, remitente_match, asunto)
            procesados += 1
            if not filtro:
                continue

            message_id = decodificar_header_email(mensaje.get("Message-ID")) or f"imap:{email_id.decode(errors='ignore')}"
            fecha_email = fecha_email_iso(mensaje.get("Date"))
            fecha_base = (fecha_email or ahora_sig().strftime("%Y-%m-%d"))[:10]
            try:
                adjuntos = extraer_adjuntos_factura_email(mensaje, asunto)
            except ValueError as error:
                errores.append(str(error))
                continue
            if not adjuntos:
                omitidas += 1
                continue

            for index, adjunto in enumerate(adjuntos, start=1):
                source_key = hashlib.sha256(
                    f"{config['user']}|{message_id}|{index}|{adjunto['filename']}|{len(adjunto['content'])}".encode("utf-8", errors="ignore")
                ).hexdigest()
                existe = conn.execute(
                    "SELECT id FROM facturas_recibidas WHERE source_key = %s",
                    (source_key,),
                ).fetchone()
                if existe:
                    omitidas += 1
                    continue
                factura_drive = subir_factura_recibida_a_drive(adjunto, filtro["proveedor"], fecha_base=fecha_base)
                conn.execute("""
                    INSERT INTO facturas_recibidas (
                        source_key, cuenta_email, message_id, filtro_id, proveedor, remitente, asunto, fecha_email,
                        archivo_nombre, archivo_mime_type, archivo_tamano,
                        drive_file_id, drive_folder_id, archivo_web_url,
                        estado, creado_por
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pendiente', %s)
                """, (
                    source_key,
                    config["user"],
                    message_id,
                    filtro["id"],
                    filtro["proveedor"],
                    remitente_email or remitente_header,
                    asunto,
                    fecha_email,
                    factura_drive["nombre"],
                    factura_drive["mime_type"],
                    factura_drive["tamano"],
                    factura_drive["file_id"],
                    factura_drive["folder_id"],
                    factura_drive["web_url"],
                    usuario,
                ))
                nuevas += 1
    finally:
        if cliente is not None:
            try:
                cliente.logout()
            except Exception:
                pass

    return {"procesados": procesados, "nuevas": nuevas, "omitidas": omitidas, "errores": errores}


def sincronizar_facturas_email(conn, usuario=None):
    configs = obtener_factura_email_configs(conn, incluir_password=True, solo_configuradas=True)
    if not configs:
        raise RuntimeError("Falta configurar al menos una casilla IMAP de facturas.")

    filtros = facturas_email_filtros_activos(conn)
    if not filtros:
        raise RuntimeError("No hay filtros activos para facturas recibidas.")

    resultado_total = {
        "procesados": 0,
        "nuevas": 0,
        "omitidas": 0,
        "errores": [],
        "cuentas": [],
    }

    for config in configs:
        try:
            resultado = sincronizar_facturas_email_cuenta(conn, config, filtros, usuario)
        except Exception as error:
            resultado_total["errores"].append(f"{config['label']} ({config['user']}): {error}")
            continue

        resultado_total["procesados"] += resultado["procesados"]
        resultado_total["nuevas"] += resultado["nuevas"]
        resultado_total["omitidas"] += resultado["omitidas"]
        resultado_total["errores"].extend(
            f"{config['label']} ({config['user']}): {error}"
            for error in resultado["errores"]
        )
        resultado_total["cuentas"].append({
            "indice": config["indice"],
            "user": config["user"],
            "procesados": resultado["procesados"],
            "nuevas": resultado["nuevas"],
            "omitidas": resultado["omitidas"],
        })

    if not resultado_total["cuentas"] and resultado_total["errores"]:
        raise RuntimeError("; ".join(resultado_total["errores"]))

    guardar_app_setting(conn, "facturas_email_sync_en", ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), usuario)
    guardar_app_setting(conn, "facturas_email_sync_por", usuario or "", usuario)
    return resultado_total


def limpiar_numero_operacion(valor):
    valor = re.sub(r"\s+", " ", valor or "").strip(" .:-#\t")
    valor = re.split(r"\s{2,}|(?:fecha|monto|importe|estado|destino|origen)\b", valor, flags=re.I)[0]
    match = re.search(r"[A-Z0-9][A-Z0-9._/-]{3,40}", valor, flags=re.I)
    return match.group(0).strip(" .:-#") if match else ""


def texto_ocr_sin_acentos(texto):
    texto = normalizar_ocr_texto(texto).lower()
    return "".join(
        char for char in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(char)
    )


def extraer_numero_operacion_comprobante(texto):
    texto = texto_ocr_sin_acentos(texto)
    patrones = [
        r"(?:nro|numero|num|n\W*)\s*(?:de\s*)?operacion\s*[:#-]?\s*([a-z0-9][a-z0-9 ._/-]{3,60})",
        r"(?:codigo|id)\s+(?:de\s+)?operacion\s*[:#-]?\s*([a-z0-9][a-z0-9 ._/-]{3,60})",
        r"operacion\s*(?:nro|numero|num|n\W*)?\s*[:#-]?\s*([a-z0-9][a-z0-9 ._/-]{3,60})",
    ]
    for linea in texto.splitlines():
        for patron in patrones:
            match = re.search(patron, linea, flags=re.I)
            if match:
                numero = limpiar_numero_operacion(match.group(1))
                if numero:
                    return numero
    for patron in patrones:
        match = re.search(patron, texto, flags=re.I)
        if match:
            numero = limpiar_numero_operacion(match.group(1))
            if numero:
                return numero
    return ""


def normalizar_monto_comprobante(valor):
    valor = re.sub(r"[^0-9,.]", "", valor or "")
    if not valor:
        return ""
    if "," in valor and "." in valor:
        if valor.rfind(",") > valor.rfind("."):
            valor = valor.replace(".", "").replace(",", ".")
        else:
            valor = valor.replace(",", "")
    elif "," in valor:
        partes = valor.split(",")
        valor = "".join(partes[:-1]).replace(".", "") + "." + partes[-1]
    elif valor.count(".") > 1:
        partes = valor.split(".")
        valor = "".join(partes[:-1]) + "." + partes[-1]
    try:
        monto = float(valor)
    except ValueError:
        return ""
    if monto <= 0:
        return ""
    return f"{monto:.2f}".rstrip("0").rstrip(".")


def extraer_monto_comprobante(texto):
    texto = texto_ocr_sin_acentos(texto)
    patrones = [
        r"(?:importe|monto|total)\s*(?:pagado|abonado|transferido|de\s+la\s+operacion)?\s*[:$-]?\s*\$?\s*([0-9][0-9.,]{1,20})",
        r"\$\s*([0-9][0-9.,]{1,20})",
    ]
    candidatos = []
    for patron in patrones:
        for match in re.finditer(patron, texto, flags=re.I):
            monto = normalizar_monto_comprobante(match.group(1))
            if monto:
                candidatos.append(monto)
    return candidatos[-1] if candidatos else ""


def procesar_comprobante_movimiento(file_storage, movimiento, existing_file_id=None):
    validado = validar_comprobante_upload_dict(file_storage)
    if not validado:
        return None, "", "", ""

    service = drive_service()
    folder_id = get_drive_caja_folder(service, movimiento.get("fecha"))
    ocr_texto = normalizar_ocr_texto(
        extraer_texto_ocr_drive(
            validado,
            {"id": "caja", "apellido": "caja", "nombre": movimiento.get("id") or "movimiento"},
            folder_id,
        )
    )
    numero_operacion = extraer_numero_operacion_comprobante(ocr_texto)
    monto = extraer_monto_comprobante(ocr_texto)
    comprobante_info = subir_comprobante_movimiento_a_drive(validado, movimiento, existing_file_id=existing_file_id)
    return comprobante_info, numero_operacion, monto, ocr_texto


def validar_ficha_medica_upload(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    filename = secure_filename(file_storage.filename)
    ext = Path(filename).suffix.lower()
    if ext not in FICHA_MEDICA_EXTENSIONS:
        raise ValueError("La ficha m?dica debe ser PDF, JPG o PNG.")

    content = file_storage.read()
    if not content:
        raise ValueError("La ficha m?dica est? vac?a.")
    if len(content) > FICHA_MEDICA_MAX_BYTES:
        max_mb = max(1, FICHA_MEDICA_MAX_BYTES // (1024 * 1024))
        raise ValueError(f"La ficha m?dica supera el tama?o m?ximo permitido ({max_mb} MB).")

    mime_type = FICHA_MEDICA_EXTENSIONS[ext]
    guessed_mime = mimetypes.guess_type(filename)[0]
    if guessed_mime in FICHA_MEDICA_EXTENSIONS.values():
        mime_type = guessed_mime

    return {
        "filename": filename,
        "ext": ext,
        "content": content,
        "mime_type": mime_type,
    }


def subir_ficha_medica_a_drive(validado, jugador, ficha=None):
    if not validado:
        return None

    service = drive_service()
    root_folder = get_drive_fichas_medicas_base_folder(service)
    folder_id = get_drive_jugador_folder(service, root_folder, jugador)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    jugador_slug = secure_filename(f"{jugador['apellido']}_{jugador['nombre']}") or f"jugador_{jugador['id']}"
    drive_name = f"ficha_medica_{jugador['id']}_{jugador_slug}_{timestamp}{validado['ext']}"
    media = MediaIoBaseUpload(
        io.BytesIO(validado["content"]),
        mimetype=validado["mime_type"],
        resumable=False,
    )

    existing_file_id = ficha.get("documento_drive_file_id") if ficha else None
    if existing_file_id:
        uploaded = service.files().update(
            fileId=existing_file_id,
            body={"name": drive_name},
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()
    else:
        uploaded = service.files().create(
            body={"name": drive_name, "parents": [folder_id]},
            media_body=media,
            fields="id, name, mimeType, size, webViewLink",
            supportsAllDrives=True,
        ).execute()

    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or validado["filename"],
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": folder_id,
    }


def extraer_texto_ocr_drive(validado, jugador, folder_id=None):
    if not validado:
        return ""

    service = drive_service()
    if folder_id is None:
        root_folder = get_drive_fichas_medicas_base_folder(service)
        folder_id = get_drive_jugador_folder(service, root_folder, jugador)

    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    doc_name = f"ocr_tmp_ficha_medica_{jugador['id']}_{timestamp}"
    media = MediaIoBaseUpload(
        io.BytesIO(validado["content"]),
        mimetype=validado["mime_type"],
        resumable=False,
    )

    doc_id = None
    try:
        doc = service.files().create(
            body={
                "name": doc_name,
                "mimeType": "application/vnd.google-apps.document",
                "parents": [folder_id],
            },
            media_body=media,
            fields="id",
            supportsAllDrives=True,
            ocrLanguage=FICHA_MEDICA_OCR_LANGUAGE,
        ).execute()
        doc_id = doc["id"]
        exported = service.files().export(
            fileId=doc_id,
            mimeType="text/plain",
        ).execute()
        if isinstance(exported, bytes):
            return exported.decode("utf-8", errors="replace").strip()
        return str(exported or "").strip()
    finally:
        if doc_id:
            try:
                service.files().delete(
                    fileId=doc_id,
                    supportsAllDrives=True,
                ).execute()
            except Exception as cleanup_error:
                status = getattr(getattr(cleanup_error, "resp", None), "status", None)
                if status != 404:
                    app.logger.warning("No se pudo eliminar el documento temporal OCR %s.", doc_id)


def normalizar_ocr_texto(texto):
    texto = (texto or "").replace("\r\n", "\n").replace("\r", "\n")
    texto = re.sub(r"[ \t]+", " ", texto)
    return texto.strip()


def normalizar_fecha_ocr(valor):
    valor = (valor or "").strip()
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(valor, formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def buscar_fecha_cercana(texto, palabras_clave):
    patron_fecha = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})"
    lineas = [linea.strip() for linea in texto.splitlines() if linea.strip()]
    for linea in lineas:
        linea_min = linea.lower()
        if any(palabra in linea_min for palabra in palabras_clave):
            match = re.search(patron_fecha, linea)
            if match:
                fecha = normalizar_fecha_ocr(match.group(1))
                if fecha:
                    return fecha
    return ""


def extraer_anio_bduar(texto):
    texto_min = texto.lower()
    patrones = (
        r"evaluaci[o?]n\s+pre\s*competitiva\s+(20\d{2})",
        r"evaluaci[o?]n\s+precompetitiva\s+(20\d{2})",
        r"pre\s*competitiva\s+(20\d{2})",
        r"\bbduar\b.*?(20\d{2})",
    )
    for patron in patrones:
        match = re.search(patron, texto_min, re.DOTALL)
        if match:
            return match.group(1)
    return ""


def extraer_telefono_ocr(texto):
    lineas = [linea.strip() for linea in texto.splitlines() if linea.strip()]
    for linea in lineas:
        linea_min = linea.lower()
        if any(palabra in linea_min for palabra in ("telefono", "tel", "celular", "emergencia")):
            match = re.search(r"(\+?\d[\d\s().-]{6,}\d)", linea)
            if match:
                return re.sub(r"\s+", " ", match.group(1)).strip()
    return ""


def extraer_contacto_emergencia_ocr(texto):
    lineas = [linea.strip() for linea in texto.splitlines() if linea.strip()]
    for index, linea in enumerate(lineas):
        linea_min = linea.lower()
        if "contacto" in linea_min and "emergencia" in linea_min:
            partes = re.split(r":|-", linea, maxsplit=1)
            if len(partes) > 1:
                contacto = re.sub(r"\+?\d[\d\s().-]{6,}\d", "", partes[1]).strip(" -:")
                if contacto:
                    return contacto[:120]
            if index + 1 < len(lineas):
                return lineas[index + 1][:120]
    return ""


def datos_ficha_desde_ocr(texto):
    texto = normalizar_ocr_texto(texto)
    texto_min = texto.lower()
    apto_detectado = None
    if re.search(r"\bno\s+apto\b", texto_min):
        apto_detectado = 0
    elif re.search(r"\bapto\b", texto_min):
        apto_detectado = 1

    fecha_vencimiento = buscar_fecha_cercana(
        texto,
        ("vencimiento", "vence", "vigencia", "validez", "valido hasta", "valida hasta"),
    )
    if not fecha_vencimiento:
        anio_bduar = extraer_anio_bduar(texto)
        if anio_bduar:
            fecha_vencimiento = f"{anio_bduar}-12-31"

    return {
        "fecha_vencimiento": fecha_vencimiento,
        "apto_fisico": apto_detectado,
        "contacto_emergencia": extraer_contacto_emergencia_ocr(texto),
        "telefono_emergencia": extraer_telefono_ocr(texto),
    }


def normalizar_texto_match(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    texto = texto.casefold()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_documento(valor):
    digitos = re.sub(r"\D+", "", str(valor or ""))
    return digitos.lstrip("0")


def extraer_dni_ocr(texto):
    texto = normalizar_ocr_texto(texto)
    lineas = [linea.strip() for linea in texto.splitlines() if linea.strip()]
    patron = r"(\d{1,2}\.?\d{3}\.?\d{3}|\d{7,9})"
    for linea in lineas:
        linea_min = linea.lower()
        if any(palabra in linea_min for palabra in ("dni", "documento", "d.n.i", "du ")):
            match = re.search(patron, linea)
            if match:
                return normalizar_documento(match.group(1))

    for match in re.finditer(patron, texto):
        documento = normalizar_documento(match.group(1))
        if 7 <= len(documento) <= 9:
            return documento
    return ""


def sugerir_jugador_ficha_ocr(texto, jugadores):
    texto_norm = normalizar_texto_match(texto)
    dni_detectado = extraer_dni_ocr(texto)
    if dni_detectado:
        for jugador in jugadores:
            if normalizar_documento(jugador.get("dni")) == dni_detectado:
                return jugador, "alta", f"DNI {dni_detectado}"

    candidatos = []
    for jugador in jugadores:
        nombre_tokens = normalizar_texto_match(jugador.get("nombre")).split()
        apellido_tokens = normalizar_texto_match(jugador.get("apellido")).split()
        tokens = [token for token in apellido_tokens + nombre_tokens if len(token) >= 3]
        if not tokens:
            continue

        coincidencias = sum(1 for token in tokens if token in texto_norm)
        if coincidencias == len(tokens):
            candidatos.append((90, jugador, "media", "Nombre y apellido detectados"))
        elif apellido_tokens and all(token in texto_norm for token in apellido_tokens):
            candidatos.append((70, jugador, "baja", "Apellido detectado"))

    if not candidatos:
        return None, "sin_coincidencia", "Sin coincidencia"

    candidatos.sort(key=lambda item: item[0], reverse=True)
    mejor = candidatos[0]
    if len(candidatos) > 1 and candidatos[1][0] == mejor[0]:
        return None, "baja", "Coincidencia ambigua"
    return mejor[1], mejor[2], mejor[3]


def obtener_jugadores_selector(conn):
    return conn.execute("""
        SELECT id, apellido, nombre, dni, categoria, estado, tipo_miembro
        FROM jugadores
        WHERE COALESCE(estado, 'Activo') <> 'Baja'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY apellido, nombre
    """).fetchall()


def obtener_fichas_medicas_batch_recientes(conn):
    return conn.execute("""
        SELECT
            batch_id,
            MIN(creado_en) AS creado_en,
            MAX(creado_por) AS creado_por,
            COUNT(*) AS total,
            SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN estado = 'procesado' THEN 1 ELSE 0 END) AS procesadas,
            MAX(id) AS ultimo_id
        FROM fichas_medicas_batch
        GROUP BY batch_id
        ORDER BY ultimo_id DESC
        LIMIT 10
    """).fetchall()


def procesar_ocr_ficha_medica_batch_item(conn, item):
    archivo = descargar_drive_file(item["drive_file_id"])
    content = archivo.getvalue()
    ext = item["extension"] or Path(item["archivo_original"] or item["documento_nombre"] or "").suffix.lower()
    mime_type = item["documento_mime_type"] or FICHA_MEDICA_EXTENSIONS.get(ext) or mimetypes.guess_type(item["documento_nombre"] or "")[0] or "application/pdf"
    filename = item["archivo_original"] or item["documento_nombre"] or f"ficha_medica_batch_{item['id']}{ext}"
    validado = {
        "filename": filename,
        "ext": ext,
        "content": content,
        "mime_type": mime_type,
    }

    ocr_texto = normalizar_ocr_texto(
        extraer_texto_ocr_drive(
            validado,
            {"id": "batch", "apellido": "batch", "nombre": item["batch_id"]},
            item["drive_folder_id"],
        )
    )
    datos_ocr = datos_ficha_desde_ocr(ocr_texto)
    jugadores = obtener_jugadores_selector(conn)
    jugador_sugerido, confianza, motivo = sugerir_jugador_ficha_ocr(ocr_texto, jugadores)

    conn.execute("""
        UPDATE fichas_medicas_batch
        SET ocr_texto = %s,
            ocr_fecha = %s,
            ocr_usuario = %s,
            jugador_sugerido_id = %s,
            confianza = %s,
            motivo = %s,
            fecha_vencimiento_sugerida = %s,
            apto_sugerido = %s,
            contacto_emergencia_sugerido = %s,
            telefono_emergencia_sugerido = %s,
            error = NULL
        WHERE id = %s
    """, (
        ocr_texto or None,
        ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if ocr_texto else None,
        session.get("username") if ocr_texto else None,
        jugador_sugerido["id"] if jugador_sugerido else None,
        confianza,
        motivo,
        datos_ocr.get("fecha_vencimiento") or None,
        datos_ocr.get("apto_fisico"),
        datos_ocr.get("contacto_emergencia") or None,
        datos_ocr.get("telefono_emergencia") or None,
        item["id"],
    ))
    return ocr_texto


def descargar_drive_file(file_id):
    service = drive_service()
    drive_request = service.files().get_media(
        fileId=file_id,
        supportsAllDrives=True,
    )
    file_buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(file_buffer, drive_request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    file_buffer.seek(0)
    return file_buffer


def eliminar_drive_file(file_id):
    if not file_id:
        return
    service = drive_service()
    service.files().delete(
        fileId=file_id,
        supportsAllDrives=True,
    ).execute()


def es_preview_pdf(mime_type):
    return (mime_type or "").lower() == "application/pdf"


def es_preview_imagen(mime_type):
    return (mime_type or "").lower().startswith("image/")


def obtener_preview_tipo(mime_type):
    if es_preview_pdf(mime_type):
        return "pdf"
    if es_preview_imagen(mime_type):
        return "image"
    return "other"


def resumir_resultados_tests(resultados):
    agrupados = {}
    for item in resultados or []:
        clave = item.get("test_nombre") or item.get("nombre") or "Test"
        agrupados.setdefault(clave, []).append(item)

    resumen = []
    for nombre, items in agrupados.items():
        ordenados = sorted(
            items,
            key=lambda fila: ((fila.get("fecha") or ""), fila.get("id") or 0),
            reverse=True,
        )
        actual = ordenados[0]
        previo = ordenados[1] if len(ordenados) > 1 else None
        delta = None
        if previo is not None:
            try:
                delta = round(float(actual.get("puntaje") or 0) - float(previo.get("puntaje") or 0), 2)
            except (TypeError, ValueError):
                delta = None
        resumen.append({
            "nombre": nombre,
            "actual": actual,
            "previo": previo,
            "delta": delta,
        })

    resumen.sort(key=lambda item: item["nombre"].lower())
    return resumen


def get_drive_lesion_folder(service, jugador, lesion):
    root_folder = get_drive_fichas_medicas_base_folder(service)
    jugador_folder = get_drive_jugador_folder(service, root_folder, jugador)
    lesiones_folder = get_or_create_drive_subfolder(service, jugador_folder, "Lesiones")
    lesion_slug = secure_filename(
        f"lesion_{lesion['id']}_{lesion.get('tipo_lesion') or 'adjuntos'}"
    ) or f"lesion_{lesion['id']}"
    return get_or_create_drive_subfolder(service, lesiones_folder, lesion_slug)


def subir_documento_lesion_a_drive(validado, jugador, lesion):
    if not validado:
        return None

    service = drive_service()
    folder_id = get_drive_lesion_folder(service, jugador, lesion)
    timestamp = ahora_sig().strftime("%Y%m%d_%H%M%S")
    drive_name = f"lesion_{lesion['id']}_{timestamp}_{validado['filename']}"
    media = MediaIoBaseUpload(
        io.BytesIO(validado["content"]),
        mimetype=validado["mime_type"],
        resumable=False,
    )
    uploaded = service.files().create(
        body={"name": drive_name, "parents": [folder_id]},
        media_body=media,
        fields="id, name, mimeType, size, webViewLink",
        supportsAllDrives=True,
    ).execute()
    return {
        "file_id": uploaded["id"],
        "nombre": uploaded.get("name") or validado["filename"],
        "mime_type": uploaded.get("mimeType") or validado["mime_type"],
        "tamano": int(uploaded.get("size") or len(validado["content"])),
        "web_url": uploaded.get("webViewLink"),
        "folder_id": folder_id,
    }


def guardar_documentos_lesion(conn, jugador, lesion, archivos, descripcion=""):
    guardados = 0
    for archivo in archivos or []:
        if not archivo or not archivo.filename:
            continue
        validado = validar_ficha_medica_upload(archivo)
        documento = subir_documento_lesion_a_drive(validado, jugador, lesion)
        conn.execute("""
            INSERT INTO lesiones_documentos (
                lesion_id, jugador_id, nombre, mime_type, tamano,
                drive_file_id, drive_folder_id, web_url, descripcion,
                creado_en, creado_por
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            lesion["id"],
            jugador["id"],
            documento["nombre"],
            documento["mime_type"],
            documento["tamano"],
            documento["file_id"],
            documento["folder_id"],
            documento["web_url"],
            descripcion or None,
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("username"),
        ))
        guardados += 1
    return guardados


def mensaje_error_drive(error, carpeta="Cuotas", accion="subir el comprobante"):
    if isinstance(error, RuntimeError):
        detalle = str(error).strip()
        detalle_lower = detalle.lower()
        if detalle_lower.startswith("falta configurar"):
            return f"{detalle} No se pudo {accion} en Google Drive."
        if "no se encontro la carpeta" in detalle_lower or "no se encontró la carpeta" in detalle_lower:
            carpeta_lower = str(carpeta or "").lower()
            id_hint = (
                "GOOGLE_DRIVE_SECRETARIA_FOLDER_ID con el ID exacto"
                if "secretaria" in carpeta_lower or "secretaría" in carpeta_lower
                else "el ID exacto de la carpeta"
            )
            return (
                f"{detalle} Revisá el nombre de la carpeta {carpeta} o configurá "
                f"{id_hint}."
            )
        if detalle:
            return detalle

    if HttpError is not None and isinstance(error, HttpError):
        status = getattr(getattr(error, "resp", None), "status", None)
        detalle = ""
        try:
            data = json.loads(error.content.decode("utf-8"))
            detalle = data.get("error", {}).get("message", "")
        except Exception:
            detalle = str(error)

        detalle_lower = detalle.lower()
        if "storage quota" in detalle_lower or "service accounts do not have storage quota" in detalle_lower:
            return (
                "Google Drive rechazó la subida porque la cuenta de servicio no puede "
                "guardar archivos en una carpeta de Mi unidad. Us? una unidad compartida "
                "de Google Drive o una integraci?n OAuth con un usuario real."
            )
        if status == 403:
            return (
                f"Google Drive rechazó el acceso. Revisá que la carpeta {carpeta} esté compartida "
                "como editor con la service account de Cloud Run."
            )
        if status == 404:
            return (
                f"Google Drive no encontr? la carpeta configurada. Revis? el ID de la carpeta {carpeta}."
            )
        if detalle:
            return f"Google Drive rechazó la operación: {truncate_audit_value(detalle, 180)}"

    return f"No se pudo {accion} en Google Drive."


def fecha_movimiento_default(mes=None):
    hoy = ahora_sig().strftime("%Y-%m-%d")
    if not mes:
        return hoy

    try:
        datetime.strptime(mes, "%Y-%m")
    except ValueError:
        return hoy

    if mes == hoy[:7]:
        return hoy

    return f"{mes}-01"


def validar_fecha_movimiento(fecha):
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except (TypeError, ValueError):
        return None
    return fecha


def validar_mes_beca(valor):
    valor = (valor or "").strip()
    if not valor:
        return ""
    try:
        datetime.strptime(valor, "%Y-%m")
    except ValueError:
        return None
    return valor


def porcentaje_beca(valor):
    try:
        porcentaje = float(valor or 0)
    except (TypeError, ValueError):
        return None
    if porcentaje < 0 or porcentaje > 100:
        return None
    return round(porcentaje, 2)


def datos_beca_form():
    activa = 1 if request.form.get("beca_activa") == "on" else 0

    if not activa:
        return {
            "beca_activa": 0,
            "beca_porcentaje": 0,
            "beca_desde": "",
            "beca_hasta": "",
            "beca_motivo": "",
        }, None

    porcentaje = porcentaje_beca(request.form.get("beca_porcentaje", "0"))
    desde = validar_mes_beca(request.form.get("beca_desde", ""))
    hasta = validar_mes_beca(request.form.get("beca_hasta", ""))
    motivo = request.form.get("beca_motivo", "").strip()
    if porcentaje is None:
        return None, "El porcentaje de beca debe estar entre 0 y 100."
    if desde is None or hasta is None:
        return None, "La vigencia de la beca debe tener formato YYYY-MM."
    if desde and hasta and hasta < desde:
        return None, "La fecha hasta de la beca no puede ser anterior a la fecha desde."
    if activa and porcentaje <= 0:
        return None, "Para activar una beca indic? un porcentaje mayor a 0."

    return {
        "beca_activa": activa,
        "beca_porcentaje": porcentaje,
        "beca_desde": desde,
        "beca_hasta": hasta,
        "beca_motivo": motivo,
    }, None


def beca_vigente(jugador, periodo):
    if not jugador or not jugador.get("beca_activa"):
        return False

    porcentaje = porcentaje_beca(jugador.get("beca_porcentaje"))
    if not porcentaje or porcentaje <= 0:
        return False

    periodo = (periodo or "").strip()
    try:
        datetime.strptime(periodo, "%Y-%m")
    except ValueError:
        return False

    desde = (jugador.get("beca_desde") or "").strip()
    hasta = (jugador.get("beca_hasta") or "").strip()
    if desde and periodo < desde:
        return False
    if hasta and periodo > hasta:
        return False
    return True


def calcular_importe_con_beca(jugador, periodo, importe_base):
    importe_original = round(float(importe_base or 0), 2)
    if not beca_vigente(jugador, periodo):
        return {
            "importe_original": importe_original,
            "importe": importe_original,
            "descuento_beca": 0,
            "beca_porcentaje": 0,
            "beca_motivo": "",
            "becada": 0,
            "beca_total": 0,
        }

    porcentaje = porcentaje_beca(jugador.get("beca_porcentaje")) or 0
    descuento = round(importe_original * porcentaje / 100, 2)
    importe = max(0, round(importe_original - descuento, 2))
    return {
        "importe_original": importe_original,
        "importe": importe,
        "descuento_beca": descuento,
        "beca_porcentaje": porcentaje,
        "beca_motivo": jugador.get("beca_motivo") or "",
        "becada": 1,
        "beca_total": 1 if importe <= 0 else 0,
    }


def indice_periodo(periodo):
    try:
        fecha = datetime.strptime(str(periodo or "")[:7], "%Y-%m")
    except ValueError:
        return None
    return fecha.year * 12 + fecha.month


def periodo_inicio_plan(plan):
    fecha = plan.get("fecha_inicio") or plan.get("creado_en") or ""
    if isinstance(fecha, datetime):
        return fecha.strftime("%Y-%m")
    texto = str(fecha or "").strip()
    return texto[:7] if re.match(r"^[0-9]{4}-[0-9]{2}", texto) else None


def periodo_minimo(*periodos):
    validos = [periodo for periodo in periodos if indice_periodo(periodo) is not None]
    if not validos:
        return None
    return min(validos, key=lambda periodo: indice_periodo(periodo))


def adicional_plan_pago_para_periodo(conn, jugador_id, periodo):
    periodo_idx = indice_periodo(periodo)
    if periodo_idx is None:
        return {"monto": 0, "detalle": ""}

    planes = conn.execute("""
        SELECT id, fecha_inicio, monto_total, cantidad_cuotas, monto_cuota,
               descripcion, creado_en
        FROM planes_pago
        WHERE jugador_id = %s
          AND estado = 'Activo'
          AND COALESCE(monto_total, 0) > 0
          AND COALESCE(cantidad_cuotas, 0) > 0
        ORDER BY fecha_inicio ASC, id ASC
    """, (jugador_id,)).fetchall()

    monto_total_periodo = 0
    detalles = []
    for plan in planes:
        inicio = periodo_inicio_plan(plan)
        inicio_idx = indice_periodo(inicio)
        cantidad = int(plan["cantidad_cuotas"] or 0)
        if inicio_idx is None or cantidad <= 0:
            continue

        numero_cuota = periodo_idx - inicio_idx + 1
        if numero_cuota < 1 or numero_cuota > cantidad:
            continue

        monto_plan = round(float(plan["monto_cuota"] or 0), 2)
        if numero_cuota == cantidad:
            monto_total = round(float(plan["monto_total"] or 0), 2)
            monto_plan = round(monto_total - (monto_plan * (cantidad - 1)), 2)
        if monto_plan <= 0:
            continue

        monto_total_periodo = round(monto_total_periodo + monto_plan, 2)
        etiqueta = f"Plan #{plan['id']} cuota {numero_cuota}/{cantidad}"
        if plan.get("descripcion"):
            etiqueta = f"{etiqueta} - {plan['descripcion']}"
        detalles.append(etiqueta)

    return {
        "monto": monto_total_periodo,
        "detalle": "; ".join(detalles),
    }


def calcular_importe_cuota_mensual(conn, jugador, periodo, importe_base):
    cuota = calcular_importe_con_beca(jugador, periodo, importe_base)
    adicional_plan = adicional_plan_pago_para_periodo(conn, jugador["id"], periodo)
    monto_plan = round(float(adicional_plan["monto"] or 0), 2)
    cuota["plan_pago_monto"] = monto_plan
    cuota["plan_pago_detalle"] = adicional_plan["detalle"]
    if monto_plan:
        cuota["importe_original"] = round(cuota["importe_original"] + monto_plan, 2)
        cuota["importe"] = round(cuota["importe"] + monto_plan, 2)
        cuota["beca_total"] = 1 if cuota["importe"] <= 0 else 0
    return cuota


def recalcular_cuotas_planes_pago(conn, jugador_id, periodo_desde=None):
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        return {"revisadas": 0, "actualizadas": 0}

    condiciones = ["jugador_id = %s", "pagado = 0", "COALESCE(anulada, 0) = 0"]
    parametros = [jugador_id]
    if periodo_desde:
        condiciones.append("periodo >= %s")
        parametros.append(periodo_desde)

    cuotas = conn.execute(f"""
        SELECT *
        FROM cuotas
        WHERE {" AND ".join(condiciones)}
        ORDER BY periodo ASC, id ASC
    """, parametros).fetchall()

    resultado = {"revisadas": len(cuotas), "actualizadas": 0}
    hoy = ahora_sig().strftime("%Y-%m-%d")

    for cuota in cuotas:
        plan_pago_monto = round(float(cuota.get("plan_pago_monto") or 0), 2)
        importe_base = cuota.get("importe_original")
        if importe_base is None:
            importe_base = cuota.get("importe") or 0
        importe_base = max(0, round(float(importe_base or 0) - plan_pago_monto, 2))

        cuota_calculada = calcular_importe_cuota_mensual(conn, jugador, cuota["periodo"], importe_base)
        pagado = 1 if cuota_calculada["beca_total"] else 0
        fecha_pago = hoy if pagado else cuota.get("fecha_pago")
        metodo_pago = "Beca" if pagado else cuota.get("metodo_pago")
        referencia_pago = (
            f"Beca total {cuota_calculada['beca_porcentaje']:g}%"
            if pagado else cuota.get("referencia_pago")
        )

        conn.execute("""
            UPDATE cuotas
            SET importe = %s,
                importe_original = %s,
                descuento_beca = %s,
                beca_porcentaje = %s,
                beca_motivo = %s,
                becada = %s,
                pagado = %s,
                fecha_pago = %s,
                metodo_pago = %s,
                referencia_pago = %s,
                plan_pago_monto = %s,
                plan_pago_detalle = %s
            WHERE id = %s
        """, (
            cuota_calculada["importe"],
            cuota_calculada["importe_original"],
            cuota_calculada["descuento_beca"],
            cuota_calculada["beca_porcentaje"],
            cuota_calculada["beca_motivo"],
            cuota_calculada["becada"],
            pagado,
            fecha_pago,
            metodo_pago,
            referencia_pago,
            cuota_calculada["plan_pago_monto"],
            cuota_calculada["plan_pago_detalle"] or None,
            cuota["id"],
        ))
        resultado["actualizadas"] += 1

    return resultado


def snapshot_beca(jugador):
    if not jugador:
        return {
            "beca_activa": 0,
            "beca_porcentaje": 0,
            "beca_desde": "",
            "beca_hasta": "",
            "beca_motivo": "",
        }
    return {
        "beca_activa": 1 if jugador.get("beca_activa") else 0,
        "beca_porcentaje": porcentaje_beca(jugador.get("beca_porcentaje")) or 0,
        "beca_desde": jugador.get("beca_desde") or "",
        "beca_hasta": jugador.get("beca_hasta") or "",
        "beca_motivo": jugador.get("beca_motivo") or "",
    }


def beca_modificada(jugador, data):
    anterior = snapshot_beca(jugador)
    nuevo = snapshot_beca(data)
    return anterior != nuevo


def registrar_historial_beca(conn, jugador_id, data, accion, detalle=None):
    snapshot = snapshot_beca(data)
    conn.execute("""
        INSERT INTO becas_historial (
            jugador_id, accion, beca_activa, beca_porcentaje, beca_desde,
            beca_hasta, beca_motivo, detalle, creado_en, creado_por
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s)
    """, (
        jugador_id,
        accion,
        snapshot["beca_activa"],
        snapshot["beca_porcentaje"],
        snapshot["beca_desde"] or None,
        snapshot["beca_hasta"] or None,
        snapshot["beca_motivo"] or None,
        json.dumps(detalle or {}, ensure_ascii=False, default=str),
        session.get("username") if has_request_context() else "sistema",
    ))


def recalcular_cuotas_becadas(conn, jugador, periodo_desde="", periodo_hasta=""):
    condiciones = ["jugador_id = %s", "pagado = 0", "COALESCE(anulada, 0) = 0"]
    parametros = [jugador["id"]]

    if periodo_desde:
        condiciones.append("periodo >= %s")
        parametros.append(periodo_desde)
    if periodo_hasta:
        condiciones.append("periodo <= %s")
        parametros.append(periodo_hasta)

    cuotas = conn.execute(f"""
        SELECT *
        FROM cuotas
        WHERE {" AND ".join(condiciones)}
        ORDER BY periodo ASC, id ASC
    """, parametros).fetchall()

    resultado = {
        "revisadas": len(cuotas),
        "actualizadas": 0,
        "becas_totales": 0,
        "becas_parciales": 0,
        "sin_beca": 0,
    }

    hoy = ahora_sig().strftime("%Y-%m-%d")
    for cuota in cuotas:
        plan_pago_monto = round(float(cuota.get("plan_pago_monto") or 0), 2)
        importe_base = cuota.get("importe_original")
        if importe_base is None:
            importe_base = cuota.get("importe") or 0
        importe_base = max(0, round(float(importe_base or 0) - plan_pago_monto, 2))

        cuota_calculada = calcular_importe_con_beca(jugador, cuota["periodo"], importe_base)
        if plan_pago_monto:
            cuota_calculada["importe_original"] = round(cuota_calculada["importe_original"] + plan_pago_monto, 2)
            cuota_calculada["importe"] = round(cuota_calculada["importe"] + plan_pago_monto, 2)
            cuota_calculada["beca_total"] = 1 if cuota_calculada["importe"] <= 0 else 0
        pagado = 1 if cuota_calculada["beca_total"] else 0
        fecha_pago = hoy if pagado else None
        metodo_pago = "Beca" if pagado else None
        referencia_pago = (
            f"Beca total {cuota_calculada['beca_porcentaje']:g}%"
            if pagado else None
        )

        conn.execute("""
            UPDATE cuotas
            SET importe = %s,
                importe_original = %s,
                descuento_beca = %s,
                beca_porcentaje = %s,
                beca_motivo = %s,
                becada = %s,
                pagado = %s,
                fecha_pago = %s,
                metodo_pago = %s,
                referencia_pago = %s
            WHERE id = %s
        """, (
            cuota_calculada["importe"],
            cuota_calculada["importe_original"],
            cuota_calculada["descuento_beca"],
            cuota_calculada["beca_porcentaje"],
            cuota_calculada["beca_motivo"],
            cuota_calculada["becada"],
            pagado,
            fecha_pago,
            metodo_pago,
            referencia_pago,
            cuota["id"],
        ))

        resultado["actualizadas"] += 1
        if cuota_calculada["beca_total"]:
            resultado["becas_totales"] += 1
        elif cuota_calculada["becada"]:
            resultado["becas_parciales"] += 1
        else:
            resultado["sin_beca"] += 1

    return resultado


def generar_portal_token():
    return secrets.token_urlsafe(32)


def normalizar_identificador_portal(valor):
    return "".join(ch for ch in str(valor or "").strip() if ch.isalnum() or ch in {"@", ".", "_", "-", "+"})


def normalizar_tipo_miembro(valor):
    valor = (valor or "Jugador").strip()
    return valor if valor in TIPOS_MIEMBRO else "Jugador"


def normalizar_mes(valor, default):
    valor = (valor or "").strip()
    try:
        datetime.strptime(valor, "%Y-%m")
        return valor
    except ValueError:
        return default


def meses_entre(desde, hasta):
    actual = datetime.strptime(desde, "%Y-%m")
    final = datetime.strptime(hasta, "%Y-%m")
    meses = []

    while actual <= final:
        meses.append(actual.strftime("%Y-%m"))
        if actual.month == 12:
            actual = actual.replace(year=actual.year + 1, month=1)
        else:
            actual = actual.replace(month=actual.month + 1)

    return meses


def sumar_meses(mes, cantidad):
    fecha = datetime.strptime(mes, "%Y-%m")
    total = fecha.year * 12 + fecha.month - 1 + cantidad
    year = total // 12
    month = total % 12 + 1
    return f"{year:04d}-{month:02d}"


def filtros_reportes():
    hoy = ahora_sig()
    default_desde = f"{hoy.year}-01"
    default_hasta = hoy.strftime("%Y-%m")

    desde = normalizar_mes(request.args.get("desde"), default_desde)
    hasta = normalizar_mes(request.args.get("hasta"), default_hasta)

    if desde > hasta:
        desde, hasta = hasta, desde

    return {
        "desde": desde,
        "hasta": hasta,
    }


def obtener_reportes(desde, hasta):
    conn = get_connection()

    resumen_caja = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
            COALESCE(SUM(CASE WHEN tipo = 'egreso' THEN monto ELSE 0 END), 0) AS egresos
        FROM movimientos
        WHERE COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) BETWEEN %s AND %s
    """, (desde, hasta)).fetchone()

    resumen_cuotas = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN pagado = 1 THEN importe ELSE 0 END), 0) AS cuotas_cobradas,
            COUNT(CASE WHEN pagado = 1 THEN 1 END) AS cuotas_pagadas,
            COUNT(CASE WHEN pagado = 0 AND COALESCE(importe, 0) > 0 THEN 1 END) AS cuotas_pendientes_periodo
        FROM cuotas
        WHERE periodo BETWEEN %s AND %s
    """, (desde, hasta)).fetchone()

    deuda_total = conn.execute("""
        SELECT
            COALESCE(SUM(importe), 0) AS deuda,
            COALESCE(SUM(
                CASE
                    WHEN fecha_vencimiento IS NOT NULL
                     AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
                     AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND fecha_vencimiento::date < CURRENT_DATE
                    THEN importe
                    ELSE 0
                END
            ), 0) AS deuda_vencida,
            COUNT(*) AS cuotas_pendientes
        FROM cuotas
        WHERE pagado = 0
          AND COALESCE(importe, 0) > 0
    """).fetchone()

    jugadores = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(CASE WHEN estado = 'Activo' THEN 1 END) AS activos
        FROM jugadores
    """).fetchone()

    asistencia_resumen = conn.execute("""
        SELECT
            COUNT(DISTINCT e.id) AS eventos,
            COUNT(a.id) AS registros,
            COALESCE(SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END), 0) AS presentes
        FROM eventos_asistencia e
        LEFT JOIN asistencias a ON a.evento_id = e.id
        WHERE substring(e.fecha from 1 for 7) BETWEEN %s AND %s
    """, (desde, hasta)).fetchone()

    movimientos_mensuales = conn.execute("""
        SELECT
            substring(fecha from 1 for 7) AS mes,
            COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
            COALESCE(SUM(CASE WHEN tipo = 'egreso' THEN monto ELSE 0 END), 0) AS egresos,
            COUNT(*) AS movimientos
        FROM movimientos
        WHERE COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) BETWEEN %s AND %s
        GROUP BY substring(fecha from 1 for 7)
        ORDER BY mes ASC
    """, (desde, hasta)).fetchall()

    cuotas_mensuales = conn.execute("""
        SELECT
            periodo AS mes,
            COUNT(*) AS cuotas_emitidas,
            COUNT(CASE WHEN pagado = 1 THEN 1 END) AS cuotas_pagadas,
            COUNT(CASE WHEN pagado = 0 AND COALESCE(importe, 0) > 0 THEN 1 END) AS cuotas_pendientes,
            COALESCE(SUM(importe), 0) AS total_emitido,
            COALESCE(SUM(CASE WHEN pagado = 1 THEN importe ELSE 0 END), 0) AS total_cobrado
        FROM cuotas
        WHERE periodo BETWEEN %s AND %s
        GROUP BY periodo
        ORDER BY periodo ASC
    """, (desde, hasta)).fetchall()

    becas_resumen = conn.execute("""
        SELECT
            COUNT(CASE WHEN COALESCE(c.becada, 0) = 1 THEN 1 END) AS cuotas_becadas,
            COUNT(CASE WHEN COALESCE(c.becada, 0) = 1 AND COALESCE(c.importe, 0) = 0 THEN 1 END) AS becas_totales,
            COUNT(CASE WHEN COALESCE(c.becada, 0) = 1 AND COALESCE(c.importe, 0) > 0 THEN 1 END) AS becas_parciales,
            COALESCE(SUM(CASE WHEN COALESCE(c.becada, 0) = 1 THEN c.descuento_beca ELSE 0 END), 0) AS total_bonificado,
            COALESCE(SUM(CASE WHEN COALESCE(c.becada, 0) = 1 THEN COALESCE(c.importe_original, c.importe) ELSE 0 END), 0) AS total_original_becado,
            COALESCE(SUM(CASE WHEN COALESCE(c.becada, 0) = 1 THEN c.importe ELSE 0 END), 0) AS total_neto_becado
        FROM cuotas c
        WHERE c.periodo BETWEEN %s AND %s
    """, (desde, hasta)).fetchone()

    becas_mensuales = conn.execute("""
        SELECT
            periodo AS mes,
            COUNT(*) AS cuotas_becadas,
            COUNT(CASE WHEN COALESCE(importe, 0) = 0 THEN 1 END) AS becas_totales,
            COUNT(CASE WHEN COALESCE(importe, 0) > 0 THEN 1 END) AS becas_parciales,
            COALESCE(SUM(descuento_beca), 0) AS total_bonificado
        FROM cuotas
        WHERE periodo BETWEEN %s AND %s
          AND COALESCE(becada, 0) = 1
        GROUP BY periodo
        ORDER BY periodo ASC
    """, (desde, hasta)).fetchall()

    becas_jugadores = conn.execute("""
        SELECT
            j.id,
            j.apellido,
            j.nombre,
            j.categoria,
            j.beca_porcentaje,
            j.beca_desde,
            j.beca_hasta,
            j.beca_motivo,
            COUNT(c.id) AS cuotas_becadas,
            COALESCE(SUM(c.descuento_beca), 0) AS total_bonificado
        FROM jugadores j
        LEFT JOIN cuotas c
          ON c.jugador_id = j.id
         AND c.periodo BETWEEN %s AND %s
         AND COALESCE(c.becada, 0) = 1
        WHERE COALESCE(j.beca_activa, 0) = 1
        GROUP BY
            j.id, j.apellido, j.nombre, j.categoria, j.beca_porcentaje,
            j.beca_desde, j.beca_hasta, j.beca_motivo
        ORDER BY j.apellido, j.nombre
    """, (desde, hasta)).fetchall()

    deuda_por_categoria = conn.execute("""
        SELECT
            COALESCE(NULLIF(j.categoria, ''), 'Sin categoria') AS categoria,
            COUNT(DISTINCT j.id) AS jugadores,
            COUNT(c.id) AS cuotas_pendientes,
            COALESCE(SUM(c.importe), 0) AS deuda,
            COALESCE(SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN c.importe
                    ELSE 0
                END
            ), 0) AS deuda_vencida
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY COALESCE(NULLIF(j.categoria, ''), 'Sin categoria')
        ORDER BY deuda DESC, categoria ASC
    """).fetchall()

    egresos_por_concepto = conn.execute("""
        SELECT
            COALESCE(NULLIF(concepto, ''), 'Sin concepto') AS concepto,
            COUNT(*) AS cantidad,
            COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'egreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) BETWEEN %s AND %s
        GROUP BY COALESCE(NULLIF(concepto, ''), 'Sin concepto')
        ORDER BY total DESC, concepto ASC
        LIMIT 15
    """, (desde, hasta)).fetchall()

    morosos_recurrentes = conn.execute("""
        SELECT
            j.id,
            j.apellido,
            j.nombre,
            j.categoria,
            j.telefono,
            j.email,
            COUNT(c.id) AS cuotas_pendientes,
            SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1
                    ELSE 0
                END
            ) AS cuotas_vencidas,
            COALESCE(SUM(c.importe), 0) AS deuda,
            MIN(NULLIF(c.fecha_vencimiento::text, '')) AS primer_vencimiento
        FROM jugadores j
        JOIN cuotas c ON c.jugador_id = j.id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY j.id, j.apellido, j.nombre, j.categoria, j.telefono, j.email
        HAVING COUNT(c.id) >= 2
        ORDER BY deuda DESC, cuotas_vencidas DESC, j.apellido, j.nombre
        LIMIT 20
    """).fetchall()

    asistencia_por_categoria = conn.execute("""
        SELECT
            COALESCE(NULLIF(j.categoria, ''), 'Sin categoria') AS categoria,
            COUNT(DISTINCT e.id) AS eventos,
            COUNT(a.id) AS registros,
            COALESCE(SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END), 0) AS presentes,
            COALESCE(SUM(CASE WHEN a.presente = 0 THEN 1 ELSE 0 END), 0) AS ausentes
        FROM eventos_asistencia e
        JOIN asistencias a ON a.evento_id = e.id
        JOIN jugadores j ON j.id = a.jugador_id
        WHERE substring(e.fecha from 1 for 7) BETWEEN %s AND %s
        GROUP BY COALESCE(NULLIF(j.categoria, ''), 'Sin categoria')
        ORDER BY categoria ASC
    """, (desde, hasta)).fetchall()

    gastos_compartidos_resumen = conn.execute("""
        SELECT
            COUNT(DISTINCT g.id) AS gastos,
            COUNT(i.id) FILTER (WHERE i.estado = 'pendiente') AS items_pendientes,
            COUNT(i.id) FILTER (WHERE i.estado = 'pagado') AS items_pagados,
            COALESCE(SUM(i.importe) FILTER (WHERE i.estado = 'pendiente'), 0) AS pendiente,
            COALESCE(SUM(i.importe) FILTER (WHERE i.estado = 'pagado'), 0) AS cobrado
        FROM gastos_compartidos g
        LEFT JOIN gasto_compartido_items i ON i.gasto_id = g.id
        WHERE substring(COALESCE(g.fecha_evento::text, g.fecha_vencimiento::text, g.creado_en::text) from 1 for 7) BETWEEN %s AND %s
    """, (desde, hasta)).fetchone()

    antiguedad_deuda = conn.execute("""
        SELECT
            COALESCE(SUM(importe) FILTER (WHERE dias <= 0), 0) AS por_vencer,
            COALESCE(SUM(importe) FILTER (WHERE dias BETWEEN 1 AND 30), 0) AS dias_1_30,
            COALESCE(SUM(importe) FILTER (WHERE dias BETWEEN 31 AND 60), 0) AS dias_31_60,
            COALESCE(SUM(importe) FILTER (WHERE dias BETWEEN 61 AND 90), 0) AS dias_61_90,
            COALESCE(SUM(importe) FILTER (WHERE dias > 90), 0) AS mas_90
        FROM (
            SELECT c.importe,
                CASE WHEN c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     THEN CURRENT_DATE - c.fecha_vencimiento::date ELSE 0 END AS dias
            FROM cuotas c
            WHERE c.pagado = 0
              AND COALESCE(c.anulada, 0) = 0
              AND COALESCE(c.importe, 0) > 0
            UNION ALL
            SELECT i.importe,
                CASE WHEN g.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     THEN CURRENT_DATE - g.fecha_vencimiento::date ELSE 0 END AS dias
            FROM gasto_compartido_items i
            JOIN gastos_compartidos g ON g.id = i.gasto_id
            WHERE i.estado = 'pendiente'
              AND COALESCE(i.importe, 0) > 0
        ) deuda
    """).fetchone()

    conn.close()

    movimientos_por_mes = {fila["mes"]: fila for fila in movimientos_mensuales}
    cuotas_por_mes = {fila["mes"]: fila for fila in cuotas_mensuales}
    becas_por_mes = {fila["mes"]: fila for fila in becas_mensuales}
    mensual = []

    for mes in meses_entre(desde, hasta):
        caja = movimientos_por_mes.get(mes, {})
        cuotas = cuotas_por_mes.get(mes, {})
        becas = becas_por_mes.get(mes, {})
        ingresos = caja.get("ingresos", 0) or 0
        egresos = caja.get("egresos", 0) or 0

        mensual.append({
            "mes": mes,
            "ingresos": ingresos,
            "egresos": egresos,
            "resultado": ingresos - egresos,
            "movimientos": caja.get("movimientos", 0) or 0,
            "cuotas_emitidas": cuotas.get("cuotas_emitidas", 0) or 0,
            "cuotas_pagadas": cuotas.get("cuotas_pagadas", 0) or 0,
            "cuotas_pendientes": cuotas.get("cuotas_pendientes", 0) or 0,
            "total_emitido": cuotas.get("total_emitido", 0) or 0,
            "total_cobrado": cuotas.get("total_cobrado", 0) or 0,
            "cuotas_becadas": becas.get("cuotas_becadas", 0) or 0,
            "becas_totales": becas.get("becas_totales", 0) or 0,
            "becas_parciales": becas.get("becas_parciales", 0) or 0,
            "total_bonificado": becas.get("total_bonificado", 0) or 0,
        })

    asistencia_registros = asistencia_resumen["registros"] or 0
    asistencia_presentes = asistencia_resumen["presentes"] or 0
    asistencia_porcentaje = (
        round((asistencia_presentes / asistencia_registros) * 100, 1)
        if asistencia_registros else 0
    )

    for fila in asistencia_por_categoria:
        registros = fila["registros"] or 0
        presentes = fila["presentes"] or 0
        fila["porcentaje"] = round((presentes / registros) * 100, 1) if registros else 0

    ingresos = resumen_caja["ingresos"] or 0
    egresos = resumen_caja["egresos"] or 0

    return {
        "resumen": {
            "ingresos": ingresos,
            "egresos": egresos,
            "resultado": ingresos - egresos,
            "cuotas_cobradas": resumen_cuotas["cuotas_cobradas"] or 0,
            "cuotas_pagadas": resumen_cuotas["cuotas_pagadas"] or 0,
            "cuotas_pendientes_periodo": resumen_cuotas["cuotas_pendientes_periodo"] or 0,
            "deuda": deuda_total["deuda"] or 0,
            "deuda_vencida": deuda_total["deuda_vencida"] or 0,
            "cuotas_pendientes": deuda_total["cuotas_pendientes"] or 0,
            "cuotas_becadas": becas_resumen["cuotas_becadas"] or 0,
            "becas_totales": becas_resumen["becas_totales"] or 0,
            "becas_parciales": becas_resumen["becas_parciales"] or 0,
            "total_bonificado_becas": becas_resumen["total_bonificado"] or 0,
            "total_original_becado": becas_resumen["total_original_becado"] or 0,
            "total_neto_becado": becas_resumen["total_neto_becado"] or 0,
            "jugadores_total": jugadores["total"] or 0,
            "jugadores_activos": jugadores["activos"] or 0,
            "asistencia_eventos": asistencia_resumen["eventos"] or 0,
            "asistencia_presentes": asistencia_presentes,
            "asistencia_registros": asistencia_registros,
            "asistencia_porcentaje": asistencia_porcentaje,
            "gastos_compartidos": gastos_compartidos_resumen["gastos"] or 0,
            "gastos_compartidos_cobrado": gastos_compartidos_resumen["cobrado"] or 0,
            "gastos_compartidos_pendiente": gastos_compartidos_resumen["pendiente"] or 0,
            "gastos_compartidos_items_pagados": gastos_compartidos_resumen["items_pagados"] or 0,
            "gastos_compartidos_items_pendientes": gastos_compartidos_resumen["items_pendientes"] or 0,
            "deuda_total_unificada": sum(
                antiguedad_deuda[clave] or 0
                for clave in ("por_vencer", "dias_1_30", "dias_31_60", "dias_61_90", "mas_90")
            ),
        },
        "mensual": mensual,
        "deuda_por_categoria": deuda_por_categoria,
        "egresos_por_concepto": egresos_por_concepto,
        "morosos_recurrentes": morosos_recurrentes,
        "asistencia_por_categoria": asistencia_por_categoria,
        "becas_jugadores": becas_jugadores,
        "antiguedad_deuda": antiguedad_deuda,
    }


def obtener_alertas():
    conn = get_connection()

    cuotas_vencidas = conn.execute("""
        SELECT
            c.id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            CURRENT_DATE - c.fecha_vencimiento::date AS dias_vencida,
            j.apellido,
            j.nombre,
            j.categoria
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
          AND c.fecha_vencimiento IS NOT NULL
          AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
          AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND c.fecha_vencimiento::date < CURRENT_DATE
        ORDER BY c.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 25
    """).fetchall()

    cuotas_por_vencer = conn.execute("""
        SELECT
            c.id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            c.fecha_vencimiento::date - CURRENT_DATE AS dias_restantes,
            j.apellido,
            j.nombre,
            j.categoria
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
          AND c.fecha_vencimiento IS NOT NULL
          AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
          AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND c.fecha_vencimiento::date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days'
        ORDER BY c.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 25
    """).fetchall()

    fichas_vencidas = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.categoria,
            f.fecha_vencimiento,
            CURRENT_DATE - f.fecha_vencimiento::date AS dias_vencida
        FROM fichas_medicas f
        JOIN jugadores j ON j.id = f.jugador_id
        WHERE f.fecha_vencimiento IS NOT NULL
          AND NULLIF(f.fecha_vencimiento::text, '') IS NOT NULL
          AND f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND f.fecha_vencimiento::date < CURRENT_DATE
        ORDER BY f.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 25
    """).fetchall()

    fichas_por_vencer = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.categoria,
            f.fecha_vencimiento,
            f.fecha_vencimiento::date - CURRENT_DATE AS dias_restantes
        FROM fichas_medicas f
        JOIN jugadores j ON j.id = f.jugador_id
        WHERE f.fecha_vencimiento IS NOT NULL
          AND NULLIF(f.fecha_vencimiento::text, '') IS NOT NULL
          AND f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND f.fecha_vencimiento::date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days'
        ORDER BY f.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 25
    """).fetchall()

    lesiones_activas = conn.execute("""
        SELECT
            l.id,
            l.jugador_id,
            l.fecha_lesion,
            l.tipo_lesion,
            l.zona_cuerpo,
            l.estado,
            j.apellido,
            j.nombre,
            j.categoria
        FROM lesiones l
        JOIN jugadores j ON j.id = l.jugador_id
        WHERE l.estado IN ('Activa', 'En recuperaci?n')
        ORDER BY
            CASE
                WHEN l.estado = 'Activa' THEN 0
                ELSE 1
            END,
            l.fecha_lesion DESC,
            j.apellido,
            j.nombre
        LIMIT 25
    """).fetchall()

    meses_sin_cerrar = conn.execute("""
        WITH mensual AS (
            SELECT
                substring(m.fecha from 1 for 7) AS mes,
                COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                COALESCE(SUM(CASE WHEN m.tipo = 'egreso' THEN m.monto ELSE 0 END), 0) AS egresos,
                COUNT(*) AS movimientos
            FROM movimientos m
            WHERE COALESCE(m.anulado, 0) = 0
              AND substring(m.fecha from 1 for 7) < to_char(CURRENT_DATE, 'YYYY-MM')
            GROUP BY substring(m.fecha from 1 for 7)
        )
        SELECT *
        FROM mensual
        WHERE NOT EXISTS (
            SELECT 1
            FROM cierres_mensuales c
            WHERE c.mes = mensual.mes
        )
        ORDER BY mes DESC
        LIMIT 12
    """).fetchall()

    movimientos_altos = conn.execute("""
        WITH stats AS (
            SELECT COALESCE(AVG(monto), 0) AS promedio
            FROM movimientos
            WHERE COALESCE(anulado, 0) = 0
              AND fecha ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
              AND fecha::date >= CURRENT_DATE - INTERVAL '90 days'
        )
        SELECT
            m.id,
            m.fecha,
            m.tipo,
            m.concepto,
            m.referencia,
            m.monto,
            stats.promedio
        FROM movimientos m
        CROSS JOIN stats
        WHERE COALESCE(m.anulado, 0) = 0
          AND m.fecha ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND m.fecha::date >= CURRENT_DATE - INTERVAL '90 days'
          AND stats.promedio > 0
          AND m.monto >= stats.promedio * 2
        ORDER BY m.monto DESC, m.fecha DESC
        LIMIT 10
    """).fetchall()

    conn.close()

    for fila in meses_sin_cerrar:
        fila["resultado"] = (fila["ingresos"] or 0) - (fila["egresos"] or 0)

    deuda_vencida = sum((fila["importe"] or 0) for fila in cuotas_vencidas)
    deuda_proxima = sum((fila["importe"] or 0) for fila in cuotas_por_vencer)

    criticas = (
        len(cuotas_vencidas)
        + len(fichas_vencidas)
        + len(meses_sin_cerrar)
        + len(movimientos_altos)
    )

    proximas = len(cuotas_por_vencer) + len(fichas_por_vencer)
    seguimiento = len(lesiones_activas)

    return {
        "resumen": {
            "criticas": criticas,
            "proximas": proximas,
            "seguimiento": seguimiento,
            "total": criticas + proximas + seguimiento,
            "deuda_vencida": deuda_vencida,
            "deuda_proxima": deuda_proxima,
        },
        "cuotas_vencidas": cuotas_vencidas,
        "cuotas_por_vencer": cuotas_por_vencer,
        "fichas_vencidas": fichas_vencidas,
        "fichas_por_vencer": fichas_por_vencer,
        "lesiones_activas": lesiones_activas,
        "meses_sin_cerrar": meses_sin_cerrar,
        "movimientos_altos": movimientos_altos,
    }


def filtrar_alertas_por_permisos(alertas, puede_ver_finanzas, puede_ver_salud):
    if not puede_ver_finanzas:
        alertas["cuotas_vencidas"] = []
        alertas["cuotas_por_vencer"] = []
        alertas["meses_sin_cerrar"] = []
        alertas["movimientos_altos"] = []

    if not puede_ver_salud:
        alertas["fichas_vencidas"] = []
        alertas["fichas_por_vencer"] = []
        alertas["lesiones_activas"] = []

    deuda_vencida = sum((fila["importe"] or 0) for fila in alertas["cuotas_vencidas"])
    deuda_proxima = sum((fila["importe"] or 0) for fila in alertas["cuotas_por_vencer"])
    criticas = (
        len(alertas["cuotas_vencidas"])
        + len(alertas["fichas_vencidas"])
        + len(alertas["meses_sin_cerrar"])
        + len(alertas["movimientos_altos"])
    )
    proximas = len(alertas["cuotas_por_vencer"]) + len(alertas["fichas_por_vencer"])
    seguimiento = len(alertas["lesiones_activas"])

    alertas["resumen"] = {
        "criticas": criticas,
        "proximas": proximas,
        "seguimiento": seguimiento,
        "total": criticas + proximas + seguimiento,
        "deuda_vencida": deuda_vencida,
        "deuda_proxima": deuda_proxima,
    }
    return alertas


def obtener_panel_salud():
    conn = get_connection()

    resumen = conn.execute("""
        SELECT
            COUNT(*) AS jugadores_activos,
            SUM(CASE WHEN f.id IS NULL OR COALESCE(f.presentada, 0) = 0 THEN 1 ELSE 0 END) AS fichas_faltantes,
            SUM(CASE WHEN f.id IS NOT NULL AND (COALESCE(f.apto_fisico, 0) = 1 OR NULLIF(COALESCE(f.documento_drive_file_id, ''), '') IS NOT NULL) THEN 1 ELSE 0 END) AS aptos,
            SUM(CASE WHEN f.id IS NOT NULL AND COALESCE(f.presentada, 0) = 1 AND NOT (COALESCE(f.apto_fisico, 0) = 1 OR NULLIF(COALESCE(f.documento_drive_file_id, ''), '') IS NOT NULL) THEN 1 ELSE 0 END) AS no_aptos,
            SUM(
                CASE
                    WHEN f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND f.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1 ELSE 0
                END
            ) AS vencidas,
            SUM(
                CASE
                    WHEN f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND f.fecha_vencimiento::date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days'
                    THEN 1 ELSE 0
                END
            ) AS por_vencer
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.estado = 'Activo'
    """).fetchone()

    fichas_por_categoria = conn.execute("""
        SELECT
            COALESCE(NULLIF(j.categoria, ''), 'Sin categoria') AS categoria,
            COUNT(*) AS jugadores,
            SUM(CASE WHEN f.id IS NOT NULL AND (COALESCE(f.apto_fisico, 0) = 1 OR NULLIF(COALESCE(f.documento_drive_file_id, ''), '') IS NOT NULL) THEN 1 ELSE 0 END) AS aptos,
            SUM(CASE WHEN f.id IS NOT NULL AND COALESCE(f.presentada, 0) = 1 AND NOT (COALESCE(f.apto_fisico, 0) = 1 OR NULLIF(COALESCE(f.documento_drive_file_id, ''), '') IS NOT NULL) THEN 1 ELSE 0 END) AS no_aptos,
            SUM(CASE WHEN f.id IS NULL OR COALESCE(f.presentada, 0) = 0 THEN 1 ELSE 0 END) AS faltantes,
            SUM(
                CASE
                    WHEN f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND f.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1 ELSE 0
                END
            ) AS vencidas,
            SUM(
                CASE
                    WHEN f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND f.fecha_vencimiento::date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days'
                    THEN 1 ELSE 0
                END
            ) AS por_vencer
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.estado = 'Activo'
        GROUP BY COALESCE(NULLIF(j.categoria, ''), 'Sin categoria')
        ORDER BY categoria
    """).fetchall()

    lesiones_por_estado = conn.execute("""
        SELECT estado, COUNT(*) AS cantidad
        FROM lesiones
        GROUP BY estado
        ORDER BY
            CASE estado
                WHEN 'Activa' THEN 0
                ELSE CASE WHEN estado ILIKE 'En recuperaci%%' THEN 1 WHEN estado ILIKE 'Alta%%' THEN 2 ELSE 3 END
            END,
            estado
    """).fetchall()

    seguimiento_retorno = conn.execute("""
        SELECT
            l.id,
            l.jugador_id,
            l.fecha_lesion,
            l.fecha_alta,
            l.etapa_recuperacion,
            l.proximo_control,
            l.fecha_retorno_estimada,
            l.tratamiento_hasta,
            l.tipo_lesion,
            l.zona_cuerpo,
            l.estado,
            j.apellido,
            j.nombre,
            j.categoria
        FROM lesiones l
        JOIN jugadores j ON j.id = l.jugador_id
        WHERE l.estado = 'Activa'
           OR l.estado ILIKE 'En recuperaci%%'
        ORDER BY
            CASE WHEN NULLIF(l.fecha_alta::text, '') IS NULL THEN 1 ELSE 0 END,
            CASE
                WHEN l.fecha_alta::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                THEN l.fecha_alta::date
                ELSE NULL
            END ASC NULLS LAST,
            l.fecha_lesion DESC
        LIMIT 40
    """).fetchall()

    documentos_por_vencer = conn.execute("""
        SELECT
            d.id,
            d.jugador_id,
            d.tipo,
            d.nombre,
            d.fecha_vencimiento,
            j.apellido,
            j.nombre,
            j.categoria
        FROM documentos_jugadores d
        JOIN jugadores j ON j.id = d.jugador_id
        WHERE d.fecha_vencimiento IS NOT NULL
          AND NULLIF(d.fecha_vencimiento::text, '') IS NOT NULL
          AND d.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND d.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
        ORDER BY d.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 40
    """).fetchall()

    conn.close()

    seguimiento_retorno = [dict(lesion) for lesion in seguimiento_retorno]
    for lesion in seguimiento_retorno:
        lesion["semaforo"] = semaforo_lesion(lesion)

    return {
        "resumen": resumen,
        "fichas_por_categoria": fichas_por_categoria,
        "lesiones_por_estado": lesiones_por_estado,
        "seguimiento_retorno": seguimiento_retorno,
        "documentos_por_vencer": documentos_por_vencer,
    }


def obtener_estado_sistema_admin():
    estado = {
        "db_ok": False,
        "db_error": None,
        "db_time": None,
        "conteos": {},
        "mantenimiento": {
            "activo": False,
            "mensaje": MAINTENANCE_DEFAULT_MESSAGE,
            "actualizado_en": None,
            "actualizado_por": None,
        },
        "cloud_run": {
            "service": os.environ.get("K_SERVICE", "local"),
            "revision": os.environ.get("K_REVISION", APP_VERSION),
            "configuration": os.environ.get("K_CONFIGURATION", "local"),
            "project": os.environ.get("GOOGLE_CLOUD_PROJECT", ""),
            "region": os.environ.get("CLOUD_RUN_REGION", "us-central1"),
            "version": APP_VERSION,
        },
        "backup": {
            "nivel": "Cloud SQL automatico",
            "retencion_dias": CLOUD_SQL_BACKUP_RETENTION_DAYS,
            "pitr_dias": CLOUD_SQL_PITR_DAYS,
            "ventana": CLOUD_SQL_BACKUP_WINDOW,
            "ultimo_backup": os.environ.get("CLOUD_SQL_LAST_BACKUP", "Visible desde Google Cloud SQL"),
            "cloud_sql": obtener_info_backups_cloud_sql(),
        },
        "integraciones": {
            "smtp_ok": smtp_configurado(),
            "smtp_from": SMTP_FROM or "",
            "smtp_user": SMTP_USER or "",
            "facturas_email_ok": factura_email_configurado(),
            "facturas_email_user": obtener_factura_email_config().get("user") or "",
            "drive_shared": bool(DRIVE_SHARED_DRIVE_ID),
            "drive_comprobantes": bool(DRIVE_COMPROBANTES_FOLDER_ID or DRIVE_SHARED_DRIVE_ID),
            "drive_secretaria": bool(DRIVE_SECRETARIA_FOLDER_ID or DRIVE_COMPROBANTES_FOLDER_ID or DRIVE_SHARED_DRIVE_ID),
            "drive_fichas": bool(
                DRIVE_FICHAS_MEDICAS_FOLDER_ID
                or DRIVE_COMPROBANTES_FOLDER_ID
                or DRIVE_SHARED_DRIVE_ID
            ),
            "cloud_sql_ok": bool(CLOUD_SQL_CONNECTION_NAME or (CLOUD_SQL_PROJECT and CLOUD_SQL_INSTANCE)),
            "whatsapp_ok": bool(WHATSAPP_ACCESS_TOKEN and WHATSAPP_PHONE_NUMBER_ID),
            "automation_token_ok": bool(os.environ.get("AUTOMATION_TOKEN", "").strip()),
            "secret_key_default": app.secret_key == "cambiar-esto-por-una-clave-segura",
        },
        "automatizaciones": {},
    }

    conn = None
    try:
        conn = get_connection()
        estado["db_ok"] = True
        estado["db_time"] = conn.execute("SELECT CURRENT_TIMESTAMP AS ahora").fetchone()["ahora"]
        estado["mantenimiento"] = obtener_config_mantenimiento(conn)
        estado["automatizaciones"] = obtener_config_automatizaciones(conn)
        estado["conteos"] = conn.execute("""
            SELECT
                (SELECT COUNT(*) FROM jugadores) AS jugadores,
                (SELECT COUNT(*) FROM cuotas WHERE pagado = 0 AND COALESCE(importe, 0) > 0) AS cuotas_pendientes,
                (
                    SELECT COUNT(*)
                    FROM cuotas
                    WHERE comprobante_drive_file_id IS NOT NULL
                      AND COALESCE(anulada, 0) = 0
                      AND COALESCE(NULLIF(comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
                ) AS comprobantes_pendientes,
                (
                    SELECT COUNT(*)
                    FROM fichas_medicas
                    WHERE fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                      AND fecha_vencimiento::date < CURRENT_DATE
                ) AS fichas_vencidas,
                (SELECT COUNT(*) FROM auditoria) AS auditoria_registros
        """).fetchone()
    except Exception as error:
        app.logger.exception("No se pudo obtener el estado del sistema.")
        estado["db_error"] = str(error)
    finally:
        if conn is not None:
            conn.close()

    return estado


def rango_mes(mes):
    inicio = datetime.strptime(mes, "%Y-%m")
    if inicio.month == 12:
        siguiente = inicio.replace(year=inicio.year + 1, month=1)
    else:
        siguiente = inicio.replace(month=inicio.month + 1)

    return inicio.strftime("%Y-%m-%d"), siguiente.strftime("%Y-%m-%d")


def normalizar_hora_evento(valor):
    valor = (valor or "").strip()
    if not valor:
        return ""
    for formato in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(valor, formato).strftime("%H:%M")
        except ValueError:
            continue
    return ""


def normalizar_duracion_evento(valor, default=90):
    try:
        duracion = int(valor or default)
    except (TypeError, ValueError):
        duracion = default
    return min(max(duracion, 15), 24 * 60)


def calendario_evento_es_deportivo(tipo):
    return (tipo or "").strip() in CALENDARIO_DEPORTIVO_TIPOS


def calendario_evento_requiere_asistencia(tipo):
    return (tipo or "").strip() in CALENDARIO_ASISTENCIA_TIPOS


def normalizar_categoria_calendario(categoria):
    return (categoria or "").strip()


def categoria_evento_aplica(categoria_evento, categoria_jugador):
    categoria_evento = normalizar_categoria_calendario(categoria_evento).lower()
    if not categoria_evento or categoria_evento in {"todo", "todos", "todo el club", "club", "general"}:
        return True
    categoria_jugador = normalizar_categoria_calendario(categoria_jugador).lower()
    return bool(categoria_jugador and categoria_jugador in categoria_evento)


def fecha_hora_evento(evento):
    fecha = evento.get("fecha")
    hora = normalizar_hora_evento(evento.get("hora_inicio") or evento.get("hora"))
    if hora:
        return f"{fecha} {hora}"
    return fecha


def formato_fecha_hora_evento(evento):
    fecha = evento.get("fecha") or ""
    hora = normalizar_hora_evento(evento.get("hora_inicio") or evento.get("hora"))
    return f"{fecha} {hora}" if hora else fecha


def crear_evento_asistencia_desde_calendario(conn, data):
    descripcion = data.get("descripcion") or data.get("titulo") or ""
    if data.get("ubicacion"):
        descripcion = f"{descripcion}\nLugar: {data['ubicacion']}".strip()
    if data.get("categoria"):
        descripcion = f"{descripcion}\nCategoria: {data['categoria']}".strip()

    fila = conn.execute("""
        INSERT INTO eventos_asistencia (fecha, tipo, descripcion)
        VALUES (%s, %s, %s)
        RETURNING id
    """, (data["fecha"], data["tipo"], descripcion)).fetchone()
    return fila["id"]


def obtener_eventos_deportivos_portal(jugador, limit=8):
    conn = get_connection()
    eventos = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE COALESCE(publicar_portal, 0) = 1
          AND fecha >= CURRENT_DATE::text
        ORDER BY fecha ASC, COALESCE(hora_inicio, '') ASC, id ASC
        LIMIT 80
    """).fetchall()
    conn.close()

    filtrados = []
    for evento in eventos:
        if categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria")):
            filtrados.append(evento)
        if len(filtrados) >= limit:
            break
    return filtrados


def horas_sueno_score(valor):
    return BIENESTAR_HORAS_SCORE.get((valor or "").strip(), 0)


def resumen_bienestar_confirmacion(confirmacion):
    if not confirmacion or confirmacion.get("sueno_calidad") is None:
        return None
    valores = []
    for clave in ("sueno_calidad", "doms", "fatiga", "estres", "animo", "motivacion", "recuperacion"):
        try:
            valores.append(int(confirmacion.get(clave) or 0))
        except (TypeError, ValueError):
            valores.append(0)
    horas_score = horas_sueno_score(confirmacion.get("horas_sueno"))
    if horas_score:
        valores.append(horas_score)
    promedio = round(sum(valores) / len([v for v in valores if v]), 1) if any(valores) else 0
    zonas = confirmacion.get("dolor_zonas_lista") or []
    zonas_alerta = [zona for zona in zonas if zona and zona not in {"No", "Otro"}]
    if confirmacion.get("dolor_otro"):
        zonas_alerta.append("Otro")
    if promedio and promedio < 2.6 or len(zonas_alerta) >= 2:
        nivel = "danger"
        label = "Alerta roja"
    elif promedio and promedio < 3.6 or len(zonas_alerta) == 1:
        nivel = "warning"
        label = "Alerta amarilla"
    else:
        nivel = "success"
        label = "Bienestar ok"
    return {
        "promedio": promedio,
        "nivel": nivel,
        "label": label,
        "dolores": zonas_alerta,
    }


def guardar_confirmacion_portal_sin_bienestar(conn, evento, jugador, estado):
    conn.execute("""
        INSERT INTO portal_asistencia_confirmaciones (
            evento_id, jugador_id, estado, sueno_calidad, horas_sueno, doms, fatiga, estres, animo,
            motivacion, recuperacion, dolor_zonas, dolor_otro, comentarios, creado_en, actualizado_en
        )
        VALUES (%s, %s, %s, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        ON CONFLICT (evento_id, jugador_id)
        DO UPDATE SET
            estado = excluded.estado,
            sueno_calidad = NULL,
            horas_sueno = NULL,
            doms = NULL,
            fatiga = NULL,
            estres = NULL,
            animo = NULL,
            motivacion = NULL,
            recuperacion = NULL,
            dolor_zonas = NULL,
            dolor_otro = NULL,
            comentarios = NULL,
            actualizado_en = CURRENT_TIMESTAMP
    """, (evento["asistencia_evento_id"], jugador["id"], estado))


def obtener_confirmaciones_portal(conn, evento_ids, jugador_id=None):
    if not evento_ids:
        return {}
    filtros = ["evento_id = ANY(%s)"]
    params = [evento_ids]
    if jugador_id is not None:
        filtros.append("jugador_id = %s")
        params.append(jugador_id)
    where = " AND ".join(filtros)
    rows = conn.execute(f"""
        SELECT *
        FROM portal_asistencia_confirmaciones
        WHERE {where}
    """, params).fetchall()
    confirmaciones = {}
    for row in rows:
        item = dict(row)
        try:
            item["dolor_zonas_lista"] = json.loads(item.get("dolor_zonas") or "[]")
        except (TypeError, ValueError):
            item["dolor_zonas_lista"] = []
        item["bienestar_resumen"] = resumen_bienestar_confirmacion(item)
        item["bienestar_completo"] = item["bienestar_resumen"] is not None
        confirmaciones[(item["evento_id"], item["jugador_id"])] = item
    return confirmaciones


def obtener_eventos_deportivos_ics(token):
    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()
    if jugador is None:
        conn.close()
        return None, []

    eventos = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE COALESCE(publicar_portal, 0) = 1
          AND fecha >= (CURRENT_DATE - INTERVAL '30 days')::text
        ORDER BY fecha ASC, COALESCE(hora_inicio, '') ASC, id ASC
    """).fetchall()
    conn.close()
    return jugador, [evento for evento in eventos if categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria"))]


def ics_escape(valor):
    valor = str(valor or "")
    return (
        valor.replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\r\n", "\\n")
        .replace("\n", "\\n")
    )


def ics_fold(linea):
    if len(linea) <= 74:
        return linea
    partes = []
    while len(linea) > 74:
        partes.append(linea[:74])
        linea = " " + linea[74:]
    partes.append(linea)
    return "\r\n".join(partes)


def ics_fecha_evento(fecha, hora=None, duracion_minutos=90):
    hora = normalizar_hora_evento(hora)
    if not hora:
        inicio = datetime.strptime(fecha, "%Y-%m-%d")
        fin = inicio + timedelta(days=1)
        return (
            f"DTSTART;VALUE=DATE:{inicio.strftime('%Y%m%d')}",
            f"DTEND;VALUE=DATE:{fin.strftime('%Y%m%d')}",
        )
    inicio = datetime.strptime(f"{fecha} {hora}", "%Y-%m-%d %H:%M")
    fin = inicio + timedelta(minutes=normalizar_duracion_evento(duracion_minutos))
    return (
        f"DTSTART;TZID={CALENDARIO_TZ}:{inicio.strftime('%Y%m%dT%H%M%S')}",
        f"DTEND;TZID={CALENDARIO_TZ}:{fin.strftime('%Y%m%dT%H%M%S')}",
    )


def generar_ics_calendario(jugador, eventos, feed_url):
    ahora = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lineas = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Ruda Macho Rugby//SIG//ES",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:SIG - {ics_escape(jugador['apellido'])}, {ics_escape(jugador['nombre'])}",
        f"X-WR-TIMEZONE:{CALENDARIO_TZ}",
        f"URL:{feed_url}",
    ]
    for evento in eventos:
        uid = f"calendario-{evento['id']}@sig.rudamachorugby.com"
        descripcion = evento.get("descripcion") or ""
        if evento.get("categoria"):
            descripcion = f"{descripcion}\nCategoria: {evento['categoria']}".strip()
        dtstart, dtend = ics_fecha_evento(evento["fecha"], evento.get("hora_inicio"), evento.get("duracion_minutos") or 90)
        lineas.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{ahora}",
            dtstart,
            dtend,
            f"SUMMARY:{ics_escape(evento['titulo'])}",
            f"DESCRIPTION:{ics_escape(descripcion)}",
            f"LOCATION:{ics_escape(evento.get('ubicacion') or '')}",
            f"CATEGORIES:{ics_escape(evento.get('tipo') or 'Club')}",
            "END:VEVENT",
        ])
    lineas.append("END:VCALENDAR")
    return "\r\n".join(ics_fold(linea) for linea in lineas) + "\r\n"


def obtener_calendario(mes):
    desde, hasta = rango_mes(mes)
    conn = get_connection()

    eventos_manuales = conn.execute("""
        SELECT
            id,
            fecha,
            tipo,
            titulo,
            descripcion,
            ubicacion,
            categoria,
            hora_inicio,
            duracion_minutos,
            publicar_portal,
            asistencia_evento_id,
            convocatoria_texto,
            convocatoria_cierre,
            minuta_post_evento
        FROM calendario_eventos
        WHERE fecha >= %s AND fecha < %s
        ORDER BY fecha ASC, id ASC
    """, (desde, hasta)).fetchall()

    eventos_asistencia = conn.execute("""
        SELECT
            id,
            fecha,
            tipo,
            descripcion
        FROM eventos_asistencia e
        WHERE fecha >= %s AND fecha < %s
          AND NOT EXISTS (
              SELECT 1
              FROM calendario_eventos ce
              WHERE ce.asistencia_evento_id = e.id
          )
        ORDER BY fecha ASC, id ASC
    """, (desde, hasta)).fetchall()

    cuotas = conn.execute("""
        SELECT
            c.id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            j.apellido,
            j.nombre,
            j.categoria
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
          AND c.fecha_vencimiento IS NOT NULL
          AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
          AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND c.fecha_vencimiento >= %s
          AND c.fecha_vencimiento < %s
        ORDER BY c.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 80
    """, (desde, hasta)).fetchall()

    fichas = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.categoria,
            f.fecha_vencimiento
        FROM fichas_medicas f
        JOIN jugadores j ON j.id = f.jugador_id
        WHERE f.fecha_vencimiento IS NOT NULL
          AND NULLIF(f.fecha_vencimiento::text, '') IS NOT NULL
          AND f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND f.fecha_vencimiento >= %s
          AND f.fecha_vencimiento < %s
        ORDER BY f.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 80
    """, (desde, hasta)).fetchall()

    confirmaciones_resumen = {}
    asistencia_evento_ids = [evento["asistencia_evento_id"] for evento in eventos_manuales if evento.get("asistencia_evento_id")]
    if asistencia_evento_ids:
        rows = conn.execute("""
            SELECT
                evento_id,
                SUM(CASE WHEN estado = 'confirmado' THEN 1 ELSE 0 END) AS confirmados,
                SUM(CASE WHEN estado = 'dudoso' THEN 1 ELSE 0 END) AS dudosos,
                SUM(CASE WHEN estado = 'no_asiste' THEN 1 ELSE 0 END) AS no_asisten
            FROM portal_asistencia_confirmaciones
            WHERE evento_id = ANY(%s)
            GROUP BY evento_id
        """, (asistencia_evento_ids,)).fetchall()
        confirmaciones_resumen = {row["evento_id"]: row for row in rows}

    conn.close()

    eventos = []

    for evento in eventos_manuales:
        eventos.append({
            "id": evento["id"],
            "fecha": evento["fecha"],
            "hora": evento.get("hora_inicio") or "",
            "fecha_hora": formato_fecha_hora_evento(evento),
            "tipo": evento["tipo"],
            "titulo": evento["titulo"],
            "detalle": evento["descripcion"] or "",
            "ubicacion": evento["ubicacion"] or "",
            "categoria": evento["categoria"] or "",
            "origen": "calendario",
            "url": url_for("tomar_asistencia", evento_id=evento["asistencia_evento_id"]) if evento.get("asistencia_evento_id") else None,
            "edit_url": url_for("editar_evento_calendario", evento_id=evento["id"]),
            "delete_url": url_for("eliminar_evento_calendario", evento_id=evento["id"]),
            "convocatoria_texto": evento.get("convocatoria_texto") or "",
            "convocatoria_cierre": evento.get("convocatoria_cierre") or "",
            "minuta_post_evento": evento.get("minuta_post_evento") or "",
            "prioridad": "normal",
        })
        resumen_confirmacion = confirmaciones_resumen.get(evento.get("asistencia_evento_id"))
        if resumen_confirmacion:
            if es_evento_partido(evento):
                eventos[-1]["disponibilidad_resumen"] = (
                    f"{resumen_confirmacion['confirmados'] or 0} juegan / "
                    f"{resumen_confirmacion['dudosos'] or 0} no juegan / "
                    f"{resumen_confirmacion['no_asisten'] or 0} no van"
                )
            else:
                eventos[-1]["disponibilidad_resumen"] = (
                    f"{resumen_confirmacion['confirmados'] or 0} conf. / "
                    f"{resumen_confirmacion['dudosos'] or 0} dud. / "
                    f"{resumen_confirmacion['no_asisten'] or 0} no asisten"
                )
        else:
            eventos[-1]["disponibilidad_resumen"] = ""

    for evento in eventos_asistencia:
        eventos.append({
            "id": evento["id"],
            "fecha": evento["fecha"],
            "hora": "",
            "fecha_hora": evento["fecha"],
            "tipo": evento["tipo"],
            "titulo": f"Asistencia: {evento['tipo']}",
            "detalle": evento["descripcion"] or "",
            "ubicacion": "",
            "categoria": "",
            "origen": "asistencia",
            "url": url_for("tomar_asistencia", evento_id=evento["id"]),
            "edit_url": None,
            "delete_url": None,
            "prioridad": "normal",
        })

    for cuota in cuotas:
        eventos.append({
            "id": cuota["id"],
            "fecha": cuota["fecha_vencimiento"],
            "tipo": "Cuota",
            "titulo": f"Vence cuota {cuota['periodo']}",
            "detalle": f"{cuota['apellido']}, {cuota['nombre']} - {formato_moneda(cuota['importe'])}",
            "ubicacion": "",
            "categoria": cuota["categoria"] or "",
            "origen": "cuota",
            "url": url_for("ver_cuotas", jugador_id=cuota["jugador_id"]),
            "edit_url": None,
            "delete_url": None,
            "prioridad": "warning",
        })

    for ficha in fichas:
        eventos.append({
            "id": ficha["jugador_id"],
            "fecha": ficha["fecha_vencimiento"],
            "tipo": "Ficha m?dica",
            "titulo": "Vence ficha médica",
            "detalle": f"{ficha['apellido']}, {ficha['nombre']}",
            "ubicacion": "",
            "categoria": ficha["categoria"] or "",
            "origen": "ficha",
            "url": url_for("ver_ficha_medica", jugador_id=ficha["jugador_id"]),
            "edit_url": None,
            "delete_url": None,
            "prioridad": "danger",
        })

    eventos.sort(key=lambda item: (item["fecha"], item.get("hora") or "", item["tipo"], item["titulo"]))

    eventos_por_dia = {}
    for evento in eventos:
        eventos_por_dia.setdefault(evento["fecha"], []).append(evento)

    return {
        "mes": mes,
        "desde": desde,
        "hasta": hasta,
        "eventos": eventos,
        "eventos_por_dia": eventos_por_dia,
        "resumen": {
            "total": len(eventos),
            "manuales": len(eventos_manuales),
            "asistencia": len(eventos_asistencia),
            "cuotas": len(cuotas),
            "fichas": len(fichas),
        },
    }


def normalizar_telefono_whatsapp(telefono):
    digitos = re.sub(r"\D+", "", telefono or "")
    if not digitos:
        return ""

    if digitos.startswith("00"):
        digitos = digitos[2:]

    if digitos.startswith("549"):
        return digitos

    if digitos.startswith("54"):
        resto = digitos[2:]
        if resto.startswith("9"):
            return digitos
        return "549" + resto.lstrip("0")

    if digitos.startswith("0"):
        digitos = digitos.lstrip("0")

    if 10 <= len(digitos) <= 11:
        return "549" + digitos

    return digitos


def whatsapp_api_disponible():
    return bool(
        WHATSAPP_ENABLED
        and WHATSAPP_ACCESS_TOKEN
        and WHATSAPP_PHONE_NUMBER_ID
        and WHATSAPP_WABA_ID
    )


def valor_texto_contacto(valor):
    texto = str(valor or "").strip()
    if texto.lower() in {"", "none", "null", "n/a", "na", "-"}:
        return ""
    return texto


def telefono_jugador_preferido(jugador):
    return normalizar_telefono_whatsapp(
        valor_texto_contacto((jugador or {}).get("telefono"))
        or valor_texto_contacto((jugador or {}).get("telefono_tutor"))
        or ""
    )


def variantes_telefono_whatsapp(telefono):
    normalizado = normalizar_telefono_whatsapp(telefono)
    if not normalizado:
        return []

    variantes = [normalizado]

    if normalizado.startswith("549") and len(normalizado) > 3:
        variantes.append("54" + normalizado[3:])
    elif normalizado.startswith("54") and not normalizado.startswith("549") and len(normalizado) > 2:
        variantes.append("549" + normalizado[2:])

    vistas = set()
    resultado = []
    for variante in variantes:
        if variante and variante not in vistas:
            resultado.append(variante)
            vistas.add(variante)
    return resultado


def registrar_whatsapp_envio(
    telefono,
    tipo,
    entidad,
    entidad_id=None,
    jugador_id=None,
    mensaje=None,
    payload=None,
    respuesta=None,
    estado="pendiente",
    meta_message_id=None,
    error_codigo=None,
    error_mensaje=None,
):
    conn = None
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO whatsapp_envios (
                jugador_id, telefono, destino_normalizado, tipo, entidad, entidad_id,
                mensaje, estado, meta_message_id, error_codigo, error_mensaje,
                payload, respuesta, creado_por, enviado_en
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            jugador_id,
            telefono or None,
            normalizar_telefono_whatsapp(telefono),
            tipo,
            entidad,
            entidad_id,
            mensaje or None,
            estado,
            meta_message_id,
            error_codigo,
            error_mensaje,
            json.dumps(payload or {}, ensure_ascii=False, default=str),
            json.dumps(respuesta or {}, ensure_ascii=False, default=str),
            session.get("username") if has_request_context() else None,
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if estado in {"enviado", "sent", "accepted"} else None,
        ))
        conn.commit()
        conn.close()
    except Exception:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


def registrar_whatsapp_webhook(event_type, payload, procesado=False):
    conn = None
    try:
        conn = get_connection()
        conn.execute("""
            INSERT INTO whatsapp_webhook_eventos (event_type, meta_object, payload, procesado)
            VALUES (%s, %s, %s, %s)
        """, (
            event_type,
            (payload or {}).get("object"),
            json.dumps(payload or {}, ensure_ascii=False, default=str),
            1 if procesado else 0,
        ))
        conn.commit()
        conn.close()
    except Exception:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


def resumir_contenido_whatsapp(tipo, mensaje):
    tipo = (tipo or "").strip().lower()
    if tipo == "text":
        return ((mensaje or {}).get("text") or {}).get("body") or ""
    if tipo == "button":
        button = (mensaje or {}).get("button") or {}
        return button.get("text") or button.get("payload") or "[Botón]"
    if tipo == "interactive":
        interactive = (mensaje or {}).get("interactive") or {}
        if interactive.get("button_reply"):
            reply = interactive["button_reply"]
            return reply.get("title") or reply.get("id") or "[Botón]"
        if interactive.get("list_reply"):
            reply = interactive["list_reply"]
            return reply.get("title") or reply.get("description") or reply.get("id") or "[Lista]"
        return "[Interactivo]"
    if tipo == "reaction":
        reaction = (mensaje or {}).get("reaction") or {}
        emoji = reaction.get("emoji")
        return f"[Reacción] {emoji}" if emoji else "[Reacción]"
    if tipo == "location":
        location = (mensaje or {}).get("location") or {}
        partes = [
            location.get("name"),
            location.get("address"),
            ", ".join(
                str(location.get(campo))
                for campo in ("latitude", "longitude")
                if location.get(campo) is not None
            ),
        ]
        detalle = " - ".join(parte for parte in partes if parte)
        return f"[Ubicación] {detalle}" if detalle else "[Ubicación]"
    if tipo == "contacts":
        contactos = (mensaje or {}).get("contacts") or []
        nombres = []
        for contacto in contactos:
            nombre = (contacto.get("name") or {}).get("formatted_name")
            if nombre:
                nombres.append(nombre)
        return f"[Contacto] {', '.join(nombres)}" if nombres else "[Contacto]"
    if tipo in {"image", "video", "audio", "document", "sticker"}:
        media = (mensaje or {}).get(tipo) or {}
        caption = media.get("caption")
        filename = media.get("filename")
        etiqueta = {
            "image": "Imagen",
            "video": "Video",
            "audio": "Audio",
            "document": "Documento",
            "sticker": "Sticker",
        }[tipo]
        detalle = caption or filename
        return f"[{etiqueta}] {detalle}" if detalle else f"[{etiqueta}]"
    if tipo == "unsupported":
        errores = (mensaje or {}).get("errors") or []
        if errores:
            error = errores[0] or {}
            detalle = (
                error.get("title")
                or error.get("message")
                or error.get("details")
                or error.get("code")
            )
            return f"[Mensaje no compatible] {detalle}" if detalle else "[Mensaje no compatible]"
        return "[Mensaje no compatible]"
    etiquetas = {
        "image": "[Imagen]",
        "video": "[Video]",
        "audio": "[Audio]",
        "document": "[Documento]",
        "sticker": "[Sticker]",
        "location": "[Ubicación]",
        "contacts": "[Contacto]",
        "button": "[Botón]",
        "interactive": "[Interactivo]",
        "reaction": "[Reacción]",
    }
    return etiquetas.get(tipo, f"[{tipo or 'mensaje'}]")


def buscar_jugador_por_whatsapp(telefono_normalizado):
    telefono_normalizado = normalizar_telefono_whatsapp(telefono_normalizado)
    if not telefono_normalizado:
        return None
    conn = get_connection()
    jugadores = conn.execute("""
        SELECT id, nombre, apellido, categoria, telefono, telefono_tutor
        FROM jugadores
        WHERE COALESCE(NULLIF(telefono, ''), NULLIF(telefono_tutor, '')) IS NOT NULL
    """).fetchall()
    conn.close()
    for jugador in jugadores:
        if normalizar_telefono_whatsapp(jugador.get("telefono")) == telefono_normalizado:
            return jugador
        if normalizar_telefono_whatsapp(jugador.get("telefono_tutor")) == telefono_normalizado:
            return jugador
    return None


def registrar_whatsapp_mensaje(
    *,
    telefono,
    wa_id=None,
    jugador_id=None,
    direccion,
    tipo,
    texto=None,
    meta_message_id=None,
    estado=None,
    payload=None,
    respuesta=None,
    creado_por=None,
):
    conn = None
    try:
        telefono_normalizado = normalizar_telefono_whatsapp(telefono or wa_id)
        conn = get_connection()
        wa_normalizado = normalizar_telefono_whatsapp(wa_id) or telefono_normalizado or None
        estado_final = estado or ("recibido" if direccion == "in" else "enviado")
        payload_json = json.dumps(payload or {}, ensure_ascii=False, default=str)
        respuesta_json = json.dumps(respuesta or {}, ensure_ascii=False, default=str)
        creado_por_final = creado_por or (session.get("username") if has_request_context() else None)

        mensaje_existente = None
        if meta_message_id:
            mensaje_existente = conn.execute(
                "SELECT id FROM whatsapp_mensajes WHERE meta_message_id = %s LIMIT 1",
                (meta_message_id,)
            ).fetchone()

        if mensaje_existente:
            conn.execute("""
                UPDATE whatsapp_mensajes
                SET
                    jugador_id = COALESCE(%s, jugador_id),
                    telefono = COALESCE(%s, telefono),
                    wa_id = COALESCE(%s, wa_id),
                    direccion = %s,
                    tipo = %s,
                    texto = COALESCE(%s, texto),
                    estado = COALESCE(%s, estado),
                    payload = COALESCE(%s, payload),
                    respuesta = COALESCE(%s, respuesta),
                    creado_por = COALESCE(%s, creado_por)
                WHERE id = %s
            """, (
                jugador_id,
                telefono_normalizado or None,
                wa_normalizado,
                direccion,
                tipo,
                texto or None,
                estado_final,
                payload_json,
                respuesta_json,
                creado_por_final,
                mensaje_existente["id"],
            ))
            resultado = "updated"
        else:
            conn.execute("""
                INSERT INTO whatsapp_mensajes (
                    jugador_id, telefono, wa_id, direccion, tipo, texto,
                    meta_message_id, estado, payload, respuesta, creado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                jugador_id,
                telefono_normalizado or None,
                wa_normalizado,
                direccion,
                tipo,
                texto or None,
                meta_message_id or None,
                estado_final,
                payload_json,
                respuesta_json,
                creado_por_final,
            ))
            resultado = "inserted"
        conn.commit()
        conn.close()
        return resultado
    except Exception:
        app.logger.exception(
            "No se pudo registrar mensaje de WhatsApp. telefono=%s wa_id=%s direccion=%s tipo=%s meta_message_id=%s",
            telefono,
            wa_id,
            direccion,
            tipo,
            meta_message_id,
        )
        try:
            if conn:
                conn.close()
        except Exception:
            pass
        return False


def destinatarios_notificacion_whatsapp_inbox():
    destinatarios = [normalizar_email(email) for email in WHATSAPP_INBOX_NOTIFY_EMAILS]
    destinatarios = [email for email in destinatarios if email]
    if destinatarios:
        return list(dict.fromkeys(destinatarios))
    fallback = normalizar_email(SMTP_FROM)
    return [fallback] if fallback else []


def enviar_notificacion_whatsapp_inbox_email(*, mensaje, telefono, jugador=None):
    if suprimir_email_whatsapp_por_presencia():
        app.logger.info("Notificacion WhatsApp por email suprimida por usuario activo en SIG.")
        return False

    destinatarios = destinatarios_notificacion_whatsapp_inbox()
    if not destinatarios:
        return False

    jugador_nombre = ""
    if jugador:
        jugador_nombre = (
            f"{(jugador.get('apellido') or '').strip()}, {(jugador.get('nombre') or '').strip()}".strip(", ")
        )
    telefono = normalizar_telefono_whatsapp(telefono) or str(telefono or "").strip()
    texto = (mensaje.get("texto") or "").strip() or "[Sin contenido]"
    tipo = (mensaje.get("tipo") or "mensaje").strip()
    inbox_url = url_for("ver_whatsapp_inbox", telefono=telefono, _external=True) if has_request_context() else ""

    asunto = "Nueva respuesta por WhatsApp en SIG"
    if jugador_nombre:
        asunto = f"Nueva respuesta WhatsApp - {jugador_nombre}"
    cuerpo = (
        "Entró una nueva respuesta por WhatsApp en SIG.\n\n"
        f"Contacto: {telefono or '-'}\n"
        f"Jugador vinculado: {jugador_nombre or 'Sin vincular'}\n"
        f"Tipo: {tipo}\n"
        f"Mensaje: {texto}\n"
    )
    if inbox_url:
        cuerpo += f"\nAbrir conversación: {inbox_url}\n"

    enviado_alguno = False
    for destinatario in destinatarios:
        try:
            enviado, motivo = enviar_email(destinatario, asunto, cuerpo)
            enviado_alguno = enviado_alguno or bool(enviado)
            if not enviado:
                app.logger.warning(
                    "No se pudo notificar respuesta WhatsApp por email a %s. motivo=%s",
                    destinatario,
                    motivo,
                )
        except Exception:
            app.logger.exception("No se pudo notificar respuesta WhatsApp por email a %s.", destinatario)
    return enviado_alguno


def listar_whatsapp_conversaciones():
    conn = get_connection()
    mensajes = conn.execute("""
        SELECT
            wm.id,
            wm.jugador_id,
            wm.telefono,
            wm.wa_id,
            wm.direccion,
            wm.tipo,
            wm.texto,
            wm.meta_message_id,
            wm.estado,
            wm.leido,
            wm.creado_en,
            j.nombre,
            j.apellido,
            j.categoria
        FROM whatsapp_mensajes wm
        LEFT JOIN jugadores j ON j.id = wm.jugador_id
        ORDER BY wm.creado_en DESC, wm.id DESC
    """).fetchall()
    conn.close()

    conversaciones = []
    vistos = set()
    for mensaje in mensajes:
        telefono = mensaje.get("wa_id") or mensaje.get("telefono")
        if not telefono or telefono in vistos:
            continue
        vistos.add(telefono)
        conversaciones.append({
            "telefono": telefono,
            "jugador_id": mensaje.get("jugador_id"),
            "nombre": (
                f"{(mensaje.get('apellido') or '').strip()}, {(mensaje.get('nombre') or '').strip()}".strip(", ")
                or "Sin vincular"
            ),
            "categoria": mensaje.get("categoria") or "",
            "ultimo_texto": mensaje.get("texto") or "",
            "ultimo_tipo": mensaje.get("tipo") or "text",
            "ultimo_direccion": mensaje.get("direccion") or "in",
            "ultimo_estado": mensaje.get("estado") or "",
            "ultimo_en": mensaje.get("creado_en"),
            "sin_leer": sum(
                1
                for item in mensajes
                if (item.get("wa_id") or item.get("telefono")) == telefono
                and item.get("direccion") == "in"
                and not item.get("leido")
            ),
        })
    return conversaciones


def listar_jugadores_basicos_para_whatsapp():
    conn = get_connection()
    jugadores = conn.execute("""
        SELECT id, nombre, apellido, categoria
        FROM jugadores
        ORDER BY apellido ASC, nombre ASC
    """).fetchall()
    conn.close()
    return jugadores


def obtener_whatsapp_conversacion(telefono):
    telefono = normalizar_telefono_whatsapp(telefono)
    if not telefono:
        return []
    conn = get_connection()
    mensajes = conn.execute("""
        SELECT
            wm.*,
            j.nombre,
            j.apellido,
            j.categoria
        FROM whatsapp_mensajes wm
        LEFT JOIN jugadores j ON j.id = wm.jugador_id
        WHERE wm.telefono = %s OR wm.wa_id = %s
        ORDER BY wm.creado_en ASC, wm.id ASC
    """, (telefono, telefono)).fetchall()
    conn.close()
    return mensajes


def obtener_contador_whatsapp_inbox():
    try:
        conn = get_connection()
        fila = conn.execute("""
            SELECT COUNT(*) AS total
            FROM whatsapp_mensajes
            WHERE direccion = 'in' AND COALESCE(leido, 0) = 0
        """).fetchone()
        conn.close()
        return int((fila or {}).get("total") or 0)
    except Exception:
        return 0


def obtener_estado_whatsapp_inbox():
    try:
        conn = get_connection()
        fila = conn.execute("""
            SELECT
                COUNT(*) FILTER (
                    WHERE direccion = 'in'
                      AND COALESCE(leido, 0) = 0
                ) AS sin_leer,
                COALESCE(MAX(id), 0) AS ultimo_id
            FROM whatsapp_mensajes
        """).fetchone()
        conn.close()
        return {
            "sin_leer": int((fila or {}).get("sin_leer") or 0),
            "ultimo_id": int((fila or {}).get("ultimo_id") or 0),
        }
    except Exception:
        return {"sin_leer": 0, "ultimo_id": 0}


def resumir_evento_webhook_whatsapp(payload_texto):
    try:
        payload = json.loads(payload_texto or "{}")
    except Exception:
        return {
            "messages_count": 0,
            "statuses_count": 0,
            "contacts_count": 0,
            "from_values": [],
            "status_values": [],
        }

    messages_count = 0
    statuses_count = 0
    contacts_count = 0
    from_values = []
    status_values = []
    for entry in payload.get("entry", []) or []:
        for change in entry.get("changes", []) or []:
            value = change.get("value") or {}
            messages = value.get("messages", []) or []
            statuses = value.get("statuses", []) or []
            contacts = value.get("contacts", []) or []
            messages_count += len(messages)
            statuses_count += len(statuses)
            contacts_count += len(contacts)
            from_values.extend(
                normalizar_telefono_whatsapp(item.get("from"))
                for item in messages
                if item.get("from")
            )
            status_values.extend(
                str(item.get("status") or "").strip()
                for item in statuses
                if item.get("status")
            )
    return {
        "messages_count": messages_count,
        "statuses_count": statuses_count,
        "contacts_count": contacts_count,
        "from_values": [item for item in from_values if item],
        "status_values": [item for item in status_values if item],
    }


def listar_whatsapp_webhook_eventos(limit=15):
    conn = get_connection()
    filas = conn.execute("""
        SELECT id, event_type, meta_object, procesado, payload, recibido_en
        FROM whatsapp_webhook_eventos
        ORDER BY id DESC
        LIMIT %s
    """, (limit,)).fetchall()
    conn.close()

    eventos = []
    for fila in filas:
        resumen = resumir_evento_webhook_whatsapp(fila.get("payload"))
        eventos.append({
            **fila,
            **resumen,
        })
    return eventos


def parse_meta_signed_request(signed_request):
    partes = str(signed_request or "").split(".", 1)
    if len(partes) != 2:
        return None
    firma_codificada, payload_codificado = partes
    try:
        payload = json.loads(base64url_decode(payload_codificado).decode("utf-8"))
    except Exception:
        return None

    if WHATSAPP_APP_SECRET:
        try:
            firma = base64url_decode(firma_codificada)
            esperado = hmac.new(
                WHATSAPP_APP_SECRET.encode("utf-8"),
                payload_codificado.encode("utf-8"),
                hashlib.sha256,
            ).digest()
            if not hmac.compare_digest(firma, esperado):
                return None
        except Exception:
            return None
    return payload


def mensaje_fallo_whatsapp(motivo, destinatario=None, detalle=None):
    if motivo == "sin_telefono":
        return "No hay un teléfono de WhatsApp configurado para este jugador."
    if motivo == "sin_configuracion":
        return "WhatsApp API no está configurado en el servidor."
    if motivo == "verificacion":
        return "WhatsApp API no tiene token de verificación configurado."
    if motivo == "http_error" and detalle:
        return f"No se pudo enviar el WhatsApp a {destinatario or 'este jugador'}: {detalle}"
    if motivo == "network_error":
        return "No se pudo conectar con WhatsApp API. Intentá nuevamente."
    if destinatario:
        return f"No se pudo enviar el WhatsApp a {destinatario}. Revisá la configuración o intentá nuevamente."
    return "No se pudo enviar el WhatsApp. Revisá la configuración o intentá nuevamente."


def resumir_envio_masivo_whatsapp(resultados, etiqueta):
    enviados = sum(1 for ok, _, _, _ in resultados if ok)
    sin_telefono = sum(1 for ok, _, motivo, _ in resultados if not ok and motivo == "sin_telefono")
    sin_configuracion = sum(1 for ok, _, motivo, _ in resultados if not ok and motivo == "sin_configuracion")
    otros = sum(1 for ok, _, motivo, _ in resultados if not ok and motivo not in ("sin_telefono", "sin_configuracion"))
    if enviados:
        partes = [f"Se enviaron {enviados} {etiqueta}."]
        if sin_telefono:
            partes.append(f"{sin_telefono} sin teléfono.")
        if sin_configuracion:
            partes.append(f"{sin_configuracion} sin configuración.")
        if otros:
            partes.append(f"{otros} con error.")
        return " ".join(partes), "ok"
    detalles = []
    if sin_telefono:
        detalles.append(f"{sin_telefono} sin teléfono")
    if sin_configuracion:
        detalles.append(f"{sin_configuracion} sin configuración")
    if otros:
        detalles.append(f"{otros} con error")
    if detalles:
        return f"No se enviaron {etiqueta}: " + ", ".join(detalles) + ".", "error"
    return f"No se enviaron {etiqueta}.", "error"


def _enviar_whatsapp_payload(
    telefono,
    payload_builder,
    *,
    tipo="general",
    entidad="sistema",
    entidad_id=None,
    jugador_id=None,
    mensaje_log=None,
):
    variantes = variantes_telefono_whatsapp(telefono)
    if not variantes:
        return False, None, "sin_telefono", None
    if not whatsapp_api_disponible():
        return False, variantes[0], "sin_configuracion", None

    headers = {
        "Authorization": f"Bearer {WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    url = f"{WHATSAPP_GRAPH_BASE}/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    ultimo_destino = variantes[0]
    ultimo_detalle = None
    ultimo_motivo = None

    for idx, telefono_normalizado in enumerate(variantes):
        payload = payload_builder(telefono_normalizado)
        data = json.dumps(payload).encode("utf-8")
        req = UrlRequest(url, data=data, headers=headers, method="POST")

        try:
            with urlopen(req, timeout=20) as resp:
                body = resp.read().decode("utf-8")
                respuesta = json.loads(body) if body else {}
            message_id = None
            mensajes = respuesta.get("messages") or []
            if mensajes and isinstance(mensajes, list):
                message_id = mensajes[0].get("id")
            registrar_whatsapp_envio(
                telefono=telefono,
                tipo=tipo,
                entidad=entidad,
                entidad_id=entidad_id,
                jugador_id=jugador_id,
                mensaje=mensaje_log,
                payload=payload,
                respuesta=respuesta,
                estado="enviado",
                meta_message_id=message_id,
            )
            texto_conversacion = ""
            if payload.get("type") == "text":
                texto_conversacion = ((payload.get("text") or {}).get("body") or "").strip()
            elif payload.get("type") == "template":
                nombre_template = ((payload.get("template") or {}).get("name") or "").strip()
                texto_conversacion = f"[Template] {nombre_template}" if nombre_template else "[Template]"
            registrar_whatsapp_mensaje(
                telefono=telefono_normalizado,
                wa_id=telefono_normalizado,
                jugador_id=jugador_id,
                direccion="out",
                tipo=payload.get("type") or "text",
                texto=texto_conversacion or (mensaje_log or ""),
                meta_message_id=message_id,
                estado="enviado",
                payload=payload,
                respuesta=respuesta,
            )
            return True, telefono_normalizado, None, respuesta
        except HTTPError as error:
            raw = error.read().decode("utf-8", errors="replace")
            try:
                detalle = json.loads(raw) if raw else {}
            except ValueError:
                detalle = {"raw": raw}
            error_data = detalle.get("error") if isinstance(detalle, dict) else {}
            codigo = str(error_data.get("code") or error.code)
            mensaje_error = error_data.get("message") or raw
            registrar_whatsapp_envio(
                telefono=telefono,
                tipo=tipo,
                entidad=entidad,
                entidad_id=entidad_id,
                jugador_id=jugador_id,
                mensaje=mensaje_log,
                payload=payload,
                respuesta=detalle,
                estado="error",
                error_codigo=codigo,
                error_mensaje=mensaje_error,
            )
            ultimo_destino = telefono_normalizado
            ultimo_detalle = mensaje_error
            ultimo_motivo = "http_error"
            if codigo == "133010" and idx < len(variantes) - 1:
                continue
            return False, ultimo_destino, ultimo_motivo, ultimo_detalle
        except URLError:
            registrar_whatsapp_envio(
                telefono=telefono,
                tipo=tipo,
                entidad=entidad,
                entidad_id=entidad_id,
                jugador_id=jugador_id,
                mensaje=mensaje_log,
                payload=payload,
                respuesta={},
                estado="error",
                error_mensaje="network_error",
            )
            return False, telefono_normalizado, "network_error", None

    return False, ultimo_destino, ultimo_motivo or "http_error", ultimo_detalle


def enviar_whatsapp_meta(telefono, mensaje, *, tipo="general", entidad="sistema", entidad_id=None, jugador_id=None):
    return _enviar_whatsapp_payload(
        telefono,
        lambda telefono_normalizado: {
            "messaging_product": "whatsapp",
            "to": telefono_normalizado,
            "type": "text",
            "text": {
                "preview_url": False,
                "body": mensaje,
            },
        },
        tipo=tipo,
        entidad=entidad,
        entidad_id=entidad_id,
        jugador_id=jugador_id,
        mensaje_log=mensaje,
    )


def enviar_whatsapp_template_meta(
    telefono,
    template_name,
    language_code,
    body_params,
    *,
    tipo="general",
    entidad="sistema",
    entidad_id=None,
    jugador_id=None,
    mensaje_log=None,
):
    parametros = [
        {"type": "text", "text": str(valor or "").strip() or "-"}
        for valor in (body_params or [])
    ]
    return _enviar_whatsapp_payload(
        telefono,
        lambda telefono_normalizado: {
            "messaging_product": "whatsapp",
            "to": telefono_normalizado,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {"code": language_code},
                "components": [
                    {
                        "type": "body",
                        "parameters": parametros,
                    }
                ] if parametros else [],
            },
        },
        tipo=tipo,
        entidad=entidad,
        entidad_id=entidad_id,
        jugador_id=jugador_id,
        mensaje_log=mensaje_log or f"template:{template_name}",
    )


def enviar_whatsapp_jugador(jugador, mensaje, *, tipo="general", entidad="jugador", entidad_id=None):
    return enviar_whatsapp_meta(
        valor_texto_contacto((jugador or {}).get("telefono"))
        or valor_texto_contacto((jugador or {}).get("telefono_tutor")),
        mensaje,
        tipo=tipo,
        entidad=entidad,
        entidad_id=entidad_id,
        jugador_id=(jugador or {}).get("id"),
    )


def enviar_whatsapp_recordatorio_cuota_template(
    jugador,
    *,
    tipo="recordatorio_cuota",
    entidad="cuota",
    entidad_id=None,
):
    nombre = nombre_jugador_corto(jugador)
    telefono = (
        valor_texto_contacto((jugador or {}).get("telefono"))
        or valor_texto_contacto((jugador or {}).get("telefono_tutor"))
    )
    return enviar_whatsapp_template_meta(
        telefono,
        WHATSAPP_TEMPLATE_CUOTA,
        WHATSAPP_TEMPLATE_CUOTA_LANG,
        [nombre],
        tipo=tipo,
        entidad=entidad,
        entidad_id=entidad_id,
        jugador_id=(jugador or {}).get("id") or (jugador or {}).get("jugador_id"),
        mensaje_log=f"template:{WHATSAPP_TEMPLATE_CUOTA}:{nombre}",
    )


def mensaje_moroso(template, jugador):
    contexto = {
        "nombre": jugador["nombre"] or "",
        "apellido": jugador["apellido"] or "",
        "deuda": formato_moneda(jugador["deuda"] or 0),
        "cuotas_pendientes": jugador["cuotas_pendientes"] or 0,
        "cuotas_vencidas": jugador["cuotas_vencidas"] or 0,
        "primer_vencimiento": jugador["primer_vencimiento"] or "-",
    }

    mensaje = template
    for clave, valor in contexto.items():
        mensaje = mensaje.replace("{" + clave + "}", str(valor))
    return mensaje


def obtener_morosos_para_comunicacion():
    conn = get_connection()
    morosos = conn.execute("""
        SELECT
            j.id,
            j.apellido,
            j.nombre,
            j.categoria,
            j.telefono,
            j.email,
            j.telefono_tutor,
            j.email_tutor,
            j.contacto_tutor,
            COUNT(c.id) AS cuotas_pendientes,
            SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1
                    ELSE 0
                END
            ) AS cuotas_vencidas,
            COALESCE(SUM(c.importe), 0) AS deuda,
            MIN(NULLIF(c.fecha_vencimiento::text, '')) AS primer_vencimiento
        FROM jugadores j
        JOIN cuotas c ON c.jugador_id = j.id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY
            j.id, j.apellido, j.nombre, j.categoria, j.telefono, j.email,
            j.telefono_tutor, j.email_tutor, j.contacto_tutor
        HAVING COALESCE(SUM(c.importe), 0) > 0
        ORDER BY deuda DESC, cuotas_vencidas DESC, j.apellido, j.nombre
    """).fetchall()
    conn.close()
    return morosos


def obtener_notificaciones_operativas():
    descartadas = obtener_notificaciones_descartadas()
    conn = get_connection()
    cuotas_vencidas = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            j.apellido,
            j.nombre,
            j.telefono,
            j.telefono_tutor
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
          AND c.fecha_vencimiento IS NOT NULL
          AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
          AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND c.fecha_vencimiento::date < CURRENT_DATE
        ORDER BY c.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 50
    """).fetchall()

    cuotas_por_vencer = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            j.apellido,
            j.nombre,
            j.telefono,
            j.telefono_tutor
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
          AND c.fecha_vencimiento IS NOT NULL
          AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
          AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND c.fecha_vencimiento::date >= CURRENT_DATE
          AND c.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '7 days'
        ORDER BY c.fecha_vencimiento ASC, j.apellido, j.nombre
        LIMIT 50
    """).fetchall()

    fichas = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.telefono,
            j.telefono_tutor,
            f.fecha_vencimiento,
            CASE
                WHEN NULLIF(f.fecha_vencimiento::text, '') IS NULL THEN 'faltante'
                WHEN f.fecha_vencimiento::text !~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN 'faltante'
                WHEN f.fecha_vencimiento::date < CURRENT_DATE THEN 'vencida'
                WHEN f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days' THEN 'por_vencer'
                ELSE 'vigente'
            END AS estado_documento
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.estado = 'Activo'
          AND (
              f.id IS NULL
              OR f.fecha_vencimiento IS NULL
              OR NULLIF(f.fecha_vencimiento::text, '') IS NULL
              OR (
                  f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
              )
          )
        ORDER BY estado_documento, f.fecha_vencimiento ASC NULLS FIRST, j.apellido, j.nombre
        LIMIT 80
    """).fetchall()

    asistencia_baja = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.telefono,
            j.telefono_tutor,
            COUNT(a.id) AS registros,
            SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END) AS presentes
        FROM jugadores j
        JOIN asistencias a ON a.jugador_id = j.id
        JOIN eventos_asistencia e ON e.id = a.evento_id
        WHERE j.estado = 'Activo'
          AND e.fecha::date >= CURRENT_DATE - INTERVAL '60 days'
        GROUP BY j.id, j.apellido, j.nombre, j.telefono, j.telefono_tutor
        HAVING COUNT(a.id) >= 3
           AND (SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END)::numeric / COUNT(a.id)) < 0.6
        ORDER BY (SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END)::numeric / COUNT(a.id)) ASC,
                 j.apellido, j.nombre
        LIMIT 50
    """).fetchall()

    comprobantes_pendientes = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.comprobante_fecha,
            c.comprobante_usuario,
            c.comprobante_nombre,
            c.comprobante_web_url,
            j.apellido,
            j.nombre
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.comprobante_drive_file_id IS NOT NULL
          AND COALESCE(c.anulada, 0) = 0
          AND COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
        ORDER BY c.comprobante_fecha DESC NULLS LAST, c.id DESC
        LIMIT 50
    """).fetchall()

    whatsapp_conversaciones = []
    if tiene_permiso("comunicaciones_ver"):
        whatsapp_conversaciones = conn.execute("""
            WITH mensajes AS (
                SELECT
                    wm.id,
                    wm.jugador_id,
                    COALESCE(NULLIF(wm.wa_id, ''), NULLIF(wm.telefono, '')) AS telefono,
                    wm.tipo,
                    wm.texto,
                    wm.creado_en,
                    j.nombre,
                    j.apellido,
                    j.categoria
                FROM whatsapp_mensajes wm
                LEFT JOIN jugadores j ON j.id = wm.jugador_id
                WHERE wm.direccion = 'in'
                  AND COALESCE(wm.leido, 0) = 0
                  AND COALESCE(NULLIF(wm.wa_id, ''), NULLIF(wm.telefono, '')) IS NOT NULL
            ),
            conteos AS (
                SELECT telefono, COUNT(*) AS sin_leer
                FROM mensajes
                GROUP BY telefono
            ),
            ultimos AS (
                SELECT DISTINCT ON (telefono)
                    id,
                    jugador_id,
                    telefono,
                    tipo,
                    texto,
                    creado_en,
                    nombre,
                    apellido,
                    categoria
                FROM mensajes
                ORDER BY telefono, creado_en DESC, id DESC
            )
            SELECT u.*, c.sin_leer
            FROM ultimos u
            JOIN conteos c ON c.telefono = u.telefono
            ORDER BY u.creado_en DESC, u.id DESC
            LIMIT 20
        """).fetchall()

    secretaria_documentos = []
    if tiene_permiso("secretaria_ver", "secretaria_gestionar"):
        secretaria_documentos = conn.execute("""
            SELECT
                id,
                categoria,
                titulo,
                descripcion,
                fecha_documento,
                fecha_vencimiento,
                archivo_nombre,
                creado_por,
                CASE
                    WHEN fecha_vencimiento::date < CURRENT_DATE THEN 'vencido'
                    ELSE 'por_vencer'
                END AS estado_vencimiento,
                CASE
                    WHEN fecha_vencimiento::date < CURRENT_DATE THEN CURRENT_DATE - fecha_vencimiento::date
                    ELSE fecha_vencimiento::date - CURRENT_DATE
                END AS dias
            FROM secretaria_documentos
            WHERE fecha_vencimiento IS NOT NULL
              AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
              AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
              AND fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
            ORDER BY fecha_vencimiento ASC, categoria, titulo
            LIMIT 50
        """).fetchall()

    ahijadxs_objetivo = conn.execute("""
        SELECT
            a.id AS aspirante_id,
            a.apellido,
            a.nombre,
            a.categoria,
            COALESCE(a.entrenamientos_objetivo, %s) AS entrenamientos_objetivo,
            COUNT(aa.id) FILTER (WHERE COALESCE(aa.presente, 0) = 1) AS entrenamientos_presentes
        FROM aspirantes a
        LEFT JOIN aspirante_asistencias aa ON aa.aspirante_id = a.id
        WHERE a.estado = 'Aspirante'
        GROUP BY a.id, a.apellido, a.nombre, a.categoria, a.entrenamientos_objetivo
        HAVING COUNT(aa.id) FILTER (WHERE COALESCE(aa.presente, 0) = 1) >= COALESCE(a.entrenamientos_objetivo, %s)
        ORDER BY a.apellido, a.nombre
        LIMIT 50
    """, (ASPIRANTE_ENTRENAMIENTOS_OBJETIVO, ASPIRANTE_ENTRENAMIENTOS_OBJETIVO)).fetchall()

    cambios_portal = []
    if tiene_permiso("alertas_portal", "auditoria_ver", "portal_jugador_gestionar"):
        cambios_portal = conn.execute("""
            SELECT
                a.id AS auditoria_id,
                a.fecha,
                a.detalle,
                j.id AS jugador_id,
                j.apellido,
                j.nombre,
                j.dni,
                j.email,
                j.telefono
            FROM auditoria a
            JOIN jugadores j ON j.id::text = a.entidad_id
            WHERE a.entidad = 'portal_jugador'
              AND a.accion = 'actualizar_contacto'
              AND a.fecha >= CURRENT_TIMESTAMP - INTERVAL '14 days'
            ORDER BY a.fecha DESC, a.id DESC
            LIMIT 50
        """).fetchall()
    conn.close()
    cuotas_vencidas = preparar_notificaciones_para_usuario(cuotas_vencidas, "cuota_vencida", lambda fila: fila["cuota_id"], descartadas)
    cuotas_por_vencer = preparar_notificaciones_para_usuario(cuotas_por_vencer, "cuota_por_vencer", lambda fila: fila["cuota_id"], descartadas)
    fichas = preparar_notificaciones_para_usuario(
        fichas,
        "ficha",
        lambda fila: f"{fila['jugador_id']}:{fila.get('estado_documento') or ''}:{fila.get('fecha_vencimiento') or ''}",
        descartadas,
    )
    asistencia_baja = preparar_notificaciones_para_usuario(asistencia_baja, "asistencia_baja", lambda fila: fila["jugador_id"], descartadas)
    comprobantes_pendientes = preparar_notificaciones_para_usuario(
        comprobantes_pendientes,
        "comprobante",
        lambda fila: f"{fila['cuota_id']}:{fila.get('comprobante_fecha') or ''}",
        descartadas,
    )
    whatsapp_conversaciones = preparar_notificaciones_para_usuario(
        whatsapp_conversaciones,
        "whatsapp",
        lambda fila: f"{fila['telefono']}:{fila['id']}",
        descartadas,
    )
    secretaria_documentos = preparar_notificaciones_para_usuario(
        secretaria_documentos,
        "secretaria_documento",
        lambda fila: f"{fila['id']}:{fila.get('fecha_vencimiento') or ''}",
        descartadas,
    )
    ahijadxs_objetivo = preparar_notificaciones_para_usuario(
        ahijadxs_objetivo,
        "ahijadx_objetivo",
        lambda fila: f"{fila['aspirante_id']}:{fila.get('entrenamientos_presentes') or 0}",
        descartadas,
    )
    cambios_portal = preparar_notificaciones_para_usuario(cambios_portal, "cambio_portal", lambda fila: fila["auditoria_id"], descartadas)
    for cambio in cambios_portal:
        cambio["detalle_resumen"] = resumen_auditoria_portal(cambio.get("detalle"))

    return {
        "cuotas_vencidas": cuotas_vencidas,
        "cuotas_por_vencer": cuotas_por_vencer,
        "fichas": fichas,
        "asistencia_baja": asistencia_baja,
        "comprobantes_pendientes": comprobantes_pendientes,
        "whatsapp_conversaciones": whatsapp_conversaciones,
        "secretaria_documentos": secretaria_documentos,
        "ahijadxs_objetivo": ahijadxs_objetivo,
        "cambios_portal": cambios_portal,
    }


def obtener_contador_notificaciones():
    if "user_id" not in session or not tiene_permiso("comunicaciones_ver"):
        return 0

    try:
        resumen = obtener_notificaciones_operativas()
    except Exception:
        app.logger.exception("No se pudo calcular el contador de notificaciones.")
        return 0

    return sum(
        len(resumen[campo])
        for campo in (
            "cuotas_vencidas",
            "cuotas_por_vencer",
            "fichas",
            "asistencia_baja",
            "comprobantes_pendientes",
            "whatsapp_conversaciones",
            "secretaria_documentos",
            "ahijadxs_objetivo",
            "cambios_portal",
        )
    )


def listar_tareas_sig(estado="pendiente", limite=80):
    conn = get_connection()
    tareas = conn.execute("""
        SELECT
            t.*,
            j.apellido,
            j.nombre,
            j.categoria
        FROM tareas_sig t
        LEFT JOIN jugadores j ON j.id = t.jugador_id
        WHERE (%s = 'todas' OR t.estado = %s)
        ORDER BY
            CASE t.prioridad WHEN 'alta' THEN 0 WHEN 'media' THEN 1 ELSE 2 END,
            NULLIF(t.fecha_vencimiento, '') ASC NULLS LAST,
            t.creado_en DESC
        LIMIT %s
    """, (estado, estado, limite)).fetchall()
    conn.close()
    return tareas


def obtener_revision_diaria():
    notificaciones = obtener_notificaciones_operativas() if tiene_permiso("comunicaciones_ver") else {}
    conn = get_connection()
    proximos_eventos = []
    if tiene_permiso("calendario_ver", "asistencia_ver"):
        proximos_eventos = conn.execute("""
            SELECT id, fecha, hora_inicio, tipo, descripcion
            FROM calendario_eventos
            WHERE fecha >= CURRENT_DATE::text
            ORDER BY fecha ASC, hora_inicio ASC NULLS LAST
            LIMIT 5
        """).fetchall()

    tareas_vencidas = conn.execute("""
        SELECT COUNT(*) AS total
        FROM tareas_sig
        WHERE estado = 'pendiente'
          AND NULLIF(fecha_vencimiento, '') IS NOT NULL
          AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND fecha_vencimiento::date < CURRENT_DATE
    """).fetchone()
    conn.close()

    def cantidad(clave):
        return len(notificaciones.get(clave, []))

    return {
        "whatsapp": cantidad("whatsapp_conversaciones"),
        "comprobantes": cantidad("comprobantes_pendientes"),
        "cambios_portal": cantidad("cambios_portal"),
        "cuotas": cantidad("cuotas_vencidas") + cantidad("cuotas_por_vencer"),
        "fichas": cantidad("fichas"),
        "asistencia_baja": cantidad("asistencia_baja"),
        "secretaria": cantidad("secretaria_documentos"),
        "ahijadxs": cantidad("ahijadxs_objetivo"),
        "proximos_eventos": proximos_eventos,
        "tareas_vencidas": int((tareas_vencidas or {}).get("total") or 0),
    }


def obtener_panel_cobranzas():
    conn = get_connection()
    resumen = conn.execute("""
        SELECT
            COUNT(*) FILTER (WHERE COALESCE(anulada, 0) = 0) AS emitidas,
            COUNT(*) FILTER (WHERE COALESCE(anulada, 0) = 0 AND pagado = 1) AS pagadas,
            COUNT(*) FILTER (
                WHERE COALESCE(anulada, 0) = 0
                  AND pagado = 0
                  AND COALESCE(importe, 0) > 0
            ) AS pendientes,
            COUNT(*) FILTER (
                WHERE COALESCE(anulada, 0) = 0
                  AND pagado = 0
                  AND COALESCE(importe, 0) > 0
                  AND fecha_vencimiento IS NOT NULL
                  AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
                  AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND fecha_vencimiento::date < CURRENT_DATE
            ) AS vencidas,
            COUNT(*) FILTER (
                WHERE comprobante_drive_file_id IS NOT NULL
                  AND COALESCE(anulada, 0) = 0
                  AND COALESCE(NULLIF(comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
            ) AS comprobantes,
            COALESCE(SUM(CASE WHEN COALESCE(anulada, 0) = 0 AND pagado = 1 THEN importe ELSE 0 END), 0) AS cobrado,
            COALESCE(SUM(CASE WHEN COALESCE(anulada, 0) = 0 AND pagado = 0 THEN importe ELSE 0 END), 0) AS pendiente_importe
        FROM cuotas
    """).fetchone()
    por_categoria = conn.execute("""
        SELECT
            COALESCE(j.categoria, 'Sin categoria') AS categoria,
            COUNT(c.id) FILTER (WHERE c.pagado = 0 AND COALESCE(c.importe, 0) > 0 AND COALESCE(c.anulada, 0) = 0) AS pendientes,
            COALESCE(SUM(CASE WHEN c.pagado = 0 AND COALESCE(c.anulada, 0) = 0 THEN c.importe ELSE 0 END), 0) AS deuda
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        GROUP BY COALESCE(j.categoria, 'Sin categoria')
        HAVING COUNT(c.id) FILTER (WHERE c.pagado = 0 AND COALESCE(c.importe, 0) > 0 AND COALESCE(c.anulada, 0) = 0) > 0
        ORDER BY deuda DESC, categoria
        LIMIT 20
    """).fetchall()
    recientes = conn.execute("""
        SELECT c.id, c.periodo, c.importe, c.comprobante_fecha, c.comprobante_usuario, j.id AS jugador_id, j.apellido, j.nombre
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.comprobante_drive_file_id IS NOT NULL
          AND COALESCE(c.anulada, 0) = 0
          AND COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
        ORDER BY c.comprobante_fecha DESC NULLS LAST, c.id DESC
        LIMIT 12
    """).fetchall()
    conn.close()
    emitidas = int((resumen or {}).get("emitidas") or 0)
    pagadas = int((resumen or {}).get("pagadas") or 0)
    avance = round((pagadas / emitidas) * 100, 1) if emitidas else None
    return {
        "resumen": resumen or {},
        "avance": avance,
        "por_categoria": por_categoria,
        "comprobantes_recientes": recientes,
    }


def inicio_mes_actual():
    return date.today().replace(day=1).strftime("%Y-%m-%d")


def fin_mes_actual():
    hoy = date.today()
    if hoy.month == 12:
        siguiente = date(hoy.year + 1, 1, 1)
    else:
        siguiente = date(hoy.year, hoy.month + 1, 1)
    return (siguiente - timedelta(days=1)).strftime("%Y-%m-%d")


def whatsapp_mensaje(telefono, mensaje):
    telefono_whatsapp = normalizar_telefono_whatsapp(telefono)
    if telefono_whatsapp:
        return f"https://wa.me/{telefono_whatsapp}?text={quote(mensaje)}"
    return f"https://wa.me/?text={quote(mensaje)}"


def normalizar_header_simple(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().lower())
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "_", texto).strip("_")


def extraer_valor_fila(fila, opciones):
    for opcion in opciones:
        if opcion in fila and str(fila[opcion] or "").strip():
            return str(fila[opcion] or "").strip()
    return ""


def parsear_importe(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    texto = re.sub(r"[^0-9,.-]", "", texto)
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return round(float(texto), 2)
    except ValueError:
        return None


def obtener_presupuesto_mensual(mes, meses_proyeccion=6):
    mes = normalizar_mes(mes, ahora_sig().strftime("%Y-%m"))
    meses = meses_entre(mes, sumar_meses(mes, max(1, min(meses_proyeccion, 24)) - 1))
    conn = get_connection()

    items = conn.execute("""
        SELECT *
        FROM presupuesto_items
        ORDER BY activo DESC, tipo ASC, categoria ASC, concepto ASC, id ASC
    """).fetchall()

    movimientos = conn.execute("""
        SELECT
            substring(fecha from 1 for 7) AS mes,
            COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
            COALESCE(SUM(CASE WHEN tipo = 'egreso' THEN monto ELSE 0 END), 0) AS egresos
        FROM movimientos
        WHERE COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = ANY(%s)
        GROUP BY substring(fecha from 1 for 7)
    """, (meses,)).fetchall()

    cuotas = conn.execute("""
        SELECT
            periodo AS mes,
            COUNT(*) AS cuotas_emitidas,
            COALESCE(SUM(importe), 0) AS total_emitido,
            COALESCE(SUM(CASE WHEN pagado = 1 THEN importe ELSE 0 END), 0) AS total_cobrado,
            COALESCE(SUM(CASE WHEN pagado = 0 AND COALESCE(anulada, 0) = 0 THEN importe ELSE 0 END), 0) AS total_pendiente
        FROM cuotas
        WHERE periodo = ANY(%s)
        GROUP BY periodo
    """, (meses,)).fetchall()
    conn.close()

    movimientos_por_mes = {fila["mes"]: fila for fila in movimientos}
    cuotas_por_mes = {fila["mes"]: fila for fila in cuotas}
    proyeccion = []
    for mes_item in meses:
        ingresos_presupuestados = 0
        egresos_presupuestados = 0
        activos = []
        for item in items:
            if not item.get("activo"):
                continue
            if item.get("mes_inicio") and item["mes_inicio"] > mes_item:
                continue
            if item.get("mes_fin") and item["mes_fin"] < mes_item:
                continue
            monto = float(item.get("monto") or 0)
            activos.append(item)
            if item["tipo"] == "ingreso":
                ingresos_presupuestados += monto
            else:
                egresos_presupuestados += monto

        caja = movimientos_por_mes.get(mes_item, {})
        cuota = cuotas_por_mes.get(mes_item, {})
        ingresos_reales = float(caja.get("ingresos") or 0)
        egresos_reales = float(caja.get("egresos") or 0)
        resultado_presupuestado = ingresos_presupuestados - egresos_presupuestados
        resultado_real = ingresos_reales - egresos_reales
        proyeccion.append({
            "mes": mes_item,
            "ingresos_presupuestados": ingresos_presupuestados,
            "egresos_presupuestados": egresos_presupuestados,
            "resultado_presupuestado": resultado_presupuestado,
            "ingresos_reales": ingresos_reales,
            "egresos_reales": egresos_reales,
            "resultado_real": resultado_real,
            "desvio": resultado_real - resultado_presupuestado,
            "cuotas_emitidas": cuota.get("cuotas_emitidas", 0) or 0,
            "cuotas_emitidas_total": cuota.get("total_emitido", 0) or 0,
            "cuotas_cobradas": cuota.get("total_cobrado", 0) or 0,
            "cuotas_pendientes": cuota.get("total_pendiente", 0) or 0,
            "items": activos,
        })

    mes_actual = proyeccion[0] if proyeccion else {}
    return {
        "mes": mes,
        "meses": meses,
        "items": items,
        "items_activos": mes_actual.get("items", []),
        "proyeccion": proyeccion,
        "resumen": mes_actual,
    }


def leer_csv_conciliacion(archivo):
    nombre = (archivo.filename or "").lower()
    if nombre.endswith(".xlsx"):
        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
        except Exception as error:
            raise ValueError("No se pudo leer el archivo Excel.") from error

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [normalizar_header_simple(valor) for valor in rows[0]]
        filas = []
        for row in rows[1:]:
            normalizada = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                valor = row[index] if index < len(row) else None
                normalizada[header] = limpiar_valor_excel(valor)
            if any(str(v or "").strip() for v in normalizada.values()):
                filas.append(normalizada)
        return filas

    raw = archivo.read()
    texto = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            texto = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ValueError("No se pudo leer el archivo CSV.")

    sample = texto[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(io.StringIO(texto), dialect=dialect)
    filas = []
    for row in reader:
        normalizada = {normalizar_header_simple(k): v for k, v in row.items() if k is not None}
        if any(str(v or "").strip() for v in normalizada.values()):
            filas.append(normalizada)
    return filas


def buscar_match_conciliacion(conn, fila):
    dni = extraer_valor_fila(fila, ["dni", "documento", "nro_documento", "numero_documento"])
    periodo = extraer_valor_fila(fila, ["periodo", "mes", "cuota", "periodo_cuota"])
    referencia = extraer_valor_fila(fila, ["referencia", "comprobante", "id", "operacion", "transaccion"])
    fecha_pago = validar_fecha_movimiento(extraer_valor_fila(fila, ["fecha", "fecha_pago", "fecha_de_pago"]))
    importe = parsear_importe(extraer_valor_fila(fila, ["importe", "monto", "valor", "total"]))

    if periodo:
        periodo = periodo[:7]

    if not dni or importe is None:
        return {
            "estado": "error",
            "motivo": "Falta DNI o importe.",
            "dni": dni,
            "periodo": periodo,
            "importe": importe,
            "referencia": referencia,
            "fecha_pago": fecha_pago,
        }

    condiciones = [
        "j.dni = %s",
        "c.pagado = 0",
        "COALESCE(c.importe, 0) > 0",
        "ABS(COALESCE(c.importe, 0) - %s) < 1",
    ]
    parametros = [dni, importe]
    if periodo:
        condiciones.append("c.periodo = %s")
        parametros.append(periodo)

    matches = conn.execute(f"""
        SELECT
            c.id AS cuota_id,
            c.jugador_id,
            c.periodo,
            c.importe,
            j.apellido,
            j.nombre,
            j.dni
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE {" AND ".join(condiciones)}
        ORDER BY c.periodo ASC, c.id ASC
        LIMIT 3
    """, parametros).fetchall()

    base = {
        "dni": dni,
        "periodo": periodo,
        "importe": importe,
        "referencia": referencia,
        "fecha_pago": fecha_pago,
    }
    if len(matches) == 1:
        match = matches[0]
        return {
            **base,
            "estado": "match",
            "cuota_id": match["cuota_id"],
            "jugador_id": match["jugador_id"],
            "jugador": f"{match['apellido']}, {match['nombre']}",
            "cuota_periodo": match["periodo"],
            "cuota_importe": match["importe"],
        }
    if len(matches) > 1:
        return {**base, "estado": "multiple", "motivo": "Hay mas de una cuota posible."}
    return {**base, "estado": "sin_match", "motivo": "No se encontro cuota pendiente compatible."}


def siguiente_numero_recibo(conn):
    conn.execute("SELECT pg_advisory_xact_lock(hashtext('sig_numero_recibo'))")
    ultimo = conn.execute("""
        SELECT MAX(numero_recibo) AS maximo
        FROM cuotas
    """).fetchone()["maximo"] or 0
    return ultimo + 1


def aplicar_matches_conciliacion(conn, matches):
    aplicados = 0
    omitidos = 0

    for match in matches:
        cuota_id = match.get("cuota_id")
        if match.get("estado") != "match" or not cuota_id:
            omitidos += 1
            continue

        revalidado = buscar_match_conciliacion(conn, match)
        try:
            cuota_id = int(cuota_id)
        except (TypeError, ValueError):
            omitidos += 1
            continue

        if revalidado.get("estado") != "match" or revalidado.get("cuota_id") != cuota_id:
            omitidos += 1
            continue

        cuota = conn.execute("""
            SELECT
                c.*,
                j.apellido,
                j.nombre
            FROM cuotas c
            JOIN jugadores j ON j.id = c.jugador_id
            WHERE c.id = %s
            FOR UPDATE
        """, (cuota_id,)).fetchone()
        if not cuota or cuota["pagado"]:
            omitidos += 1
            continue

        fecha_pago = match.get("fecha_pago") or ahora_sig().strftime("%Y-%m-%d")
        referencia = match.get("referencia") or "Conciliacion importada"
        numero_recibo = cuota["numero_recibo"] or siguiente_numero_recibo(conn)
        conn.execute("""
            UPDATE cuotas
            SET pagado = 1,
                fecha_pago = %s,
                numero_recibo = %s,
                metodo_pago = 'Conciliacion',
                referencia_pago = %s
            WHERE id = %s
        """, (fecha_pago, numero_recibo, referencia, cuota["id"]))

        conn.execute("""
            INSERT INTO movimientos (tipo, concepto, monto, fecha, referencia)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            "ingreso",
            f"Cuota {cuota['periodo']} - {cuota['apellido']}, {cuota['nombre']}",
            cuota["importe"],
            fecha_pago,
            "Cuota Social (Conciliacion)",
        ))
        aplicados += 1

    return {"aplicados": aplicados, "omitidos": omitidos}


def normalizar_header_excel(valor):
    texto = str(valor or "").strip().lower()
    reemplazos = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto


def limpiar_valor_excel(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip()


def mapear_fila_jugador(headers, row):
    aliases = {
        "nombre": "nombre",
        "apellido": "apellido",
        "dni": "dni",
        "documento": "dni",
        "fecha_nacimiento": "fecha_nacimiento",
        "nacimiento": "fecha_nacimiento",
        "telefono": "telefono",
        "telefono_jugador": "telefono",
        "tel_fono": "telefono",
        "celular": "telefono",
        "email": "email",
        "mail": "email",
        "categoria": "categoria",
        "fecha_ingreso": "fecha_ingreso",
        "ingreso": "fecha_ingreso",
        "estado": "estado",
        "tipo": "tipo_miembro",
        "tipo_miembro": "tipo_miembro",
        "tipo_de_miembro": "tipo_miembro",
        "cobra_cuota": "cobra_cuota",
        "cuota": "cobra_cuota",
        "contacto_tutor": "contacto_tutor",
        "tutor": "contacto_tutor",
        "responsable": "contacto_tutor",
        "parentesco_tutor": "parentesco_tutor",
        "parentesco": "parentesco_tutor",
        "telefono_tutor": "telefono_tutor",
        "telefono_del_tutor": "telefono_tutor",
        "tel_tutor": "telefono_tutor",
        "tel_fono_tutor": "telefono_tutor",
        "email_tutor": "email_tutor",
        "direccion": "direccion",
        "obra_social": "obra_social",
        "numero_afiliado_obra_social": "numero_afiliado_obra_social",
        "nro_afiliado_obra_social": "numero_afiliado_obra_social",
        "numero_obra_social": "numero_afiliado_obra_social",
        "nro_obra_social": "numero_afiliado_obra_social",
        "numero_socio_obra_social": "numero_afiliado_obra_social",
        "numero_socio": "numero_socio",
        "nro_socio": "numero_socio",
        "documentos": "documentos",
        "observaciones": "observaciones",
    }

    data = {
        "nombre": "",
        "apellido": "",
        "dni": "",
        "fecha_nacimiento": "",
        "telefono": "",
        "email": "",
        "categoria": "",
        "fecha_ingreso": "",
        "estado": "Activo",
        "tipo_miembro": "Jugador",
        "cobra_cuota": "1",
        "contacto_tutor": "",
        "parentesco_tutor": "",
        "telefono_tutor": "",
        "email_tutor": "",
        "direccion": "",
        "obra_social": "",
        "numero_afiliado_obra_social": "",
        "numero_socio": "",
        "documentos": "",
        "observaciones": "",
    }

    for index, header in enumerate(headers):
        campo = aliases.get(header)
        if campo:
            data[campo] = limpiar_valor_excel(row[index] if index < len(row) else None)

    if not data["estado"]:
        data["estado"] = "Activo"
    data["tipo_miembro"] = normalizar_tipo_miembro(data["tipo_miembro"])
    data["cobra_cuota"] = 0 if str(data["cobra_cuota"]).strip().lower() in {"0", "no", "false", "off"} else 1

    return data


def parsear_puntaje_test(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    texto = re.sub(r"[^0-9,.-]", "", texto)
    if not texto:
        return None
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def normalizar_fecha_test(valor, usar_hoy_si_vacia=True):
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor or "").strip()
    if not texto:
        return ahora_sig().strftime("%Y-%m-%d") if usar_hoy_si_vacia else None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto[:10], formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return validar_fecha_movimiento(texto)


def obtener_test_tipos(conn, solo_activos=True):
    where = "WHERE activo = 1" if solo_activos else ""
    return conn.execute(f"""
        SELECT *
        FROM test_tipos
        {where}
        ORDER BY nombre
    """).fetchall()


def sugerir_jugador_por_nombre_test(datos, jugadores):
    nombre_completo = str(
        datos.get("jugador") or
        datos.get("nombre_completo") or
        datos.get("jugador_nombre") or
        ""
    ).strip()
    nombre = str(datos.get("nombre") or "").strip()
    apellido = str(datos.get("apellido") or "").strip()
    texto = " ".join(parte for parte in (nombre_completo, apellido, nombre) if parte).strip()

    if not texto:
        return None, "sin_coincidencia", "Sin nombre"

    texto_norm = normalizar_texto_match(texto)
    candidatos = []
    for jugador in jugadores:
        nombre_tokens = normalizar_texto_match(jugador.get("nombre")).split()
        apellido_tokens = normalizar_texto_match(jugador.get("apellido")).split()
        tokens = [token for token in apellido_tokens + nombre_tokens if len(token) >= 3]
        if not tokens:
            continue

        coincidencias = sum(1 for token in tokens if token in texto_norm)
        jugador_nombre_completo = normalizar_texto_match(f"{jugador.get('apellido')} {jugador.get('nombre')}")
        jugador_nombre_inverso = normalizar_texto_match(f"{jugador.get('nombre')} {jugador.get('apellido')}")

        if texto_norm in {jugador_nombre_completo, jugador_nombre_inverso}:
            candidatos.append((100, jugador, "alta", "Nombre completo exacto"))
        elif coincidencias == len(tokens):
            candidatos.append((90, jugador, "media", "Nombre y apellido detectados"))
        elif apellido_tokens and all(token in texto_norm for token in apellido_tokens):
            candidatos.append((70, jugador, "baja", "Apellido detectado"))

    if not candidatos:
        return None, "sin_coincidencia", "Sin coincidencia"

    candidatos.sort(key=lambda item: item[0], reverse=True)
    mejor = candidatos[0]
    if len(candidatos) > 1 and candidatos[1][0] == mejor[0]:
        return None, "baja", "Coincidencia ambigua"
    return mejor[1], mejor[2], mejor[3]


def obtener_test_importaciones_batch_recientes(conn):
    return conn.execute("""
        SELECT
            batch_id,
            MIN(creado_en) AS creado_en,
            MAX(creado_por) AS creado_por,
            COUNT(*) AS total,
            SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN estado = 'procesado' THEN 1 ELSE 0 END) AS procesadas,
            MAX(id) AS ultimo_id
        FROM test_importaciones_batch
        GROUP BY batch_id
        ORDER BY ultimo_id DESC
        LIMIT 10
    """).fetchall()


def construir_grafico_tests(resultados):
    if not resultados:
        return {
            "series": [],
            "fechas": [],
            "minimo": 0,
            "maximo": 0,
            "svg_width": 920,
            "svg_height": 320,
            "eje_y": [],
        }

    fechas = sorted({fila["fecha"] for fila in resultados if fila["fecha"]})
    valores = [float(fila["puntaje"] or 0) for fila in resultados]
    minimo = min(valores)
    maximo = max(valores)
    if minimo == maximo:
        minimo -= 1
        maximo += 1

    width = 920
    height = 320
    pad_left = 54
    pad_right = 24
    pad_top = 22
    pad_bottom = 42
    inner_w = width - pad_left - pad_right
    inner_h = height - pad_top - pad_bottom
    palette = ["#2563eb", "#16a34a", "#dc2626", "#9333ea", "#ea580c", "#0891b2", "#be123c", "#4f46e5"]

    fecha_x = {}
    total_fechas = max(1, len(fechas) - 1)
    for index, fecha in enumerate(fechas):
        fecha_x[fecha] = pad_left + (inner_w * index / total_fechas if len(fechas) > 1 else inner_w / 2)

    agrupado = {}
    for fila in resultados:
        clave = fila["jugador_id"]
        agrupado.setdefault(clave, {
            "jugador_id": clave,
            "nombre": f"{fila['apellido']}, {fila['nombre']}",
            "categoria": fila["categoria"] or "-",
            "puntos": [],
        })
        valor = float(fila["puntaje"] or 0)
        y = pad_top + inner_h - ((valor - minimo) / (maximo - minimo) * inner_h)
        agrupado[clave]["puntos"].append({
            "fecha": fila["fecha"],
            "valor": valor,
            "x": round(fecha_x[fila["fecha"]], 2),
            "y": round(y, 2),
        })

    series = []
    for index, serie in enumerate(agrupado.values()):
        serie["puntos"].sort(key=lambda item: item["fecha"])
        serie["polyline"] = " ".join(f"{p['x']},{p['y']}" for p in serie["puntos"])
        serie["color"] = palette[index % len(palette)]
        serie["ultimo"] = serie["puntos"][-1]["valor"] if serie["puntos"] else None
        series.append(serie)

    eje_y = []
    for index in range(5):
        valor = maximo - ((maximo - minimo) * index / 4)
        y = pad_top + inner_h * index / 4
        eje_y.append({"valor": round(valor, 2), "y": round(y, 2)})

    return {
        "series": series,
        "fechas": fechas,
        "minimo": round(minimo, 2),
        "maximo": round(maximo, 2),
        "svg_width": width,
        "svg_height": height,
        "pad_left": pad_left,
        "pad_right": pad_right,
        "pad_top": pad_top,
        "pad_bottom": pad_bottom,
        "inner_w": inner_w,
        "inner_h": inner_h,
        "eje_y": eje_y,
    }


def construir_comparativo_tests(resultados, test_actual):
    fechas = sorted({fila["fecha"] for fila in resultados if fila["fecha"]})
    mayor_es_mejor = True
    if test_actual is not None:
        mayor_es_mejor = bool(test_actual["mayor_es_mejor"])

    agrupado = {}
    for fila in resultados:
        jugador_id = fila["jugador_id"]
        jugador = agrupado.setdefault(jugador_id, {
            "jugador_id": jugador_id,
            "nombre": f"{fila['apellido']}, {fila['nombre']}",
            "categoria": fila["categoria"] or "-",
            "valores": {},
            "ordenados": [],
            "primero": None,
            "ultimo": None,
            "delta": None,
            "estado": "sin_datos",
            "estado_label": "Sin datos",
        })
        valor = float(fila["puntaje"] or 0)
        punto = {"fecha": fila["fecha"], "valor": valor}
        jugador["valores"][fila["fecha"]] = valor
        jugador["ordenados"].append(punto)

    filas = []
    for jugador in agrupado.values():
        jugador["ordenados"].sort(key=lambda item: item["fecha"])
        if jugador["ordenados"]:
            jugador["primero"] = jugador["ordenados"][0]["valor"]
            jugador["ultimo"] = jugador["ordenados"][-1]["valor"]

        if len(jugador["ordenados"]) >= 2:
            anterior = jugador["ordenados"][-2]["valor"]
            ultimo = jugador["ultimo"]
            delta = ultimo - anterior
            jugador["delta"] = round(delta, 2)
            if delta == 0:
                jugador["estado"] = "igual"
                jugador["estado_label"] = "Sin cambio"
            elif (delta > 0 and mayor_es_mejor) or (delta < 0 and not mayor_es_mejor):
                jugador["estado"] = "mejora"
                jugador["estado_label"] = "Mejora"
            else:
                jugador["estado"] = "empeora"
                jugador["estado_label"] = "Empeora"
        elif jugador["ordenados"]:
            jugador["estado"] = "unico"
            jugador["estado_label"] = "Sin comparar"

        filas.append(jugador)

    filas.sort(key=lambda item: (item["categoria"], item["nombre"]))
    return {
        "fechas": fechas,
        "filas": filas,
        "mayor_es_mejor": mayor_es_mejor,
    }


AUDIT_ENDPOINTS = {
    "nuevo_movimiento": ("crear", "movimiento_caja"),
    "editar_movimiento": ("editar", "movimiento_caja"),
    "eliminar_movimiento": ("anular", "movimiento_caja"),
    "cerrar_mes": ("cerrar", "caja"),
    "nuevo_jugador": ("crear", "jugador"),
    "editar_jugador": ("editar", "jugador"),
    "eliminar_jugador": ("eliminar", "jugador"),
    "nueva_bitacora_jugador": ("crear", "bitacora_jugador"),
    "importar_jugadores": ("importar", "jugadores"),
    "acciones_masivas_jugadores": ("accion_masiva", "jugadores"),
    "nuevo_aspirante": ("crear", "ahijadx"),
    "editar_aspirante": ("editar", "ahijadx"),
    "convertir_aspirante": ("convertir", "ahijadx"),
    "eliminar_aspirante": ("eliminar", "ahijadx"),
    "nueva_cuota": ("crear", "cuota"),
    "pagar_cuota": ("pagar", "cuota"),
    "subir_comprobante_cuota": ("subir", "comprobante_cuota"),
    "revisar_comprobante_cuota": ("revisar", "comprobante_cuota"),
    "eliminar_cuota": ("eliminar", "cuota"),
    "generar_cuotas": ("generar", "cuotas"),
    "recalcular_becas_jugador": ("recalcular_beca", "cuotas"),
    "nuevo_plan_pago": ("crear", "plan_pago"),
    "editar_plan_pago": ("editar", "plan_pago"),
    "actualizar_plan_pago": ("actualizar", "plan_pago"),
    "eliminar_plan_pago": ("eliminar", "plan_pago"),
    "conciliar_pagos": ("conciliar", "cuotas"),
    "nuevo_test_tipo": ("crear", "test_deportivo"),
    "cargar_test_resultados": ("cargar", "test_deportivo"),
    "importar_test_resultados": ("importar", "test_deportivo"),
    "revisar_test_importacion": ("confirmar_importacion", "test_deportivo"),
    "exportar_tests": ("exportar", "test_deportivo"),
    "editar_ficha_medica": ("editar", "ficha_medica"),
    "cargar_fichas_medicas_batch": ("cargar_batch", "ficha_medica"),
    "revisar_fichas_medicas_batch": ("confirmar_batch", "ficha_medica"),
    "nuevo_documento_jugador": ("crear", "documento"),
    "eliminar_documento_jugador": ("eliminar", "documento"),
    "nueva_lesion": ("crear", "lesion"),
    "editar_lesion": ("editar", "lesion"),
    "eliminar_lesion": ("eliminar", "lesion"),
    "nuevo_usuario": ("crear", "usuario"),
    "editar_usuario": ("editar", "usuario"),
    "eliminar_usuario": ("eliminar", "usuario"),
    "nuevo_rol": ("crear", "rol"),
    "editar_rol": ("editar", "rol"),
    "eliminar_rol": ("eliminar", "rol"),
    "configurar_mantenimiento": ("configurar", "mantenimiento"),
    "actualizar_sugerencia_recomendacion": ("actualizar", "sugerencia_recomendacion"),
    "reenviar_sugerencia_recomendacion": ("reenviar", "sugerencia_recomendacion"),
    "configurar_sugerencias_recomendaciones": ("configurar", "sugerencia_recomendacion"),
    "cambiar_mi_password": ("cambiar_password", "usuario"),
    "resetear_password_usuario": ("resetear_password", "usuario"),
    "solicitar_recuperacion_password": ("solicitar_reset", "usuario"),
    "restablecer_password": ("restablecer_password", "usuario"),
    "descartar_onboarding": ("descartar", "onboarding"),
    "generar_portal_jugador": ("generar", "portal_jugador"),
    "desactivar_portal_jugador": ("desactivar", "portal_jugador"),
    "nuevo_evento_asistencia": ("crear", "evento_asistencia"),
    "tomar_asistencia": ("registrar", "asistencia"),
    "nuevo_evento_calendario": ("crear", "evento_calendario"),
}

SENSITIVE_AUDIT_FIELDS = {
    "password",
    "pass",
    "clave",
    "contrase?a",
    "contrasena",
    "admin_password",
    "db_pass",
    "secret_key",
}


def audit_request_ip():
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr


def truncate_audit_value(value, max_length=300):
    if value is None:
        return None
    value = str(value)
    if len(value) <= max_length:
        return value
    return value[:max_length] + "...[truncado]"


def sanitized_audit_form():
    data = {}
    for key, values in request.form.lists():
        key_lower = key.lower()
        if any(sensitive in key_lower for sensitive in SENSITIVE_AUDIT_FIELDS):
            data[key] = "[redactado]"
            continue

        clean_values = [truncate_audit_value(value) for value in values]
        data[key] = clean_values[0] if len(clean_values) == 1 else clean_values
    return data


def audit_entity_id():
    if not request.view_args:
        return None
    for key in (
        "jugador_id", "cuota_id", "plan_id", "documento_id", "lesion_id",
        "usuario_id", "rol_id", "movimiento_id", "evento_id"
    ):
        if key in request.view_args:
            return str(request.view_args[key])
    return None


def registrar_auditoria(
    accion,
    entidad="sistema",
    entidad_id=None,
    detalle=None,
    usuario_id=None,
    username=None,
    rol=None,
):
    conn = None
    try:
        if has_request_context():
            usuario_id = usuario_id if usuario_id is not None else session.get("user_id")
            username = username if username is not None else session.get("username")
            rol = rol if rol is not None else session.get("rol")
            ip = audit_request_ip()
            user_agent = request.headers.get("User-Agent", "")
        else:
            ip = None
            user_agent = None

        if isinstance(detalle, str):
            detalle_texto = detalle
        else:
            detalle_texto = json.dumps(detalle or {}, ensure_ascii=False, default=str)

        conn = get_connection()
        conn.execute("""
            INSERT INTO auditoria (
                usuario_id, username, rol, accion, entidad, entidad_id,
                detalle, ip, user_agent
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            usuario_id,
            username,
            rol,
            accion,
            entidad,
            entidad_id,
            detalle_texto,
            ip,
            truncate_audit_value(user_agent, 500),
        ))
        conn.commit()
        conn.close()
    except Exception:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


def init_db():
    conn = get_connection()
    conn.execute("SELECT pg_advisory_xact_lock(hashtext('sig_init_db'))")

    conn.execute("""
    CREATE TABLE IF NOT EXISTS jugadores (
        id SERIAL PRIMARY KEY,
        nombre TEXT NOT NULL,
        apellido TEXT NOT NULL,
        dni TEXT,
        fecha_nacimiento TEXT,
        telefono TEXT,
        email TEXT,
        categoria TEXT,
        fecha_ingreso TEXT,
        estado TEXT NOT NULL DEFAULT 'Activo',
        observaciones TEXT
    )
""")

    columnas_jugadores = get_columns(conn, "jugadores")
    columnas_extra_jugador = {
        "telefono_tutor": "TEXT",
        "email_tutor": "TEXT",
        "contacto_tutor": "TEXT",
        "parentesco_tutor": "TEXT",
        "direccion": "TEXT",
        "obra_social": "TEXT",
        "numero_afiliado_obra_social": "TEXT",
        "numero_socio": "TEXT",
        "documentos": "TEXT",
        "beca_activa": "INTEGER DEFAULT 0",
        "beca_porcentaje": "REAL DEFAULT 0",
        "beca_desde": "TEXT",
        "beca_hasta": "TEXT",
        "beca_motivo": "TEXT",
        "portal_token": "TEXT",
        "portal_activo": "INTEGER DEFAULT 0",
        "portal_actualizado_en": "TEXT",
        "tipo_miembro": "TEXT DEFAULT 'Jugador'",
        "cobra_cuota": "INTEGER DEFAULT 1",
    }

    for columna, tipo_columna in columnas_extra_jugador.items():
        if columna not in columnas_jugadores:
            conn.execute(f"ALTER TABLE jugadores ADD COLUMN {columna} {tipo_columna}")

    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_jugadores_portal_token
        ON jugadores (portal_token)
        WHERE portal_token IS NOT NULL
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS aspirantes (
            id SERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            apellido TEXT NOT NULL,
            dni TEXT,
            fecha_nacimiento TEXT,
            telefono TEXT,
            email TEXT,
            categoria TEXT,
            fecha_postulacion TEXT,
            estado TEXT NOT NULL DEFAULT 'Aspirante',
            madrina_jugador_id INTEGER,
            entrenamientos_objetivo INTEGER DEFAULT 8,
            fecha_ingreso_club TEXT,
            jugador_id INTEGER,
            observaciones TEXT,
            FOREIGN KEY (madrina_jugador_id) REFERENCES jugadores(id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    columnas_aspirantes = get_columns(conn, "aspirantes")
    columnas_extra_aspirante = {
        "dni": "TEXT",
        "fecha_nacimiento": "TEXT",
        "telefono": "TEXT",
        "email": "TEXT",
        "categoria": "TEXT",
        "fecha_postulacion": "TEXT",
        "estado": "TEXT NOT NULL DEFAULT 'Aspirante'",
        "madrina_jugador_id": "INTEGER",
        "entrenamientos_objetivo": "INTEGER DEFAULT 8",
        "fecha_ingreso_club": "TEXT",
        "jugador_id": "INTEGER",
        "observaciones": "TEXT",
    }

    for columna, tipo_columna in columnas_extra_aspirante.items():
        if columna not in columnas_aspirantes:
            conn.execute(f"ALTER TABLE aspirantes ADD COLUMN {columna} {tipo_columna}")

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_aspirantes_estado
        ON aspirantes (estado)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            clave TEXT PRIMARY KEY,
            valor TEXT,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_por TEXT
        )
    """)

    columnas_app_settings = get_columns(conn, "app_settings")
    if "actualizado_en" not in columnas_app_settings:
        conn.execute("ALTER TABLE app_settings ADD COLUMN actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP")
    if "actualizado_por" not in columnas_app_settings:
        conn.execute("ALTER TABLE app_settings ADD COLUMN actualizado_por TEXT")

    conn.execute("""
        INSERT INTO app_settings (clave, valor, actualizado_por)
        VALUES ('maintenance_mode', 'false', 'sistema')
        ON CONFLICT(clave) DO NOTHING
    """)

    conn.execute("""
        INSERT INTO app_settings (clave, valor, actualizado_por)
        VALUES ('maintenance_message', %s, 'sistema')
        ON CONFLICT(clave) DO NOTHING
    """, (MAINTENANCE_DEFAULT_MESSAGE,))

    conn.execute("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            version TEXT PRIMARY KEY,
            aplicado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        INSERT INTO schema_migrations (version)
        VALUES ('2026-06-14-paquete-gestion-v1')
        ON CONFLICT(version) DO NOTHING
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS automatizacion_ejecuciones (
            id SERIAL PRIMARY KEY,
            tipo TEXT NOT NULL,
            clave TEXT NOT NULL,
            estado TEXT NOT NULL,
            detalle TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(tipo, clave)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_automatizacion_ejecuciones_tipo_fecha
        ON automatizacion_ejecuciones (tipo, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS public_rate_limits (
            id SERIAL PRIMARY KEY,
            endpoint TEXT NOT NULL,
            ip TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_public_rate_limits_endpoint_ip_fecha
        ON public_rate_limits (endpoint, ip, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS pwa_push_subscriptions (
            id SERIAL PRIMARY KEY,
            endpoint TEXT NOT NULL UNIQUE,
            actor_tipo TEXT NOT NULL,
            usuario_id INTEGER,
            jugador_id INTEGER,
            portal_token_hash TEXT,
            subscription_json TEXT NOT NULL,
            user_agent TEXT,
            enabled INTEGER DEFAULT 1,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_pwa_push_subscriptions_actor
        ON pwa_push_subscriptions (actor_tipo, usuario_id, jugador_id, enabled)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS pwa_push_envios (
            id SERIAL PRIMARY KEY,
            titulo TEXT NOT NULL,
            mensaje TEXT NOT NULL,
            destino TEXT NOT NULL,
            categoria TEXT,
            jugador_id INTEGER,
            url TEXT,
            enviados INTEGER DEFAULT 0,
            errores INTEGER DEFAULT 0,
            detalle TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_pwa_push_envios_fecha
        ON pwa_push_envios (creado_en DESC)
    """)

    columnas_pwa_envios = get_columns(conn, "pwa_push_envios")
    columnas_extra_pwa_envios = {
        "mostrar_portal": "INTEGER DEFAULT 0",
        "visible_hasta": "TEXT",
    }
    for columna, definicion in columnas_extra_pwa_envios.items():
        if columna not in columnas_pwa_envios:
            conn.execute(f"ALTER TABLE pwa_push_envios ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_pwa_push_envios_portal
        ON pwa_push_envios (mostrar_portal, visible_hasta, destino, categoria, jugador_id)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS cuotas (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER,
            periodo TEXT,
            importe REAL,
            pagado INTEGER DEFAULT 0,
            fecha_vencimiento TEXT,
            fecha_pago TEXT
        )
    """)


    columnas_cuotas = get_columns(conn, "cuotas")
    if "numero_recibo" not in columnas_cuotas:
        conn.execute("ALTER TABLE cuotas ADD COLUMN numero_recibo INTEGER")



    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_cuotas_jugador_periodo
        ON cuotas (jugador_id, periodo)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS gastos_compartidos (
            id SERIAL PRIMARY KEY,
            titulo TEXT NOT NULL,
            concepto TEXT,
            calendario_evento_id INTEGER,
            fecha_evento TEXT,
            fecha_vencimiento TEXT,
            modo_importe TEXT NOT NULL DEFAULT 'por_jugador',
            monto_total REAL,
            monto_por_jugador REAL,
            estado TEXT NOT NULL DEFAULT 'Activo',
            observaciones TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_gastos_compartidos_fecha
        ON gastos_compartidos (fecha_vencimiento, creado_en DESC)
    """)

    columnas_gastos_compartidos = get_columns(conn, "gastos_compartidos")
    columnas_extra_gastos_compartidos = {
        "fuente_jugadores": "TEXT DEFAULT 'manual'",
        "actualizado_en": "TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP",
        "cerrado_en": "TIMESTAMPTZ",
        "cerrado_por": "TEXT",
        "cierre_movimiento_id": "INTEGER",
        "cierre_monto": "REAL",
    }
    for columna, definicion in columnas_extra_gastos_compartidos.items():
        if columna not in columnas_gastos_compartidos:
            conn.execute(f"ALTER TABLE gastos_compartidos ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS gasto_compartido_items (
            id SERIAL PRIMARY KEY,
            gasto_id INTEGER NOT NULL,
            jugador_id INTEGER NOT NULL,
            importe REAL NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            fecha_pago TEXT,
            observaciones TEXT,
            comprobante_drive_file_id TEXT,
            comprobante_nombre TEXT,
            comprobante_mime_type TEXT,
            comprobante_tamano INTEGER,
            comprobante_fecha TEXT,
            comprobante_usuario TEXT,
            comprobante_web_url TEXT,
            comprobante_estado TEXT DEFAULT 'sin_comprobante',
            comprobante_revisado_en TEXT,
            comprobante_revisado_por TEXT,
            comprobante_observaciones TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (gasto_id) REFERENCES gastos_compartidos(id) ON DELETE CASCADE,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id),
            UNIQUE(gasto_id, jugador_id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_gasto_compartido_items_gasto
        ON gasto_compartido_items (gasto_id, estado)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_gasto_compartido_items_jugador
        ON gasto_compartido_items (jugador_id, estado, fecha_pago)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS urba_circulares (
            id SERIAL PRIMARY KEY,
            anio INTEGER NOT NULL,
            titulo TEXT NOT NULL,
            url TEXT NOT NULL,
            origen_url TEXT,
            orden_fuente INTEGER,
            nueva INTEGER NOT NULL DEFAULT 1,
            detectada_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizada_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            notificada_en TIMESTAMPTZ
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_urba_circulares_anio_url
        ON urba_circulares (anio, url)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS secretaria_documentos (
            id SERIAL PRIMARY KEY,
            categoria TEXT NOT NULL,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            fecha_documento TEXT,
            drive_file_id TEXT NOT NULL,
            drive_folder_id TEXT,
            archivo_nombre TEXT,
            archivo_mime_type TEXT,
            archivo_tamano INTEGER,
            archivo_web_url TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_secretaria_documentos_categoria
        ON secretaria_documentos (categoria, creado_en DESC)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_secretaria_documentos_fecha
        ON secretaria_documentos (creado_en DESC)
    """)
    columnas_secretaria_documentos = get_columns(conn, "secretaria_documentos")
    if "fecha_vencimiento" not in columnas_secretaria_documentos:
        conn.execute("ALTER TABLE secretaria_documentos ADD COLUMN fecha_vencimiento TEXT")
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_secretaria_documentos_vencimiento
        ON secretaria_documentos (fecha_vencimiento)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_envios (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER,
            telefono TEXT,
            destino_normalizado TEXT,
            tipo TEXT NOT NULL DEFAULT 'general',
            entidad TEXT NOT NULL DEFAULT 'sistema',
            entidad_id TEXT,
            mensaje TEXT,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            meta_message_id TEXT,
            error_codigo TEXT,
            error_mensaje TEXT,
            payload TEXT,
            respuesta TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            enviado_en TEXT,
            creado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_envios_jugador
        ON whatsapp_envios (jugador_id, creado_en DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_envios_estado
        ON whatsapp_envios (estado, creado_en DESC)
    """)

    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_whatsapp_envios_meta_message_id
        ON whatsapp_envios (meta_message_id)
        WHERE meta_message_id IS NOT NULL
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_webhook_eventos (
            id SERIAL PRIMARY KEY,
            event_type TEXT,
            meta_object TEXT,
            payload TEXT,
            procesado INTEGER NOT NULL DEFAULT 0,
            recibido_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_webhook_eventos_recibido
        ON whatsapp_webhook_eventos (recibido_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_mensajes (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER,
            telefono TEXT,
            wa_id TEXT,
            direccion TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'text',
            texto TEXT,
            meta_message_id TEXT,
            estado TEXT,
            leido INTEGER NOT NULL DEFAULT 0,
            payload TEXT,
            respuesta TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_mensajes_telefono
        ON whatsapp_mensajes (COALESCE(wa_id, telefono), creado_en DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_mensajes_leido
        ON whatsapp_mensajes (leido, direccion, creado_en DESC)
    """)

    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_whatsapp_mensajes_meta_message_id
        ON whatsapp_mensajes (meta_message_id)
        WHERE meta_message_id IS NOT NULL
    """)

    columnas_cuotas = get_columns(conn, "cuotas")

    if "fecha_vencimiento" not in columnas_cuotas:
        conn.execute("ALTER TABLE cuotas ADD COLUMN fecha_vencimiento TEXT")

    if "referencia_pago" not in columnas_cuotas:
        conn.execute("ALTER TABLE cuotas ADD COLUMN referencia_pago TEXT")

    if "metodo_pago" not in columnas_cuotas:
        conn.execute("ALTER TABLE cuotas ADD COLUMN metodo_pago TEXT")

    columnas_cuotas = get_columns(conn, "cuotas")
    columnas_beca_cuota = {
        "importe_original": "REAL",
        "descuento_beca": "REAL DEFAULT 0",
        "beca_porcentaje": "REAL DEFAULT 0",
        "beca_motivo": "TEXT",
        "becada": "INTEGER DEFAULT 0",
    }

    for columna, tipo_columna in columnas_beca_cuota.items():
        if columna not in columnas_cuotas:
            conn.execute(f"ALTER TABLE cuotas ADD COLUMN {columna} {tipo_columna}")

    columnas_cuotas = get_columns(conn, "cuotas")
    columnas_plan_cuota = {
        "plan_pago_monto": "REAL DEFAULT 0",
        "plan_pago_detalle": "TEXT",
        "plan_pago_id": "INTEGER",
        "anulada": "INTEGER DEFAULT 0",
        "anulada_en": "TEXT",
        "anulada_por": "TEXT",
        "anulacion_motivo": "TEXT",
        "importe_anulado": "REAL",
    }
    for columna, tipo_columna in columnas_plan_cuota.items():
        if columna not in columnas_cuotas:
            conn.execute(f"ALTER TABLE cuotas ADD COLUMN {columna} {tipo_columna}")

    conn.execute("""
        UPDATE cuotas
        SET importe_original = importe
        WHERE importe_original IS NULL
          AND importe IS NOT NULL
    """)

    columnas_cuotas = get_columns(conn, "cuotas")
    columnas_comprobante_cuota = {
        "comprobante_drive_file_id": "TEXT",
        "comprobante_nombre": "TEXT",
        "comprobante_mime_type": "TEXT",
        "comprobante_tamano": "INTEGER",
        "comprobante_fecha": "TEXT",
        "comprobante_usuario": "TEXT",
        "comprobante_web_url": "TEXT",
        "comprobante_estado": "TEXT DEFAULT 'sin_comprobante'",
        "comprobante_revisado_en": "TEXT",
        "comprobante_revisado_por": "TEXT",
        "comprobante_observaciones": "TEXT",
    }

    for columna, tipo_columna in columnas_comprobante_cuota.items():
        if columna not in columnas_cuotas:
            conn.execute(f"ALTER TABLE cuotas ADD COLUMN {columna} {tipo_columna}")

    conn.execute("""
        UPDATE cuotas
        SET comprobante_estado = CASE
            WHEN comprobante_drive_file_id IS NOT NULL AND pagado = 1 THEN 'aceptado'
            WHEN comprobante_drive_file_id IS NOT NULL THEN 'pendiente'
            ELSE 'sin_comprobante'
        END
        WHERE comprobante_estado IS NULL
           OR comprobante_estado = ''
           OR (
               comprobante_drive_file_id IS NOT NULL
               AND comprobante_estado = 'sin_comprobante'
           )
    """)

    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_cuotas_jugador_periodo
        ON cuotas (jugador_id, periodo)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS becas_historial (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            beca_activa INTEGER DEFAULT 0,
            beca_porcentaje REAL DEFAULT 0,
            beca_desde TEXT,
            beca_hasta TEXT,
            beca_motivo TEXT,
            detalle TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_becas_historial_jugador
        ON becas_historial (jugador_id, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS jugador_bitacora (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'general',
            nota TEXT NOT NULL,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id) ON DELETE CASCADE
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_jugador_bitacora_jugador
        ON jugador_bitacora (jugador_id, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS tareas_sig (
            id SERIAL PRIMARY KEY,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            modulo TEXT NOT NULL DEFAULT 'general',
            prioridad TEXT NOT NULL DEFAULT 'media',
            estado TEXT NOT NULL DEFAULT 'pendiente',
            responsable TEXT,
            fecha_vencimiento TEXT,
            jugador_id INTEGER,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_tareas_sig_estado
        ON tareas_sig (estado, fecha_vencimiento, creado_en DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_tareas_sig_jugador
        ON tareas_sig (jugador_id, estado)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS planes_pago (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL,
            fecha_inicio TEXT,
            monto_total REAL NOT NULL,
            cantidad_cuotas INTEGER DEFAULT 1,
            monto_cuota REAL,
            estado TEXT NOT NULL DEFAULT 'Activo',
            descripcion TEXT,
            observaciones TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            cerrado_en TEXT,
            cerrado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_planes_pago_jugador
        ON planes_pago (jugador_id, estado)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS documentos_jugadores (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            nombre TEXT,
            fecha_presentacion TEXT,
            fecha_vencimiento TEXT,
            url TEXT,
            observaciones TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_documentos_jugadores_vencimiento
        ON documentos_jugadores (fecha_vencimiento)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS fichas_medicas (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL UNIQUE,
            presentada INTEGER DEFAULT 0,
            fecha_vencimiento TEXT,
            apto_fisico INTEGER DEFAULT 0,
            contacto_emergencia TEXT,
            telefono_emergencia TEXT,
            observaciones TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores (id)
        )
    """)

    columnas_fichas_medicas = get_columns(conn, "fichas_medicas")
    columnas_extra_ficha_medica = {
        "documento_drive_file_id": "TEXT",
        "documento_nombre": "TEXT",
        "documento_mime_type": "TEXT",
        "documento_tamano": "INTEGER",
        "documento_fecha": "TEXT",
        "documento_usuario": "TEXT",
        "documento_web_url": "TEXT",
        "ocr_texto": "TEXT",
        "ocr_fecha": "TEXT",
        "ocr_usuario": "TEXT",
    }

    for columna, tipo_columna in columnas_extra_ficha_medica.items():
        if columna not in columnas_fichas_medicas:
            conn.execute(f"ALTER TABLE fichas_medicas ADD COLUMN {columna} {tipo_columna}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS fichas_medicas_batch (
            id SERIAL PRIMARY KEY,
            batch_id TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            archivo_original TEXT,
            drive_file_id TEXT,
            drive_folder_id TEXT,
            documento_nombre TEXT,
            documento_mime_type TEXT,
            documento_tamano INTEGER,
            documento_web_url TEXT,
            extension TEXT,
            ocr_texto TEXT,
            ocr_fecha TEXT,
            ocr_usuario TEXT,
            jugador_sugerido_id INTEGER,
            confianza TEXT,
            motivo TEXT,
            fecha_vencimiento_sugerida TEXT,
            apto_sugerido INTEGER,
            contacto_emergencia_sugerido TEXT,
            telefono_emergencia_sugerido TEXT,
            jugador_id INTEGER,
            procesado_en TEXT,
            procesado_por TEXT,
            error TEXT,
            creado_en TEXT,
            creado_por TEXT,
            FOREIGN KEY (jugador_sugerido_id) REFERENCES jugadores (id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores (id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_fichas_medicas_batch_batch_id
        ON fichas_medicas_batch (batch_id)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_fichas_medicas_batch_estado
        ON fichas_medicas_batch (estado)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS lesiones (
            id SERIAL PRIMARY KEY,
            jugador_id INTEGER NOT NULL,
            fecha_lesion TEXT,
            tipo_lesion TEXT,
            zona_cuerpo TEXT,
            diagnostico TEXT,
            tratamiento TEXT,
            estado TEXT NOT NULL DEFAULT 'Activa',
            fecha_alta TEXT,
            observaciones TEXT,
            FOREIGN KEY (jugador_id) REFERENCES jugadores (id)
        )
    """)

    columnas_lesiones = get_columns(conn, "lesiones")
    columnas_extra_lesiones = {
        "etapa_recuperacion": "TEXT",
        "proximo_control": "TEXT",
        "fecha_retorno_estimada": "TEXT",
        "tratamiento_hasta": "TEXT",
    }
    for columna, definicion in columnas_extra_lesiones.items():
        if columna not in columnas_lesiones:
            conn.execute(f"ALTER TABLE lesiones ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS lesiones_documentos (
            id SERIAL PRIMARY KEY,
            lesion_id INTEGER NOT NULL,
            jugador_id INTEGER NOT NULL,
            nombre TEXT,
            mime_type TEXT,
            tamano INTEGER,
            drive_file_id TEXT NOT NULL,
            drive_folder_id TEXT,
            web_url TEXT,
            descripcion TEXT,
            creado_en TEXT,
            creado_por TEXT,
            FOREIGN KEY (lesion_id) REFERENCES lesiones (id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores (id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_lesiones_documentos_lesion
        ON lesiones_documentos (lesion_id, id DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS roles (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL,
            descripcion TEXT,
            permisos TEXT NOT NULL DEFAULT '[]',
            sistema INTEGER DEFAULT 0
        )
    """)

    columnas_roles = get_columns(conn, "roles")
    if "descripcion" not in columnas_roles:
        conn.execute("ALTER TABLE roles ADD COLUMN descripcion TEXT")
    if "permisos" not in columnas_roles:
        conn.execute("ALTER TABLE roles ADD COLUMN permisos TEXT NOT NULL DEFAULT '[]'")
    if "sistema" not in columnas_roles:
        conn.execute("ALTER TABLE roles ADD COLUMN sistema INTEGER DEFAULT 0")

    roles_base = {
        "admin": "Acceso completo al sistema.",
        "tesorero": "Gesti?n financiera, cuotas, caja y reportes.",
        "medico": "Gesti?n de fichas m?dicas y lesiones.",
        "entrenador": "Gesti?n deportiva, jugadores, calendario y asistencia.",
    }

    for nombre_rol, descripcion in roles_base.items():
        conn.execute("""
            INSERT INTO roles (nombre, descripcion, permisos, sistema)
            VALUES (%s, %s, %s, 1)
            ON CONFLICT(nombre) DO UPDATE SET
                descripcion = EXCLUDED.descripcion,
                permisos = CASE
                    WHEN roles.permisos IS NULL OR roles.permisos = '' OR roles.permisos = '[]' THEN EXCLUDED.permisos
                    ELSE roles.permisos
                END,
                sistema = 1
        """, (
            nombre_rol,
            descripcion,
            serializar_permisos(permisos_default_rol(nombre_rol)),
        ))

    columnas_usuarios = get_columns(conn, "usuarios")
    if "rol" not in columnas_usuarios:
        conn.execute("ALTER TABLE usuarios ADD COLUMN rol TEXT DEFAULT 'admin'")
    columnas_usuarios = get_columns(conn, "usuarios")
    onboarding_columna_creada = "onboarding_visto" not in columnas_usuarios
    columnas_extra_usuarios = {
        "email": "TEXT",
        "debe_cambiar_password": "INTEGER DEFAULT 0",
        "onboarding_visto": "INTEGER DEFAULT 0",
        "creado_en": "TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP",
        "ultimo_login": "TIMESTAMPTZ",
    }
    for columna, tipo_columna in columnas_extra_usuarios.items():
        if columna not in columnas_usuarios:
            conn.execute(f"ALTER TABLE usuarios ADD COLUMN {columna} {tipo_columna}")

    if onboarding_columna_creada:
        conn.execute("""
            UPDATE usuarios
            SET onboarding_visto = 1
            WHERE onboarding_visto IS NULL
               OR onboarding_visto = 0
        """)

    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_usuarios_email
        ON usuarios (email)
        WHERE email IS NOT NULL AND email <> ''
    """)

    usuario_admin = conn.execute("""
        SELECT * FROM usuarios WHERE username = 'admin'
    """).fetchone()

    if not usuario_admin:
        admin_password = ADMIN_PASSWORD or "admin123"
        conn.execute("""
    INSERT INTO usuarios (username, password, rol, debe_cambiar_password, onboarding_visto)
    VALUES (%s, %s, %s, %s, %s)
""", ("admin", generate_password_hash(admin_password), "admin", 1, 0))
    elif ADMIN_PASSWORD and (
        FORCE_ADMIN_PASSWORD_UPDATE
        or check_password_hash(usuario_admin["password"], "admin123")
    ):
        conn.execute("""
            UPDATE usuarios
            SET password = %s, rol = 'admin'
            WHERE username = 'admin'
        """, (generate_password_hash(ADMIN_PASSWORD),))

    conn.execute("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id SERIAL PRIMARY KEY,
            username TEXT,
            ip TEXT,
            success INTEGER DEFAULT 0,
            fecha TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_login_attempts_user_ip_fecha
        ON login_attempts (username, ip, fecha DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS sugerencias_denuncias (
            id SERIAL PRIMARY KEY,
            tipo TEXT NOT NULL,
            categoria TEXT,
            anonima INTEGER DEFAULT 1,
            nombre TEXT,
            contacto TEXT,
            mensaje TEXT NOT NULL,
            destinatarios TEXT,
            email_estado TEXT,
            seguimiento_estado TEXT NOT NULL DEFAULT 'nuevo',
            notas_internas TEXT,
            actualizado_en TIMESTAMPTZ,
            actualizado_por TEXT,
            notificacion_reintentos INTEGER DEFAULT 0,
            notificado_en TIMESTAMPTZ,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    columnas_sugerencias = get_columns(conn, "sugerencias_denuncias")
    columnas_extra_sugerencias = {
        "seguimiento_estado": "TEXT NOT NULL DEFAULT 'nuevo'",
        "notas_internas": "TEXT",
        "actualizado_en": "TIMESTAMPTZ",
        "actualizado_por": "TEXT",
        "notificacion_reintentos": "INTEGER DEFAULT 0",
        "notificado_en": "TIMESTAMPTZ",
    }
    for columna, definicion in columnas_extra_sugerencias.items():
        if columna not in columnas_sugerencias:
            conn.execute(f"ALTER TABLE sugerencias_denuncias ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_sugerencias_denuncias_tipo_fecha
        ON sugerencias_denuncias (tipo, creado_en DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_sugerencias_denuncias_email_estado
        ON sugerencias_denuncias (email_estado, creado_en DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_sugerencias_denuncias_seguimiento
        ON sugerencias_denuncias (seguimiento_estado, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id SERIAL PRIMARY KEY,
            usuario_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            expira_en TIMESTAMPTZ NOT NULL,
            usado INTEGER DEFAULT 0,
            usado_en TIMESTAMPTZ,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_password_reset_tokens_usuario
        ON password_reset_tokens (usuario_id, usado, expira_en DESC)
    """)

    conn.execute("""
        DO $$
        BEGIN
            IF EXISTS (
                SELECT 1
                FROM information_schema.table_constraints
                WHERE constraint_schema = 'public'
                  AND table_name = 'password_reset_tokens'
                  AND constraint_name = 'password_reset_tokens_usuario_id_fkey'
                  AND constraint_type = 'FOREIGN KEY'
            ) THEN
                ALTER TABLE password_reset_tokens
                DROP CONSTRAINT password_reset_tokens_usuario_id_fkey;
            END IF;

            ALTER TABLE password_reset_tokens
            ADD CONSTRAINT password_reset_tokens_usuario_id_fkey
            FOREIGN KEY (usuario_id)
            REFERENCES usuarios(id)
            ON DELETE CASCADE;
        END $$;
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS movimientos (
        id SERIAL PRIMARY KEY,
        tipo TEXT NOT NULL,                -- ingreso / egreso
        concepto TEXT,
        monto REAL NOT NULL,
        fecha TEXT NOT NULL,
        referencia TEXT,
        anulado INTEGER DEFAULT 0,
        fecha_anulacion TEXT,
        usuario_anulacion TEXT,
        motivo_anulacion TEXT
    )
""")

    columnas_movimientos = get_columns(conn, "movimientos")
    if "anulado" not in columnas_movimientos:
        conn.execute("ALTER TABLE movimientos ADD COLUMN anulado INTEGER DEFAULT 0")
    if "fecha_anulacion" not in columnas_movimientos:
        conn.execute("ALTER TABLE movimientos ADD COLUMN fecha_anulacion TEXT")
    if "usuario_anulacion" not in columnas_movimientos:
        conn.execute("ALTER TABLE movimientos ADD COLUMN usuario_anulacion TEXT")
    if "motivo_anulacion" not in columnas_movimientos:
        conn.execute("ALTER TABLE movimientos ADD COLUMN motivo_anulacion TEXT")
    columnas_comprobante_movimiento = {
        "comprobante_drive_file_id": "TEXT",
        "comprobante_nombre": "TEXT",
        "comprobante_mime_type": "TEXT",
        "comprobante_tamano": "INTEGER",
        "comprobante_fecha": "TEXT",
        "comprobante_usuario": "TEXT",
        "comprobante_web_url": "TEXT",
        "comprobante_operacion": "TEXT",
        "comprobante_ocr_texto": "TEXT",
    }
    for columna, definicion in columnas_comprobante_movimiento.items():
        if columna not in columnas_movimientos:
            conn.execute(f"ALTER TABLE movimientos ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS facturas_email_filtros (
            id SERIAL PRIMARY KEY,
            proveedor TEXT NOT NULL,
            remitente_patron TEXT,
            asunto_patron TEXT,
            activo INTEGER NOT NULL DEFAULT 1,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_por TEXT
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_facturas_email_filtros_unico
        ON facturas_email_filtros (
            lower(proveedor),
            lower(COALESCE(remitente_patron, '')),
            lower(COALESCE(asunto_patron, ''))
        )
    """)
    for filtro in FACTURA_EMAIL_DEFAULT_FILTERS:
        conn.execute("""
            INSERT INTO facturas_email_filtros (proveedor, remitente_patron, asunto_patron, creado_por)
            VALUES (%s, %s, %s, 'sistema')
            ON CONFLICT DO NOTHING
        """, (
            filtro["proveedor"],
            filtro["remitente_patron"],
            filtro["asunto_patron"],
        ))

    conn.execute("""
        CREATE TABLE IF NOT EXISTS facturas_recibidas (
            id SERIAL PRIMARY KEY,
            source_key TEXT UNIQUE NOT NULL,
            cuenta_email TEXT,
            message_id TEXT,
            filtro_id INTEGER,
            proveedor TEXT NOT NULL,
            remitente TEXT,
            asunto TEXT,
            fecha_email TEXT,
            archivo_nombre TEXT,
            archivo_mime_type TEXT,
            archivo_tamano INTEGER,
            drive_file_id TEXT NOT NULL,
            drive_folder_id TEXT,
            archivo_web_url TEXT,
            monto_detectado REAL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            movimiento_id INTEGER,
            notas TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_por TEXT,
            FOREIGN KEY (filtro_id) REFERENCES facturas_email_filtros(id),
            FOREIGN KEY (movimiento_id) REFERENCES movimientos(id)
        )
    """)
    columnas_facturas_recibidas = get_columns(conn, "facturas_recibidas")
    if "cuenta_email" not in columnas_facturas_recibidas:
        conn.execute("ALTER TABLE facturas_recibidas ADD COLUMN cuenta_email TEXT")
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_facturas_recibidas_estado
        ON facturas_recibidas (estado, creado_en DESC)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_facturas_recibidas_proveedor
        ON facturas_recibidas (proveedor, creado_en DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS presupuesto_items (
            id SERIAL PRIMARY KEY,
            tipo TEXT NOT NULL,
            categoria TEXT NOT NULL DEFAULT 'fijo',
            concepto TEXT NOT NULL,
            monto REAL NOT NULL,
            mes_inicio TEXT,
            mes_fin TEXT,
            activo INTEGER DEFAULT 1,
            notas TEXT,
            creado_en TEXT,
            actualizado_en TEXT
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_presupuesto_items_activo
        ON presupuesto_items (activo)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS cierres_mensuales (
            id SERIAL PRIMARY KEY,
            mes TEXT UNIQUE NOT NULL,
            ingresos REAL NOT NULL,
            egresos REAL NOT NULL,
            resultado REAL NOT NULL,
            fecha_cierre TEXT NOT NULL,
            usuario TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS eventos_asistencia (
            id SERIAL PRIMARY KEY,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL,
            descripcion TEXT
        )
    """)

    columnas_eventos_asistencia = get_columns(conn, "eventos_asistencia")
    columnas_extra_eventos_asistencia = {
        "cerrado": "INTEGER DEFAULT 0",
        "cerrado_en": "TIMESTAMPTZ",
        "cerrado_por": "TEXT",
    }
    for columna, definicion in columnas_extra_eventos_asistencia.items():
        if columna not in columnas_eventos_asistencia:
            conn.execute(f"ALTER TABLE eventos_asistencia ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS calendario_eventos (
            id SERIAL PRIMARY KEY,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            ubicacion TEXT,
            categoria TEXT
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_calendario_eventos_fecha
        ON calendario_eventos (fecha)
    """)

    columnas_calendario_eventos = get_columns(conn, "calendario_eventos")
    columnas_extra_calendario = {
        "hora_inicio": "TEXT",
        "duracion_minutos": "INTEGER DEFAULT 90",
        "publicar_portal": "INTEGER DEFAULT 1",
        "asistencia_evento_id": "INTEGER",
        "convocatoria_texto": "TEXT",
        "convocatoria_cierre": "TEXT",
        "minuta_post_evento": "TEXT",
    }
    for columna, definicion in columnas_extra_calendario.items():
        if columna not in columnas_calendario_eventos:
            conn.execute(f"ALTER TABLE calendario_eventos ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS asistencias (
            id SERIAL PRIMARY KEY,
            evento_id INTEGER NOT NULL,
            jugador_id INTEGER NOT NULL,
            presente INTEGER DEFAULT 0,
            observaciones TEXT,
            FOREIGN KEY (evento_id) REFERENCES eventos_asistencia(id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id),
            UNIQUE(evento_id, jugador_id)
        )
    """)

    columnas_asistencias = get_columns(conn, "asistencias")
    if "estado_asistencia" not in columnas_asistencias:
        conn.execute("ALTER TABLE asistencias ADD COLUMN estado_asistencia TEXT")

    conn.execute("""
        UPDATE asistencias
        SET estado_asistencia = CASE
            WHEN presente = 1 THEN 'a_tiempo'
            ELSE 'ausente'
        END
        WHERE estado_asistencia IS NULL
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS aspirante_asistencias (
            id SERIAL PRIMARY KEY,
            evento_id INTEGER NOT NULL,
            aspirante_id INTEGER NOT NULL,
            presente INTEGER DEFAULT 0,
            estado_asistencia TEXT,
            observaciones TEXT,
            FOREIGN KEY (evento_id) REFERENCES eventos_asistencia(id),
            FOREIGN KEY (aspirante_id) REFERENCES aspirantes(id),
            UNIQUE(evento_id, aspirante_id)
        )
    """)

    columnas_aspirante_asistencias = get_columns(conn, "aspirante_asistencias")
    if "estado_asistencia" not in columnas_aspirante_asistencias:
        conn.execute("ALTER TABLE aspirante_asistencias ADD COLUMN estado_asistencia TEXT")

    conn.execute("""
        UPDATE aspirante_asistencias
        SET estado_asistencia = CASE
            WHEN presente = 1 THEN 'a_tiempo'
            ELSE 'ausente'
        END
        WHERE estado_asistencia IS NULL
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS portal_asistencia_confirmaciones (
            id SERIAL PRIMARY KEY,
            evento_id INTEGER NOT NULL,
            jugador_id INTEGER NOT NULL,
            estado TEXT NOT NULL,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            actualizado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (evento_id) REFERENCES eventos_asistencia(id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id),
            UNIQUE(evento_id, jugador_id)
        )
    """)

    columnas_portal_confirmaciones = get_columns(conn, "portal_asistencia_confirmaciones")
    columnas_extra_portal_confirmaciones = {
        "sueno_calidad": "INTEGER",
        "horas_sueno": "TEXT",
        "doms": "INTEGER",
        "fatiga": "INTEGER",
        "estres": "INTEGER",
        "animo": "INTEGER",
        "motivacion": "INTEGER",
        "recuperacion": "INTEGER",
        "dolor_zonas": "TEXT",
        "dolor_otro": "TEXT",
        "comentarios": "TEXT",
    }
    for columna, definicion in columnas_extra_portal_confirmaciones.items():
        if columna not in columnas_portal_confirmaciones:
            conn.execute(f"ALTER TABLE portal_asistencia_confirmaciones ADD COLUMN {columna} {definicion}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS test_tipos (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL,
            descripcion TEXT,
            unidad TEXT,
            puntaje_min REAL,
            puntaje_max REAL,
            mayor_es_mejor INTEGER DEFAULT 1,
            activo INTEGER DEFAULT 1,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS test_resultados (
            id SERIAL PRIMARY KEY,
            test_id INTEGER NOT NULL,
            jugador_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            puntaje REAL NOT NULL,
            observaciones TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            FOREIGN KEY (test_id) REFERENCES test_tipos(id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_test_resultados_test_fecha
        ON test_resultados (test_id, fecha DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_test_resultados_jugador
        ON test_resultados (jugador_id, fecha DESC)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS test_importaciones_batch (
            id SERIAL PRIMARY KEY,
            batch_id TEXT NOT NULL,
            estado TEXT DEFAULT 'pendiente',
            fila INTEGER,
            test_id INTEGER,
            test_nombre TEXT,
            jugador_sugerido_id INTEGER,
            jugador_id INTEGER,
            confianza TEXT,
            motivo TEXT,
            nombre_excel TEXT,
            apellido_excel TEXT,
            nombre_completo_excel TEXT,
            fecha TEXT,
            puntaje REAL,
            observaciones TEXT,
            error TEXT,
            creado_en TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            creado_por TEXT,
            procesado_en TIMESTAMPTZ,
            procesado_por TEXT,
            FOREIGN KEY (test_id) REFERENCES test_tipos(id),
            FOREIGN KEY (jugador_sugerido_id) REFERENCES jugadores(id),
            FOREIGN KEY (jugador_id) REFERENCES jugadores(id)
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_test_importaciones_batch_id
        ON test_importaciones_batch (batch_id)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_test_importaciones_estado
        ON test_importaciones_batch (estado)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS auditoria (
            id SERIAL PRIMARY KEY,
            fecha TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            usuario_id INTEGER,
            username TEXT,
            rol TEXT,
            accion TEXT NOT NULL,
            entidad TEXT,
            entidad_id TEXT,
            detalle TEXT,
            ip TEXT,
            user_agent TEXT
        )
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_auditoria_fecha
        ON auditoria (fecha DESC)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_auditoria_usuario
        ON auditoria (username)
    """)

    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_auditoria_entidad
        ON auditoria (entidad)
    """)

    conn.commit()
    conn.close()

@app.route("/presupuesto")
def ver_presupuesto():
    check = permiso_requerido("presupuesto_ver", "caja_ver")
    if check:
        return check

    mes = normalizar_mes(request.args.get("mes"), ahora_sig().strftime("%Y-%m"))
    try:
        meses_proyeccion = int(request.args.get("meses", "6") or 6)
    except ValueError:
        meses_proyeccion = 6
    presupuesto = obtener_presupuesto_mensual(mes, meses_proyeccion)
    return render_template("presupuesto.html", presupuesto=presupuesto, mes=mes, meses_proyeccion=meses_proyeccion)


@app.route("/presupuesto/items", methods=["POST"])
def crear_presupuesto_item():
    check = permiso_requerido("presupuesto_gestionar", "caja_gestionar")
    if check:
        return check

    tipo = request.form.get("tipo", "").strip()
    categoria = request.form.get("categoria", "fijo").strip()
    concepto = request.form.get("concepto", "").strip()
    monto = parsear_importe(request.form.get("monto"))
    mes_inicio = normalizar_mes(request.form.get("mes_inicio"), "")
    mes_fin = normalizar_mes(request.form.get("mes_fin"), "")
    notas = request.form.get("notas", "").strip()
    mes_retorno = normalizar_mes(request.form.get("mes_retorno"), ahora_sig().strftime("%Y-%m"))

    if tipo not in {"ingreso", "egreso"}:
        flash("Selecciona si el concepto es ingreso o egreso.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))
    if categoria not in {"fijo", "variable"}:
        flash("La categoria del presupuesto no es valida.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))
    if not concepto:
        flash("El concepto es obligatorio.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))
    if monto is None or monto <= 0:
        flash("El monto debe ser mayor a cero.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))
    if mes_inicio and mes_fin and mes_fin < mes_inicio:
        flash("El mes de fin no puede ser anterior al inicio.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))

    ahora = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_connection()
    conn.execute("""
        INSERT INTO presupuesto_items (
            tipo, categoria, concepto, monto, mes_inicio, mes_fin, activo, notas, creado_en, actualizado_en
        )
        VALUES (%s, %s, %s, %s, %s, %s, 1, %s, %s, %s)
    """, (tipo, categoria, concepto, monto, mes_inicio or None, mes_fin or None, notas, ahora, ahora))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "crear",
        "presupuesto_item",
        None,
        {"tipo": tipo, "categoria": categoria, "concepto": concepto, "monto": monto},
    )
    flash("Concepto agregado al presupuesto.", "ok")
    return redirect(url_for("ver_presupuesto", mes=mes_retorno))


@app.route("/presupuesto/items/<int:item_id>/toggle", methods=["POST"])
def alternar_presupuesto_item(item_id):
    check = permiso_requerido("presupuesto_gestionar", "caja_gestionar")
    if check:
        return check

    mes_retorno = normalizar_mes(request.form.get("mes_retorno"), ahora_sig().strftime("%Y-%m"))
    conn = get_connection()
    item = conn.execute("SELECT * FROM presupuesto_items WHERE id = %s", (item_id,)).fetchone()
    if item is None:
        conn.close()
        flash("Concepto no encontrado.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))

    nuevo_estado = 0 if item.get("activo") else 1
    conn.execute("""
        UPDATE presupuesto_items
        SET activo = %s,
            actualizado_en = %s
        WHERE id = %s
    """, (nuevo_estado, ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), item_id))
    conn.commit()
    conn.close()

    registrar_auditoria("alternar", "presupuesto_item", str(item_id), {"activo": nuevo_estado})
    flash("Concepto de presupuesto actualizado.", "ok")
    return redirect(url_for("ver_presupuesto", mes=mes_retorno))


@app.route("/presupuesto/items/<int:item_id>/eliminar", methods=["POST"])
def eliminar_presupuesto_item(item_id):
    check = permiso_requerido("presupuesto_gestionar", "caja_gestionar")
    if check:
        return check

    mes_retorno = normalizar_mes(request.form.get("mes_retorno"), ahora_sig().strftime("%Y-%m"))
    conn = get_connection()
    item = conn.execute("SELECT * FROM presupuesto_items WHERE id = %s", (item_id,)).fetchone()
    if item is None:
        conn.close()
        flash("Concepto no encontrado.", "error")
        return redirect(url_for("ver_presupuesto", mes=mes_retorno))
    conn.execute("DELETE FROM presupuesto_items WHERE id = %s", (item_id,))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "eliminar",
        "presupuesto_item",
        str(item_id),
        {"concepto": item["concepto"], "monto": item["monto"]},
    )
    flash("Concepto eliminado del presupuesto.", "ok")
    return redirect(url_for("ver_presupuesto", mes=mes_retorno))


@app.route("/caja")
def ver_caja():
    check = permiso_requerido("caja_ver")
    if check:
        return check

    mes = request.args.get("mes")
    mes_actual = mes or ahora_sig().strftime("%Y-%m")

    conn = get_connection()

    movimientos = conn.execute("""
        SELECT *
        FROM movimientos
        ORDER BY fecha DESC, id DESC
        LIMIT 50
    """).fetchall()

    total_ingresos = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'ingreso'
          AND COALESCE(anulado, 0) = 0
    """).fetchone()["total"]

    total_egresos = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'egreso'
          AND COALESCE(anulado, 0) = 0
    """).fetchone()["total"]

    ingresos_mes = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'ingreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    egresos_mes = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'egreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    movimientos_mes = conn.execute("""
        SELECT *
        FROM movimientos
        WHERE substring(fecha from 1 for 7) = %s
        ORDER BY fecha DESC, id DESC
    """, (mes_actual,)).fetchall()

    cierre_mes = conn.execute("""
        SELECT *
        FROM cierres_mensuales
        WHERE mes = %s
    """, (mes_actual,)).fetchone()

    conn.close()

    saldo = total_ingresos - total_egresos
    resultado_mes = ingresos_mes - egresos_mes

    return render_template(
        "caja.html",
        movimientos=movimientos,
        total_ingresos=total_ingresos,
        total_egresos=total_egresos,
        saldo=saldo,
        mes_actual=mes_actual,
        ingresos_mes=ingresos_mes,
        egresos_mes=egresos_mes,
        resultado_mes=resultado_mes,
        movimientos_mes=movimientos_mes,
        cierre_mes=cierre_mes
    )

@app.route("/movimientos/nuevo", methods=["GET", "POST"])
def nuevo_movimiento():
    check = permiso_requerido("caja_gestionar")
    if check:
        return check

    if request.method == "POST":
        tipo = request.form.get("tipo")
        concepto = request.form.get("concepto")
        monto = request.form.get("monto", "").strip()
        fecha = validar_fecha_movimiento(request.form.get("fecha", "").strip())
        referencia = request.form.get("referencia", "").strip()
        comprobante_pago = request.files.get("comprobante_pago")

        movimiento_form = {
            "tipo": tipo,
            "concepto": concepto,
            "monto": monto,
            "fecha": request.form.get("fecha", "").strip(),
            "referencia": referencia,
        }

        if not fecha:
            flash("La fecha del movimiento no es valida.", "error")
            return render_template("movimiento_form.html", movimiento=movimiento_form)

        mes_movimiento = fecha[:7]
        movimiento_form["fecha"] = fecha
        if mes_esta_cerrado(mes_movimiento):
            flash("No se puede agregar un movimiento en un mes cerrado.", "error")
            return render_template("movimiento_form.html", movimiento=movimiento_form)

        comprobante_info = None
        numero_operacion = ""
        monto_ocr = ""
        ocr_texto = ""
        if comprobante_pago and comprobante_pago.filename:
            try:
                comprobante_info, numero_operacion, monto_ocr, ocr_texto = procesar_comprobante_movimiento(
                    comprobante_pago,
                    {**movimiento_form, "fecha": fecha},
                )
            except (RuntimeError, ValueError) as error:
                flash(str(error), "error")
                return render_template("movimiento_form.html", movimiento=movimiento_form)
            except Exception as error:
                app.logger.exception("No se pudo procesar comprobante de movimiento de caja.")
                flash(mensaje_error_drive(error, carpeta="Caja", accion="subir o leer el comprobante"), "error")
                return render_template("movimiento_form.html", movimiento=movimiento_form)

        monto_manual = monto
        if not referencia and numero_operacion:
            referencia = numero_operacion
        if not monto and monto_ocr:
            monto = monto_ocr
        if not monto:
            movimiento_form.update({"referencia": referencia, "monto": monto})
            flash("Debe indicar un monto o adjuntar un comprobante donde pueda leerse el monto.", "error")
            return render_template("movimiento_form.html", movimiento=movimiento_form)

        conn = get_connection()
        conn.execute("""
            INSERT INTO movimientos (
                tipo, concepto, monto, fecha, referencia,
                comprobante_drive_file_id, comprobante_nombre, comprobante_mime_type,
                comprobante_tamano, comprobante_fecha, comprobante_usuario,
                comprobante_web_url, comprobante_operacion, comprobante_ocr_texto
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            tipo, concepto, monto, fecha, referencia,
            comprobante_info["file_id"] if comprobante_info else None,
            comprobante_info["nombre"] if comprobante_info else None,
            comprobante_info["mime_type"] if comprobante_info else None,
            comprobante_info["tamano"] if comprobante_info else None,
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if comprobante_info else None,
            session.get("username") if comprobante_info else None,
            comprobante_info["web_url"] if comprobante_info else None,
            numero_operacion or None,
            ocr_texto or None,
        ))
        conn.commit()
        conn.close()

        if comprobante_info:
            detalle = []
            if numero_operacion:
                detalle.append(f"operacion {numero_operacion}")
            if monto_ocr and not monto_manual:
                detalle.append(f"monto {monto}")
            extra = f" ({', '.join(detalle)})" if detalle else ""
            flash(f"Movimiento registrado y comprobante guardado{extra}.", "ok")
        else:
            flash("Movimiento registrado.", "ok")
        return redirect(url_for("ver_caja", mes=mes_movimiento))

    fecha_default = fecha_movimiento_default(request.args.get("mes"))
    return render_template(
        "movimiento_form.html",
        movimiento={"tipo": "egreso", "fecha": fecha_default},
    )


@app.route("/finanzas/facturas-recibidas")
def listar_facturas_recibidas():
    check = permiso_requerido("facturas_recibidas_ver", "caja_ver")
    if check:
        return check

    estado = request.args.get("estado", "pendiente").strip()
    if estado not in {"pendiente", "registrada", "descartada", "todas"}:
        estado = "pendiente"

    condiciones = []
    params = []
    if estado != "todas":
        condiciones.append("fr.estado = %s")
        params.append(estado)
    where_sql = f"WHERE {' AND '.join(condiciones)}" if condiciones else ""

    conn = get_connection()
    facturas = conn.execute(f"""
        SELECT fr.*, f.remitente_patron, f.asunto_patron
        FROM facturas_recibidas fr
        LEFT JOIN facturas_email_filtros f ON f.id = fr.filtro_id
        {where_sql}
        ORDER BY fr.creado_en DESC, fr.id DESC
        LIMIT 120
    """, params).fetchall()
    filtros = conn.execute("""
        SELECT *
        FROM facturas_email_filtros
        ORDER BY activo DESC, proveedor, id
    """).fetchall()
    sync_config = conn.execute("""
        SELECT clave, valor
        FROM app_settings
        WHERE clave IN ('facturas_email_sync_en', 'facturas_email_sync_por')
    """).fetchall()
    conn.close()
    sync_info = {fila["clave"]: fila["valor"] for fila in sync_config}

    return render_template(
        "facturas_recibidas.html",
        facturas=facturas,
        filtros=filtros,
        estado=estado,
        imap_configurado=factura_email_configurado(),
        sync_info=sync_info,
    )


@app.route("/finanzas/facturas-recibidas/filtros", methods=["POST"])
def crear_filtro_factura_email():
    check = permiso_requerido("facturas_recibidas_gestionar", "caja_gestionar")
    if check:
        return check

    proveedor = request.form.get("proveedor", "").strip()
    remitente_patron = request.form.get("remitente_patron", "").strip()
    asunto_patron = request.form.get("asunto_patron", "").strip()
    if not proveedor:
        flash("Indicá el proveedor del filtro.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    if not remitente_patron and not asunto_patron:
        flash("El filtro necesita remitente o asunto.", "error")
        return redirect(url_for("listar_facturas_recibidas"))

    conn = get_connection()
    conn.execute("""
        INSERT INTO facturas_email_filtros (
            proveedor, remitente_patron, asunto_patron, creado_por, actualizado_por
        )
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (
        proveedor,
        remitente_patron or None,
        asunto_patron or None,
        session.get("username"),
        session.get("username"),
    ))
    conn.commit()
    conn.close()
    registrar_auditoria("crear", "factura_email_filtro", None, {
        "proveedor": proveedor,
        "remitente_patron": remitente_patron,
        "asunto_patron": asunto_patron,
    })
    flash("Filtro de facturas agregado.", "ok")
    return redirect(url_for("listar_facturas_recibidas"))


@app.route("/finanzas/facturas-recibidas/filtros/<int:filtro_id>/toggle", methods=["POST"])
def alternar_filtro_factura_email(filtro_id):
    check = permiso_requerido("facturas_recibidas_gestionar", "caja_gestionar")
    if check:
        return check

    conn = get_connection()
    filtro = conn.execute("SELECT * FROM facturas_email_filtros WHERE id = %s", (filtro_id,)).fetchone()
    if not filtro:
        conn.close()
        flash("Filtro no encontrado.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    activo = 0 if filtro["activo"] else 1
    conn.execute("""
        UPDATE facturas_email_filtros
        SET activo = %s,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (activo, session.get("username"), filtro_id))
    conn.commit()
    conn.close()
    registrar_auditoria("alternar", "factura_email_filtro", str(filtro_id), {"activo": activo})
    flash("Filtro actualizado.", "ok")
    return redirect(url_for("listar_facturas_recibidas"))


@app.route("/finanzas/facturas-recibidas/sync", methods=["POST"])
def sincronizar_facturas_recibidas_view():
    check = permiso_requerido("facturas_recibidas_gestionar", "caja_gestionar")
    if check:
        return check

    conn = get_connection()
    try:
        resultado = sincronizar_facturas_email(conn, session.get("username"))
        conn.commit()
    except RuntimeError as error:
        conn.rollback()
        conn.close()
        flash(str(error), "error")
        return redirect(url_for("listar_facturas_recibidas"))
    except Exception as error:
        conn.rollback()
        conn.close()
        app.logger.exception("No se pudieron sincronizar facturas por email.")
        flash(f"No se pudieron sincronizar facturas: {error}", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    conn.close()

    registrar_auditoria("sincronizar", "facturas_recibidas", None, resultado)
    cuentas = len(resultado.get("cuentas") or [])
    resumen = (
        f"Sincronizacion completada en {cuentas} cuenta(s): "
        f"{resultado['nuevas']} nueva(s), {resultado['omitidas']} omitida(s)."
    )
    if resultado.get("errores"):
        flash(f"{resumen} Errores: {'; '.join(resultado['errores'])}", "warning")
    else:
        flash(resumen, "ok")
    return redirect(url_for("listar_facturas_recibidas"))


@app.route("/finanzas/facturas-recibidas/<int:factura_id>/archivo")
def ver_factura_recibida_archivo(factura_id):
    check = permiso_requerido("facturas_recibidas_ver", "caja_ver")
    if check:
        return check

    conn = get_connection()
    factura = conn.execute("SELECT * FROM facturas_recibidas WHERE id = %s", (factura_id,)).fetchone()
    conn.close()
    if not factura:
        flash("Factura no encontrada.", "error")
        return redirect(url_for("listar_facturas_recibidas"))

    try:
        archivo = descargar_drive_file(factura["drive_file_id"])
    except Exception as error:
        flash(mensaje_error_drive(error, carpeta="Facturas recibidas", accion="descargar la factura"), "error")
        return redirect(url_for("listar_facturas_recibidas"))

    registrar_auditoria("ver", "factura_recibida", str(factura_id), {
        "archivo": factura.get("archivo_nombre"),
        "drive_file_id": factura.get("drive_file_id"),
    })
    return send_file(
        archivo,
        mimetype=factura.get("archivo_mime_type") or "application/octet-stream",
        as_attachment=False,
        download_name=factura.get("archivo_nombre") or f"factura_{factura_id}",
    )


@app.route("/finanzas/facturas-recibidas/<int:factura_id>/registrar", methods=["POST"])
def registrar_factura_recibida_en_caja(factura_id):
    check = permiso_requerido("facturas_recibidas_gestionar", "caja_gestionar")
    if check:
        return check

    fecha = validar_fecha_movimiento(request.form.get("fecha", "").strip())
    concepto = request.form.get("concepto", "").strip()
    monto = parsear_importe(request.form.get("monto", "").strip())
    referencia = request.form.get("referencia", "").strip()

    if not fecha:
        flash("La fecha del egreso no es valida.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    if mes_esta_cerrado(fecha[:7]):
        flash("No se puede registrar un egreso en un mes cerrado.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    if not concepto:
        flash("Indicá el concepto del egreso.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    if monto is None or monto <= 0:
        flash("Indicá un monto valido.", "error")
        return redirect(url_for("listar_facturas_recibidas"))

    conn = get_connection()
    factura = conn.execute("SELECT * FROM facturas_recibidas WHERE id = %s FOR UPDATE", (factura_id,)).fetchone()
    if not factura:
        conn.close()
        flash("Factura no encontrada.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    if factura["estado"] != "pendiente":
        conn.close()
        flash("Solo se pueden registrar facturas pendientes.", "error")
        return redirect(url_for("listar_facturas_recibidas"))

    movimiento = conn.execute("""
        INSERT INTO movimientos (
            tipo, concepto, monto, fecha, referencia,
            comprobante_drive_file_id, comprobante_nombre, comprobante_mime_type,
            comprobante_tamano, comprobante_fecha, comprobante_usuario,
            comprobante_web_url
        )
        VALUES ('egreso', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        concepto,
        monto,
        fecha,
        referencia or f"Factura {factura['proveedor']}",
        factura["drive_file_id"],
        factura["archivo_nombre"],
        factura["archivo_mime_type"],
        factura["archivo_tamano"],
        ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
        session.get("username"),
        factura["archivo_web_url"],
    )).fetchone()
    conn.execute("""
        UPDATE facturas_recibidas
        SET estado = 'registrada',
            movimiento_id = %s,
            monto_detectado = %s,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (movimiento["id"], monto, session.get("username"), factura_id))
    conn.commit()
    conn.close()

    registrar_auditoria("registrar_en_caja", "factura_recibida", str(factura_id), {
        "movimiento_id": movimiento["id"],
        "monto": monto,
        "fecha": fecha,
    })
    flash("Factura registrada como egreso de caja.", "ok")
    return redirect(url_for("ver_caja", mes=fecha[:7]))


@app.route("/finanzas/facturas-recibidas/<int:factura_id>/descartar", methods=["POST"])
def descartar_factura_recibida(factura_id):
    check = permiso_requerido("facturas_recibidas_gestionar", "caja_gestionar")
    if check:
        return check

    notas = request.form.get("notas", "").strip()
    conn = get_connection()
    factura = conn.execute("SELECT * FROM facturas_recibidas WHERE id = %s", (factura_id,)).fetchone()
    if not factura:
        conn.close()
        flash("Factura no encontrada.", "error")
        return redirect(url_for("listar_facturas_recibidas"))
    conn.execute("""
        UPDATE facturas_recibidas
        SET estado = 'descartada',
            notas = %s,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (notas or None, session.get("username"), factura_id))
    conn.commit()
    conn.close()
    registrar_auditoria("descartar", "factura_recibida", str(factura_id), {"notas": notas})
    flash("Factura descartada.", "ok")
    return redirect(url_for("listar_facturas_recibidas"))

@app.before_request
def proteger_rutas():

    rutas_publicas = {
        "login",
        "solicitar_recuperacion_password",
        "restablecer_password",
        "logout",
        "static",
        "meta_data_deletion",
        "meta_data_deletion_callback",
        "meta_data_deletion_callback_info",
        "meta_data_deletion_status",
        "sugerencias_recomendaciones",
        "sugerencias_denuncias_legacy",
        "portal_buscar",
        "portal_jugador",
        "portal_actualizar_contacto",
        "portal_subir_comprobante",
        "portal_ver_comprobante",
        "portal_subir_comprobante_gasto_compartido",
        "portal_ver_comprobante_gasto_compartido",
        "portal_descargar_recibo",
        "portal_descargar_constancia",
        "portal_confirmar_asistencia",
        "portal_bienestar_asistencia",
        "portal_calendario_ics",
        "pwa_manifest",
        "pwa_service_worker",
        "pwa_config",
        "pwa_push_subscribe",
        "pwa_push_unsubscribe",
        "pwa_push_test",
        "whatsapp_webhook_verify",
        "whatsapp_webhook_receive",
        "ejecutar_automatizaciones_programadas",
    }

    csrf_exentas = {
        "static",
        "whatsapp_webhook_receive",
        "meta_data_deletion_callback",
        "ejecutar_automatizaciones_programadas",
    }

    if request.method == "POST" and request.endpoint not in csrf_exentas:
        if not csrf_valido():
            abort(400)

    if request.path in {"/meta/data-deletion", "/meta/data-deletion-callback"}:
        return

    if request.endpoint in rutas_publicas:
        return

    if "user_id" not in session:
        return redirect(url_for("login"))

    if (
        session.get("debe_cambiar_password")
        and request.endpoint not in {"cambiar_mi_password", "logout", "static"}
    ):
        flash("Cambiá tu contraseña para continuar.", "warning")
        return redirect(url_for("cambiar_mi_password"))

    try:
        g.mantenimiento = obtener_config_mantenimiento()
    except Exception:
        app.logger.exception("No se pudo consultar el modo mantenimiento.")
        g.mantenimiento = {
            "activo": False,
            "mensaje": MAINTENANCE_DEFAULT_MESSAGE,
            "actualizado_en": None,
            "actualizado_por": None,
        }

    if g.mantenimiento["activo"] and session.get("rol") != "admin":
        return render_template("mantenimiento.html", mantenimiento=g.mantenimiento), 503


@app.route("/webhooks/whatsapp", methods=["GET"], endpoint="whatsapp_webhook_verify")
def whatsapp_webhook_verify():
    mode = request.args.get("hub.mode", "").strip()
    verify_token = request.args.get("hub.verify_token", "").strip()
    challenge = request.args.get("hub.challenge", "").strip()

    if mode == "subscribe" and WHATSAPP_VERIFY_TOKEN and secrets.compare_digest(verify_token, WHATSAPP_VERIFY_TOKEN):
        return Response(challenge, status=200, mimetype="text/plain")
    return Response("forbidden", status=403, mimetype="text/plain")


@app.route("/webhooks/whatsapp", methods=["POST"], endpoint="whatsapp_webhook_receive")
def whatsapp_webhook_receive():
    payload = request.get_json(silent=True) or {}

    if WHATSAPP_APP_SECRET:
        firma = request.headers.get("X-Hub-Signature-256", "")
        esperado = "sha256=" + hmac.new(
            WHATSAPP_APP_SECRET.encode("utf-8"),
            request.get_data(),
            hashlib.sha256,
        ).hexdigest()
        if not firma or not secrets.compare_digest(firma, esperado):
            registrar_whatsapp_webhook("signature_error", payload, procesado=False)
            return Response("invalid signature", status=403, mimetype="text/plain")

    procesado = False
    try:
        resumen_previo = resumir_evento_webhook_whatsapp(json.dumps(payload, ensure_ascii=False, default=str))
        app.logger.warning(
            "WhatsApp webhook recibido: messages=%s statuses=%s contacts=%s from=%s statuses_values=%s",
            resumen_previo["messages_count"],
            resumen_previo["statuses_count"],
            resumen_previo["contacts_count"],
            ",".join(resumen_previo["from_values"][:5]),
            ",".join(resumen_previo["status_values"][:5]),
        )
        for entry in payload.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value") or {}
                contactos = {
                    (item.get("wa_id") or item.get("input")): item
                    for item in (value.get("contacts") or [])
                    if item.get("wa_id") or item.get("input")
                }
                for mensaje in value.get("messages", []) or []:
                    wa_id = normalizar_telefono_whatsapp(mensaje.get("from"))
                    contacto = contactos.get(mensaje.get("from")) or contactos.get(wa_id) or {}
                    tipo = (mensaje.get("type") or "text").strip().lower()
                    texto = resumir_contenido_whatsapp(tipo, mensaje)
                    jugador = buscar_jugador_por_whatsapp(wa_id)
                    registro = registrar_whatsapp_mensaje(
                        telefono=wa_id,
                        wa_id=contacto.get("wa_id") or wa_id,
                        jugador_id=(jugador or {}).get("id"),
                        direccion="in",
                        tipo=tipo,
                        texto=texto,
                        meta_message_id=mensaje.get("id"),
                        estado="recibido",
                        payload=mensaje,
                        respuesta=contacto,
                    )
                    if registro == "inserted":
                        enviar_notificacion_whatsapp_inbox_email(
                            mensaje={"tipo": tipo, "texto": texto},
                            telefono=wa_id,
                            jugador=jugador,
                        )
                    procesado = True
                for status_item in value.get("statuses", []) or []:
                    meta_message_id = status_item.get("id")
                    estado = status_item.get("status") or "status"
                    error_items = status_item.get("errors") or []
                    error_codigo = None
                    error_mensaje = None
                    if error_items:
                        error_codigo = str(error_items[0].get("code") or "")
                        error_mensaje = error_items[0].get("title") or error_items[0].get("message")
                    if meta_message_id:
                        conn = get_connection()
                        conn.execute("""
                            UPDATE whatsapp_envios
                            SET estado = %s,
                                error_codigo = COALESCE(%s, error_codigo),
                                error_mensaje = COALESCE(%s, error_mensaje),
                                respuesta = %s
                            WHERE meta_message_id = %s
                        """, (
                            estado,
                            error_codigo,
                            error_mensaje,
                            json.dumps(status_item, ensure_ascii=False, default=str),
                            meta_message_id,
                        ))
                        conn.execute("""
                            UPDATE whatsapp_mensajes
                            SET estado = %s,
                                respuesta = %s
                            WHERE meta_message_id = %s
                        """, (
                            estado,
                            json.dumps(status_item, ensure_ascii=False, default=str),
                            meta_message_id,
                        ))
                        conn.commit()
                        conn.close()
                    procesado = True
    except Exception:
        app.logger.exception("No se pudo procesar el webhook de WhatsApp.")

    registrar_whatsapp_webhook("receive", payload, procesado=procesado)
    return Response("ok", status=200, mimetype="text/plain")


@app.after_request
def auditar_acciones(response):
    if request.method != "POST":
        return response

    if request.endpoint in {
        "login",
        "static",
        "presencia_heartbeat",
        "sugerencias_recomendaciones",
    }:
        return response

    accion_base, entidad = AUDIT_ENDPOINTS.get(
        request.endpoint,
        (request.endpoint or request.path, "sistema"),
    )
    resultado = "ok" if response.status_code < 400 else "error"

    registrar_auditoria(
        accion=f"{accion_base}_{resultado}",
        entidad=entidad,
        entidad_id=audit_entity_id(),
        detalle={
            "endpoint": request.endpoint,
            "path": request.path,
            "method": request.method,
            "status_code": response.status_code,
            "form": sanitized_audit_form(),
            "route_args": request.view_args or {},
        },
    )
    return response


def rol_requerido(*roles_permitidos):
    rol = session.get("rol")
    if rol == "admin":
        return None

    if rol not in roles_permitidos:
        flash("No tenés permiso para acceder a esa sección.", "error")
        return redirect(url_for("index"))
    return None


def tiene_rol(*roles):
    rol = session.get("rol")
    return rol == "admin" or rol in roles


def tiene_permiso(*permisos):
    if session.get("rol") == "admin":
        return True
    permisos_usuario = set(session.get("permisos") or permisos_default_rol(session.get("rol")))
    return any(permiso in permisos_usuario for permiso in permisos)


def permiso_requerido(*permisos):
    if tiene_permiso(*permisos):
        return None
    flash("No tenes permiso para acceder a esa seccion.", "error")
    return redirect(url_for("index"))


def puede_ver_bitacora_tipo(tipo):
    if tipo == "general":
        return tiene_permiso("jugadores_ver")
    if tipo == "finanzas":
        return tiene_permiso("cuotas_ver", "caja_ver")
    if tipo == "salud":
        return tiene_permiso("salud_ver")
    if tipo == "deportivo":
        return tiene_permiso("asistencia_ver", "calendario_ver", "tests_ver")
    return False


def puede_crear_bitacora_tipo(tipo):
    if tipo == "general":
        return tiene_permiso("jugadores_gestionar")
    if tipo == "finanzas":
        return tiene_permiso("cuotas_gestionar", "caja_gestionar")
    if tipo == "salud":
        return tiene_permiso("salud_gestionar")
    if tipo == "deportivo":
        return tiene_permiso("asistencia_gestionar", "calendario_gestionar", "tests_gestionar")
    return False


def tipos_bitacora_disponibles():
    return [
        {"clave": clave, "nombre": nombre}
        for clave, nombre in BITACORA_TIPOS.items()
        if puede_crear_bitacora_tipo(clave)
    ]


def filtrar_bitacora_visible(items):
    return [item for item in items if puede_ver_bitacora_tipo(item["tipo"])]


def puede_ver_operacion():
    return tiene_permiso(
        "jugadores_ver",
        "cuotas_ver",
        "salud_ver",
        "asistencia_ver",
        "comunicaciones_ver",
        "secretaria_ver",
    )


def puede_gestionar_tareas_sig():
    return tiene_permiso(
        "jugadores_gestionar",
        "cuotas_gestionar",
        "salud_gestionar",
        "asistencia_gestionar",
        "comunicaciones_ver",
        "secretaria_gestionar",
    )


def validar_password_nueva(password, confirmacion):
    if not password or not confirmacion:
        return "La contraseña y la confirmación son obligatorias."
    if password != confirmacion:
        return "La confirmación no coincide."
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres."
    return None

def normalizar_username(username):
    return (username or "").strip().lower()


def normalizar_email(email):
    return (email or "").strip().lower()


def normalizar_clave_texto(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().lower())
    return "".join(ch for ch in texto if not unicodedata.combining(ch))


def normalizar_lista_emails(valores):
    emails = []
    invalidos = []
    for valor in valores:
        email = normalizar_email(valor)
        if not email:
            continue
        if "@" not in email or email.startswith("@") or email.endswith("@"):
            invalidos.append(valor.strip())
            continue
        if email not in emails:
            emails.append(email)
    return emails, invalidos


def parsear_emails_config(texto):
    partes = re.split(r"[\s,;]+", str(texto or ""))
    return normalizar_lista_emails(partes)


def serializar_lista_config(valores):
    return json.dumps(list(valores or []), ensure_ascii=False)


def leer_lista_config(valor, default=None):
    texto = str(valor or "").strip()
    if not texto:
        return list(default or [])
    try:
        data = json.loads(texto)
    except (TypeError, ValueError):
        data = re.split(r"[\s,;]+", texto)
    if not isinstance(data, list):
        return list(default or [])
    return [str(item).strip() for item in data if str(item or "").strip()]


def obtener_sugerencias_config(conn=None):
    own_conn = conn is None
    conn = conn or get_connection()
    settings = obtener_app_settings(conn, SUGERENCIAS_CONFIG_KEYS)
    if own_conn:
        conn.close()

    directiva, _ = normalizar_lista_emails(
        leer_lista_config((settings.get(SUGERENCIAS_DIRECTIVA_EMAILS_KEY) or {}).get("valor"))
    )

    actualizado_en = None
    actualizado_por = None
    for row in settings.values():
        if row.get("actualizado_en") and (actualizado_en is None or row["actualizado_en"] > actualizado_en):
            actualizado_en = row["actualizado_en"]
            actualizado_por = row.get("actualizado_por")

    return {
        "directiva_emails": directiva,
        "actualizado_en": actualizado_en,
        "actualizado_por": actualizado_por,
    }


def obtener_destinatarios_sugerencias(conn):
    config = obtener_sugerencias_config(conn)
    return list(config["directiva_emails"])


def asunto_sugerencia_recomendacion(tipo, categoria, registro_id):
    etiqueta = "Recomendacion" if tipo == "recomendacion" else "Sugerencia"
    categoria_texto = f" - {categoria}" if categoria else ""
    return f"{etiqueta}{categoria_texto} #{registro_id} - SIG"


def cuerpo_sugerencia_recomendacion(data, registro_id):
    identidad = "Anonima" if data["anonima"] else (data["nombre"] or "Sin nombre informado")
    contacto = "No informado" if data["anonima"] else (data["contacto"] or "No informado")
    tipo = "Recomendacion" if data["tipo"] == "recomendacion" else "Sugerencia"
    return "\n".join([
        f"Se registro una nueva {tipo.lower()} en el SIG.",
        "",
        f"ID: {registro_id}",
        f"Tipo: {tipo}",
        f"Categoria: {data['categoria'] or 'General'}",
        f"Identidad: {identidad}",
        f"Contacto: {contacto}",
        "",
        "Mensaje:",
        data["mensaje"],
    ])


def enviar_notificacion_sugerencia_recomendacion(data, registro_id, destinatarios):
    if not destinatarios:
        return "sin_destinatarios", 0

    asunto = asunto_sugerencia_recomendacion(data["tipo"], data["categoria"], registro_id)
    cuerpo = cuerpo_sugerencia_recomendacion(data, registro_id)
    enviados = 0
    for destinatario in destinatarios:
        try:
            enviado, _ = enviar_email(destinatario, asunto, cuerpo)
        except Exception:
            app.logger.exception("No se pudo enviar sugerencia/recomendacion %s a %s.", registro_id, destinatario)
            enviado = False
        if enviado:
            enviados += 1

    if enviados == len(destinatarios):
        return "enviado", enviados
    if enviados:
        return "parcial", enviados
    return "fallo_email", enviados


def formatear_numero_socio(numero):
    try:
        return f"{int(numero):05d}"
    except (TypeError, ValueError):
        return ""


def siguiente_numero_socio(conn):
    fila = conn.execute("""
        SELECT COALESCE(
            MAX(
                CASE
                    WHEN NULLIF(REGEXP_REPLACE(COALESCE(numero_socio, ''), '[^0-9]', '', 'g'), '') IS NOT NULL
                    THEN CAST(REGEXP_REPLACE(numero_socio, '[^0-9]', '', 'g') AS INTEGER)
                    ELSE NULL
                END
            ),
            0
        ) AS maximo
        FROM jugadores
    """).fetchone()
    return formatear_numero_socio((fila["maximo"] or 0) + 1)


def recalcular_numeros_socio(conn):
    jugadores = conn.execute("""
        SELECT id, nombre, apellido, fecha_ingreso, numero_socio, numero_afiliado_obra_social
        FROM jugadores
        ORDER BY id ASC
        FOR UPDATE
    """).fetchall()

    ordenados = sorted(
        jugadores,
        key=lambda jugador: (
            0 if normalizar_texto_match(jugador.get("apellido")) == "del valle" and normalizar_texto_match(jugador.get("nombre")) == "eduardo" else 1,
            1 if not validar_fecha_movimiento(jugador.get("fecha_ingreso")) else 0,
            validar_fecha_movimiento(jugador.get("fecha_ingreso")) or "9999-12-31",
            jugador["id"],
        ),
    )

    migrados_obra_social = 0
    for indice, jugador in enumerate(ordenados, start=1):
        numero_socio_anterior = (jugador["numero_socio"] or "").strip()
        numero_afiliado_actual = (jugador["numero_afiliado_obra_social"] or "").strip()
        if numero_socio_anterior and not numero_afiliado_actual:
            conn.execute(
                "UPDATE jugadores SET numero_afiliado_obra_social = %s WHERE id = %s",
                (numero_socio_anterior, jugador["id"]),
            )
            migrados_obra_social += 1
        conn.execute(
            "UPDATE jugadores SET numero_socio = %s WHERE id = %s",
            (formatear_numero_socio(indice), jugador["id"]),
        )

    primero = ordenados[0] if ordenados else None
    return {
        "cantidad": len(ordenados),
        "primero": {
            "id": primero["id"],
            "nombre": primero["nombre"],
            "apellido": primero["apellido"],
        } if primero else None,
        "migrados_obra_social": migrados_obra_social,
    }


def enviar_recibo_cuota_por_email(cuota, archivo_recibo):
    if not archivo_recibo or not Path(archivo_recibo).exists():
        return False, None, "sin_archivo"
    asunto = f"Recibo de cuota {cuota.get('periodo') or '-'}"
    cuerpo = construir_texto_recibo_cuota(cuota)
    return enviar_email_jugador_con_adjuntos(
        cuota,
        asunto,
        cuerpo,
        adjuntos=[{
            "path": str(archivo_recibo),
            "maintype": "application",
            "subtype": "pdf",
            "filename": Path(archivo_recibo).name,
        }],
    )


def smtp_configurado():
    return bool(SMTP_HOST and SMTP_FROM)


def render_email_html(cuerpo, logo_cid=None):
    cuerpo_html = "<br>".join(html.escape(linea) if linea else "" for linea in str(cuerpo or "").splitlines())
    logo_html = ""
    if logo_cid:
        logo_html = (
            f'<div style="margin-top:12px;">'
            f'<img src="cid:{logo_cid}" alt="Ruda Macho Rugby Club" '
            f'style="max-width:120px;height:auto;display:block;">'
            f"</div>"
        )
    return (
        '<html><body style="font-family:Arial,Helvetica,sans-serif;font-size:14px;color:#111827;line-height:1.5;">'
        f'<div>{cuerpo_html}</div>'
        '<div style="margin-top:20px;">'
        '<strong>Tesorer\u00eda - Ruda Macho Rugby Club</strong>'
        f'{logo_html}'
        '</div>'
        '</body></html>'
    )


def enviar_email(destinatario, asunto, cuerpo, adjuntos=None):
    if not smtp_configurado():
        return False, "smtp"

    cuerpo_base = str(cuerpo or "").rstrip()
    cuerpo_texto = cuerpo_base + "\n\nTesorer\u00eda - Ruda Macho Rugby Club"
    mensaje = EmailMessage()
    mensaje["From"] = formataddr((SMTP_FROM_NAME, SMTP_FROM))
    mensaje["To"] = destinatario
    mensaje["Subject"] = asunto
    mensaje.set_content(cuerpo_texto)
    logo_path = BASE_DIR / "static" / "img" / "logo.png"
    logo_cid = None
    if logo_path.exists():
        logo_cid = make_msgid(domain="rudamachorugby.com")[1:-1]
    mensaje.add_alternative(render_email_html(cuerpo_base, logo_cid=logo_cid), subtype="html")
    if logo_cid and logo_path.exists():
        with logo_path.open("rb") as fh:
            mensaje.get_payload()[-1].add_related(
                fh.read(),
                maintype="image",
                subtype="png",
                cid=f"<{logo_cid}>",
                filename="logo.png",
            )
    for adjunto in adjuntos or []:
        ruta = adjunto.get("path")
        if not ruta:
            continue
        ruta = Path(ruta)
        if not ruta.exists():
            continue
        maintype = adjunto.get("maintype", "application")
        subtype = adjunto.get("subtype", "octet-stream")
        filename = adjunto.get("filename") or ruta.name
        with ruta.open("rb") as fh:
            mensaje.add_attachment(
                fh.read(),
                maintype=maintype,
                subtype=subtype,
                filename=filename,
            )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
        if SMTP_USE_TLS:
            smtp.starttls()
        if SMTP_USER:
            smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(mensaje)
    return True, None


def email_jugador_preferido(jugador):
    return normalizar_email(
        (jugador or {}).get("email")
        or (jugador or {}).get("email_tutor")
        or ""
    )


def nombre_jugador_corto(jugador):
    return ((jugador or {}).get("nombre") or "").strip() or ((jugador or {}).get("apellido") or "").strip() or "jugador"


def enviar_email_jugador(jugador, asunto, cuerpo):
    destinatario = email_jugador_preferido(jugador)
    if not destinatario:
        return False, None, "sin_email"
    enviado, motivo = enviar_email(destinatario, asunto, cuerpo)
    return enviado, destinatario, motivo


def enviar_email_jugador_con_adjuntos(jugador, asunto, cuerpo, adjuntos=None):
    destinatario = email_jugador_preferido(jugador)
    if not destinatario:
        return False, None, "sin_email"
    enviado, motivo = enviar_email(destinatario, asunto, cuerpo, adjuntos=adjuntos)
    return enviado, destinatario, motivo


def mensaje_fallo_email(motivo, destinatario=None):
    if motivo == "sin_email":
        return "No hay un email configurado para este jugador."
    if motivo == "smtp":
        return "SMTP no est\u00e1 configurado en el servidor."
    if destinatario:
        return f"No se pudo enviar el email a {destinatario}. Revis\u00e1 la configuraci\u00f3n SMTP o intent\u00e1 nuevamente."
    return "No se pudo enviar el email. Revis\u00e1 la configuraci\u00f3n SMTP o intent\u00e1 nuevamente."


def resumir_envio_masivo_email(resultados, etiqueta):
    enviados = sum(1 for ok, _, _ in resultados if ok)
    sin_email = sum(1 for ok, _, motivo in resultados if not ok and motivo == "sin_email")
    smtp = sum(1 for ok, _, motivo in resultados if not ok and motivo == "smtp")
    otros = sum(1 for ok, _, motivo in resultados if not ok and motivo not in ("sin_email", "smtp"))
    if enviados:
        partes = [f"Se enviaron {enviados} {etiqueta}."]
        if sin_email:
            partes.append(f"{sin_email} sin email.")
        if smtp:
            partes.append(f"{smtp} sin SMTP.")
        if otros:
            partes.append(f"{otros} con error.")
        return " ".join(partes), "ok"
    if smtp and not sin_email and not otros:
        return f"No se enviaron {etiqueta}: SMTP no est\u00e1 configurado en el servidor.", "error"
    detalles = []
    if sin_email:
        detalles.append(f"{sin_email} sin email")
    if smtp:
        detalles.append(f"{smtp} sin SMTP")
    if otros:
        detalles.append(f"{otros} con error")
    if detalles:
        return f"No se enviaron {etiqueta}: " + ", ".join(detalles) + ".", "error"
    return f"No se enviaron {etiqueta}.", "error"


def construir_texto_recibo_cuota(cuota):
    nombre = nombre_jugador_corto(cuota)
    return (
        f"Hola {nombre}, te adjuntamos el recibo correspondiente a la cuota {cuota.get('periodo') or '-'} "
        f"por {formato_moneda(cuota.get('importe') or 0)}.\n\n"
        "Gracias por colaborar con el club."
    )


def construir_texto_recordatorio_cuota(cuota):
    nombre = nombre_jugador_corto(cuota)
    estado = "venció" if (cuota.get("dias_vencida") or 0) > 0 else "vence"
    fecha = cuota.get("fecha_vencimiento") or "-"
    return (
        f"Hola {nombre}, te escribimos de Ruda Macho Rugby Club.\n\n"
        f"La cuota {cuota.get('periodo') or '-'} por {formato_moneda(cuota.get('importe') or 0)} {estado} el {fecha}.\n"
        "Si ya realizaste el pago, podés responder este correo o cargar el comprobante desde tu portal.\n\n"
        "Gracias."
    )


def construir_texto_recordatorio_ficha(ficha):
    nombre = nombre_jugador_corto(ficha)
    if ficha.get("estado_documento") == "vencida":
        estado = f"venció el {ficha.get('fecha_vencimiento') or '-'}"
    elif ficha.get("estado_documento") == "por_vencer":
        estado = f"vence el {ficha.get('fecha_vencimiento') or '-'}"
    else:
        estado = "figura pendiente de carga"
    return (
        f"Hola {nombre}, te escribimos de Ruda Macho Rugby Club.\n\n"
        f"La ficha médica {estado}. Cuando puedas, acercanos la actualización o cargala por los canales habituales.\n\n"
        "Gracias."
    )


def construir_texto_recordatorio_evento(jugador, evento, portal_url=None):
    nombre = nombre_jugador_corto(jugador)
    cuerpo = (
        f"Hola {nombre}, te recordamos el evento {evento.get('titulo') or evento.get('tipo') or 'del club'} "
        f"del {evento.get('fecha') or '-'}"
    )
    if evento.get("hora_inicio"):
        cuerpo += f" a las {evento['hora_inicio']}"
    cuerpo += "."
    if evento.get("ubicacion"):
        cuerpo += f"\nLugar: {evento['ubicacion']}."
    if evento.get("descripcion"):
        cuerpo += f"\nDetalle: {evento['descripcion']}"
    if portal_url:
        cuerpo += f"\n\nPodés confirmar asistencia desde tu portal:\n{portal_url}"
    return cuerpo + "\n\nGracias."


def construir_texto_rechazo_comprobante(cuota, observaciones=None, portal_url=None):
    nombre = nombre_jugador_corto(cuota)
    cuerpo = (
        f"Hola {nombre}, revisamos el comprobante de la cuota {cuota.get('periodo') or '-'} "
        f"por {formato_moneda(cuota.get('importe') or 0)} y qued? rechazado para correcci?n."
    )
    if observaciones:
        cuerpo += f"\nMotivo: {observaciones}"
    if portal_url:
        cuerpo += f"\n\nPod?s volver a cargarlo desde tu portal:\n{portal_url}"

def semaforo_lesion(lesion):
    estado = (lesion.get("estado") or "").strip()
    if estado == "Activa":
        return "rojo"
    if estado.lower().startswith("en recuperaci"):
        return "amarillo"
    if estado.lower().startswith("alta"):
        return "verde"
    return "gris"


def estado_ficha_portal(ficha):
    if not ficha:
        return {"label": "Sin ficha m?dica", "nivel": "warning"}
    if not ficha.get("presentada"):
        return {"label": "Ficha pendiente", "nivel": "warning"}
    fecha_vencimiento = validar_fecha_movimiento(ficha.get("fecha_vencimiento"))
    if fecha_vencimiento:
        try:
            fecha = datetime.strptime(fecha_vencimiento, "%Y-%m-%d").date()
            hoy = ahora_sig().date()
            if fecha < hoy:
                return {"label": "Ficha vencida", "nivel": "danger"}
            if fecha <= hoy + timedelta(days=30):
                return {"label": "Ficha por vencer", "nivel": "warning"}
        except ValueError:
            pass
    if ficha_tiene_apto_efectivo(ficha):
        return {"label": "Apto fisico vigente", "nivel": "success"}
    return {"label": "Ficha presentada sin apto", "nivel": "warning"}


def ficha_tiene_apto_efectivo(ficha):
    if not ficha:
        return False
    return bool(
        ficha.get("apto_fisico")
        or (ficha.get("documento_drive_file_id") or "").strip()
    )


def repartir_importe_gasto(total, cantidad):
    if cantidad <= 0:
        return []
    base = round(total / cantidad, 2)
    valores = [base for _ in range(cantidad)]
    diferencia = round(total - sum(valores), 2)
    if valores:
        valores[-1] = round(valores[-1] + diferencia, 2)
    return valores


def crear_token_recuperacion(conn, usuario_id):
    token = secrets.token_urlsafe(32)
    token_hash = generate_password_hash(token)
    conn.execute("""
        INSERT INTO password_reset_tokens (
            usuario_id, token_hash, creado_en, expira_en, usado
        )
        VALUES (
            %s,
            %s,
            CURRENT_TIMESTAMP,
            CURRENT_TIMESTAMP + (%s || ' minutes')::interval,
            0
        )
    """, (usuario_id, token_hash, str(PASSWORD_RESET_TOKEN_MINUTES)))
    return token


def buscar_token_recuperacion(conn, token):
    tokens = conn.execute("""
        SELECT
            t.*,
            u.username,
            u.email
        FROM password_reset_tokens t
        JOIN usuarios u ON u.id = t.usuario_id
        WHERE t.usado = 0
          AND t.expira_en > CURRENT_TIMESTAMP
        ORDER BY t.creado_en DESC
        LIMIT 100
    """).fetchall()

    for fila in tokens:
        if check_password_hash(fila["token_hash"], token):
            return fila
    return None


def login_bloqueado(conn, username, ip):
    intentos = conn.execute("""
        SELECT COUNT(*) AS total
        FROM login_attempts
        WHERE success = 0
          AND username = %s
          AND ip = %s
          AND fecha >= CURRENT_TIMESTAMP - (%s || ' minutes')::interval
    """, (normalizar_username(username), ip, str(LOGIN_ATTEMPT_WINDOW_MINUTES))).fetchone()
    return (intentos["total"] or 0) >= MAX_LOGIN_ATTEMPTS


def registrar_intento_login(conn, username, ip, success):
    normalized_username = normalizar_username(username)
    conn.execute("""
        INSERT INTO login_attempts (username, ip, success)
        VALUES (%s, %s, %s)
    """, (normalized_username, ip, 1 if success else 0))

    if success:
        conn.execute("""
            DELETE FROM login_attempts
            WHERE username = %s
              AND ip = %s
              AND success = 0
        """, (normalized_username, ip))
    else:
        conn.execute("""
            DELETE FROM login_attempts
            WHERE fecha < CURRENT_TIMESTAMP - INTERVAL '7 days'
        """)


SUGERENCIA_RECOMENDACION_CATEGORIAS = [
    {"clave": "general", "nombre": "General"},
    {"clave": "infraestructura", "nombre": "Infraestructura"},
    {"clave": "comunicacion", "nombre": "Comunicacion"},
    {"clave": "finanzas", "nombre": "Finanzas / cuotas"},
    {"clave": "convivencia", "nombre": "Convivencia"},
    {"clave": "deportivo", "nombre": "Deportivo"},
    {"clave": "social", "nombre": "Social"},
    {"clave": "otra", "nombre": "Otra"},
]

SUGERENCIA_EMAIL_ESTADOS = {
    "enviado": {
        "label": "Email enviado",
        "badge": "badge-success",
        "descripcion": "La notificacion salio a todos los destinatarios configurados.",
    },
    "parcial": {
        "label": "Email parcial",
        "badge": "badge-warning",
        "descripcion": "La notificacion salio solo a algunos destinatarios.",
    },
    "sin_destinatarios": {
        "label": "Sin destinatarios",
        "badge": "badge-danger",
        "descripcion": "No habia emails configurados para el area correspondiente.",
    },
    "fallo_email": {
        "label": "Fallo email",
        "badge": "badge-danger",
        "descripcion": "Habia destinatarios, pero el servidor SMTP no pudo enviar la notificacion.",
    },
    "pendiente": {
        "label": "Pendiente",
        "badge": "badge-warning",
        "descripcion": "El registro quedo guardado antes de completar el intento de notificacion.",
    },
}

SUGERENCIA_SEGUIMIENTO_ESTADOS = {
    "nuevo": {
        "label": "Nuevo",
        "badge": "badge-info",
    },
    "en_revision": {
        "label": "En revision",
        "badge": "badge-warning",
    },
    "resuelto": {
        "label": "Resuelto",
        "badge": "badge-success",
    },
    "archivado": {
        "label": "Archivado",
        "badge": "badge-muted",
    },
}


def info_email_estado_sugerencia(estado):
    return SUGERENCIA_EMAIL_ESTADOS.get(estado or "", {
        "label": estado or "Sin estado",
        "badge": "badge-muted",
        "descripcion": "Estado no reconocido.",
    })


def info_seguimiento_estado_sugerencia(estado):
    return SUGERENCIA_SEGUIMIENTO_ESTADOS.get(estado or "", {
        "label": estado or "Sin estado",
        "badge": "badge-muted",
    })


def puede_ver_tipo_sugerencia(tipo):
    return tipo in {"sugerencia", "recomendacion"} and tiene_permiso("sugerencias_ver")


def puede_gestionar_tipo_sugerencia(tipo):
    return tiene_permiso("sugerencias_gestionar") and puede_ver_tipo_sugerencia(tipo)


def normalizar_estado_seguimiento_sugerencia(estado):
    estado = (estado or "").strip().lower()
    return estado if estado in SUGERENCIA_SEGUIMIENTO_ESTADOS else "nuevo"


@app.route("/sugerencias-recomendaciones", methods=["GET", "POST"])
def sugerencias_recomendaciones():
    categorias_validas = {item["clave"] for item in SUGERENCIA_RECOMENDACION_CATEGORIAS}
    data = {
        "tipo": request.form.get("tipo", "sugerencia").strip().lower(),
        "categoria": request.form.get("categoria", "general").strip().lower(),
        "anonima": request.form.get("anonima", "1") == "1",
        "nombre": request.form.get("nombre", "").strip(),
        "contacto": request.form.get("contacto", "").strip(),
        "mensaje": request.form.get("mensaje", "").strip(),
    }

    if request.method == "POST":
        if not consumir_limite_publico("sugerencias_recomendaciones", max_intentos=5, minutos=60):
            flash("Se alcanzo el limite de envios. Intenta nuevamente mas tarde.", "error")
            return render_template(
                "sugerencias_recomendaciones.html",
                data=data,
                categorias=SUGERENCIA_RECOMENDACION_CATEGORIAS,
            ), 429
        if request.form.get("website", "").strip():
            flash("No se pudo registrar el formulario.", "error")
            return render_template(
                "sugerencias_recomendaciones.html",
                data=data,
                categorias=SUGERENCIA_RECOMENDACION_CATEGORIAS,
            )

        if data["tipo"] not in {"sugerencia", "recomendacion"}:
            data["tipo"] = "sugerencia"
        if data["categoria"] not in categorias_validas:
            data["categoria"] = "general"
        if data["anonima"]:
            data["nombre"] = ""
            data["contacto"] = ""

        if len(data["mensaje"]) < 10:
            flash("Escribi un mensaje de al menos 10 caracteres.", "error")
            return render_template(
                "sugerencias_recomendaciones.html",
                data=data,
                categorias=SUGERENCIA_RECOMENDACION_CATEGORIAS,
            )

        conn = get_connection()
        destinatarios = obtener_destinatarios_sugerencias(conn)
        registro = conn.execute("""
            INSERT INTO sugerencias_denuncias (
                tipo, categoria, anonima, nombre, contacto, mensaje, destinatarios, email_estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data["tipo"],
            data["categoria"],
            1 if data["anonima"] else 0,
            data["nombre"] or None,
            data["contacto"] or None,
            data["mensaje"],
            json.dumps(destinatarios, ensure_ascii=False),
            "pendiente",
        )).fetchone()

        email_estado, enviados = enviar_notificacion_sugerencia_recomendacion(data, registro["id"], destinatarios)
        conn.execute("""
            UPDATE sugerencias_denuncias
            SET email_estado = %s,
                notificado_en = CASE WHEN %s IN ('enviado', 'parcial') THEN CURRENT_TIMESTAMP ELSE notificado_en END
            WHERE id = %s
        """, (email_estado, email_estado, registro["id"]))
        conn.commit()
        conn.close()

        registrar_auditoria("crear", "sugerencia_recomendacion", str(registro["id"]), {
            "tipo": data["tipo"],
            "categoria": data["categoria"],
            "anonima": data["anonima"],
            "email_estado": email_estado,
            "emails_enviados": enviados,
        }, username="portal")

        if email_estado in {"enviado", "parcial"}:
            flash("Tu mensaje fue registrado y derivado al area correspondiente.", "ok")
        else:
            flash("Tu mensaje fue registrado. No pudimos enviar el aviso por email, pero quedo disponible para revision interna.", "warning")
        return redirect(url_for("sugerencias_recomendaciones"))

    return render_template(
        "sugerencias_recomendaciones.html",
        data=data,
        categorias=SUGERENCIA_RECOMENDACION_CATEGORIAS,
    )


@app.route("/sugerencias-denuncias")
def sugerencias_denuncias_legacy():
    return redirect(url_for("sugerencias_recomendaciones"), code=301)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        ip = audit_request_ip()

        conn = get_connection()
        if login_bloqueado(conn, username, ip):
            conn.close()
            registrar_auditoria(
                "login_bloqueado",
                "usuario",
                None,
                {"username": username, "ip": ip},
                username=username,
            )
            flash("Demasiados intentos fallidos. Proba nuevamente en unos minutos.", "error")
            return render_template("login.html")

        usuario = conn.execute("""
            SELECT u.*, r.permisos AS rol_permisos
            FROM usuarios u
            LEFT JOIN roles r ON r.nombre = u.rol
            WHERE lower(u.username) = %s
        """, (normalizar_username(username),)).fetchone()

        if usuario and check_password_hash(usuario["password"], password):
            registrar_intento_login(conn, username, ip, True)
            conn.execute("""
                UPDATE usuarios
                SET ultimo_login = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (usuario["id"],))
            conn.commit()
            conn.close()

            session.clear()
            session.permanent = True
            csrf_token()
            session["user_id"] = usuario["id"]
            session["username"] = usuario["username"]
            session["rol"] = usuario["rol"]
            session["debe_cambiar_password"] = bool(usuario.get("debe_cambiar_password"))
            session["onboarding_visto"] = bool(usuario.get("onboarding_visto"))
            session["permisos"] = deserializar_permisos(
                usuario["rol_permisos"],
                usuario["rol"],
            )
            registrar_auditoria(
                "login_ok",
                "usuario",
                str(usuario["id"]),
                {"username": username},
                usuario_id=usuario["id"],
                username=usuario["username"],
                rol=usuario["rol"],
            )
            if session["debe_cambiar_password"]:
                flash("Por seguridad, cambiá tu contraseña para continuar.", "warning")
                return redirect(url_for("cambiar_mi_password"))
            return redirect(url_for("index"))
        else:
            registrar_intento_login(conn, username, ip, False)
            conn.commit()
            conn.close()
            registrar_auditoria(
                "login_error",
                "usuario",
                None,
                {"username": username},
                username=username,
            )
            flash("Usuario o contraseña incorrectos", "error")

    return render_template("login.html")


@app.route("/meta/data-deletion")
def meta_data_deletion():
    return render_template("meta_data_deletion.html")


@app.route("/meta/data-deletion-callback", methods=["GET"], endpoint="meta_data_deletion_callback_info")
def meta_data_deletion_callback_info():
    return render_template("meta_data_deletion_callback.html")


@app.route("/meta/data-deletion-callback", methods=["POST"], endpoint="meta_data_deletion_callback")
def meta_data_deletion_callback():
    signed_request = request.form.get("signed_request", "")
    payload = parse_meta_signed_request(signed_request)
    if not payload:
        return Response("invalid signed_request", status=400, mimetype="text/plain")

    confirmation_code = secrets.token_hex(12)
    status_url = url_for(
        "meta_data_deletion_status",
        confirmation_code=confirmation_code,
        _external=True,
    )
    try:
        registrar_auditoria(
            "solicitud_eliminacion_datos",
            "meta_app",
            str((payload or {}).get("user_id") or ""),
            {
                "confirmation_code": confirmation_code,
                "algorithm": (payload or {}).get("algorithm"),
                "issued_at": (payload or {}).get("issued_at"),
            },
            username="meta",
            rol="sistema",
        )
    except Exception:
        pass
    return Response(
        json.dumps(
            {
                "url": status_url,
                "confirmation_code": confirmation_code,
            },
            ensure_ascii=False,
        ),
        status=200,
        mimetype="application/json",
    )


@app.route("/meta/data-deletion-status/<confirmation_code>")
def meta_data_deletion_status(confirmation_code):
    return render_template(
        "meta_data_deletion_status.html",
        confirmation_code=confirmation_code,
    )


@app.route("/password/recuperar", methods=["GET", "POST"])
def solicitar_recuperacion_password():
    if request.method == "POST":
        email = normalizar_email(request.form.get("email", ""))

        if email:
            conn = get_connection()
            usuario = conn.execute("""
                SELECT id, username, email
                FROM usuarios
                WHERE lower(email) = %s
            """, (email,)).fetchone()

            if usuario:
                token = crear_token_recuperacion(conn, usuario["id"])
                reset_url = url_for("restablecer_password", token=token, _external=True)
                enviado = False
                try:
                    enviado = enviar_email(
                        usuario["email"],
                        "Recuperar clave - SIG Ruda Macho",
                        (
                            f"Hola {usuario['username']},\n\n"
                            "Recibimos un pedido para restablecer tu clave del SIG.\n"
                            f"El enlace vence en {PASSWORD_RESET_TOKEN_MINUTES} minutos:\n\n"
                            f"{reset_url}\n\n"
                            "Si no pediste este cambio, pod?s ignorar este mensaje."
                        ),
                    )
                except Exception:
                    app.logger.exception("No se pudo enviar email de recuperacion a %s.", email)
                conn.commit()
                conn.close()

                registrar_auditoria(
                    "password_reset_solicitado",
                    "usuario",
                    str(usuario["id"]),
                    {"email": email, "email_enviado": enviado},
                    username=usuario["username"],
                )
            else:
                conn.close()

        flash("Si el email está registrado, vas a recibir un enlace de recuperación.", "ok")
        return redirect(url_for("login"))

    return render_template("password_recuperar.html")


@app.route("/password/restablecer/<token>", methods=["GET", "POST"])
def restablecer_password(token):
    conn = get_connection()
    token_row = buscar_token_recuperacion(conn, token)
    if token_row is None:
        conn.close()
        flash("El enlace de recuperación no es válido o venció.", "error")
        return redirect(url_for("login"))

    usuario = {
        "id": token_row["usuario_id"],
        "username": token_row["username"],
        "email": token_row["email"],
    }

    if request.method == "POST":
        password_nueva = request.form.get("password_nueva", "")
        password_confirmacion = request.form.get("password_confirmacion", "")

        error = validar_password_nueva(password_nueva, password_confirmacion)
        if error:
            conn.close()
            flash(error, "error")
            return render_template("password_reset_form.html", token=token, usuario=usuario)

        conn.execute("""
            UPDATE usuarios
            SET password = %s,
                debe_cambiar_password = 0
            WHERE id = %s
        """, (generate_password_hash(password_nueva), usuario["id"]))
        conn.execute("""
            UPDATE password_reset_tokens
            SET usado = 1,
                usado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (token_row["id"],))
        conn.commit()
        conn.close()

        registrar_auditoria(
            "password_reset_ok",
            "usuario",
            str(usuario["id"]),
            {"origen": "recuperacion_email"},
            username=usuario["username"],
        )
        flash("Contraseña actualizada. Ya podés iniciar sesión.", "ok")
        return redirect(url_for("login"))

    conn.close()
    return render_template("password_reset_form.html", token=token, usuario=usuario)


def login_required():
    if "user_id" not in session:
        return redirect(url_for("login"))

@app.route("/logout")
def logout():
    registrar_auditoria(
        "logout",
        "usuario",
        str(session.get("user_id")) if session.get("user_id") else None,
        {"username": session.get("username")},
    )
    session.clear()
    return redirect(url_for("login"))


@app.route("/mi-cuenta/password", methods=["GET", "POST"])
def cambiar_mi_password():
    usuario_id = session.get("user_id")
    conn = get_connection()
    usuario = conn.execute("""
        SELECT id, username, password, rol, debe_cambiar_password
        FROM usuarios
        WHERE id = %s
    """, (usuario_id,)).fetchone()

    if usuario is None:
        conn.close()
        session.clear()
        flash("Tu usuario ya no existe. Iniciá sesión nuevamente.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        password_actual = request.form.get("password_actual", "")
        password_nueva = request.form.get("password_nueva", "")
        password_confirmacion = request.form.get("password_confirmacion", "")

        error = validar_password_nueva(password_nueva, password_confirmacion)
        if error:
            conn.close()
            flash(error, "error")
            return render_template("password_form.html", usuario=usuario, modo="propio")

        if not usuario["debe_cambiar_password"] and not check_password_hash(usuario["password"], password_actual):
            conn.close()
            flash("La contraseña actual no es correcta.", "error")
            return render_template("password_form.html", usuario=usuario, modo="propio")

        conn.execute("""
            UPDATE usuarios
            SET password = %s,
                debe_cambiar_password = 0
            WHERE id = %s
        """, (generate_password_hash(password_nueva), usuario_id))
        conn.commit()
        conn.close()

        session["debe_cambiar_password"] = False
        flash("Contraseña actualizada correctamente.", "ok")
        return redirect(url_for("index"))

    conn.close()
    return render_template("password_form.html", usuario=usuario, modo="propio")


@app.route("/mi-cuenta/onboarding", methods=["POST"])
def descartar_onboarding():
    conn = get_connection()
    conn.execute("""
        UPDATE usuarios
        SET onboarding_visto = 1
        WHERE id = %s
    """, (session.get("user_id"),))
    conn.commit()
    conn.close()
    session["onboarding_visto"] = True
    return redirect(destino_interno(request.form.get("next")))


@app.route("/")
def index():
    mes = request.args.get("mes")  # formato YYYY-MM
    mes_actual = mes or ahora_sig().strftime("%Y-%m")

    conn = get_connection()

    total_jugadores = conn.execute("""
        SELECT COUNT(*) AS total
        FROM jugadores
    """).fetchone()["total"]

    jugadores_con_deuda = conn.execute("""
        SELECT
            j.id,
            j.nombre,
            j.apellido,
            COALESCE(SUM(c.importe), 0) AS deuda,
            SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1
                    ELSE 0
                END
            ) AS cuotas_vencidas
        FROM jugadores j
        JOIN cuotas c ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY j.id, j.nombre, j.apellido
        HAVING COALESCE(SUM(c.importe), 0) > 0
        ORDER BY deuda DESC, j.apellido, j.nombre
    """).fetchall()

    fichas_vencidas = conn.execute("""
        SELECT
            j.id,
            j.nombre,
            j.apellido,
            f.fecha_vencimiento
        FROM jugadores j
        JOIN fichas_medicas f ON j.id = f.jugador_id
        WHERE f.fecha_vencimiento IS NOT NULL
          AND NULLIF(f.fecha_vencimiento::text, '') IS NOT NULL
          AND f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND f.fecha_vencimiento::date < CURRENT_DATE
        ORDER BY f.fecha_vencimiento ASC, j.apellido, j.nombre
    """).fetchall()

    lesiones_activas = conn.execute("""
        SELECT
            l.id,
            j.id AS jugador_id,
            j.nombre,
            j.apellido,
            l.fecha_lesion,
            l.tipo_lesion,
            l.zona_cuerpo,
            l.estado
        FROM lesiones l
        JOIN jugadores j ON j.id = l.jugador_id
        WHERE l.estado IN ('Activa', 'En recuperaci?n')
        ORDER BY
            CASE
                WHEN l.estado = 'Activa' THEN 0
                ELSE 1
            END,
            l.fecha_lesion DESC
    """).fetchall()

    total_recaudado_mes = conn.execute("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuotas
        WHERE pagado = 1
           AND substring(fecha_pago from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    deuda_total = conn.execute("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuotas
        WHERE pagado = 0
        AND COALESCE(importe, 0) > 0
    """).fetchone()["total"]

    cuotas_pagadas_mes = conn.execute("""
    SELECT COUNT(*) AS total
    FROM cuotas
    WHERE pagado = 1
      AND substring(fecha_pago from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    deuda_vencida_total = conn.execute("""
    SELECT COALESCE(SUM(importe), 0) AS total
    FROM cuotas
    WHERE pagado = 0
      AND COALESCE(importe, 0) > 0
      AND fecha_vencimiento IS NOT NULL
      AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
      AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
      AND fecha_vencimiento::date < CURRENT_DATE
    """).fetchone()["total"]

    cuotas_pendientes = conn.execute("""
    SELECT COUNT(*) AS total
    FROM cuotas
    WHERE pagado = 0
      AND COALESCE(importe, 0) > 0
      AND substring(periodo from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    cuotas_pendientes_lista = conn.execute("""
    SELECT
        c.id,
        c.periodo,
        c.importe,
        c.fecha_vencimiento,
        j.id AS jugador_id,
        j.nombre,
        j.apellido,
        j.categoria
    FROM cuotas c
    JOIN jugadores j ON j.id = c.jugador_id
    WHERE c.pagado = 0
      AND COALESCE(c.importe, 0) > 0
    ORDER BY
        CASE
            WHEN c.fecha_vencimiento IS NOT NULL
             AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
             AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
             AND c.fecha_vencimiento::date < CURRENT_DATE
            THEN 0
            ELSE 1
        END,
        c.fecha_vencimiento ASC,
        j.apellido,
        j.nombre
    LIMIT 20
    """).fetchall()

    comprobantes_pendientes_lista = conn.execute("""
        SELECT
            c.id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.comprobante_fecha,
            c.comprobante_usuario,
            j.nombre,
            j.apellido
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.comprobante_drive_file_id IS NOT NULL
          AND COALESCE(c.anulada, 0) = 0
          AND COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
        ORDER BY c.comprobante_fecha DESC NULLS LAST, c.id DESC
        LIMIT 10
    """).fetchall()

    comprobantes_pendientes_count = conn.execute("""
        SELECT COUNT(*) AS total
        FROM cuotas
        WHERE comprobante_drive_file_id IS NOT NULL
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(NULLIF(comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
    """).fetchone()["total"]

    fichas_por_vencer_count = conn.execute("""
        SELECT COUNT(*) AS total
        FROM fichas_medicas
        WHERE fecha_vencimiento IS NOT NULL
          AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
          AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND fecha_vencimiento::date >= CURRENT_DATE
          AND fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
    """).fetchone()["total"]

    conn.close()

    total_cuotas_mes = cuotas_pagadas_mes + cuotas_pendientes
    cobranza_ratio = round((cuotas_pagadas_mes / total_cuotas_mes) * 100, 1) if total_cuotas_mes else None

    resumen_notificaciones = None
    if tiene_permiso("comunicaciones_ver"):
        notificaciones = obtener_notificaciones_operativas()
        resumen_notificaciones = {
            "cuotas_vencidas": len(notificaciones["cuotas_vencidas"]),
            "cuotas_por_vencer": len(notificaciones["cuotas_por_vencer"]),
            "fichas": len(notificaciones["fichas"]),
            "asistencia_baja": len(notificaciones["asistencia_baja"]),
            "comprobantes": len(notificaciones["comprobantes_pendientes"]),
            "cambios_portal": len(notificaciones["cambios_portal"]),
        }

    sistema_resumen = obtener_estado_sistema_admin() if session.get("rol") == "admin" else None

    return render_template(
        "dashboard.html",
        total_jugadores=total_jugadores,
        jugadores_con_deuda=jugadores_con_deuda,
        fichas_vencidas=fichas_vencidas,
        lesiones_activas=lesiones_activas,
        total_recaudado_mes=total_recaudado_mes,
        deuda_total=deuda_total,
        deuda_vencida_total=deuda_vencida_total,
        cuotas_pagadas_mes=cuotas_pagadas_mes,
        cuotas_pendientes=cuotas_pendientes,
        cobranza_ratio=cobranza_ratio,
        fichas_por_vencer_count=fichas_por_vencer_count,
        cuotas_pendientes_lista=cuotas_pendientes_lista,
        comprobantes_pendientes_count=comprobantes_pendientes_count,
        comprobantes_pendientes_lista=comprobantes_pendientes_lista,
        mes_actual=mes_actual,
        resumen_notificaciones=resumen_notificaciones,
        sistema_resumen=sistema_resumen,
        puede_ver_jugadores=tiene_permiso("jugadores_ver"),
        puede_ver_finanzas=tiene_permiso("cuotas_ver", "cuotas_gestionar"),
        puede_ver_salud=tiene_permiso("salud_ver"),
    )


@app.route("/operacion")
def ver_operacion_diaria():
    if not puede_ver_operacion():
        flash("No tenes permiso para acceder a esa seccion.", "error")
        return redirect(url_for("index"))

    estado = request.args.get("estado", "pendiente")
    if estado not in {"pendiente", "hecha", "cancelada", "todas"}:
        estado = "pendiente"
    return render_template(
        "operacion.html",
        revision=obtener_revision_diaria(),
        tareas=listar_tareas_sig(estado=estado),
        estado=estado,
        puede_gestionar_tareas=puede_gestionar_tareas_sig(),
    )


@app.route("/tareas", methods=["POST"])
def crear_tarea_sig():
    if not puede_gestionar_tareas_sig():
        flash("No tenes permiso para crear tareas.", "error")
        return redirect(url_for("ver_operacion_diaria"))

    titulo = (request.form.get("titulo") or "").strip()
    if not titulo:
        flash("La tarea necesita un titulo.", "error")
        return redirect(url_for("ver_operacion_diaria"))

    jugador_id_raw = (request.form.get("jugador_id") or "").strip()
    jugador_id = int(jugador_id_raw) if jugador_id_raw.isdigit() else None
    modulo = (request.form.get("modulo") or "general").strip().lower()
    prioridad = (request.form.get("prioridad") or "media").strip().lower()
    if prioridad not in {"alta", "media", "baja"}:
        prioridad = "media"

    conn = get_connection()
    conn.execute("""
        INSERT INTO tareas_sig (
            titulo, descripcion, modulo, prioridad, responsable,
            fecha_vencimiento, jugador_id, creado_por, actualizado_por
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        titulo,
        (request.form.get("descripcion") or "").strip(),
        modulo,
        prioridad,
        (request.form.get("responsable") or "").strip(),
        (request.form.get("fecha_vencimiento") or "").strip(),
        jugador_id,
        session.get("username"),
        session.get("username"),
    ))
    conn.commit()
    conn.close()
    registrar_auditoria("crear", "tarea_sig", None, {"titulo": titulo, "modulo": modulo, "prioridad": prioridad})
    flash("Tarea creada.", "ok")
    return redirect(url_for("ver_operacion_diaria"))


@app.route("/tareas/<int:tarea_id>/estado", methods=["POST"])
def actualizar_estado_tarea_sig(tarea_id):
    if not puede_gestionar_tareas_sig():
        flash("No tenes permiso para actualizar tareas.", "error")
        return redirect(url_for("ver_operacion_diaria"))

    estado = (request.form.get("estado") or "hecha").strip().lower()
    if estado not in {"pendiente", "hecha", "cancelada"}:
        estado = "hecha"
    conn = get_connection()
    conn.execute("""
        UPDATE tareas_sig
        SET estado = %s,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (estado, session.get("username"), tarea_id))
    conn.commit()
    conn.close()
    registrar_auditoria("actualizar_estado", "tarea_sig", str(tarea_id), {"estado": estado})
    flash("Tarea actualizada.", "ok")
    return redirect(url_for("ver_operacion_diaria"))


@app.route("/finanzas/cobranzas")
def ver_panel_cobranzas():
    check = permiso_requerido("cuotas_ver", "cuotas_gestionar")
    if check:
        return check

    return render_template("cobranzas.html", panel=obtener_panel_cobranzas())

@app.route("/jugadores/<int:jugador_id>/ficha-medica")
def ver_ficha_medica(jugador_id):
    check = permiso_requerido("salud_ver")
    if check:
        return check

    conn = get_connection()

    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    ficha = conn.execute("""
        SELECT * FROM fichas_medicas
        WHERE jugador_id = %s
    """, (jugador_id,)).fetchone()

    conn.close()

    ficha_apto_efectivo = ficha_tiene_apto_efectivo(ficha)

    return render_template(
        "ficha_medica.html",
        jugador=jugador,
        ficha=ficha,
        ficha_apto_efectivo=ficha_apto_efectivo,
    )


@app.route("/jugadores/<int:jugador_id>/ficha-medica/documento")
def descargar_ficha_medica_documento(jugador_id):
    check = permiso_requerido("salud_ver")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    ficha = conn.execute("""
        SELECT * FROM fichas_medicas
        WHERE jugador_id = %s
    """, (jugador_id,)).fetchone()
    conn.close()

    if jugador is None:
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    if not ficha or not ficha["documento_drive_file_id"]:
        flash("La ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica no tiene documento adjunto.", "error")
        return redirect(url_for("ver_ficha_medica", jugador_id=jugador_id))

    try:
        archivo = descargar_drive_file(ficha["documento_drive_file_id"])
    except RuntimeError as error:
        flash(str(error), "error")
        return redirect(url_for("ver_ficha_medica", jugador_id=jugador_id))
    except Exception as error:
        app.logger.exception("No se pudo descargar ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica del jugador %s.", jugador_id)
        flash(mensaje_error_drive(error, carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER, accion="descargar la ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica"), "error")
        return redirect(url_for("ver_ficha_medica", jugador_id=jugador_id))

    registrar_auditoria(
        "descargar_ok",
        "ficha_medica_documento",
        str(jugador_id),
        {
            "archivo": ficha["documento_nombre"],
            "drive_file_id": ficha["documento_drive_file_id"],
        },
    )

    return send_file(
        archivo,
        mimetype=ficha["documento_mime_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=ficha["documento_nombre"] or f"ficha_medica_{jugador_id}",
    )


@app.route("/jugadores/<int:jugador_id>/ficha-medica/editar", methods=["GET", "POST"])
def editar_ficha_medica(jugador_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()

    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    ficha = conn.execute("""
        SELECT * FROM fichas_medicas
        WHERE jugador_id = %s
    """, (jugador_id,)).fetchone()

    if request.method == "POST":
        presentada = 1 if request.form.get("presentada") == "on" else 0
        fecha_vencimiento = request.form.get("fecha_vencimiento", "").strip()
        apto_fisico = 1 if request.form.get("apto_fisico") == "on" else 0
        contacto_emergencia = request.form.get("contacto_emergencia", "").strip()
        telefono_emergencia = request.form.get("telefono_emergencia", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        procesar_ocr = request.form.get("procesar_ocr") == "on"
        archivo_ficha = request.files.get("ficha_archivo")
        documento_info = None
        documento_fecha = None
        documento_usuario = None
        ocr_texto = None
        ocr_fecha = None
        ocr_usuario = None

        ficha_form = {
            "presentada": presentada,
            "fecha_vencimiento": fecha_vencimiento,
            "apto_fisico": apto_fisico,
            "contacto_emergencia": contacto_emergencia,
            "telefono_emergencia": telefono_emergencia,
            "observaciones": observaciones,
            "documento_drive_file_id": ficha["documento_drive_file_id"] if ficha else None,
            "documento_nombre": ficha["documento_nombre"] if ficha else None,
            "documento_fecha": ficha["documento_fecha"] if ficha else None,
            "documento_usuario": ficha["documento_usuario"] if ficha else None,
            "ocr_texto": ficha["ocr_texto"] if ficha else None,
            "ocr_fecha": ficha["ocr_fecha"] if ficha else None,
            "ocr_usuario": ficha["ocr_usuario"] if ficha else None,
        }

        try:
            ficha_validada = validar_ficha_medica_upload(archivo_ficha)
        except ValueError as error:
            conn.close()
            flash(str(error), "error")
            return render_template(
                "ficha_medica_form.html",
                jugador=jugador,
                ficha=ficha_form,
            )

        if ficha_validada:
            try:
                documento_info = subir_ficha_medica_a_drive(ficha_validada, jugador, ficha)
            except RuntimeError as error:
                conn.close()
                flash(str(error), "error")
                return render_template(
                    "ficha_medica_form.html",
                    jugador=jugador,
                    ficha=ficha_form,
                )
            except Exception as error:
                conn.close()
                app.logger.exception("No se pudo subir ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica del jugador %s.", jugador_id)
                flash(mensaje_error_drive(error, carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER, accion="guardar la ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica"), "error")
                return render_template(
                    "ficha_medica_form.html",
                    jugador=jugador,
                    ficha=ficha_form,
                )

            presentada = 1
            apto_fisico = 1
            documento_fecha = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
            documento_usuario = session.get("username")

            if procesar_ocr:
                try:
                    ocr_texto = normalizar_ocr_texto(
                        extraer_texto_ocr_drive(
                            ficha_validada,
                            jugador,
                            documento_info.get("folder_id"),
                        )
                    )
                    if ocr_texto:
                        datos_ocr = datos_ficha_desde_ocr(ocr_texto)
                        if not fecha_vencimiento and datos_ocr.get("fecha_vencimiento"):
                            fecha_vencimiento = datos_ocr["fecha_vencimiento"]
                        if datos_ocr.get("apto_fisico") is not None:
                            apto_fisico = datos_ocr["apto_fisico"]
                        if not contacto_emergencia and datos_ocr.get("contacto_emergencia"):
                            contacto_emergencia = datos_ocr["contacto_emergencia"]
                        if not telefono_emergencia and datos_ocr.get("telefono_emergencia"):
                            telefono_emergencia = datos_ocr["telefono_emergencia"]
                        ocr_fecha = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
                        ocr_usuario = session.get("username")
                    else:
                        flash("El documento se guard?, pero OCR no devolvi? texto para completar campos.", "warning")
                except Exception:
                    app.logger.exception("No se pudo procesar OCR de ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica del jugador %s.", jugador_id)
                    flash(
                        "El documento se guard?, pero no se pudo procesar OCR. Revis? los datos manualmente.",
                        "warning",
                    )

        if ficha:
            conn.execute("""
                UPDATE fichas_medicas
                SET presentada = %s, fecha_vencimiento = %s, apto_fisico = %s,
                    contacto_emergencia = %s, telefono_emergencia = %s, observaciones = %s,
                    documento_drive_file_id = COALESCE(%s, documento_drive_file_id),
                    documento_nombre = COALESCE(%s, documento_nombre),
                    documento_mime_type = COALESCE(%s, documento_mime_type),
                    documento_tamano = COALESCE(%s, documento_tamano),
                    documento_fecha = COALESCE(%s, documento_fecha),
                    documento_usuario = COALESCE(%s, documento_usuario),
                    documento_web_url = COALESCE(%s, documento_web_url),
                    ocr_texto = COALESCE(%s, ocr_texto),
                    ocr_fecha = COALESCE(%s, ocr_fecha),
                    ocr_usuario = COALESCE(%s, ocr_usuario)
                WHERE jugador_id = %s
            """, (
                presentada, fecha_vencimiento, apto_fisico,
                contacto_emergencia, telefono_emergencia, observaciones,
                documento_info["file_id"] if documento_info else None,
                documento_info["nombre"] if documento_info else None,
                documento_info["mime_type"] if documento_info else None,
                documento_info["tamano"] if documento_info else None,
                documento_fecha,
                documento_usuario,
                documento_info["web_url"] if documento_info else None,
                ocr_texto,
                ocr_fecha,
                ocr_usuario,
                jugador_id
            ))
        else:
            conn.execute("""
                INSERT INTO fichas_medicas (
                    jugador_id, presentada, fecha_vencimiento, apto_fisico,
                    contacto_emergencia, telefono_emergencia, observaciones,
                    documento_drive_file_id, documento_nombre, documento_mime_type,
                    documento_tamano, documento_fecha, documento_usuario, documento_web_url,
                    ocr_texto, ocr_fecha, ocr_usuario
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                jugador_id, presentada, fecha_vencimiento, apto_fisico,
                contacto_emergencia, telefono_emergencia, observaciones,
                documento_info["file_id"] if documento_info else None,
                documento_info["nombre"] if documento_info else None,
                documento_info["mime_type"] if documento_info else None,
                documento_info["tamano"] if documento_info else None,
                documento_fecha,
                documento_usuario,
                documento_info["web_url"] if documento_info else None,
                ocr_texto,
                ocr_fecha,
                ocr_usuario,
            ))

        conn.commit()
        conn.close()

        flash("Ficha médica guardada correctamente.", "ok")
        if documento_info:
            flash("Documento de ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica guardado en Google Drive.", "ok")
        if ocr_texto:
            flash("OCR procesado y campos detectados aplicados.", "ok")
        return redirect(url_for("ver_ficha_medica", jugador_id=jugador_id))

    conn.close()
    return render_template(
        "ficha_medica_form.html",
        jugador=jugador,
        ficha=ficha
    )


@app.route("/fichas-medicas/batch", methods=["GET", "POST"])
def cargar_fichas_medicas_batch():
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    if request.method == "POST":
        archivos = [
            archivo
            for archivo in request.files.getlist("fichas_archivos")
            if archivo and archivo.filename
        ]
        if not archivos:
            flash("Seleccion? al menos una ficha m?dica para cargar.", "error")
            conn = get_connection()
            batches_recientes = obtener_fichas_medicas_batch_recientes(conn)
            conn.close()
            return render_template("fichas_medicas_batch.html", batches_recientes=batches_recientes)

        conn = get_connection()
        jugadores = obtener_jugadores_selector(conn)
        batch_id = f"{ahora_sig().strftime('%Y%m%d%H%M%S')}_{secrets.token_urlsafe(6)}"
        cargadas = 0
        errores = 0

        for archivo in archivos:
            archivo_original = secure_filename(archivo.filename) or archivo.filename
            try:
                validado = validar_ficha_medica_upload(archivo)
                documento_info = subir_ficha_medica_batch_pendiente(validado, batch_id)
                ocr_texto = ""
                ocr_fecha = None
                ocr_usuario = None
                error = None

                datos_ocr = datos_ficha_desde_ocr(ocr_texto)
                jugador_sugerido, confianza, motivo = sugerir_jugador_ficha_ocr(archivo_original, jugadores)
                if confianza == "sin_coincidencia":
                    error = "OCR pendiente. Procesalo desde la revisi?n o asign? manualmente."

                conn.execute("""
                    INSERT INTO fichas_medicas_batch (
                        batch_id, estado, archivo_original, drive_file_id, drive_folder_id,
                        documento_nombre, documento_mime_type, documento_tamano, documento_web_url,
                        extension, ocr_texto, ocr_fecha, ocr_usuario, jugador_sugerido_id,
                        confianza, motivo, fecha_vencimiento_sugerida, apto_sugerido,
                        contacto_emergencia_sugerido, telefono_emergencia_sugerido,
                        error, creado_en, creado_por
                    )
                    VALUES (%s, 'pendiente', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    batch_id,
                    archivo_original,
                    documento_info["file_id"],
                    documento_info["folder_id"],
                    documento_info["nombre"],
                    documento_info["mime_type"],
                    documento_info["tamano"],
                    documento_info["web_url"],
                    validado["ext"],
                    ocr_texto or None,
                    ocr_fecha,
                    ocr_usuario,
                    jugador_sugerido["id"] if jugador_sugerido else None,
                    confianza,
                    motivo,
                    datos_ocr.get("fecha_vencimiento") or None,
                    datos_ocr.get("apto_fisico"),
                    datos_ocr.get("contacto_emergencia") or None,
                    datos_ocr.get("telefono_emergencia") or None,
                    error,
                    ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
                    session.get("username"),
                ))
                conn.commit()
                cargadas += 1
            except ValueError as error:
                errores += 1
                flash(f"{archivo_original}: {error}", "error")
            except RuntimeError as error:
                errores += 1
                flash(str(error), "error")
            except Exception as error:
                conn.rollback()
                errores += 1
                app.logger.exception("No se pudo cargar ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica batch %s.", archivo_original)
                flash(mensaje_error_drive(error, carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER, accion="guardar la ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica"), "error")

        conn.commit()

        if cargadas:
            conn.close()
            flash(f"Se cargaron {cargadas} ficha(s) para revisar.", "ok")
            flash("Para evitar timeouts, el OCR se procesa por archivo desde la pantalla de revisi?n.", "warning")
            if errores:
                flash(f"{errores} archivo(s) no pudieron cargarse.", "warning")
            return redirect(url_for("revisar_fichas_medicas_batch", batch_id=batch_id))

        flash("No se pudo cargar ninguna ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica.", "error")
        batches_recientes = obtener_fichas_medicas_batch_recientes(conn)
        conn.close()
        return render_template("fichas_medicas_batch.html", batches_recientes=batches_recientes)

    conn = get_connection()
    batches_recientes = obtener_fichas_medicas_batch_recientes(conn)
    conn.close()
    return render_template("fichas_medicas_batch.html", batches_recientes=batches_recientes)


@app.route("/fichas-medicas/batch/<batch_id>/revisar", methods=["GET", "POST"])
def revisar_fichas_medicas_batch(batch_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()

    if request.method == "POST":
        item_ids = request.form.getlist("item_ids")
        procesadas = 0
        omitidas = 0
        errores = 0

        for item_id in item_ids:
            if request.form.get(f"procesar_{item_id}") != "on":
                omitidas += 1
                continue

            jugador_id = request.form.get(f"jugador_id_{item_id}", "").strip()
            fecha_vencimiento = request.form.get(f"fecha_vencimiento_{item_id}", "").strip()
            apto_fisico = 1 if request.form.get(f"apto_fisico_{item_id}") == "on" else 0
            contacto_emergencia = request.form.get(f"contacto_emergencia_{item_id}", "").strip()
            telefono_emergencia = request.form.get(f"telefono_emergencia_{item_id}", "").strip()
            observaciones = request.form.get(f"observaciones_{item_id}", "").strip()

            if not jugador_id:
                errores += 1
                flash(f"El archivo #{item_id} no tiene jugador asignado.", "error")
                continue

            if fecha_vencimiento and not validar_fecha_movimiento(fecha_vencimiento):
                errores += 1
                flash(f"El archivo #{item_id} tiene una fecha de vencimiento inv?lida.", "error")
                continue

            item = conn.execute("""
                SELECT *
                FROM fichas_medicas_batch
                WHERE id = %s AND batch_id = %s AND estado = 'pendiente'
            """, (item_id, batch_id)).fetchone()

            jugador = conn.execute("""
                SELECT *
                FROM jugadores
                WHERE id = %s
            """, (jugador_id,)).fetchone()

            if not item or not jugador:
                errores += 1
                flash(f"No se encontr? el archivo pendiente #{item_id} o el jugador asignado.", "error")
                continue

            try:
                documento_info = mover_ficha_medica_batch_a_jugador(
                    item["drive_file_id"],
                    jugador,
                    item["extension"],
                    item["drive_folder_id"],
                )

                conn.execute("""
                    INSERT INTO fichas_medicas (
                        jugador_id, presentada, fecha_vencimiento, apto_fisico,
                        contacto_emergencia, telefono_emergencia, observaciones,
                        documento_drive_file_id, documento_nombre, documento_mime_type,
                        documento_tamano, documento_fecha, documento_usuario, documento_web_url,
                        ocr_texto, ocr_fecha, ocr_usuario
                    )
                    VALUES (%s, 1, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (jugador_id) DO UPDATE SET
                        presentada = 1,
                        fecha_vencimiento = CASE
                            WHEN NULLIF(EXCLUDED.fecha_vencimiento::text, '') IS NULL
                            THEN fichas_medicas.fecha_vencimiento
                            ELSE EXCLUDED.fecha_vencimiento
                        END,
                        apto_fisico = EXCLUDED.apto_fisico,
                        contacto_emergencia = COALESCE(NULLIF(EXCLUDED.contacto_emergencia, ''), fichas_medicas.contacto_emergencia),
                        telefono_emergencia = COALESCE(NULLIF(EXCLUDED.telefono_emergencia, ''), fichas_medicas.telefono_emergencia),
                        observaciones = COALESCE(NULLIF(EXCLUDED.observaciones, ''), fichas_medicas.observaciones),
                        documento_drive_file_id = EXCLUDED.documento_drive_file_id,
                        documento_nombre = EXCLUDED.documento_nombre,
                        documento_mime_type = EXCLUDED.documento_mime_type,
                        documento_tamano = EXCLUDED.documento_tamano,
                        documento_fecha = EXCLUDED.documento_fecha,
                        documento_usuario = EXCLUDED.documento_usuario,
                        documento_web_url = EXCLUDED.documento_web_url,
                        ocr_texto = COALESCE(EXCLUDED.ocr_texto, fichas_medicas.ocr_texto),
                        ocr_fecha = COALESCE(EXCLUDED.ocr_fecha, fichas_medicas.ocr_fecha),
                        ocr_usuario = COALESCE(EXCLUDED.ocr_usuario, fichas_medicas.ocr_usuario)
                """, (
                    jugador["id"],
                    fecha_vencimiento or None,
                    apto_fisico,
                    contacto_emergencia or None,
                    telefono_emergencia or None,
                    observaciones or None,
                    documento_info["file_id"],
                    documento_info["nombre"],
                    documento_info["mime_type"] or item["documento_mime_type"],
                    documento_info["tamano"] or item["documento_tamano"],
                    ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
                    session.get("username"),
                    documento_info["web_url"],
                    item["ocr_texto"],
                    item["ocr_fecha"],
                    item["ocr_usuario"],
                ))

                conn.execute("""
                    UPDATE fichas_medicas_batch
                    SET estado = 'procesado',
                        jugador_id = %s,
                        fecha_vencimiento_sugerida = %s,
                        apto_sugerido = %s,
                        contacto_emergencia_sugerido = %s,
                        telefono_emergencia_sugerido = %s,
                        procesado_en = %s,
                        procesado_por = %s,
                        error = NULL
                    WHERE id = %s
                """, (
                    jugador["id"],
                    fecha_vencimiento or None,
                    apto_fisico,
                    contacto_emergencia or None,
                    telefono_emergencia or None,
                    ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
                    session.get("username"),
                    item["id"],
                ))
                conn.commit()
                procesadas += 1
            except Exception as error:
                conn.rollback()
                errores += 1
                app.logger.exception("No se pudo confirmar ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica batch item %s.", item_id)
                flash(mensaje_error_drive(error, carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER, accion="asignar la ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica"), "error")

        if procesadas:
            flash(f"Se asignaron {procesadas} ficha(s) m?dica(s).", "ok")
        if omitidas:
            flash(f"{omitidas} archivo(s) quedaron pendientes.", "warning")
        if errores:
            flash(f"{errores} archivo(s) requieren revisi?n.", "error")
        return redirect(url_for("revisar_fichas_medicas_batch", batch_id=batch_id))

    items = conn.execute("""
        SELECT
            b.*,
            js.apellido AS sugerido_apellido,
            js.nombre AS sugerido_nombre,
            js.dni AS sugerido_dni,
            ja.apellido AS asignado_apellido,
            ja.nombre AS asignado_nombre
        FROM fichas_medicas_batch b
        LEFT JOIN jugadores js ON js.id = b.jugador_sugerido_id
        LEFT JOIN jugadores ja ON ja.id = b.jugador_id
        WHERE b.batch_id = %s
        ORDER BY b.id
    """, (batch_id,)).fetchall()

    jugadores = obtener_jugadores_selector(conn)
    conn.close()

    if not items:
        flash("No se encontr? la tanda de fichas m?dicas.", "error")
        return redirect(url_for("cargar_fichas_medicas_batch"))

    return render_template(
        "fichas_medicas_batch_revision.html",
        batch_id=batch_id,
        items=items,
        jugadores=jugadores,
    )


@app.route("/fichas-medicas/batch/items/<int:item_id>/ocr", methods=["POST"])
def procesar_ficha_medica_batch_ocr(item_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()
    item = conn.execute("""
        SELECT *
        FROM fichas_medicas_batch
        WHERE id = %s
    """, (item_id,)).fetchone()

    if not item:
        conn.close()
        flash("No se encontr? el archivo pendiente.", "error")
        return redirect(url_for("cargar_fichas_medicas_batch"))

    if item["estado"] != "pendiente":
        conn.close()
        flash("La ficha ya fue procesada.", "warning")
        return redirect(url_for("revisar_fichas_medicas_batch", batch_id=item["batch_id"]))

    try:
        ocr_texto = procesar_ocr_ficha_medica_batch_item(conn, item)
        conn.commit()
        if ocr_texto:
            flash("OCR procesado para el archivo seleccionado.", "ok")
        else:
            flash("OCR no devolvi? texto para este archivo. Pod?s asignarlo manualmente.", "warning")
    except Exception as error:
        conn.rollback()
        app.logger.exception("No se pudo procesar OCR batch item %s.", item_id)
        conn.execute("""
            UPDATE fichas_medicas_batch
            SET error = %s
            WHERE id = %s
        """, (f"OCR no disponible: {truncate_audit_value(error, 160)}", item_id))
        conn.commit()
        flash("No se pudo procesar OCR para este archivo. Pod?s cargar los datos manualmente.", "warning")
    finally:
        conn.close()

    return redirect(url_for("revisar_fichas_medicas_batch", batch_id=item["batch_id"]))


@app.route("/fichas-medicas/batch/items/<int:item_id>/documento")
def descargar_ficha_medica_batch_documento(item_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()
    item = conn.execute("""
        SELECT *
        FROM fichas_medicas_batch
        WHERE id = %s
    """, (item_id,)).fetchone()
    conn.close()

    if not item or not item["drive_file_id"]:
        flash("No se encontr? el documento pendiente.", "error")
        return redirect(url_for("cargar_fichas_medicas_batch"))

    try:
        archivo = descargar_drive_file(item["drive_file_id"])
    except Exception as error:
        app.logger.exception("No se pudo descargar ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica batch item %s.", item_id)
        flash(mensaje_error_drive(error, carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER, accion="descargar la ficha mÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â©dica"), "error")
        return redirect(url_for("revisar_fichas_medicas_batch", batch_id=item["batch_id"]))

    return send_file(
        archivo,
        mimetype=item["documento_mime_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=item["archivo_original"] or item["documento_nombre"] or f"ficha_medica_batch_{item_id}",
    )


@app.route("/documentos")
def ver_documentos_vencidos():
    check = permiso_requerido("documentos_ver", "salud_ver")
    if check:
        return check

    conn = get_connection()
    manuales = conn.execute("""
        SELECT
            d.*,
            j.apellido,
            j.nombre,
            j.categoria
        FROM documentos_jugadores d
        JOIN jugadores j ON j.id = d.jugador_id
        WHERE d.fecha_vencimiento IS NOT NULL
          AND NULLIF(d.fecha_vencimiento::text, '') IS NOT NULL
          AND d.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND d.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
        ORDER BY d.fecha_vencimiento ASC, j.apellido, j.nombre
    """).fetchall()

    fichas = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.categoria,
            f.fecha_vencimiento,
            f.presentada,
            CASE
                WHEN f.id IS NULL THEN 'faltante'
                WHEN NULLIF(f.fecha_vencimiento::text, '') IS NULL THEN 'sin_vencimiento'
                WHEN f.fecha_vencimiento::text !~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN 'sin_vencimiento'
                WHEN f.fecha_vencimiento::date < CURRENT_DATE THEN 'vencida'
                WHEN f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days' THEN 'por_vencer'
                ELSE 'vigente'
            END AS estado_documento
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.estado = 'Activo'
          AND (
              f.id IS NULL
              OR f.fecha_vencimiento IS NULL
              OR NULLIF(f.fecha_vencimiento::text, '') IS NULL
              OR (
                  f.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
              )
          )
        ORDER BY estado_documento, f.fecha_vencimiento ASC NULLS FIRST, j.apellido, j.nombre
    """).fetchall()
    conn.close()

    return render_template("documentos.html", manuales=manuales, fichas=fichas)


@app.route("/jugadores/<int:jugador_id>/documentos/nuevo", methods=["GET", "POST"])
def nuevo_documento_jugador(jugador_id):
    check = permiso_requerido("documentos_gestionar", "salud_gestionar")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    if request.method == "POST":
        data = {
            "tipo": request.form.get("tipo", "").strip(),
            "nombre": request.form.get("nombre", "").strip(),
            "fecha_presentacion": request.form.get("fecha_presentacion", "").strip(),
            "fecha_vencimiento": request.form.get("fecha_vencimiento", "").strip(),
            "url": request.form.get("url", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }

        if not data["tipo"]:
            conn.close()
            flash("El tipo de documento es obligatorio.", "error")
            return render_template("documento_form.html", jugador=jugador, documento=data)

        if data["fecha_presentacion"] and not validar_fecha_movimiento(data["fecha_presentacion"]):
            conn.close()
            flash("La fecha de presentacion no es valida.", "error")
            return render_template("documento_form.html", jugador=jugador, documento=data)

        if data["fecha_vencimiento"] and not validar_fecha_movimiento(data["fecha_vencimiento"]):
            conn.close()
            flash("La fecha de vencimiento no es valida.", "error")
            return render_template("documento_form.html", jugador=jugador, documento=data)

        conn.execute("""
            INSERT INTO documentos_jugadores (
                jugador_id, tipo, nombre, fecha_presentacion, fecha_vencimiento,
                url, observaciones, creado_por
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            jugador_id,
            data["tipo"],
            data["nombre"],
            data["fecha_presentacion"] or None,
            data["fecha_vencimiento"] or None,
            data["url"],
            data["observaciones"],
            session.get("username"),
        ))
        conn.commit()
        conn.close()

        flash("Documento cargado correctamente.", "ok")
        return redirect(url_for("detalle_jugador", jugador_id=jugador_id))

    conn.close()
    return render_template("documento_form.html", jugador=jugador, documento={})


@app.route("/documentos/<int:documento_id>/eliminar", methods=["POST"])
def eliminar_documento_jugador(documento_id):
    check = permiso_requerido("documentos_gestionar", "salud_gestionar")
    if check:
        return check

    conn = get_connection()
    documento = conn.execute(
        "SELECT * FROM documentos_jugadores WHERE id = %s",
        (documento_id,),
    ).fetchone()
    if documento is None:
        conn.close()
        flash("Documento no encontrado.", "error")
        return redirect(url_for("ver_documentos_vencidos"))

    conn.execute("DELETE FROM documentos_jugadores WHERE id = %s", (documento_id,))
    conn.commit()
    conn.close()

    flash("Documento eliminado.", "ok")
    return redirect(url_for("detalle_jugador", jugador_id=documento["jugador_id"]))


def categorias_secretaria_disponibles(conn):
    existentes = conn.execute("""
        SELECT DISTINCT categoria
        FROM secretaria_documentos
        WHERE NULLIF(TRIM(categoria), '') IS NOT NULL
        ORDER BY categoria
    """).fetchall()
    categorias = {categoria for categoria in SECRETARIA_CATEGORIAS}
    categorias.update((fila["categoria"] or "").strip() for fila in existentes if (fila["categoria"] or "").strip())
    return sorted(categorias, key=lambda item: item.lower())


def render_documentos_secretaria(categoria_filtro="", vencimiento_filtro="", form_data=None):
    conn = get_connection()
    categorias = categorias_secretaria_disponibles(conn)

    condiciones = []
    params = []
    if categoria_filtro:
        condiciones.append("categoria = %s")
        params.append(categoria_filtro)
    if vencimiento_filtro == "vencidos":
        condiciones.extend([
            "fecha_vencimiento IS NOT NULL",
            "NULLIF(fecha_vencimiento::text, '') IS NOT NULL",
            "fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'",
            "fecha_vencimiento::date < CURRENT_DATE",
        ])
    elif vencimiento_filtro == "por_vencer":
        condiciones.extend([
            "fecha_vencimiento IS NOT NULL",
            "NULLIF(fecha_vencimiento::text, '') IS NOT NULL",
            "fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'",
            "fecha_vencimiento::date >= CURRENT_DATE",
            "fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'",
        ])
    elif vencimiento_filtro == "sin_vencimiento":
        condiciones.append("(fecha_vencimiento IS NULL OR NULLIF(fecha_vencimiento::text, '') IS NULL)")

    where_sql = f"WHERE {' AND '.join(condiciones)}" if condiciones else ""
    documentos = conn.execute(f"""
        SELECT
            *,
            CASE
                WHEN NULLIF(fecha_vencimiento::text, '') IS NULL THEN 'sin_vencimiento'
                WHEN fecha_vencimiento::text !~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}$' THEN 'sin_vencimiento'
                WHEN fecha_vencimiento::date < CURRENT_DATE THEN 'vencido'
                WHEN fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days' THEN 'por_vencer'
                ELSE 'vigente'
            END AS estado_vencimiento
        FROM secretaria_documentos
        {where_sql}
        ORDER BY
            CASE
                WHEN NULLIF(fecha_vencimiento::text, '') IS NOT NULL
                 AND fecha_vencimiento::text ~ '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}$'
                 AND fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
                THEN 0
                ELSE 1
            END,
            fecha_vencimiento ASC NULLS LAST,
            COALESCE(fecha_documento, creado_en::text) DESC,
            id DESC
    """, params).fetchall()

    resumen = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(DISTINCT categoria) AS categorias,
            COUNT(*) FILTER (
                WHERE fecha_vencimiento IS NOT NULL
                  AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
                  AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND fecha_vencimiento::date < CURRENT_DATE
            ) AS vencidos,
            COUNT(*) FILTER (
                WHERE fecha_vencimiento IS NOT NULL
                  AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
                  AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                  AND fecha_vencimiento::date >= CURRENT_DATE
                  AND fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
            ) AS por_vencer
        FROM secretaria_documentos
    """).fetchone()
    conn.close()

    documentos_preview = [dict(item, preview_tipo=obtener_preview_tipo(item.get("archivo_mime_type"))) for item in documentos]
    return render_template(
        "secretaria_documentos.html",
        documentos=documentos_preview,
        categorias=categorias,
        categoria_filtro=categoria_filtro,
        vencimiento_filtro=vencimiento_filtro,
        form_data=form_data or {},
        resumen=resumen or {"total": 0, "categorias": 0, "vencidos": 0, "por_vencer": 0},
        secretaria_accept=",".join(SECRETARIA_EXTENSIONS.keys()),
    )


@app.route("/secretaria/documentos")
def listar_documentos_secretaria():
    check = permiso_requerido("secretaria_ver", "secretaria_gestionar")
    if check:
        return check

    categoria_filtro = request.args.get("categoria", "").strip()
    vencimiento_filtro = request.args.get("vencimiento", "").strip()
    if vencimiento_filtro not in {"", "vencidos", "por_vencer", "sin_vencimiento"}:
        vencimiento_filtro = ""
    return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro)


@app.route("/secretaria/documentos/nuevo", methods=["POST"])
def nuevo_documento_secretaria():
    check = permiso_requerido("secretaria_gestionar")
    if check:
        return check

    data = {
        "categoria": request.form.get("categoria", "").strip(),
        "titulo": request.form.get("titulo", "").strip(),
        "descripcion": request.form.get("descripcion", "").strip(),
        "fecha_documento": request.form.get("fecha_documento", "").strip(),
        "fecha_vencimiento": request.form.get("fecha_vencimiento", "").strip(),
    }
    categoria_filtro = request.form.get("categoria_filtro", "").strip()
    vencimiento_filtro = request.form.get("vencimiento_filtro", "").strip()
    archivo = request.files.get("archivo")

    if not data["categoria"]:
        flash("La categoría es obligatoria.", "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)
    if not data["titulo"]:
        flash("El título del documento es obligatorio.", "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)
    if data["fecha_documento"] and not validar_fecha_movimiento(data["fecha_documento"]):
        flash("La fecha del documento no es válida.", "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)
    if data["fecha_vencimiento"] and not validar_fecha_movimiento(data["fecha_vencimiento"]):
        flash("La fecha de vencimiento no es válida.", "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)

    try:
        validado = validar_documento_secretaria_upload(archivo)
    except ValueError as error:
        flash(str(error), "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)

    if not validado:
        flash("Tenés que adjuntar un archivo.", "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)

    try:
        documento_drive = subir_documento_secretaria_a_drive(
            validado,
            data["categoria"],
            data["titulo"],
            fecha_base=data["fecha_documento"] or None,
        )
    except Exception as error:
        app.logger.exception("No se pudo guardar documento de secretaria en Drive.")
        flash(mensaje_error_drive(error, carpeta="Secretaria", accion="guardar el documento"), "error")
        return render_documentos_secretaria(categoria_filtro=categoria_filtro, vencimiento_filtro=vencimiento_filtro, form_data=data)

    conn = get_connection()
    documento = conn.execute("""
        INSERT INTO secretaria_documentos (
            categoria, titulo, descripcion, fecha_documento, fecha_vencimiento,
            drive_file_id, drive_folder_id, archivo_nombre, archivo_mime_type,
            archivo_tamano, archivo_web_url, creado_por
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        data["categoria"],
        data["titulo"],
        data["descripcion"] or None,
        data["fecha_documento"] or None,
        data["fecha_vencimiento"] or None,
        documento_drive["file_id"],
        documento_drive["folder_id"],
        documento_drive["nombre"],
        documento_drive["mime_type"],
        documento_drive["tamano"],
        documento_drive["web_url"],
        session.get("username"),
    )).fetchone()
    conn.commit()
    conn.close()

    registrar_auditoria("subir", "documento_secretaria", str(documento["id"]), {
        "categoria": data["categoria"],
        "titulo": data["titulo"],
        "fecha_vencimiento": data["fecha_vencimiento"] or None,
    })
    flash("Documento de secretaría guardado.", "ok")
    return redirect(url_for("listar_documentos_secretaria", categoria=data["categoria"]))


@app.route("/secretaria/documentos/<int:documento_id>/archivo")
def ver_documento_secretaria(documento_id):
    check = permiso_requerido("secretaria_ver", "secretaria_gestionar")
    if check:
        return check

    conn = get_connection()
    documento = conn.execute(
        "SELECT * FROM secretaria_documentos WHERE id = %s",
        (documento_id,),
    ).fetchone()
    conn.close()
    if not documento:
        flash("Documento de secretaría no encontrado.", "error")
        return redirect(url_for("listar_documentos_secretaria"))

    try:
        archivo = descargar_drive_file(documento["drive_file_id"])
    except Exception as error:
        flash(mensaje_error_drive(error, carpeta="Secretaria", accion="descargar el documento"), "error")
        return redirect(url_for("listar_documentos_secretaria", categoria=documento.get("categoria") or ""))

    registrar_auditoria("ver", "documento_secretaria", str(documento_id), {
        "archivo": documento.get("archivo_nombre"),
        "drive_file_id": documento.get("drive_file_id"),
    })
    return send_file(
        archivo,
        mimetype=documento.get("archivo_mime_type") or "application/octet-stream",
        as_attachment=False,
        download_name=documento.get("archivo_nombre") or f"secretaria_{documento_id}",
    )


@app.route("/secretaria/documentos/<int:documento_id>/eliminar", methods=["POST"])
def eliminar_documento_secretaria(documento_id):
    check = permiso_requerido("secretaria_gestionar")
    if check:
        return check

    conn = get_connection()
    documento = conn.execute(
        "SELECT * FROM secretaria_documentos WHERE id = %s",
        (documento_id,),
    ).fetchone()
    if not documento:
        conn.close()
        flash("Documento de secretaría no encontrado.", "error")
        return redirect(url_for("listar_documentos_secretaria"))

    conn.execute("DELETE FROM secretaria_documentos WHERE id = %s", (documento_id,))
    conn.commit()
    conn.close()

    if documento.get("drive_file_id"):
        try:
            eliminar_drive_file(documento["drive_file_id"])
        except Exception:
            app.logger.warning("No se pudo eliminar el archivo de Drive del documento de secretaria %s.", documento_id)

    registrar_auditoria("eliminar", "documento_secretaria", str(documento_id), {
        "categoria": documento.get("categoria"),
        "titulo": documento.get("titulo"),
    })
    flash("Documento de secretaría eliminado.", "ok")
    return redirect(url_for("listar_documentos_secretaria", categoria=documento.get("categoria") or ""))


@app.route("/jugadores")
def listar_jugadores():
    check = permiso_requerido("jugadores_ver")
    if check:
        return check

    busqueda = request.args.get("q", "").strip()
    tipo_filtro = request.args.get("tipo", "").strip()
    estado_filtro = request.args.get("estado", "").strip()
    categoria_filtro = request.args.get("categoria", "").strip()
    orden = request.args.get("orden", "apellido")

    if tipo_filtro not in TIPOS_MIEMBRO:
        tipo_filtro = ""
    if estado_filtro not in ESTADOS_JUGADOR:
        estado_filtro = ""

    ordenes_validos = {
        "apellido": "apellido ASC, nombre ASC",
        "fecha_ingreso_asc": "fecha_ingreso ASC NULLS LAST, apellido ASC, nombre ASC",
        "fecha_ingreso_desc": "fecha_ingreso DESC NULLS LAST, apellido ASC, nombre ASC",
    }
    if orden not in ordenes_validos:
        orden = "apellido"

    conn = get_connection()

    filtros = []
    parametros = []
    if tipo_filtro:
        filtros.append("COALESCE(tipo_miembro, 'Jugador') = %s")
        parametros.append(tipo_filtro)
    if estado_filtro:
        filtros.append("COALESCE(estado, 'Activo') = %s")
        parametros.append(estado_filtro)
    if categoria_filtro:
        filtros.append("COALESCE(categoria, '') ILIKE %s")
        parametros.append(f"%{categoria_filtro}%")

    if busqueda:
        terminos = [termino for termino in re.split(r"\s+", busqueda) if termino]
        for termino in terminos:
            like = f"%{termino}%"
            filtros.append("""
                (
                    nombre ILIKE %s
                    OR apellido ILIKE %s
                    OR dni ILIKE %s
                    OR categoria ILIKE %s
                    OR telefono ILIKE %s
                    OR email ILIKE %s
                    OR estado ILIKE %s
                    OR tipo_miembro ILIKE %s
                    OR concat_ws(' ', nombre, apellido) ILIKE %s
                    OR concat_ws(' ', apellido, nombre) ILIKE %s
                )
            """)
            parametros.extend([like] * 10)

    where_sql = "WHERE " + " AND ".join(filtros) if filtros else ""
    jugadores = conn.execute(
        f"""
            SELECT * FROM jugadores
            {where_sql}
            ORDER BY {ordenes_validos[orden]}
        """,
        parametros,
    ).fetchall()

    conn.close()
    return render_template(
        "jugadores.html",
        jugadores=jugadores,
        busqueda=busqueda,
        tipo_filtro=tipo_filtro,
        estado_filtro=estado_filtro,
        categoria_filtro=categoria_filtro,
        orden=orden,
        tipos_miembro=sorted(TIPOS_MIEMBRO),
        estados_jugador=ESTADOS_JUGADOR,
    )


def obtener_madrinas_disponibles(conn):
    return conn.execute("""
        SELECT id, nombre, apellido, categoria
        FROM jugadores
        WHERE estado = 'Activo'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY apellido, nombre
    """).fetchall()


def aspirante_desde_formulario(aspirante=None):
    objetivo_raw = request.form.get(
        "entrenamientos_objetivo",
        str(aspirante["entrenamientos_objetivo"] if aspirante else ASPIRANTE_ENTRENAMIENTOS_OBJETIVO),
    )
    try:
        entrenamientos_objetivo = max(0, int(objetivo_raw))
    except (TypeError, ValueError):
        entrenamientos_objetivo = ASPIRANTE_ENTRENAMIENTOS_OBJETIVO

    madrina_raw = request.form.get("madrina_jugador_id", "").strip()
    try:
        madrina_jugador_id = int(madrina_raw) if madrina_raw else None
    except ValueError:
        madrina_jugador_id = None

    estado = request.form.get("estado", "").strip() or (
        aspirante["estado"] if aspirante else "Aspirante"
    )
    if estado not in ASPIRANTE_ESTADOS:
        estado = "Aspirante"

    return {
        "nombre": request.form.get("nombre", "").strip(),
        "apellido": request.form.get("apellido", "").strip(),
        "dni": request.form.get("dni", "").strip(),
        "fecha_nacimiento": request.form.get("fecha_nacimiento", "").strip(),
        "telefono": request.form.get("telefono", "").strip(),
        "email": request.form.get("email", "").strip(),
        "categoria": request.form.get("categoria", "").strip(),
        "fecha_postulacion": request.form.get("fecha_postulacion", "").strip(),
        "estado": estado,
        "madrina_jugador_id": madrina_jugador_id,
        "entrenamientos_objetivo": entrenamientos_objetivo,
        "observaciones": request.form.get("observaciones", "").strip(),
    }


def aspirante_con_progreso(aspirante):
    fila = dict(aspirante)
    objetivo = fila.get("entrenamientos_objetivo") or ASPIRANTE_ENTRENAMIENTOS_OBJETIVO
    presentes = fila.get("entrenamientos_realizados") or 0
    fila["progreso"] = min(100, round((presentes / objetivo) * 100)) if objetivo else 100
    fila["listo_para_ingresar"] = (
        fila.get("estado") == "Aspirante"
        and presentes >= objetivo
    )
    return fila


def buscar_aspirante(conn, aspirante_id):
    return conn.execute("""
        SELECT
            a.*,
            m.nombre AS madrina_nombre,
            m.apellido AS madrina_apellido,
            m.categoria AS madrina_categoria,
            COALESCE(stats.entrenamientos_realizados, 0) AS entrenamientos_realizados
        FROM aspirantes a
        LEFT JOIN jugadores m ON m.id = a.madrina_jugador_id
        LEFT JOIN (
            SELECT
                aa.aspirante_id,
                SUM(CASE WHEN aa.presente = 1 AND LOWER(e.tipo) = 'entrenamiento' THEN 1 ELSE 0 END) AS entrenamientos_realizados
            FROM aspirante_asistencias aa
            JOIN eventos_asistencia e ON e.id = aa.evento_id
            GROUP BY aa.aspirante_id
        ) stats ON stats.aspirante_id = a.id
        WHERE a.id = %s
    """, (aspirante_id,)).fetchone()


@app.route("/ahijadxs")
def listar_aspirantes():
    check = permiso_requerido("aspirantes_ver")
    if check:
        return check

    busqueda = request.args.get("q", "").strip()
    estado = request.args.get("estado", "Aspirante").strip()
    if estado not in ASPIRANTE_ESTADOS and estado != "todos":
        estado = "Aspirante"

    condiciones = []
    parametros = []
    if estado != "todos":
        condiciones.append("a.estado = %s")
        parametros.append(estado)

    if busqueda:
        terminos = [termino for termino in re.split(r"\s+", busqueda) if termino]
        for termino in terminos:
            like = f"%{termino}%"
            condiciones.append("""
                (
                    a.nombre ILIKE %s
                    OR a.apellido ILIKE %s
                    OR a.dni ILIKE %s
                    OR a.categoria ILIKE %s
                    OR concat_ws(' ', a.nombre, a.apellido) ILIKE %s
                    OR concat_ws(' ', a.apellido, a.nombre) ILIKE %s
                    OR m.nombre ILIKE %s
                    OR m.apellido ILIKE %s
                )
            """)
            parametros.extend([like] * 8)

    where_sql = "WHERE " + " AND ".join(condiciones) if condiciones else ""

    conn = get_connection()
    aspirantes = conn.execute(f"""
        SELECT
            a.*,
            m.nombre AS madrina_nombre,
            m.apellido AS madrina_apellido,
            COALESCE(stats.entrenamientos_realizados, 0) AS entrenamientos_realizados
        FROM aspirantes a
        LEFT JOIN jugadores m ON m.id = a.madrina_jugador_id
        LEFT JOIN (
            SELECT
                aa.aspirante_id,
                SUM(CASE WHEN aa.presente = 1 AND LOWER(e.tipo) = 'entrenamiento' THEN 1 ELSE 0 END) AS entrenamientos_realizados
            FROM aspirante_asistencias aa
            JOIN eventos_asistencia e ON e.id = aa.evento_id
            GROUP BY aa.aspirante_id
        ) stats ON stats.aspirante_id = a.id
        {where_sql}
        ORDER BY a.estado, a.apellido, a.nombre
    """, parametros).fetchall()
    conn.close()

    aspirantes = [aspirante_con_progreso(aspirante) for aspirante in aspirantes]

    return render_template(
        "aspirantes.html",
        aspirantes=aspirantes,
        busqueda=busqueda,
        estado=estado,
        estados=sorted(ASPIRANTE_ESTADOS),
    )


@app.route("/ahijadxs/nuevo", methods=["GET", "POST"])
def nuevo_aspirante():
    check = permiso_requerido("aspirantes_gestionar")
    if check:
        return check

    conn = get_connection()
    madrinas = obtener_madrinas_disponibles(conn)

    if request.method == "POST":
        data = aspirante_desde_formulario()
        if not data["fecha_postulacion"]:
            data["fecha_postulacion"] = ahora_sig().strftime("%Y-%m-%d")

        if not data["nombre"] or not data["apellido"]:
            conn.close()
            flash("Nombre y apellido son obligatorios.", "error")
            return render_template("aspirante_form.html", aspirante=data, madrinas=madrinas, modo="nuevo")

        if data["dni"]:
            existente_jugador = conn.execute(
                "SELECT id FROM jugadores WHERE dni = %s",
                (data["dni"],),
            ).fetchone()
            existente_aspirante = conn.execute(
                "SELECT id FROM aspirantes WHERE dni = %s AND estado <> 'Baja'",
                (data["dni"],),
            ).fetchone()
            if existente_jugador or existente_aspirante:
                conn.close()
                flash("Ya existe un jugador o ahijadx activo con ese DNI.", "error")
                return render_template("aspirante_form.html", aspirante=data, madrinas=madrinas, modo="nuevo")

        conn.execute("""
            INSERT INTO aspirantes (
                nombre, apellido, dni, fecha_nacimiento, telefono, email, categoria,
                fecha_postulacion, estado, madrina_jugador_id, entrenamientos_objetivo,
                observaciones
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data["nombre"], data["apellido"], data["dni"], data["fecha_nacimiento"],
            data["telefono"], data["email"], data["categoria"], data["fecha_postulacion"],
            data["estado"], data["madrina_jugador_id"], data["entrenamientos_objetivo"],
            data["observaciones"],
        ))
        conn.commit()
        conn.close()

        flash("Ahijadx cargado correctamente.", "ok")
        return redirect(url_for("listar_aspirantes"))

    conn.close()
    return render_template(
        "aspirante_form.html",
        aspirante={
            "fecha_postulacion": ahora_sig().strftime("%Y-%m-%d"),
            "estado": "Aspirante",
            "entrenamientos_objetivo": ASPIRANTE_ENTRENAMIENTOS_OBJETIVO,
        },
        madrinas=madrinas,
        modo="nuevo",
    )


@app.route("/ahijadxs/<int:aspirante_id>")
def detalle_aspirante(aspirante_id):
    check = permiso_requerido("aspirantes_ver")
    if check:
        return check

    conn = get_connection()
    aspirante = buscar_aspirante(conn, aspirante_id)
    if aspirante is None:
        conn.close()
        flash("Ahijadx no encontrado.", "error")
        return redirect(url_for("listar_aspirantes"))

    asistencias = conn.execute("""
        SELECT aa.*, e.fecha, e.tipo, e.descripcion
        FROM aspirante_asistencias aa
        JOIN eventos_asistencia e ON e.id = aa.evento_id
        WHERE aa.aspirante_id = %s
        ORDER BY e.fecha DESC, e.id DESC
    """, (aspirante_id,)).fetchall()
    conn.close()

    return render_template(
        "aspirante_detalle.html",
        aspirante=aspirante_con_progreso(aspirante),
        asistencias=asistencias,
    )


@app.route("/ahijadxs/<int:aspirante_id>/editar", methods=["GET", "POST"])
def editar_aspirante(aspirante_id):
    check = permiso_requerido("aspirantes_gestionar")
    if check:
        return check

    conn = get_connection()
    aspirante = conn.execute("SELECT * FROM aspirantes WHERE id = %s", (aspirante_id,)).fetchone()
    if aspirante is None:
        conn.close()
        flash("Ahijadx no encontrado.", "error")
        return redirect(url_for("listar_aspirantes"))

    madrinas = obtener_madrinas_disponibles(conn)

    if request.method == "POST":
        data = aspirante_desde_formulario(aspirante)
        if not data["nombre"] or not data["apellido"]:
            conn.close()
            flash("Nombre y apellido son obligatorios.", "error")
            data["id"] = aspirante_id
            return render_template("aspirante_form.html", aspirante=data, madrinas=madrinas, modo="editar")

        if data["dni"]:
            jugador_actual_id = aspirante["jugador_id"] or 0
            existente_jugador = conn.execute(
                "SELECT id FROM jugadores WHERE dni = %s AND id <> %s",
                (data["dni"], jugador_actual_id),
            ).fetchone()
            existente_aspirante = conn.execute(
                "SELECT id FROM aspirantes WHERE dni = %s AND id <> %s AND estado <> 'Baja'",
                (data["dni"], aspirante_id),
            ).fetchone()
            if existente_jugador or existente_aspirante:
                conn.close()
                flash("Ya existe otro jugador o ahijadx activo con ese DNI.", "error")
                data["id"] = aspirante_id
                return render_template("aspirante_form.html", aspirante=data, madrinas=madrinas, modo="editar")

        conn.execute("""
            UPDATE aspirantes
            SET nombre = %s,
                apellido = %s,
                dni = %s,
                fecha_nacimiento = %s,
                telefono = %s,
                email = %s,
                categoria = %s,
                fecha_postulacion = %s,
                estado = %s,
                madrina_jugador_id = %s,
                entrenamientos_objetivo = %s,
                observaciones = %s
            WHERE id = %s
        """, (
            data["nombre"], data["apellido"], data["dni"], data["fecha_nacimiento"],
            data["telefono"], data["email"], data["categoria"], data["fecha_postulacion"],
            data["estado"], data["madrina_jugador_id"], data["entrenamientos_objetivo"],
            data["observaciones"], aspirante_id,
        ))
        conn.commit()
        conn.close()

        flash("Ahijadx actualizado correctamente.", "ok")
        return redirect(url_for("detalle_aspirante", aspirante_id=aspirante_id))

    conn.close()
    return render_template("aspirante_form.html", aspirante=aspirante, madrinas=madrinas, modo="editar")


@app.route("/ahijadxs/<int:aspirante_id>/convertir", methods=["POST"])
def convertir_aspirante(aspirante_id):
    check = permiso_requerido("aspirantes_gestionar")
    if check:
        return check

    conn = get_connection()
    aspirante = buscar_aspirante(conn, aspirante_id)
    if aspirante is None:
        conn.close()
        flash("Ahijadx no encontrado.", "error")
        return redirect(url_for("listar_aspirantes"))

    aspirante = aspirante_con_progreso(aspirante)
    if aspirante["estado"] != "Aspirante":
        conn.close()
        flash("Solo se pueden ingresar ahijadxs en seguimiento.", "error")
        return redirect(url_for("detalle_aspirante", aspirante_id=aspirante_id))

    if not aspirante["listo_para_ingresar"]:
        conn.close()
        flash("El ahijadx todavia no alcanzo la cantidad requerida de entrenamientos.", "error")
        return redirect(url_for("detalle_aspirante", aspirante_id=aspirante_id))

    if aspirante["dni"]:
        existente = conn.execute(
            "SELECT id FROM jugadores WHERE dni = %s",
            (aspirante["dni"],),
        ).fetchone()
        if existente:
            conn.close()
            flash("Ya existe un jugador con ese DNI.", "error")
            return redirect(url_for("detalle_aspirante", aspirante_id=aspirante_id))

    fecha_ingreso = request.form.get("fecha_ingreso", "").strip() or ahora_sig().strftime("%Y-%m-%d")
    jugador_creado = conn.execute("""
        INSERT INTO jugadores (
            nombre, apellido, dni, fecha_nacimiento, telefono, email, categoria,
            fecha_ingreso, estado, observaciones
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Activo', %s)
        RETURNING id
    """, (
        aspirante["nombre"], aspirante["apellido"], aspirante["dni"], aspirante["fecha_nacimiento"],
        aspirante["telefono"], aspirante["email"], aspirante["categoria"], fecha_ingreso,
        aspirante["observaciones"],
    )).fetchone()
    jugador_id = jugador_creado["id"]

    conn.execute("""
        UPDATE aspirantes
        SET estado = 'Ingresado',
            fecha_ingreso_club = %s,
            jugador_id = %s
        WHERE id = %s
    """, (fecha_ingreso, jugador_id, aspirante_id))
    conn.commit()
    conn.close()

    flash("Ahijadx ingresado como jugador activo.", "ok")
    return redirect(url_for("detalle_jugador", jugador_id=jugador_id))


@app.route("/ahijadxs/<int:aspirante_id>/eliminar", methods=["POST"])
def eliminar_aspirante(aspirante_id):
    check = permiso_requerido("aspirantes_gestionar")
    if check:
        return check

    conn = get_connection()
    conn.execute("""
        UPDATE aspirantes
        SET estado = 'Baja'
        WHERE id = %s
    """, (aspirante_id,))
    conn.commit()
    conn.close()

    flash("Ahijadx dado de baja.", "ok")
    return redirect(url_for("listar_aspirantes"))


@app.route("/jugadores/nuevo", methods=["GET", "POST"])
def nuevo_jugador():
    check = permiso_requerido("jugadores_gestionar")
    if check:
        return check

    if request.method == "POST":
        data = {
            "nombre": request.form.get("nombre", "").strip(),
            "apellido": request.form.get("apellido", "").strip(),
            "dni": request.form.get("dni", "").strip(),
            "fecha_nacimiento": request.form.get("fecha_nacimiento", "").strip(),
            "telefono": request.form.get("telefono", "").strip(),
            "email": request.form.get("email", "").strip(),
            "categoria": request.form.get("categoria", "").strip(),
            "fecha_ingreso": request.form.get("fecha_ingreso", "").strip(),
            "estado": request.form.get("estado", "").strip() or "Activo",
            "contacto_tutor": request.form.get("contacto_tutor", "").strip(),
            "parentesco_tutor": request.form.get("parentesco_tutor", "").strip(),
            "telefono_tutor": request.form.get("telefono_tutor", "").strip(),
            "email_tutor": request.form.get("email_tutor", "").strip(),
            "direccion": request.form.get("direccion", "").strip(),
            "obra_social": request.form.get("obra_social", "").strip(),
            "numero_afiliado_obra_social": request.form.get("numero_afiliado_obra_social", "").strip(),
            "numero_socio": request.form.get("numero_socio", "").strip(),
            "tipo_miembro": normalizar_tipo_miembro(request.form.get("tipo_miembro")),
            "cobra_cuota": 1 if request.form.get("cobra_cuota", "on") == "on" else 0,
            "documentos": request.form.get("documentos", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        beca_data, beca_error = datos_beca_form()
        data.update(beca_data or {
            "beca_activa": 1 if request.form.get("beca_activa") == "on" else 0,
            "beca_porcentaje": request.form.get("beca_porcentaje", "").strip(),
            "beca_desde": request.form.get("beca_desde", "").strip(),
            "beca_hasta": request.form.get("beca_hasta", "").strip(),
            "beca_motivo": request.form.get("beca_motivo", "").strip(),
        })

        if not data["nombre"] or not data["apellido"]:
            flash("Nombre y apellido son obligatorios.", "error")
            return render_template("jugador_form.html", jugador=data, modo="nuevo")
        if beca_error:
            flash(beca_error, "error")
            return render_template("jugador_form.html", jugador=data, modo="nuevo")

        conn = get_connection()
        if not data["numero_socio"]:
            data["numero_socio"] = siguiente_numero_socio(conn)
        else:
            data["numero_socio"] = formatear_numero_socio(re.sub(r"\D+", "", data["numero_socio"]) or data["numero_socio"])
        creado = conn.execute("""
            INSERT INTO jugadores
            (
                nombre, apellido, dni, fecha_nacimiento, telefono, email, categoria,
                fecha_ingreso, estado, contacto_tutor, parentesco_tutor, telefono_tutor,
                email_tutor, direccion, obra_social, numero_afiliado_obra_social, numero_socio, documentos, observaciones,
                beca_activa, beca_porcentaje, beca_desde, beca_hasta, beca_motivo,
                tipo_miembro, cobra_cuota
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data["nombre"], data["apellido"], data["dni"], data["fecha_nacimiento"],
            data["telefono"], data["email"], data["categoria"], data["fecha_ingreso"],
            data["estado"], data["contacto_tutor"], data["parentesco_tutor"],
            data["telefono_tutor"], data["email_tutor"], data["direccion"],
            data["obra_social"], data["numero_afiliado_obra_social"], data["numero_socio"], data["documentos"], data["observaciones"],
            data["beca_activa"], data["beca_porcentaje"], data["beca_desde"],
            data["beca_hasta"], data["beca_motivo"], data["tipo_miembro"], data["cobra_cuota"]
        )).fetchone()

        if data["beca_activa"]:
            registrar_historial_beca(
                conn,
                creado["id"],
                data,
                "alta",
                {"origen": "alta_jugador"},
            )
        conn.commit()
        conn.close()

        flash("Registro cargado correctamente.", "ok")
        return redirect(url_for("listar_jugadores"))

    return render_template("jugador_form.html", jugador=None, modo="nuevo")


@app.route("/jugadores/importar", methods=["GET", "POST"])
def importar_jugadores():
    check = permiso_requerido("jugadores_gestionar")
    if check:
        return check

    resultado = None

    if request.method == "POST":
        archivo = request.files.get("archivo")

        if not archivo or not archivo.filename:
            flash("Deb?s seleccionar un archivo Excel.", "error")
            return render_template("importar_jugadores.html", resultado=resultado)

        if not archivo.filename.lower().endswith(".xlsx"):
            flash("El archivo debe ser .xlsx.", "error")
            return render_template("importar_jugadores.html", resultado=resultado)

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            flash("No se pudo leer el archivo Excel.", "error")
            return render_template("importar_jugadores.html", resultado=resultado)

        if not rows:
            flash("El archivo est? vac?o.", "error")
            return render_template("importar_jugadores.html", resultado=resultado)

        headers = [normalizar_header_excel(valor) for valor in rows[0]]
        creados = 0
        omitidos = 0
        errores = []

        conn = get_connection()

        for numero_fila, row in enumerate(rows[1:], start=2):
            if not any(limpiar_valor_excel(valor) for valor in row):
                continue

            data = mapear_fila_jugador(headers, row)

            if not data["nombre"] or not data["apellido"]:
                omitidos += 1
                errores.append(f"Fila {numero_fila}: falta nombre o apellido.")
                continue

            if data["dni"]:
                existente = conn.execute(
                    "SELECT id FROM jugadores WHERE dni = %s",
                    (data["dni"],)
                ).fetchone()
                if existente:
                    omitidos += 1
                    errores.append(f"Fila {numero_fila}: DNI ya existente ({data['dni']}).")
                    continue

            if not data["numero_socio"]:
                data["numero_socio"] = siguiente_numero_socio(conn)
            else:
                data["numero_socio"] = formatear_numero_socio(re.sub(r"\D+", "", data["numero_socio"]) or data["numero_socio"])

            conn.execute("""
                INSERT INTO jugadores (
                    nombre, apellido, dni, fecha_nacimiento, telefono, email, categoria,
                    fecha_ingreso, estado, contacto_tutor, parentesco_tutor, telefono_tutor,
                    email_tutor, direccion, obra_social, numero_afiliado_obra_social, numero_socio,
                    documentos, observaciones, tipo_miembro, cobra_cuota
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data["nombre"], data["apellido"], data["dni"], data["fecha_nacimiento"],
                data["telefono"], data["email"], data["categoria"], data["fecha_ingreso"],
                data["estado"], data["contacto_tutor"], data["parentesco_tutor"],
                data["telefono_tutor"], data["email_tutor"], data["direccion"],
                data["obra_social"], data["numero_afiliado_obra_social"], data["numero_socio"], data["documentos"],
                data["observaciones"], data["tipo_miembro"], data["cobra_cuota"],
            ))
            creados += 1

        conn.commit()
        conn.close()

        resultado = {
            "creados": creados,
            "omitidos": omitidos,
            "errores": errores[:30],
        }

        registrar_auditoria(
            "importar_ok",
            "jugadores",
            None,
            {
                "archivo": archivo.filename,
                "creados": creados,
                "omitidos": omitidos,
                "errores": len(errores),
            },
        )

        flash(f"Importación terminada. Creados: {creados}. Omitidos: {omitidos}.", "ok")

    return render_template("importar_jugadores.html", resultado=resultado)


@app.route("/jugadores/acciones-masivas", methods=["POST"])
def acciones_masivas_jugadores():
    check = permiso_requerido("jugadores_gestionar")
    if check:
        return check

    jugador_ids = request.form.getlist("jugador_ids")
    accion = request.form.get("accion", "").strip()
    nuevo_estado = request.form.get("estado", "").strip()
    nueva_categoria = request.form.get("categoria", "").strip()

    if not jugador_ids:
        flash("Seleccioná al menos un jugador.", "error")
        return redirect(url_for("listar_jugadores"))

    ids = []
    for valor in jugador_ids:
        try:
            ids.append(int(valor))
        except ValueError:
            continue

    if not ids:
        flash("La selección no es válida.", "error")
        return redirect(url_for("listar_jugadores"))

    conn = get_connection()

    if accion == "estado":
        if nuevo_estado not in {"Activo", "Inactivo", "Suspendido", "Baja"}:
            conn.close()
            flash("El estado seleccionado no es válido.", "error")
            return redirect(url_for("listar_jugadores"))

        conn.execute(
            "UPDATE jugadores SET estado = %s WHERE id = ANY(%s)",
            (nuevo_estado, ids)
        )
        detalle = {"accion": "estado", "estado": nuevo_estado, "cantidad": len(ids)}
        mensaje = f"Estado actualizado para {len(ids)} jugador(es)."
    elif accion == "categoria":
        if not nueva_categoria:
            conn.close()
            flash("Ingresá una categoría.", "error")
            return redirect(url_for("listar_jugadores"))

        conn.execute(
            "UPDATE jugadores SET categoria = %s WHERE id = ANY(%s)",
            (nueva_categoria, ids)
        )
        detalle = {"accion": "categoria", "categoria": nueva_categoria, "cantidad": len(ids)}
        mensaje = f"Categoría actualizada para {len(ids)} jugador(es)."
    elif accion == "portal_activar":
        if not tiene_permiso("portal_jugador_gestionar"):
            conn.close()
            flash("No tenes permiso para gestionar portales de jugadores.", "error")
            return redirect(url_for("listar_jugadores"))

        jugadores = conn.execute("""
            SELECT id, portal_token
            FROM jugadores
            WHERE id = ANY(%s)
            FOR UPDATE
        """, (ids,)).fetchall()
        ahora = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
        activados = 0
        tokens_generados = 0
        for jugador in jugadores:
            token = jugador.get("portal_token")
            if not token:
                token = generar_portal_token()
                tokens_generados += 1
            conn.execute("""
                UPDATE jugadores
                SET portal_token = %s,
                    portal_activo = 1,
                    portal_actualizado_en = %s
                WHERE id = %s
            """, (token, ahora, jugador["id"]))
            activados += 1

        detalle = {
            "accion": "portal_activar",
            "cantidad": activados,
            "tokens_generados": tokens_generados,
        }
        mensaje = f"Portal activado para {activados} jugador(es)."
    else:
        conn.close()
        flash("La acción masiva no es válida.", "error")
        return redirect(url_for("listar_jugadores"))

    conn.commit()
    conn.close()

    registrar_auditoria("accion_masiva", "jugadores", None, detalle)

    flash(mensaje, "ok")
    return redirect(url_for("listar_jugadores"))


@app.route("/jugadores/recalcular-numeros-socio", methods=["POST"])
def recalcular_numeros_socio_route():
    check = permiso_requerido("jugadores_gestionar")
    if check:
        return check

    conn = get_connection()
    resultado = recalcular_numeros_socio(conn)
    conn.commit()
    conn.close()

    registrar_auditoria(
        "recalcular",
        "numero_socio",
        None,
        resultado,
    )

    if resultado["cantidad"]:
        primero = resultado["primero"]
        detalle_migracion = ""
        if resultado.get("migrados_obra_social"):
            detalle_migracion = (
                f" Se preservaron {resultado['migrados_obra_social']} numeros previos como afiliado de obra social."
            )
        flash(
            f"Numeros de socio del club regenerados para {resultado['cantidad']} registro(s). "
            f"El numero 00001 quedo asignado a {primero['apellido']}, {primero['nombre']}."
            f"{detalle_migracion}",
            "ok",
        )
    else:
        flash("No habia registros para numerar.", "warning")
    return redirect(url_for("listar_jugadores"))


@app.route("/jugadores/<int:jugador_id>/editar", methods=["GET", "POST"])
def editar_jugador(jugador_id):
    check = permiso_requerido("jugadores_gestionar")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()

    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    if request.method == "POST":
        data = {
            "nombre": request.form.get("nombre", "").strip(),
            "apellido": request.form.get("apellido", "").strip(),
            "dni": request.form.get("dni", "").strip(),
            "fecha_nacimiento": request.form.get("fecha_nacimiento", "").strip(),
            "telefono": request.form.get("telefono", "").strip(),
            "email": request.form.get("email", "").strip(),
            "categoria": request.form.get("categoria", "").strip(),
            "fecha_ingreso": request.form.get("fecha_ingreso", "").strip(),
            "estado": request.form.get("estado", "").strip() or "Activo",
            "contacto_tutor": request.form.get("contacto_tutor", "").strip(),
            "parentesco_tutor": request.form.get("parentesco_tutor", "").strip(),
            "telefono_tutor": request.form.get("telefono_tutor", "").strip(),
            "email_tutor": request.form.get("email_tutor", "").strip(),
            "direccion": request.form.get("direccion", "").strip(),
            "obra_social": request.form.get("obra_social", "").strip(),
            "numero_afiliado_obra_social": request.form.get("numero_afiliado_obra_social", "").strip(),
            "numero_socio": request.form.get("numero_socio", "").strip(),
            "tipo_miembro": normalizar_tipo_miembro(request.form.get("tipo_miembro")),
            "cobra_cuota": 1 if request.form.get("cobra_cuota", "on") == "on" else 0,
            "documentos": request.form.get("documentos", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        beca_data, beca_error = datos_beca_form()
        data.update(beca_data or {
            "beca_activa": 1 if request.form.get("beca_activa") == "on" else 0,
            "beca_porcentaje": request.form.get("beca_porcentaje", "").strip(),
            "beca_desde": request.form.get("beca_desde", "").strip(),
            "beca_hasta": request.form.get("beca_hasta", "").strip(),
            "beca_motivo": request.form.get("beca_motivo", "").strip(),
        })

        if not data["nombre"] or not data["apellido"]:
            conn.close()
            flash("Nombre y apellido son obligatorios.", "error")
            jugador_dict = dict(data)
            jugador_dict["id"] = jugador_id
            return render_template("jugador_form.html", jugador=jugador_dict, modo="editar")
        if beca_error:
            conn.close()
            flash(beca_error, "error")
            jugador_dict = dict(data)
            jugador_dict["id"] = jugador_id
            return render_template("jugador_form.html", jugador=jugador_dict, modo="editar")

        if data["numero_socio"]:
            data["numero_socio"] = formatear_numero_socio(re.sub(r"\D+", "", data["numero_socio"]) or data["numero_socio"])
        registro_beca = beca_modificada(jugador, data)
        conn.execute("""
            UPDATE jugadores
            SET nombre = %s, apellido = %s, dni = %s, fecha_nacimiento = %s, telefono = %s,
                email = %s, categoria = %s, fecha_ingreso = %s, estado = %s,
                contacto_tutor = %s, parentesco_tutor = %s, telefono_tutor = %s,
                email_tutor = %s, direccion = %s, obra_social = %s, numero_afiliado_obra_social = %s, numero_socio = %s,
                tipo_miembro = %s, cobra_cuota = %s, documentos = %s, observaciones = %s,
                beca_activa = %s, beca_porcentaje = %s, beca_desde = %s,
                beca_hasta = %s, beca_motivo = %s
            WHERE id = %s
        """, (
            data["nombre"], data["apellido"], data["dni"], data["fecha_nacimiento"],
            data["telefono"], data["email"], data["categoria"], data["fecha_ingreso"],
            data["estado"], data["contacto_tutor"], data["parentesco_tutor"],
            data["telefono_tutor"], data["email_tutor"], data["direccion"],
            data["obra_social"], data["numero_afiliado_obra_social"], data["numero_socio"], data["tipo_miembro"], data["cobra_cuota"],
            data["documentos"], data["observaciones"], data["beca_activa"], data["beca_porcentaje"],
            data["beca_desde"], data["beca_hasta"], data["beca_motivo"], jugador_id
        ))
        if registro_beca:
            registrar_historial_beca(
                conn,
                jugador_id,
                data,
                "actualizacion",
                {"anterior": snapshot_beca(jugador), "nuevo": snapshot_beca(data)},
            )
        conn.commit()
        conn.close()

        flash("Registro actualizado correctamente.", "ok")
        return redirect(url_for("listar_jugadores"))

    conn.close()
    return render_template("jugador_form.html", jugador=jugador, modo="editar")


@app.route("/jugadores/<int:jugador_id>/eliminar", methods=["POST"])
def eliminar_jugador(jugador_id):
    check = permiso_requerido("jugadores_eliminar")
    if check:
        return check

    conn = get_connection()
    conn.execute("DELETE FROM jugadores WHERE id = %s", (jugador_id,))
    conn.commit()
    conn.close()
    flash("Jugador eliminado.", "ok")
    return redirect(url_for("listar_jugadores"))

@app.route("/jugadores/<int:jugador_id>/cuotas")
def ver_cuotas(jugador_id):
    check = permiso_requerido("cuotas_ver", "cuotas_gestionar")
    if check:
        return check

    conn = get_connection()

    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    cuotas = conn.execute("""
        SELECT * FROM cuotas
        WHERE jugador_id = %s
        ORDER BY periodo DESC
    """, (jugador_id,)).fetchall()

    deuda = conn.execute("""
        SELECT SUM(importe) AS total
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(importe, 0) > 0
    """, (jugador_id,)).fetchone()["total"] or 0

    becas_historial = conn.execute("""
        SELECT *
        FROM becas_historial
        WHERE jugador_id = %s
        ORDER BY creado_en DESC, id DESC
        LIMIT 10
    """, (jugador_id,)).fetchall()

    planes_pago = conn.execute("""
        SELECT
            p.*,
            COALESCE(incluidas.cuotas_incluidas, 0) AS cuotas_incluidas,
            COALESCE(incluidas.monto_incluido, 0) AS monto_incluido
        FROM planes_pago p
        LEFT JOIN (
            SELECT plan_pago_id, COUNT(*) AS cuotas_incluidas, SUM(COALESCE(NULLIF(importe_anulado, 0), NULLIF(importe, 0), importe_original, 0)) AS monto_incluido
            FROM cuotas
            WHERE COALESCE(anulada, 0) = 1
              AND plan_pago_id IS NOT NULL
            GROUP BY plan_pago_id
        ) incluidas ON incluidas.plan_pago_id = p.id
        WHERE p.jugador_id = %s
        ORDER BY
            CASE WHEN estado = 'Activo' THEN 0 ELSE 1 END,
            fecha_inicio DESC,
            id DESC
    """, (jugador_id,)).fetchall()

    conn.close()

    return render_template(
        "cuotas.html",
        jugador=jugador,
        cuotas=cuotas,
        deuda=deuda,
        becas_historial=becas_historial,
        planes_pago=planes_pago,
    )


@app.route("/jugadores/<int:jugador_id>/cuotas/recalcular-becas", methods=["POST"])
def recalcular_becas_jugador(jugador_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    periodo_desde = validar_mes_beca(request.form.get("periodo_desde", ""))
    periodo_hasta = validar_mes_beca(request.form.get("periodo_hasta", ""))
    if periodo_desde is None or periodo_hasta is None:
        flash("Los periodos de recalculo deben tener formato YYYY-MM.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=jugador_id))
    if periodo_desde and periodo_hasta and periodo_hasta < periodo_desde:
        flash("El periodo hasta no puede ser anterior al periodo desde.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=jugador_id))

    conn = get_connection()
    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,),
    ).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    resultado = recalcular_cuotas_becadas(conn, jugador, periodo_desde or "", periodo_hasta or "")
    registrar_historial_beca(
        conn,
        jugador_id,
        jugador,
        "recalculo_cuotas",
        {
            "periodo_desde": periodo_desde or None,
            "periodo_hasta": periodo_hasta or None,
            **resultado,
        },
    )
    conn.commit()
    conn.close()

    flash(
        "Recalculo terminado. "
        f"Revisadas: {resultado['revisadas']}. "
        f"Actualizadas: {resultado['actualizadas']}. "
        f"Becas totales: {resultado['becas_totales']}. "
        f"Becas parciales: {resultado['becas_parciales']}.",
        "ok",
    )
    return redirect(url_for("ver_cuotas", jugador_id=jugador_id))


@app.route("/planes-pago")
def listar_planes_pago():
    check = permiso_requerido("planes_pago_ver")
    if check:
        return check

    estado = request.args.get("estado", "Activo").strip()
    if estado not in {"Activo", "Cumplido", "Caido", "todos"}:
        estado = "Activo"

    condiciones = []
    parametros = []
    if estado != "todos":
        condiciones.append("p.estado = %s")
        parametros.append(estado)

    where_sql = "WHERE " + " AND ".join(condiciones) if condiciones else ""

    conn = get_connection()
    planes = conn.execute(f"""
        SELECT
            p.*,
            j.apellido,
            j.nombre,
            j.categoria,
            COALESCE(deuda.deuda, 0) AS deuda_actual,
            COALESCE(incluidas.cuotas_incluidas, 0) AS cuotas_incluidas,
            COALESCE(incluidas.monto_incluido, 0) AS monto_incluido
        FROM planes_pago p
        JOIN jugadores j ON j.id = p.jugador_id
        LEFT JOIN (
            SELECT jugador_id, SUM(importe) AS deuda
            FROM cuotas
            WHERE pagado = 0
              AND COALESCE(anulada, 0) = 0
              AND COALESCE(importe, 0) > 0
            GROUP BY jugador_id
        ) deuda ON deuda.jugador_id = j.id
        LEFT JOIN (
            SELECT plan_pago_id, COUNT(*) AS cuotas_incluidas, SUM(COALESCE(NULLIF(importe_anulado, 0), NULLIF(importe, 0), importe_original, 0)) AS monto_incluido
            FROM cuotas
            WHERE COALESCE(anulada, 0) = 1
              AND plan_pago_id IS NOT NULL
            GROUP BY plan_pago_id
        ) incluidas ON incluidas.plan_pago_id = p.id
        {where_sql}
        ORDER BY
            CASE WHEN p.estado = 'Activo' THEN 0 ELSE 1 END,
            p.fecha_inicio DESC,
            j.apellido,
            j.nombre
    """, parametros).fetchall()
    conn.close()

    return render_template("planes_pago.html", planes=planes, estado=estado)


def datos_plan_pago_form(require_monto=True):
    fecha_inicio = request.form.get("fecha_inicio", "").strip() or ahora_sig().strftime("%Y-%m-%d")
    descripcion = request.form.get("descripcion", "").strip()
    observaciones = request.form.get("observaciones", "").strip()
    estado = request.form.get("estado", "Activo").strip() or "Activo"
    if estado not in {"Activo", "Cumplido", "Caido"}:
        estado = "Activo"

    try:
        monto_total = float(request.form.get("monto_total", "0") or 0)
        cantidad_cuotas = max(1, int(request.form.get("cantidad_cuotas", "1") or 1))
    except ValueError:
        raise ValueError("El monto y la cantidad de cuotas deben ser numericos.")

    if monto_total < 0:
        raise ValueError("El monto total del plan no puede ser negativo.")
    if require_monto and monto_total <= 0:
        raise ValueError("El monto total del plan debe ser mayor a cero.")

    monto_total = round(monto_total, 2)
    return {
        "fecha_inicio": fecha_inicio,
        "monto_total": monto_total,
        "cantidad_cuotas": cantidad_cuotas,
        "monto_cuota": round(monto_total / cantidad_cuotas, 2) if monto_total else 0,
        "estado": estado,
        "descripcion": descripcion,
        "observaciones": observaciones,
    }


def obtener_cuotas_impagas_para_plan(conn, jugador_id):
    return conn.execute("""
        SELECT id, periodo, importe, importe_original, fecha_vencimiento
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(importe, 0) > 0
        ORDER BY periodo ASC, id ASC
    """, (jugador_id,)).fetchall()


def total_cuotas_plan(cuotas):
    total = 0
    for cuota in cuotas:
        try:
            total += float(cuota["importe"] or 0)
        except (TypeError, ValueError):
            continue
    return round(total, 2)


@app.route("/jugadores/<int:jugador_id>/planes/nuevo", methods=["GET", "POST"])
def nuevo_plan_pago(jugador_id):
    check = permiso_requerido("planes_pago_gestionar")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    cuotas_impagas = obtener_cuotas_impagas_para_plan(conn, jugador_id)
    deuda = total_cuotas_plan(cuotas_impagas)

    if request.method == "POST":
        try:
            data = datos_plan_pago_form(require_monto=False)
        except ValueError as error:
            conn.close()
            flash(str(error), "error")
            return render_template("plan_pago_form.html", jugador=jugador, deuda=deuda, cuotas_impagas=cuotas_impagas, modo="nuevo")

        cuotas_ids = []
        for valor in request.form.getlist("cuotas_incluidas"):
            try:
                cuotas_ids.append(int(valor))
            except (TypeError, ValueError):
                continue
        if not cuotas_ids:
            conn.close()
            flash("Selecciona al menos una cuota impaga para incluir en el plan.", "error")
            return render_template("plan_pago_form.html", jugador=jugador, deuda=deuda, cuotas_impagas=cuotas_impagas, modo="nuevo")

        cuotas_incluidas = conn.execute("""
            SELECT id, importe, importe_original
            FROM cuotas
            WHERE jugador_id = %s
              AND id = ANY(%s)
              AND pagado = 0
              AND COALESCE(anulada, 0) = 0
              AND COALESCE(importe, 0) > 0
            FOR UPDATE
        """, (jugador_id, cuotas_ids)).fetchall()
        if len(cuotas_incluidas) != len(set(cuotas_ids)):
            conn.close()
            flash("Una o mas cuotas seleccionadas ya no estan disponibles para incluir.", "error")
            return render_template("plan_pago_form.html", jugador=jugador, deuda=deuda, cuotas_impagas=cuotas_impagas, modo="nuevo")

        monto_total = total_cuotas_plan(cuotas_incluidas)
        data["monto_total"] = monto_total
        data["monto_cuota"] = round(monto_total / data["cantidad_cuotas"], 2)

        plan = conn.execute("""
            INSERT INTO planes_pago (
                jugador_id, fecha_inicio, monto_total, cantidad_cuotas, monto_cuota,
                estado, descripcion, observaciones, creado_por
            )
            VALUES (%s, %s, %s, %s, %s, 'Activo', %s, %s, %s)
            RETURNING id
        """, (
            jugador_id,
            data["fecha_inicio"],
            data["monto_total"],
            data["cantidad_cuotas"],
            data["monto_cuota"],
            data["descripcion"],
            data["observaciones"],
            session.get("username"),
        )).fetchone()
        plan_id = plan["id"]

        conn.execute("""
            UPDATE cuotas
            SET anulada = 1,
                anulada_en = %s,
                anulada_por = %s,
                anulacion_motivo = %s,
                plan_pago_id = %s,
                importe_anulado = importe,
                importe_original = COALESCE(importe_original, importe),
                importe = 0,
                metodo_pago = 'Plan de pago',
                referencia_pago = %s,
                comprobante_estado = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN 'aceptado'
                    ELSE comprobante_estado
                END,
                comprobante_revisado_en = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN %s
                    ELSE comprobante_revisado_en
                END,
                comprobante_revisado_por = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN %s
                    ELSE comprobante_revisado_por
                END,
                comprobante_observaciones = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN %s
                    ELSE comprobante_observaciones
                END
            WHERE jugador_id = %s
              AND id = ANY(%s)
        """, (
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("username"),
            f"Incluida en plan de pago #{plan_id}",
            plan_id,
            f"Cuota anulada e incluida en plan de pago #{plan_id}",
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("username"),
            f"Archivado por plan de pago #{plan_id}",
            jugador_id,
            cuotas_ids,
        ))
        recalcular_cuotas_planes_pago(conn, jugador_id, periodo_inicio_plan(data))
        conn.commit()
        conn.close()

        flash(f"Plan de pago creado. Cuotas incluidas/anuladas: {len(cuotas_incluidas)}.", "ok")
        return redirect(url_for("ver_cuotas", jugador_id=jugador_id))

    conn.close()
    return render_template("plan_pago_form.html", jugador=jugador, deuda=deuda, cuotas_impagas=cuotas_impagas, modo="nuevo")


@app.route("/planes-pago/<int:plan_id>/editar", methods=["GET", "POST"])
def editar_plan_pago(plan_id):
    check = permiso_requerido("planes_pago_gestionar")
    if check:
        return check

    conn = get_connection()
    plan = conn.execute("""
        SELECT p.*, j.apellido, j.nombre, j.categoria
        FROM planes_pago p
        JOIN jugadores j ON j.id = p.jugador_id
        WHERE p.id = %s
    """, (plan_id,)).fetchone()
    if plan is None:
        conn.close()
        flash("Plan de pago no encontrado.", "error")
        return redirect(url_for("listar_planes_pago"))

    deuda = conn.execute("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(importe, 0) > 0
    """, (plan["jugador_id"],)).fetchone()["total"]

    if request.method == "POST":
        try:
            data = datos_plan_pago_form()
        except ValueError as error:
            conn.close()
            flash(str(error), "error")
            return render_template("plan_pago_form.html", jugador=plan, deuda=deuda, plan=plan, modo="editar")

        cerrado_en = plan["cerrado_en"]
        cerrado_por = plan["cerrado_por"]
        if data["estado"] != plan["estado"]:
            cerrado_en = ahora_sig().strftime("%Y-%m-%d") if data["estado"] != "Activo" else None
            cerrado_por = session.get("username") if data["estado"] != "Activo" else None

        periodo_recalculo = periodo_minimo(periodo_inicio_plan(plan), periodo_inicio_plan(data))

        conn.execute("""
            UPDATE planes_pago
            SET fecha_inicio = %s,
                monto_total = %s,
                cantidad_cuotas = %s,
                monto_cuota = %s,
                estado = %s,
                descripcion = %s,
                observaciones = %s,
                cerrado_en = %s,
                cerrado_por = %s
            WHERE id = %s
        """, (
            data["fecha_inicio"],
            data["monto_total"],
            data["cantidad_cuotas"],
            data["monto_cuota"],
            data["estado"],
            data["descripcion"],
            data["observaciones"],
            cerrado_en,
            cerrado_por,
            plan_id,
        ))
        recalcular_cuotas_planes_pago(conn, plan["jugador_id"], periodo_recalculo)
        conn.commit()
        conn.close()

        flash("Plan de pago editado correctamente.", "ok")
        return redirect(url_for("ver_cuotas", jugador_id=plan["jugador_id"]))

    conn.close()
    return render_template("plan_pago_form.html", jugador=plan, deuda=deuda, plan=plan, modo="editar")


@app.route("/planes-pago/<int:plan_id>/eliminar", methods=["POST"])
def eliminar_plan_pago(plan_id):
    check = permiso_requerido("planes_pago_gestionar")
    if check:
        return check

    conn = get_connection()
    plan = conn.execute("SELECT * FROM planes_pago WHERE id = %s", (plan_id,)).fetchone()
    if plan is None:
        conn.close()
        flash("Plan de pago no encontrado.", "error")
        return redirect(url_for("listar_planes_pago"))

    conn.execute("""
        UPDATE cuotas
        SET anulada = 0,
            anulada_en = NULL,
            anulada_por = NULL,
            anulacion_motivo = NULL,
            plan_pago_id = NULL,
            importe = COALESCE(NULLIF(importe_anulado, 0), NULLIF(importe, 0), importe_original, importe),
            importe_anulado = NULL,
            metodo_pago = NULL,
            referencia_pago = NULL,
            comprobante_estado = CASE
                WHEN comprobante_drive_file_id IS NOT NULL AND pagado = 0 THEN 'pendiente'
                ELSE comprobante_estado
            END,
            comprobante_revisado_en = CASE
                WHEN comprobante_drive_file_id IS NOT NULL AND pagado = 0 THEN NULL
                ELSE comprobante_revisado_en
            END,
            comprobante_revisado_por = CASE
                WHEN comprobante_drive_file_id IS NOT NULL AND pagado = 0 THEN NULL
                ELSE comprobante_revisado_por
            END,
            comprobante_observaciones = CASE
                WHEN comprobante_drive_file_id IS NOT NULL AND pagado = 0 THEN NULL
                ELSE comprobante_observaciones
            END
        WHERE plan_pago_id = %s
          AND COALESCE(anulada, 0) = 1
    """, (plan_id,))
    conn.execute("DELETE FROM planes_pago WHERE id = %s", (plan_id,))
    recalcular_cuotas_planes_pago(conn, plan["jugador_id"], periodo_inicio_plan(plan))
    conn.commit()
    conn.close()

    flash("Plan de pago eliminado.", "ok")
    return redirect(request.form.get("next") or url_for("ver_cuotas", jugador_id=plan["jugador_id"]))


@app.route("/planes-pago/<int:plan_id>/actualizar", methods=["POST"])
def actualizar_plan_pago(plan_id):
    check = permiso_requerido("planes_pago_gestionar")
    if check:
        return check

    estado = request.form.get("estado", "").strip()
    if estado not in {"Activo", "Cumplido", "Caido"}:
        flash("Estado de plan no valido.", "error")
        return redirect(url_for("listar_planes_pago"))

    conn = get_connection()
    plan = conn.execute("SELECT * FROM planes_pago WHERE id = %s", (plan_id,)).fetchone()
    if plan is None:
        conn.close()
        flash("Plan de pago no encontrado.", "error")
        return redirect(url_for("listar_planes_pago"))

    cerrado_en = ahora_sig().strftime("%Y-%m-%d") if estado != "Activo" else None
    cerrado_por = session.get("username") if estado != "Activo" else None
    conn.execute("""
        UPDATE planes_pago
        SET estado = %s,
            cerrado_en = %s,
            cerrado_por = %s
        WHERE id = %s
    """, (estado, cerrado_en, cerrado_por, plan_id))
    recalcular_cuotas_planes_pago(conn, plan["jugador_id"], periodo_inicio_plan(plan))
    conn.commit()
    conn.close()

    flash("Plan de pago actualizado.", "ok")
    return redirect(request.form.get("next") or url_for("ver_cuotas", jugador_id=plan["jugador_id"]))

@app.route("/jugadores/<int:jugador_id>/cuotas/nueva", methods=["GET", "POST"])
def nueva_cuota(jugador_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    if request.method == "POST":
        periodo = request.form.get("periodo", "").strip()
        importe = request.form.get("importe", "").strip()
        fecha_vencimiento = request.form.get("fecha_vencimiento", "").strip()

        if not periodo or not importe:
            flash("Período e importe son obligatorios.", "error")
            return render_template("cuota_form.html", jugador_id=jugador_id)

        try:
            importe_valor = float(importe)
        except ValueError:
            flash("El importe debe ser numérico.", "error")
            return render_template(
                "cuota_form.html",
                jugador_id=jugador_id,
                periodo=periodo,
                importe=importe,
                fecha_vencimiento=fecha_vencimiento
            )

        conn = get_connection()
        jugador = conn.execute(
            "SELECT * FROM jugadores WHERE id = %s",
            (jugador_id,)
        ).fetchone()

        if jugador is None:
            conn.close()
            flash("Jugador no encontrado.", "error")
            return redirect(url_for("listar_jugadores"))

        existente = conn.execute("""
            SELECT id
            FROM cuotas
            WHERE jugador_id = %s AND periodo = %s
        """, (jugador_id, periodo)).fetchone()

        if existente:
            conn.close()
            flash("Ese jugador ya tiene una cuota cargada para ese período.", "error")
            return render_template(
                "cuota_form.html",
                jugador_id=jugador_id,
                periodo=periodo,
                importe=importe,
                fecha_vencimiento=fecha_vencimiento
            )

        cuota_calculada = calcular_importe_cuota_mensual(conn, jugador, periodo, importe_valor)
        pagado_inicial = 1 if cuota_calculada["beca_total"] else 0
        fecha_pago_inicial = ahora_sig().strftime("%Y-%m-%d") if pagado_inicial else None
        metodo_inicial = "Beca" if pagado_inicial else None
        referencia_inicial = (
            f"Beca total {cuota_calculada['beca_porcentaje']:g}%"
            if pagado_inicial else None
        )

        conn.execute("""
            INSERT INTO cuotas (
                jugador_id, periodo, importe, pagado, fecha_pago, fecha_vencimiento,
                importe_original, descuento_beca, beca_porcentaje, beca_motivo,
                becada, metodo_pago, referencia_pago, plan_pago_monto, plan_pago_detalle
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            jugador_id,
            periodo,
            cuota_calculada["importe"],
            pagado_inicial,
            fecha_pago_inicial,
            fecha_vencimiento or None,
            cuota_calculada["importe_original"],
            cuota_calculada["descuento_beca"],
            cuota_calculada["beca_porcentaje"],
            cuota_calculada["beca_motivo"],
            cuota_calculada["becada"],
            metodo_inicial,
            referencia_inicial,
            cuota_calculada["plan_pago_monto"],
            cuota_calculada["plan_pago_detalle"] or None,
        ))

        conn.commit()
        conn.close()

        if cuota_calculada["beca_total"]:
            flash("Cuota cargada como beca total.", "ok")
        elif cuota_calculada["becada"]:
            flash("Cuota cargada con beca parcial aplicada.", "ok")
        else:
            flash("Cuota cargada correctamente.", "ok")
        return redirect(url_for("ver_cuotas", jugador_id=jugador_id))

    return render_template("cuota_form.html", jugador_id=jugador_id)

@app.route("/cuotas/<int:cuota_id>/pagar", methods=["GET", "POST"])
def pagar_cuota(cuota_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    conn = get_connection()

    cuota = conn.execute("""
        SELECT
            c.*,
            j.nombre,
            j.apellido,
            j.email,
            j.email_tutor
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
        FOR UPDATE
    """, (cuota_id,)).fetchone()

    if cuota is None:
        conn.close()
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    if cuota["pagado"]:
        conn.close()
        flash("La cuota ya estaba marcada como pagada.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    if cuota.get("anulada"):
        conn.close()
        flash("La cuota esta anulada por un plan de pago.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    if request.method == "POST":
        metodo_pago = request.form.get("metodo_pago", "").strip()
        referencia_pago = request.form.get("referencia_pago", "").strip()
        comprobante_pago = request.files.get("comprobante_pago")

        if not metodo_pago:
            conn.close()
            flash("Debe seleccionar un método de pago.", "error")
            return render_template("pagar_cuota.html", cuota=cuota)

        try:
            comprobante_info = subir_comprobante_a_drive(comprobante_pago, cuota)
        except (RuntimeError, ValueError) as error:
            conn.close()
            flash(str(error), "error")
            return render_template("pagar_cuota.html", cuota=cuota)
        except Exception as error:
            app.logger.exception("No se pudo subir comprobante de cuota %s al registrar pago.", cuota_id)
            conn.close()
            flash(f"{mensaje_error_drive(error)} La cuota no fue marcada como pagada.", "error")
            return render_template("pagar_cuota.html", cuota=cuota)

        nuevo_numero = cuota["numero_recibo"] or siguiente_numero_recibo(conn)
        comprobante_fecha = ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if comprobante_info else None
        comprobante_usuario = session.get("username") if comprobante_info else None
        hay_comprobante = bool(comprobante_info or cuota.get("comprobante_drive_file_id"))
        revisado_en = ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if hay_comprobante else None
        revisado_por = session.get("username") if hay_comprobante else None

        conn.execute("""
            UPDATE cuotas
            SET pagado = 1,
                fecha_pago = CURRENT_DATE,
                numero_recibo = %s,
                metodo_pago = %s,
                referencia_pago = %s,
                comprobante_drive_file_id = COALESCE(%s, comprobante_drive_file_id),
                comprobante_nombre = COALESCE(%s, comprobante_nombre),
                comprobante_mime_type = COALESCE(%s, comprobante_mime_type),
                comprobante_tamano = COALESCE(%s, comprobante_tamano),
                comprobante_fecha = COALESCE(%s, comprobante_fecha),
                comprobante_usuario = COALESCE(%s, comprobante_usuario),
                comprobante_web_url = COALESCE(%s, comprobante_web_url),
                comprobante_estado = CASE
                    WHEN COALESCE(%s, comprobante_drive_file_id) IS NOT NULL THEN 'aceptado'
                    ELSE comprobante_estado
                END,
                comprobante_revisado_en = CASE
                    WHEN COALESCE(%s, comprobante_drive_file_id) IS NOT NULL THEN %s
                    ELSE comprobante_revisado_en
                END,
                comprobante_revisado_por = CASE
                    WHEN COALESCE(%s, comprobante_drive_file_id) IS NOT NULL THEN %s
                    ELSE comprobante_revisado_por
                END,
                comprobante_observaciones = CASE
                    WHEN COALESCE(%s, comprobante_drive_file_id) IS NOT NULL THEN NULL
                    ELSE comprobante_observaciones
                END
            WHERE id = %s
        """, (
            nuevo_numero,
            metodo_pago,
            referencia_pago,
            comprobante_info["file_id"] if comprobante_info else None,
            comprobante_info["nombre"] if comprobante_info else None,
            comprobante_info["mime_type"] if comprobante_info else None,
            comprobante_info["tamano"] if comprobante_info else None,
            comprobante_fecha,
            comprobante_usuario,
            comprobante_info["web_url"] if comprobante_info else None,
            comprobante_info["file_id"] if comprobante_info else None,
            comprobante_info["file_id"] if comprobante_info else None,
            revisado_en,
            comprobante_info["file_id"] if comprobante_info else None,
            revisado_por,
            comprobante_info["file_id"] if comprobante_info else None,
            cuota_id,
        ))

        conn.execute("""
            INSERT INTO movimientos (tipo, concepto, monto, fecha, referencia)
            VALUES (%s, %s, %s, CURRENT_DATE, %s)
        """, (
            "ingreso",
            f"Cuota {cuota['periodo']} - {cuota['apellido']}, {cuota['nombre']}",
            cuota["importe"],
            f"Cuota Social ({metodo_pago})"
        ))

        conn.commit()
        conn.close()

        archivo_recibo = generar_recibo_pdf(cuota_id)
        recibo_enviado = False
        motivo_envio = None
        try:
            recibo_enviado, _, motivo_envio = enviar_recibo_cuota_por_email(cuota, archivo_recibo)
        except Exception:
            motivo_envio = "error"
            app.logger.exception("No se pudo enviar por email el recibo de cuota %s.", cuota_id)

        if comprobante_info:
            flash("Cuota marcada como pagada, registrada en caja y comprobante guardado en Drive.", "ok")
        else:
            flash("Cuota marcada como pagada y registrada en caja.", "ok")
        if recibo_enviado:
            flash("Recibo enviado por email al jugador.", "ok")
        elif motivo_envio == "sin_email":
            flash("El recibo se gener\u00f3, pero el jugador no tiene email cargado.", "warning")
        elif motivo_envio == "smtp":
            flash("El recibo se gener\u00f3, pero el email no est\u00e1 configurado en el sistema.", "warning")
        else:
            flash("El recibo se gener\u00f3, pero no se pudo enviar por email autom\u00e1ticamente.", "warning")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    conn.close()
    return render_template("pagar_cuota.html", cuota=cuota)

@app.route("/cuotas/<int:cuota_id>/comprobante/subir", methods=["GET", "POST"])
def subir_comprobante_cuota(cuota_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.*,
            j.nombre,
            j.apellido
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
    """, (cuota_id,)).fetchone()

    if cuota is None:
        conn.close()
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    if cuota.get("anulada"):
        conn.close()
        flash("La cuota esta anulada por un plan de pago.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    if not cuota["pagado"]:
        conn.close()
        flash("Solo se pueden adjuntar comprobantes en cuotas pagadas.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    if request.method == "POST":
        comprobante_pago = request.files.get("comprobante_pago")
        try:
            comprobante_info = subir_comprobante_a_drive(comprobante_pago, cuota)
        except (RuntimeError, ValueError) as error:
            conn.close()
            flash(str(error), "error")
            return render_template("comprobante_form.html", cuota=cuota)
        except Exception as error:
            app.logger.exception("No se pudo subir comprobante de cuota %s.", cuota_id)
            conn.close()
            flash(mensaje_error_drive(error), "error")
            return render_template("comprobante_form.html", cuota=cuota)

        if not comprobante_info:
            conn.close()
            flash("Debe seleccionar un archivo para adjuntar.", "error")
            return render_template("comprobante_form.html", cuota=cuota)

        conn.execute("""
            UPDATE cuotas
            SET comprobante_drive_file_id = %s,
                comprobante_nombre = %s,
                comprobante_mime_type = %s,
                comprobante_tamano = %s,
                comprobante_fecha = %s,
                comprobante_usuario = %s,
                comprobante_web_url = %s,
                comprobante_estado = 'aceptado',
                comprobante_revisado_en = %s,
                comprobante_revisado_por = %s,
                comprobante_observaciones = NULL
            WHERE id = %s
        """, (
            comprobante_info["file_id"],
            comprobante_info["nombre"],
            comprobante_info["mime_type"],
            comprobante_info["tamano"],
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("username"),
            comprobante_info["web_url"],
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
            session.get("username"),
            cuota_id,
        ))

        conn.commit()
        conn.close()

        flash("Comprobante guardado en Google Drive.", "ok")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    conn.close()
    return render_template("comprobante_form.html", cuota=cuota)


@app.route("/cuotas/comprobantes")
def listar_comprobantes_revision():
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    estado = request.args.get("estado", "pendiente").strip() or "pendiente"
    if estado not in COMPROBANTE_ESTADOS and estado != "todos":
        estado = "pendiente"

    filtros = [
        "c.comprobante_drive_file_id IS NOT NULL",
        "COALESCE(c.anulada, 0) = 0",
    ]
    params = []
    if estado != "todos":
        if estado == "pendiente":
            filtros.append("COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')")
        else:
            filtros.append("COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') = %s")
            params.append(estado)

    where_sql = " AND ".join(filtros)

    conn = get_connection()
    comprobantes = conn.execute(f"""
        SELECT
            c.*,
            CASE
                WHEN c.comprobante_drive_file_id IS NOT NULL
                 AND COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante') = 'sin_comprobante'
                THEN 'pendiente'
                ELSE COALESCE(NULLIF(c.comprobante_estado, ''), 'sin_comprobante')
            END AS comprobante_estado_resuelto,
            j.apellido,
            j.nombre
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE {where_sql}
        ORDER BY
            CASE COALESCE(NULLIF(c.comprobante_estado, ''), 'pendiente')
                WHEN 'pendiente' THEN 0
                WHEN 'sin_comprobante' THEN 0
                WHEN 'rechazado' THEN 1
                WHEN 'aceptado' THEN 2
                ELSE 3
            END,
            c.comprobante_fecha DESC NULLS LAST,
            c.id DESC
        LIMIT 150
    """, tuple(params)).fetchall()

    resumen = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (
                WHERE COALESCE(NULLIF(comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'sin_comprobante')
                  AND comprobante_drive_file_id IS NOT NULL
            ) AS pendientes,
            COUNT(*) FILTER (
                WHERE COALESCE(NULLIF(comprobante_estado, ''), 'pendiente') = 'aceptado'
            ) AS aceptados,
            COUNT(*) FILTER (
                WHERE COALESCE(NULLIF(comprobante_estado, ''), 'pendiente') = 'rechazado'
            ) AS rechazados
        FROM cuotas
        WHERE comprobante_drive_file_id IS NOT NULL
          AND COALESCE(anulada, 0) = 0
    """).fetchone()
    conn.close()

    return render_template(
        "comprobantes_revision.html",
        comprobantes=comprobantes,
        estado=estado,
        resumen=resumen,
    )


@app.route("/cuotas/<int:cuota_id>/comprobante/revisar", methods=["POST"])
def revisar_comprobante_cuota(cuota_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    accion = request.form.get("accion", "").strip()
    observaciones = request.form.get("observaciones", "").strip()
    next_url = destino_interno(request.form.get("next"), fallback="listar_comprobantes_revision")

    if accion not in {"aceptar", "rechazar"}:
        flash("La accion de revision no es valida.", "error")
        return redirect(next_url)

    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.*,
            j.apellido,
            j.nombre,
            j.email,
            j.email_tutor,
            j.portal_token,
            j.portal_activo
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
        FOR UPDATE
    """, (cuota_id,)).fetchone()

    if cuota is None:
        conn.close()
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    if not cuota["comprobante_drive_file_id"]:
        conn.close()
        flash("La cuota no tiene comprobante adjunto.", "error")
        return redirect(next_url)

    if cuota.get("anulada"):
        conn.close()
        flash("La cuota esta anulada por un plan de pago.", "error")
        return redirect(next_url)

    revisado_en = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
    revisado_por = session.get("username")

    if accion == "rechazar":
        conn.execute("""
            UPDATE cuotas
            SET comprobante_estado = 'rechazado',
                comprobante_revisado_en = %s,
                comprobante_revisado_por = %s,
                comprobante_observaciones = %s
            WHERE id = %s
        """, (revisado_en, revisado_por, observaciones or None, cuota_id))
        conn.commit()
        conn.close()
        portal_url = None
        if cuota.get("portal_token") and cuota.get("portal_activo"):
            portal_url = url_for("portal_jugador", token=cuota["portal_token"], _external=True)
        cuerpo = construir_texto_rechazo_comprobante(cuota, observaciones=observaciones, portal_url=portal_url)
        enviar_email_jugador(cuota, f"Comprobante rechazado - cuota {cuota['periodo']}", cuerpo)
        flash("Comprobante rechazado.", "ok")
        return redirect(next_url)

    numero_recibo = cuota["numero_recibo"] or siguiente_numero_recibo(conn)
    fecha_pago = cuota["fecha_pago"] or ahora_sig().strftime("%Y-%m-%d")
    metodo_pago = cuota["metodo_pago"] or "Comprobante portal"
    referencia_pago = cuota["referencia_pago"] or "Comprobante validado"

    conn.execute("""
        UPDATE cuotas
        SET pagado = 1,
            fecha_pago = %s,
            numero_recibo = %s,
            metodo_pago = %s,
            referencia_pago = %s,
            comprobante_estado = 'aceptado',
            comprobante_revisado_en = %s,
            comprobante_revisado_por = %s,
            comprobante_observaciones = %s
        WHERE id = %s
    """, (
        fecha_pago,
        numero_recibo,
        metodo_pago,
        referencia_pago,
        revisado_en,
        revisado_por,
        observaciones or None,
        cuota_id,
    ))

    try:
        importe = float(cuota["importe"] or 0)
    except (TypeError, ValueError):
        importe = 0

    if not cuota["pagado"] and importe > 0:
        conn.execute("""
            INSERT INTO movimientos (tipo, concepto, monto, fecha, referencia)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            "ingreso",
            f"Cuota {cuota['periodo']} - {cuota['apellido']}, {cuota['nombre']}",
            cuota["importe"],
            fecha_pago,
            f"Cuota Social ({metodo_pago})",
        ))

    conn.commit()
    conn.close()

    archivo_recibo = generar_recibo_pdf(cuota_id)
    recibo_enviado = False
    motivo_envio = None
    try:
        recibo_enviado, _, motivo_envio = enviar_recibo_cuota_por_email(cuota, archivo_recibo)
    except Exception:
        motivo_envio = "error"
        app.logger.exception("No se pudo enviar por email el recibo de cuota %s tras validar comprobante.", cuota_id)

    flash("Comprobante aceptado. Cuota marcada como pagada y recibo generado.", "ok")
    if recibo_enviado:
        flash("Recibo enviado por email al jugador.", "ok")
    elif motivo_envio == "sin_email":
        flash("El recibo se gener\u00f3, pero el jugador no tiene email cargado.", "warning")
    elif motivo_envio == "smtp":
        flash("El recibo se gener\u00f3, pero el email no est\u00e1 configurado en el sistema.", "warning")
    else:
        flash("El recibo se gener\u00f3, pero no se pudo enviar por email autom\u00e1ticamente.", "warning")
    return redirect(next_url)


@app.route("/cuotas/<int:cuota_id>/comprobante")
def descargar_comprobante_cuota(cuota_id):
    check = permiso_requerido("cuotas_ver", "cuotas_gestionar")
    if check:
        return check

    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.*,
            j.nombre,
            j.apellido
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
    """, (cuota_id,)).fetchone()
    conn.close()

    if cuota is None:
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    if not cuota["comprobante_drive_file_id"]:
        flash("La cuota no tiene comprobante adjunto.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    try:
        archivo = descargar_drive_file(cuota["comprobante_drive_file_id"])
    except RuntimeError as error:
        flash(str(error), "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))
    except Exception as error:
        app.logger.exception("No se pudo descargar comprobante de cuota %s.", cuota_id)
        flash(mensaje_error_drive(error), "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    registrar_auditoria(
        "descargar_ok",
        "comprobante_cuota",
        str(cuota_id),
        {
            "archivo": cuota["comprobante_nombre"],
            "drive_file_id": cuota["comprobante_drive_file_id"],
        },
    )

    return send_file(
        archivo,
        mimetype=cuota["comprobante_mime_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=cuota["comprobante_nombre"] or f"comprobante_cuota_{cuota_id}",
    )


@app.route("/cuotas/conciliacion", methods=["GET", "POST"])
def conciliar_pagos():
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    resultado = None
    matches = []
    matches_json = ""

    if request.method == "POST":
        accion = request.form.get("accion", "preview")
        if accion == "aplicar":
            try:
                matches = json.loads(request.form.get("matches_json", "[]"))
            except ValueError:
                matches = []

            conn = get_connection()
            resultado = aplicar_matches_conciliacion(conn, matches)
            conn.commit()
            conn.close()
            registrar_auditoria("conciliar_ok", "cuotas", None, resultado)
            flash(
                f"Conciliacion aplicada. Pagos registrados: {resultado['aplicados']}. "
                f"Omitidos: {resultado['omitidos']}.",
                "ok",
            )
            return redirect(url_for("conciliar_pagos"))

        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename:
            flash("Selecciona un archivo CSV o Excel.", "error")
            return render_template(
                "conciliacion.html",
                resultado=resultado,
                matches=matches,
                matches_json=matches_json,
            )

        try:
            filas = leer_csv_conciliacion(archivo)
        except ValueError as error:
            flash(str(error), "error")
            return render_template(
                "conciliacion.html",
                resultado=resultado,
                matches=matches,
                matches_json=matches_json,
            )

        conn = get_connection()
        matches = [buscar_match_conciliacion(conn, fila) for fila in filas]
        conn.close()

        resultado = {
            "filas": len(matches),
            "matches": sum(1 for item in matches if item["estado"] == "match"),
            "sin_match": sum(1 for item in matches if item["estado"] == "sin_match"),
            "multiples": sum(1 for item in matches if item["estado"] == "multiple"),
            "errores": sum(1 for item in matches if item["estado"] == "error"),
        }
        matches_json = json.dumps(matches, ensure_ascii=False, default=str)

    return render_template(
        "conciliacion.html",
        resultado=resultado,
        matches=matches,
        matches_json=matches_json,
    )


@app.route("/jugadores/<int:jugador_id>/lesiones")
def ver_lesiones(jugador_id):
    check = permiso_requerido("salud_ver")
    if check:
        return check

    conn = get_connection()

    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    lesiones = conn.execute("""
        SELECT * FROM lesiones
        WHERE jugador_id = %s
        ORDER BY fecha_lesion DESC, id DESC
    """, (jugador_id,)).fetchall()

    documentos = conn.execute("""
        SELECT *
        FROM lesiones_documentos
        WHERE jugador_id = %s
        ORDER BY id DESC
    """, (jugador_id,)).fetchall()

    conn.close()

    lesiones = [dict(lesion) for lesion in lesiones]
    for lesion in lesiones:
        lesion["semaforo"] = semaforo_lesion(lesion)

    documentos_por_lesion = {}
    for documento in documentos:
        documento = dict(documento)
        documento["preview_tipo"] = obtener_preview_tipo(documento.get("mime_type"))
        documentos_por_lesion.setdefault(documento["lesion_id"], []).append(documento)

    return render_template(
        "lesiones.html",
        jugador=jugador,
        lesiones=lesiones,
        documentos_por_lesion=documentos_por_lesion,
    )

@app.route("/jugadores/<int:jugador_id>/lesiones/nueva", methods=["GET", "POST"])
def nueva_lesion(jugador_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()

    jugador = conn.execute(
        "SELECT * FROM jugadores WHERE id = %s",
        (jugador_id,)
    ).fetchone()

    if request.method == "POST":
        fecha_lesion = request.form.get("fecha_lesion", "").strip()
        tipo_lesion = request.form.get("tipo_lesion", "").strip()
        zona_cuerpo = request.form.get("zona_cuerpo", "").strip()
        diagnostico = request.form.get("diagnostico", "").strip()
        tratamiento = request.form.get("tratamiento", "").strip()
        estado = request.form.get("estado", "").strip() or "Activa"
        etapa_recuperacion = request.form.get("etapa_recuperacion", "").strip()
        proximo_control = request.form.get("proximo_control", "").strip()
        fecha_retorno_estimada = request.form.get("fecha_retorno_estimada", "").strip()
        tratamiento_hasta = request.form.get("tratamiento_hasta", "").strip()
        fecha_alta = request.form.get("fecha_alta", "").strip()
        observaciones = request.form.get("observaciones", "").strip()

        lesion_id = conn.execute("""
            INSERT INTO lesiones (
                jugador_id, fecha_lesion, tipo_lesion, zona_cuerpo,
                diagnostico, tratamiento, estado, etapa_recuperacion,
                proximo_control, fecha_retorno_estimada, tratamiento_hasta,
                fecha_alta, observaciones
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            jugador_id, fecha_lesion, tipo_lesion, zona_cuerpo,
            diagnostico, tratamiento, estado, etapa_recuperacion,
            proximo_control, fecha_retorno_estimada, tratamiento_hasta,
            fecha_alta, observaciones
        )).fetchone()["id"]

        archivos = request.files.getlist("documentos_adjuntos")
        descripcion_adjuntos = request.form.get("descripcion_adjuntos", "").strip()
        try:
            guardar_documentos_lesion(
                conn,
                jugador,
                {"id": lesion_id, "tipo_lesion": tipo_lesion},
                archivos,
                descripcion_adjuntos,
            )
        except (RuntimeError, ValueError) as error:
            conn.rollback()
            conn.close()
            flash(str(error), "error")
            return render_template("lesion_form.html", jugador=jugador, lesion=request.form, documentos=[])
        except Exception as error:
            conn.rollback()
            conn.close()
            app.logger.exception("No se pudieron guardar adjuntos de lesion del jugador %s.", jugador_id)
            flash(
                mensaje_error_drive(
                    error,
                    carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER,
                    accion="guardar adjuntos de la lesion",
                ),
                "error",
            )
            return render_template("lesion_form.html", jugador=jugador, lesion=request.form, documentos=[])

        conn.commit()
        conn.close()

        flash("Lesión cargada correctamente.", "ok")
        return redirect(url_for("ver_lesiones", jugador_id=jugador_id))

    conn.close()
    return render_template("lesion_form.html", jugador=jugador, documentos=[])

@app.route("/lesiones/<int:lesion_id>/editar", methods=["GET", "POST"])
def editar_lesion(lesion_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check
    conn = get_connection()

    lesion = conn.execute("""
        SELECT * FROM lesiones
        WHERE id = %s
    """, (lesion_id,)).fetchone()

    if lesion is None:
        conn.close()
        flash("Lesión no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    jugador = conn.execute("""
        SELECT * FROM jugadores
        WHERE id = %s
    """, (lesion["jugador_id"],)).fetchone()

    documentos = conn.execute("""
        SELECT *
        FROM lesiones_documentos
        WHERE lesion_id = %s
        ORDER BY id DESC
    """, (lesion_id,)).fetchall()

    if request.method == "POST":
        fecha_lesion = request.form.get("fecha_lesion", "").strip()
        tipo_lesion = request.form.get("tipo_lesion", "").strip()
        zona_cuerpo = request.form.get("zona_cuerpo", "").strip()
        diagnostico = request.form.get("diagnostico", "").strip()
        tratamiento = request.form.get("tratamiento", "").strip()
        estado = request.form.get("estado", "").strip() or "Activa"
        etapa_recuperacion = request.form.get("etapa_recuperacion", "").strip()
        proximo_control = request.form.get("proximo_control", "").strip()
        fecha_retorno_estimada = request.form.get("fecha_retorno_estimada", "").strip()
        tratamiento_hasta = request.form.get("tratamiento_hasta", "").strip()
        fecha_alta = request.form.get("fecha_alta", "").strip()
        observaciones = request.form.get("observaciones", "").strip()

        conn.execute("""
            UPDATE lesiones
            SET fecha_lesion = %s, tipo_lesion = %s, zona_cuerpo = %s,
                diagnostico = %s, tratamiento = %s, estado = %s,
                etapa_recuperacion = %s, proximo_control = %s,
                fecha_retorno_estimada = %s, tratamiento_hasta = %s,
                fecha_alta = %s, observaciones = %s
            WHERE id = %s
        """, (
            fecha_lesion, tipo_lesion, zona_cuerpo,
            diagnostico, tratamiento, estado,
            etapa_recuperacion, proximo_control, fecha_retorno_estimada,
            tratamiento_hasta, fecha_alta, observaciones, lesion_id
        ))

        archivos = request.files.getlist("documentos_adjuntos")
        descripcion_adjuntos = request.form.get("descripcion_adjuntos", "").strip()
        try:
            guardar_documentos_lesion(
                conn,
                jugador,
                {"id": lesion_id, "tipo_lesion": tipo_lesion},
                archivos,
                descripcion_adjuntos,
            )
        except (RuntimeError, ValueError) as error:
            conn.rollback()
            conn.close()
            flash(str(error), "error")
            lesion_data = dict(lesion)
            lesion_data.update({
                "fecha_lesion": fecha_lesion,
                "tipo_lesion": tipo_lesion,
                "zona_cuerpo": zona_cuerpo,
                "diagnostico": diagnostico,
                "tratamiento": tratamiento,
                "estado": estado,
                "etapa_recuperacion": etapa_recuperacion,
                "proximo_control": proximo_control,
                "fecha_retorno_estimada": fecha_retorno_estimada,
                "tratamiento_hasta": tratamiento_hasta,
                "fecha_alta": fecha_alta,
                "observaciones": observaciones,
            })
            documentos_preview = [dict(item, preview_tipo=obtener_preview_tipo(item.get("mime_type"))) for item in documentos]
            return render_template("lesion_form.html", jugador=jugador, lesion=lesion_data, documentos=documentos_preview)
        except Exception as error:
            conn.rollback()
            conn.close()
            app.logger.exception("No se pudieron guardar adjuntos de la lesion %s.", lesion_id)
            flash(
                mensaje_error_drive(
                    error,
                    carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER,
                    accion="guardar adjuntos de la lesion",
                ),
                "error",
            )
            lesion_data = dict(lesion)
            lesion_data.update({
                "fecha_lesion": fecha_lesion,
                "tipo_lesion": tipo_lesion,
                "zona_cuerpo": zona_cuerpo,
                "diagnostico": diagnostico,
                "tratamiento": tratamiento,
                "estado": estado,
                "etapa_recuperacion": etapa_recuperacion,
                "proximo_control": proximo_control,
                "fecha_retorno_estimada": fecha_retorno_estimada,
                "tratamiento_hasta": tratamiento_hasta,
                "fecha_alta": fecha_alta,
                "observaciones": observaciones,
            })
            documentos_preview = [dict(item, preview_tipo=obtener_preview_tipo(item.get("mime_type"))) for item in documentos]
            return render_template("lesion_form.html", jugador=jugador, lesion=lesion_data, documentos=documentos_preview)

        conn.commit()
        conn.close()

        flash("Lesión actualizada correctamente.", "ok")
        return redirect(url_for("ver_lesiones", jugador_id=jugador["id"]))

    conn.close()
    documentos = [dict(item, preview_tipo=obtener_preview_tipo(item.get("mime_type"))) for item in documentos]
    return render_template("lesion_form.html", jugador=jugador, lesion=lesion, documentos=documentos)

@app.route("/lesiones/<int:lesion_id>/eliminar", methods=["POST"])
def eliminar_lesion(lesion_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check
    conn = get_connection()

    lesion = conn.execute("""
        SELECT jugador_id FROM lesiones
        WHERE id = %s
    """, (lesion_id,)).fetchone()

    if lesion is None:
        conn.close()
        flash("Lesión no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    documentos = conn.execute("""
        SELECT drive_file_id
        FROM lesiones_documentos
        WHERE lesion_id = %s
    """, (lesion_id,)).fetchall()

    for documento in documentos:
        try:
            eliminar_drive_file(documento["drive_file_id"])
        except Exception:
            app.logger.warning("No se pudo eliminar archivo de Drive de la lesion %s.", lesion_id)

    conn.execute("DELETE FROM lesiones_documentos WHERE lesion_id = %s", (lesion_id,))
    conn.execute("DELETE FROM lesiones WHERE id = %s", (lesion_id,))
    conn.commit()
    conn.close()

    flash("Lesión eliminada.", "ok")
    return redirect(url_for("ver_lesiones", jugador_id=lesion["jugador_id"]))

@app.route("/lesiones/documentos/<int:documento_id>")
def ver_documento_lesion(documento_id):
    check = permiso_requerido("salud_ver")
    if check:
        return check

    conn = get_connection()
    documento = conn.execute("""
        SELECT *
        FROM lesiones_documentos
        WHERE id = %s
    """, (documento_id,)).fetchone()
    conn.close()

    if documento is None:
        flash("Adjunto de lesion no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    try:
        archivo = descargar_drive_file(documento["drive_file_id"])
    except RuntimeError as error:
        flash(str(error), "error")
        return redirect(url_for("ver_lesiones", jugador_id=documento["jugador_id"]))
    except Exception as error:
        app.logger.exception("No se pudo descargar adjunto de lesion %s.", documento_id)
        flash(
            mensaje_error_drive(
                error,
                carpeta=DRIVE_FICHAS_MEDICAS_SUBFOLDER,
                accion="descargar el adjunto de la lesion",
            ),
            "error",
        )
        return redirect(url_for("ver_lesiones", jugador_id=documento["jugador_id"]))

    return send_file(
        archivo,
        mimetype=documento["mime_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=documento["nombre"] or f"lesion_documento_{documento_id}",
    )


@app.route("/lesiones/documentos/<int:documento_id>/eliminar", methods=["POST"])
def eliminar_documento_lesion(documento_id):
    check = permiso_requerido("salud_gestionar")
    if check:
        return check

    conn = get_connection()
    documento = conn.execute("""
        SELECT *
        FROM lesiones_documentos
        WHERE id = %s
    """, (documento_id,)).fetchone()
    if documento is None:
        conn.close()
        flash("Adjunto de lesion no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    try:
        eliminar_drive_file(documento["drive_file_id"])
    except Exception:
        app.logger.warning("No se pudo eliminar archivo de Drive del documento de lesion %s.", documento_id)

    conn.execute("DELETE FROM lesiones_documentos WHERE id = %s", (documento_id,))
    conn.commit()
    conn.close()

    flash("Adjunto eliminado.", "ok")
    return redirect(url_for("editar_lesion", lesion_id=documento["lesion_id"]))


@app.route("/cuotas/generar", methods=["GET", "POST"])
def generar_cuotas():
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check
    if request.method == "POST":
        periodo = request.form.get("periodo", "").strip()
        importe = request.form.get("importe", "").strip()
        fecha_vencimiento = request.form.get("fecha_vencimiento", "").strip()
        categoria = request.form.get("categoria", "").strip()

        if not periodo or not importe:
            flash("Período e importe son obligatorios.", "error")
            return render_template("generar_cuotas.html")

        try:
            importe_valor = float(importe)
        except ValueError:
            flash("El importe debe ser numérico.", "error")
            return render_template(
                "generar_cuotas.html",
                periodo=periodo,
                importe=importe,
                fecha_vencimiento=fecha_vencimiento,
                categoria=categoria
            )

        conn = get_connection()

        if categoria:
            jugadores = conn.execute("""
                SELECT *
                FROM jugadores
                WHERE estado = 'Activo'
                  AND COALESCE(cobra_cuota, 1) = 1
                  AND categoria = %s
                ORDER BY apellido, nombre
            """, (categoria,)).fetchall()
        else:
            jugadores = conn.execute("""
                SELECT *
                FROM jugadores
                WHERE estado = 'Activo'
                  AND COALESCE(cobra_cuota, 1) = 1
                ORDER BY apellido, nombre
            """).fetchall()

        creadas = 0
        omitidas = 0
        becadas_totales = 0
        becadas_parciales = 0

        for jugador in jugadores:
            existente = conn.execute("""
                SELECT id
                FROM cuotas
                WHERE jugador_id = %s AND periodo = %s
            """, (jugador["id"], periodo)).fetchone()

            if existente:
                omitidas += 1
                continue

            cuota_calculada = calcular_importe_cuota_mensual(conn, jugador, periodo, importe_valor)
            pagado_inicial = 1 if cuota_calculada["beca_total"] else 0
            fecha_pago_inicial = ahora_sig().strftime("%Y-%m-%d") if pagado_inicial else None
            metodo_inicial = "Beca" if pagado_inicial else None
            referencia_inicial = (
                f"Beca total {cuota_calculada['beca_porcentaje']:g}%"
                if pagado_inicial else None
            )

            conn.execute("""
                INSERT INTO cuotas (
                    jugador_id, periodo, importe, pagado, fecha_pago, fecha_vencimiento,
                    importe_original, descuento_beca, beca_porcentaje, beca_motivo,
                    becada, metodo_pago, referencia_pago, plan_pago_monto, plan_pago_detalle
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                jugador["id"],
                periodo,
                cuota_calculada["importe"],
                pagado_inicial,
                fecha_pago_inicial,
                fecha_vencimiento or None,
                cuota_calculada["importe_original"],
                cuota_calculada["descuento_beca"],
                cuota_calculada["beca_porcentaje"],
                cuota_calculada["beca_motivo"],
                cuota_calculada["becada"],
                metodo_inicial,
                referencia_inicial,
                cuota_calculada["plan_pago_monto"],
                cuota_calculada["plan_pago_detalle"] or None,
            ))
            creadas += 1
            if cuota_calculada["beca_total"]:
                becadas_totales += 1
            elif cuota_calculada["becada"]:
                becadas_parciales += 1

        conn.commit()
        conn.close()

        flash(
            f"Generación terminada. Cuotas creadas: {creadas}. Becas totales: {becadas_totales}. "
            f"Becas parciales: {becadas_parciales}. Ya existentes: {omitidas}.",
            "ok",
        )
        return redirect(url_for("index"))

    return render_template("generar_cuotas.html")


@app.route("/jugadores/<int:jugador_id>/timeline")
def timeline_jugador(jugador_id):
    check = permiso_requerido("jugadores_ver")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    eventos = []

    if jugador["fecha_ingreso"]:
        eventos.append({
            "fecha": jugador["fecha_ingreso"],
            "tipo": "Alta",
            "titulo": "Ingreso al club",
            "detalle": jugador["estado"],
        })

    if tiene_permiso("cuotas_ver"):
        cuotas = conn.execute("""
            SELECT *
            FROM cuotas
            WHERE jugador_id = %s
            ORDER BY periodo DESC, id DESC
            LIMIT 80
        """, (jugador_id,)).fetchall()
        for cuota in cuotas:
            eventos.append({
                "fecha": cuota["fecha_pago"] or cuota["fecha_vencimiento"] or f"{cuota['periodo']}-01",
                "tipo": "Cuota",
                "titulo": f"Cuota {cuota['periodo']}",
                "detalle": (
                    f"{'Pagada' if cuota['pagado'] else 'Pendiente'} - "
                    f"{formato_moneda(cuota['importe'])}"
                    + (f" - Beca {cuota['beca_porcentaje']}%" if cuota.get("becada") else "")
                ),
            })

        becas = conn.execute("""
            SELECT *
            FROM becas_historial
            WHERE jugador_id = %s
            ORDER BY creado_en DESC
            LIMIT 40
        """, (jugador_id,)).fetchall()
        for beca in becas:
            eventos.append({
                "fecha": str(beca["creado_en"])[:10],
                "tipo": "Beca",
                "titulo": beca["accion"],
                "detalle": f"{beca['beca_porcentaje'] or 0}% - {beca['beca_motivo'] or '-'}",
            })

    if tiene_permiso("salud_ver"):
        ficha = conn.execute("SELECT * FROM fichas_medicas WHERE jugador_id = %s", (jugador_id,)).fetchone()
        if ficha:
            eventos.append({
                "fecha": ficha["fecha_vencimiento"] or str(ficha["id"]),
                "tipo": "Ficha médica",
                "titulo": "Ficha médica",
                "detalle": f"Vencimiento: {ficha['fecha_vencimiento'] or '-'}",
            })

        lesiones = conn.execute("""
            SELECT *
            FROM lesiones
            WHERE jugador_id = %s
            ORDER BY fecha_lesion DESC, id DESC
            LIMIT 50
        """, (jugador_id,)).fetchall()
        for lesion in lesiones:
            eventos.append({
                "fecha": lesion["fecha_lesion"] or lesion["fecha_alta"] or "",
                "tipo": "Lesion",
                "titulo": lesion["tipo_lesion"] or "Lesion",
                "detalle": f"{lesion['estado']} - {lesion['zona_cuerpo'] or '-'}",
            })

        documentos = conn.execute("""
            SELECT *
            FROM documentos_jugadores
            WHERE jugador_id = %s
            ORDER BY COALESCE(fecha_presentacion, fecha_vencimiento) DESC, id DESC
            LIMIT 50
        """, (jugador_id,)).fetchall()
        for documento in documentos:
            eventos.append({
                "fecha": documento["fecha_presentacion"] or documento["fecha_vencimiento"] or "",
                "tipo": "Documento",
                "titulo": documento["tipo"],
                "detalle": f"Vence: {documento['fecha_vencimiento'] or '-'}",
            })

    if tiene_permiso("asistencia_ver"):
        asistencias = conn.execute("""
            SELECT a.*, e.fecha, e.tipo, e.descripcion
            FROM asistencias a
            JOIN eventos_asistencia e ON e.id = a.evento_id
            WHERE a.jugador_id = %s
            ORDER BY e.fecha DESC, e.id DESC
            LIMIT 80
        """, (jugador_id,)).fetchall()
        for asistencia in asistencias:
            eventos.append({
                "fecha": asistencia["fecha"],
                "tipo": "Asistencia",
                "titulo": asistencia["tipo"],
                "detalle": asistencia["estado_asistencia"] or ("presente" if asistencia["presente"] else "ausente"),
            })

    bitacora = conn.execute("""
        SELECT *
        FROM jugador_bitacora
        WHERE jugador_id = %s
        ORDER BY creado_en DESC, id DESC
        LIMIT 80
    """, (jugador_id,)).fetchall()
    for nota in filtrar_bitacora_visible(bitacora):
        eventos.append({
            "fecha": str(nota["creado_en"])[:10],
            "tipo": f"Bitacora {BITACORA_TIPOS.get(nota['tipo'], nota['tipo'])}",
            "titulo": nota["creado_por"] or "SIG",
            "detalle": nota["nota"],
        })

    if tiene_permiso("comunicaciones_ver"):
        mensajes = conn.execute("""
            SELECT direccion, tipo, texto, creado_en, estado
            FROM whatsapp_mensajes
            WHERE jugador_id = %s
            ORDER BY creado_en DESC, id DESC
            LIMIT 40
        """, (jugador_id,)).fetchall()
        for mensaje in mensajes:
            eventos.append({
                "fecha": str(mensaje["creado_en"])[:10],
                "tipo": "WhatsApp",
                "titulo": "Entrante" if mensaje["direccion"] == "in" else "Saliente",
                "detalle": mensaje["texto"] or f"[{mensaje['tipo']}] {mensaje['estado'] or ''}".strip(),
            })

    tareas = conn.execute("""
        SELECT titulo, descripcion, estado, prioridad, creado_en
        FROM tareas_sig
        WHERE jugador_id = %s
        ORDER BY creado_en DESC, id DESC
        LIMIT 40
    """, (jugador_id,)).fetchall()
    for tarea in tareas:
        eventos.append({
            "fecha": str(tarea["creado_en"])[:10],
            "tipo": "Tarea",
            "titulo": f"{tarea['titulo']} ({tarea['estado']})",
            "detalle": f"{tarea['prioridad']} - {tarea['descripcion'] or '-'}",
        })

    conn.close()
    eventos.sort(key=lambda item: item["fecha"] or "", reverse=True)

    return render_template("timeline_jugador.html", jugador=jugador, eventos=eventos)


@app.route("/jugadores/<int:jugador_id>")
def detalle_jugador(jugador_id):
    check = permiso_requerido("jugadores_ver")
    if check:
        return check

    puede_ver_finanzas = tiene_permiso("cuotas_ver")
    puede_ver_salud = tiene_permiso("salud_ver")
    puede_ver_asistencia = tiene_permiso("asistencia_ver")
    puede_ver_tests = tiene_permiso("tests_ver")
    puede_gestionar_portal = tiene_permiso("portal_jugador_gestionar")
    puede_ver_cambios_portal = tiene_permiso("alertas_portal", "auditoria_ver", "portal_jugador_gestionar")

    conn = get_connection()

    jugador = conn.execute("""
        SELECT * FROM jugadores
        WHERE id = %s
    """, (jugador_id,)).fetchone()

    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    es_jugador_deportivo = (jugador.get("tipo_miembro") or "Jugador") == "Jugador"
    puede_ver_salud = puede_ver_salud and es_jugador_deportivo
    puede_ver_asistencia = puede_ver_asistencia and es_jugador_deportivo
    puede_ver_tests = puede_ver_tests and es_jugador_deportivo

    deuda_cuotas = conn.execute("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(importe, 0) > 0
    """, (jugador_id,)).fetchone()["total"]

    gastos_compartidos_pendientes = conn.execute("""
        SELECT
            i.id,
            i.importe,
            i.estado,
            g.id AS gasto_id,
            g.titulo,
            g.concepto,
            g.fecha_evento,
            g.fecha_vencimiento,
            g.estado AS gasto_estado
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        WHERE i.jugador_id = %s
          AND i.estado = 'pendiente'
          AND COALESCE(i.importe, 0) > 0
        ORDER BY
            COALESCE(g.fecha_vencimiento, g.fecha_evento, g.creado_en::text) DESC,
            i.id DESC
    """, (jugador_id,)).fetchall()
    deuda_gastos_compartidos = round(
        sum(float(item.get("importe") or 0) for item in gastos_compartidos_pendientes),
        2,
    )
    deuda = round(float(deuda_cuotas or 0) + deuda_gastos_compartidos, 2)

    resumen_cuotas = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(CASE WHEN pagado = 1 THEN 1 END) AS pagadas,
            COUNT(CASE WHEN pagado = 0 AND COALESCE(importe, 0) > 0 THEN 1 END) AS pendientes,
            COALESCE(SUM(CASE WHEN pagado = 1 THEN importe ELSE 0 END), 0) AS total_pagado,
            COALESCE(SUM(CASE WHEN pagado = 0 AND COALESCE(importe, 0) > 0 THEN importe ELSE 0 END), 0) AS total_pendiente
        FROM cuotas
        WHERE jugador_id = %s
    """, (jugador_id,)).fetchone()

    ultimas_cuotas = conn.execute("""
        SELECT *
        FROM cuotas
        WHERE jugador_id = %s
        ORDER BY periodo DESC, id DESC
        LIMIT 5
    """, (jugador_id,)).fetchall()

    ficha = conn.execute("""
        SELECT *
        FROM fichas_medicas
        WHERE jugador_id = %s
    """, (jugador_id,)).fetchone()

    lesiones = conn.execute("""
        SELECT *
        FROM lesiones
        WHERE jugador_id = %s
        ORDER BY
            CASE
                WHEN estado = 'Activa' THEN 0
                WHEN estado = 'En recuperación' THEN 1
                ELSE 2
            END,
            fecha_lesion DESC,
            id DESC
        LIMIT 5
    """, (jugador_id,)).fetchall()

    ficha_vencida = False
    if ficha and validar_fecha_movimiento(ficha["fecha_vencimiento"]):
        check_vencimiento = conn.execute("""
            SELECT CASE
                WHEN %s::date < CURRENT_DATE THEN 1
                ELSE 0
            END AS vencida
        """, (ficha["fecha_vencimiento"],)).fetchone()
        ficha_vencida = bool(check_vencimiento["vencida"])

    lesiones_activas_count = conn.execute("""
        SELECT COUNT(*) AS total
        FROM lesiones
        WHERE jugador_id = %s
          AND estado IN ('Activa', 'En recuperación')
    """, (jugador_id,)).fetchone()["total"]

    cuotas_pendientes_count = conn.execute("""
        SELECT COUNT(*) AS total
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(importe, 0) > 0
    """, (jugador_id,)).fetchone()["total"]

    asistencia_resumen = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COALESCE(SUM(CASE WHEN a.presente = 1 THEN 1 ELSE 0 END), 0) AS presentes,
            COALESCE(SUM(CASE WHEN a.presente = 0 THEN 1 ELSE 0 END), 0) AS ausentes
        FROM asistencias a
        JOIN eventos_asistencia e ON e.id = a.evento_id
        WHERE a.jugador_id = %s
    """, (jugador_id,)).fetchone()

    ultimas_asistencias = conn.execute("""
        SELECT
            e.fecha,
            e.tipo,
            e.descripcion,
            a.presente,
            a.estado_asistencia,
            a.observaciones
        FROM asistencias a
        JOIN eventos_asistencia e ON e.id = a.evento_id
        WHERE a.jugador_id = %s
        ORDER BY e.fecha DESC, e.id DESC
        LIMIT 8
    """, (jugador_id,)).fetchall()

    bienestar_reciente = conn.execute("""
        SELECT
            p.*,
            e.fecha,
            e.tipo,
            e.descripcion
        FROM portal_asistencia_confirmaciones p
        JOIN eventos_asistencia e ON e.id = p.evento_id
        WHERE p.jugador_id = %s
          AND p.sueno_calidad IS NOT NULL
        ORDER BY e.fecha DESC, p.actualizado_en DESC
        LIMIT 8
    """, (jugador_id,)).fetchall()

    documentos_manual = conn.execute("""
        SELECT *
        FROM documentos_jugadores
        WHERE jugador_id = %s
        ORDER BY COALESCE(fecha_vencimiento, fecha_presentacion) DESC NULLS LAST, id DESC
    """, (jugador_id,)).fetchall()

    planes_pago = conn.execute("""
        SELECT *
        FROM planes_pago
        WHERE jugador_id = %s
        ORDER BY
            CASE WHEN estado = 'Activo' THEN 0 ELSE 1 END,
            fecha_inicio DESC,
            id DESC
        LIMIT 5
    """, (jugador_id,)).fetchall()

    bitacora_raw = conn.execute("""
        SELECT *
        FROM jugador_bitacora
        WHERE jugador_id = %s
        ORDER BY creado_en DESC, id DESC
        LIMIT 40
    """, (jugador_id,)).fetchall()

    tests_recientes = []
    if puede_ver_tests:
        tests_recientes = conn.execute("""
            SELECT
                r.*,
                t.nombre AS test_nombre,
                t.unidad
            FROM test_resultados r
            JOIN test_tipos t ON t.id = r.test_id
            WHERE r.jugador_id = %s
            ORDER BY r.fecha DESC, r.id DESC
            LIMIT 8
        """, (jugador_id,)).fetchall()

    proximo_evento = None
    eventos_candidatos = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE fecha >= CURRENT_DATE::text
        ORDER BY fecha ASC, COALESCE(hora_inicio, '') ASC, id ASC
        LIMIT 40
    """).fetchall()
    for evento in eventos_candidatos:
        if categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria")):
            proximo_evento = evento
            break

    plan_activo = next((plan for plan in planes_pago if (plan.get("estado") or "").strip() == "Activo"), None)
    ultimo_test_resumen = tests_recientes[0] if tests_recientes else None
    documento_alertas = conn.execute("""
        SELECT COUNT(*) AS total
        FROM documentos_jugadores
        WHERE jugador_id = %s
          AND fecha_vencimiento IS NOT NULL
          AND NULLIF(fecha_vencimiento::text, '') IS NOT NULL
          AND fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
          AND fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days'
    """, (jugador_id,)).fetchone()["total"]

    ficha_estado = "Sin ficha"
    if ficha and validar_fecha_movimiento(ficha.get("fecha_vencimiento")):
        if ficha_vencida:
            ficha_estado = "Vencida"
        else:
            ficha_estado = "Vigente"
    elif ficha and ficha.get("presentada"):
        ficha_estado = "Presentada"

    ficha_apto_efectivo = ficha_tiene_apto_efectivo(ficha)

    cambios_portal = []
    if puede_ver_cambios_portal:
        cambios_portal = conn.execute("""
            SELECT fecha, detalle
            FROM auditoria
            WHERE entidad = 'portal_jugador'
              AND entidad_id = %s
              AND accion = 'actualizar_contacto'
            ORDER BY fecha DESC, id DESC
            LIMIT 8
        """, (str(jugador_id),)).fetchall()
        cambios_portal = [dict(cambio) for cambio in cambios_portal]
        for cambio in cambios_portal:
            cambio["detalle_resumen"] = resumen_auditoria_portal(cambio.get("detalle"))

    cuenta_corriente = obtener_cuenta_corriente_jugador(conn, jugador_id, limite=20)
    notificaciones_portal = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COALESCE(SUM(CASE WHEN enabled = 1 THEN 1 ELSE 0 END), 0) AS activas,
            MAX(CASE WHEN enabled = 1 THEN actualizado_en ELSE NULL END) AS ultima_activacion
        FROM pwa_push_subscriptions
        WHERE actor_tipo = 'portal'
          AND jugador_id = %s
    """, (jugador_id,)).fetchone()

    for item in bienestar_reciente:
        try:
            item["dolor_zonas_lista"] = json.loads(item.get("dolor_zonas") or "[]")
        except (TypeError, ValueError):
            item["dolor_zonas_lista"] = []
        item["bienestar_resumen"] = resumen_bienestar_confirmacion(item)

    conn.close()

    asistencia_total = asistencia_resumen["total"] or 0
    asistencia_presentes = asistencia_resumen["presentes"] or 0
    asistencia_porcentaje = (
        round((asistencia_presentes / asistencia_total) * 100, 1)
        if asistencia_total else 0
    )

    documentos = []
    if jugador["documentos"]:
        documentos = [
            linea.strip()
            for linea in jugador["documentos"].splitlines()
            if linea.strip()
        ]

    portal_url = None
    if jugador.get("portal_activo") and jugador.get("portal_token"):
        portal_url = url_for("portal_jugador", token=jugador["portal_token"], _external=True)

    bitacora = filtrar_bitacora_visible(bitacora_raw)

    return render_template(
        "jugador_detalle.html",
        jugador=jugador,
        deuda=deuda,
        deuda_cuotas=deuda_cuotas,
        deuda_gastos_compartidos=deuda_gastos_compartidos,
        gastos_compartidos_pendientes=gastos_compartidos_pendientes,
        resumen_cuotas=resumen_cuotas,
        ultimas_cuotas=ultimas_cuotas,
        ficha=ficha,
        ficha_apto_efectivo=ficha_apto_efectivo,
        lesiones=lesiones,
        ficha_vencida=ficha_vencida,
        lesiones_activas_count=lesiones_activas_count,
        cuotas_pendientes_count=cuotas_pendientes_count,
        asistencia_resumen=asistencia_resumen,
        asistencia_porcentaje=asistencia_porcentaje,
        ultimas_asistencias=ultimas_asistencias,
        documentos=documentos,
        documentos_manual=documentos_manual,
        planes_pago=planes_pago,
        bitacora=bitacora,
        bitacora_tipos=tipos_bitacora_disponibles(),
        bitacora_labels=BITACORA_TIPOS,
        portal_url=portal_url,
        puede_ver_finanzas=puede_ver_finanzas,
        puede_ver_salud=puede_ver_salud,
        puede_ver_asistencia=puede_ver_asistencia,
        puede_ver_tests=puede_ver_tests,
        tests_recientes=tests_recientes,
        proximo_evento=proximo_evento,
        plan_activo=plan_activo,
        ultimo_test_resumen=ultimo_test_resumen,
        documento_alertas=documento_alertas,
        ficha_estado=ficha_estado,
        puede_gestionar_portal=puede_gestionar_portal,
        puede_ver_cambios_portal=puede_ver_cambios_portal,
        cambios_portal=cambios_portal,
        cuenta_corriente=cuenta_corriente,
        notificaciones_portal=notificaciones_portal,
        )


@app.route("/jugadores/<int:jugador_id>/bitacora", methods=["POST"])
def nueva_bitacora_jugador(jugador_id):
    tipos_disponibles = {item["clave"] for item in tipos_bitacora_disponibles()}
    if not tipos_disponibles:
        flash("No tenes permiso para agregar notas en la bitacora.", "error")
        return redirect(url_for("detalle_jugador", jugador_id=jugador_id))

    tipo = request.form.get("tipo", "general").strip()
    nota = request.form.get("nota", "").strip()

    if tipo not in tipos_disponibles:
        flash("No tenes permiso para ese tipo de nota.", "error")
        return redirect(url_for("detalle_jugador", jugador_id=jugador_id))

    if not nota:
        flash("La nota de bitacora no puede estar vacia.", "error")
        return redirect(url_for("detalle_jugador", jugador_id=jugador_id))

    conn = get_connection()
    jugador = conn.execute("SELECT id FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    conn.execute("""
        INSERT INTO jugador_bitacora (jugador_id, tipo, nota, creado_por)
        VALUES (%s, %s, %s, %s)
    """, (jugador_id, tipo, nota, session.get("username")))
    conn.commit()
    conn.close()

    flash("Nota agregada a la bitacora.", "ok")
    return redirect(url_for("detalle_jugador", jugador_id=jugador_id))


@app.route("/jugadores/<int:jugador_id>/portal/generar", methods=["POST"])
def generar_portal_jugador(jugador_id):
    check = permiso_requerido("portal_jugador_gestionar")
    if check:
        return check

    conn = get_connection()
    jugador = conn.execute("SELECT * FROM jugadores WHERE id = %s", (jugador_id,)).fetchone()
    if jugador is None:
        conn.close()
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("listar_jugadores"))

    token = jugador.get("portal_token") or generar_portal_token()
    conn.execute("""
        UPDATE jugadores
        SET portal_token = %s,
            portal_activo = 1,
            portal_actualizado_en = %s
        WHERE id = %s
    """, (token, ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), jugador_id))
    conn.commit()
    conn.close()

    flash("Portal externo habilitado para el jugador.", "ok")
    return redirect(url_for("detalle_jugador", jugador_id=jugador_id))


@app.route("/jugadores/<int:jugador_id>/portal/desactivar", methods=["POST"])
def desactivar_portal_jugador(jugador_id):
    check = permiso_requerido("portal_jugador_gestionar")
    if check:
        return check

    conn = get_connection()
    conn.execute("""
        UPDATE jugadores
        SET portal_activo = 0,
            portal_actualizado_en = %s
        WHERE id = %s
    """, (ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), jugador_id))
    conn.commit()
    conn.close()

    flash("Portal externo desactivado.", "ok")
    return redirect(url_for("detalle_jugador", jugador_id=jugador_id))


def opciones_eventos_gastos_compartidos(conn):
    eventos = conn.execute("""
        SELECT
            id,
            fecha,
            titulo,
            categoria,
            asistencia_evento_id
        FROM calendario_eventos
        WHERE fecha IS NOT NULL
        ORDER BY fecha DESC, id DESC
        LIMIT 80
    """).fetchall()
    opciones = []
    for evento in eventos:
        etiqueta = f"{evento['fecha'] or '-'} · {evento['titulo']}"
        if evento.get("categoria"):
            etiqueta = f"{etiqueta} ({evento['categoria']})"
        opciones.append({
            "id": evento["id"],
            "label": etiqueta,
            "fecha": evento["fecha"],
            "asistencia_evento_id": evento.get("asistencia_evento_id"),
        })
    return opciones


def jugadores_gasto_desde_fuente(conn, fuente, calendario_evento_id, jugadores_ids):
    if fuente == "manual":
        ids = [int(valor) for valor in jugadores_ids if str(valor).isdigit()]
        if not ids:
            return []
        return conn.execute("""
            SELECT id, apellido, nombre, categoria
            FROM jugadores
            WHERE id = ANY(%s)
            ORDER BY apellido, nombre
        """, (ids,)).fetchall()

    if not calendario_evento_id:
        return []

    evento = conn.execute("""
        SELECT id, asistencia_evento_id
        FROM calendario_eventos
        WHERE id = %s
    """, (calendario_evento_id,)).fetchone()
    if not evento or not evento.get("asistencia_evento_id"):
        return []

    if fuente == "presentes_evento":
        return conn.execute("""
            SELECT j.id, j.apellido, j.nombre, j.categoria
            FROM asistencias a
            JOIN jugadores j ON j.id = a.jugador_id
            WHERE a.evento_id = %s
              AND a.presente = 1
            ORDER BY j.apellido, j.nombre
        """, (evento["asistencia_evento_id"],)).fetchall()

    if fuente == "confirmados_portal":
        return conn.execute("""
            SELECT j.id, j.apellido, j.nombre, j.categoria
            FROM portal_asistencia_confirmaciones p
            JOIN jugadores j ON j.id = p.jugador_id
            WHERE p.evento_id = %s
              AND p.estado = 'confirmado'
            ORDER BY j.apellido, j.nombre
        """, (evento["asistencia_evento_id"],)).fetchall()

    return []


def construir_texto_gasto_compartido(item):
    nombre = nombre_jugador_corto(item)
    titulo = item.get("titulo") or "gasto compartido"
    importe = formato_moneda(item.get("importe") or 0)
    vencimiento = item.get("fecha_vencimiento") or "sin vencimiento"
    concepto = item.get("concepto") or "Sin detalle adicional."
    return (
        f"Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        f"Tenes pendiente el gasto compartido \"{titulo}\" por {importe}. "
        f"Vencimiento: {vencimiento}. "
        f"Detalle: {concepto}"
    )


def gasto_compartido_esta_cerrado(gasto):
    return (gasto.get("estado") or "").strip().lower() == "cerrado"


class UrbaCircularesParser(HTMLParser):
    def __init__(self, base_url):
        super().__init__()
        self.base_url = base_url
        self.current_href = None
        self.current_text = []
        self.items = []
        self.seen = set()

    def handle_starttag(self, tag, attrs):
        if tag.lower() != "a":
            return
        attrs_map = dict(attrs)
        self.current_href = attrs_map.get("href")
        self.current_text = []

    def handle_data(self, data):
        if self.current_href is not None:
            self.current_text.append(data)

    def handle_endtag(self, tag):
        if tag.lower() != "a" or self.current_href is None:
            return
        titulo = " ".join(parte.strip() for parte in self.current_text if parte and parte.strip()).strip()
        href = self.current_href.strip()
        self.current_href = None
        self.current_text = []
        if not titulo or titulo.lower() == "image":
            return
        url = urljoin(self.base_url, href)
        if "azureedge.net" not in url and not url.lower().endswith(".pdf"):
            return
        clave = (titulo, url)
        if clave in self.seen:
            return
        self.seen.add(clave)
        self.items.append({
            "titulo": html.unescape(titulo),
            "url": url,
        })


def urba_circulares_url_anio(anio):
    anio = int(anio)
    actual = ahora_sig().year
    if anio == actual:
        return f"https://urba.org.ar/circulares-{anio}"
    return f"https://urba.org.ar/circulares-{anio}-copy"


def anios_circulares_urba():
    actual = ahora_sig().year
    return list(range(actual, 2001, -1))


def obtener_config_circulares_urba(conn=None):
    own_conn = conn is None
    conn = conn or get_connection()
    rows = conn.execute("""
        SELECT clave, valor, actualizado_en, actualizado_por
        FROM app_settings
        WHERE clave IN (
            'urba_circulares_notify_user_ids',
            'urba_circulares_sync_en',
            'urba_circulares_sync_por'
        )
    """).fetchall()
    if own_conn:
        conn.close()
    data = {row["clave"]: row for row in rows}
    try:
        notify_user_ids = json.loads((data.get("urba_circulares_notify_user_ids") or {}).get("valor") or "[]")
    except (TypeError, ValueError):
        notify_user_ids = []
    notify_user_ids = [int(valor) for valor in notify_user_ids if str(valor).isdigit()]
    return {
        "notify_user_ids": notify_user_ids,
        "sync_en": (data.get("urba_circulares_sync_en") or {}).get("valor"),
        "sync_por": (data.get("urba_circulares_sync_por") or {}).get("valor"),
        "actualizado_en": (data.get("urba_circulares_notify_user_ids") or {}).get("actualizado_en"),
        "actualizado_por": (data.get("urba_circulares_notify_user_ids") or {}).get("actualizado_por"),
    }


def usuarios_notificar_circulares_urba(conn, user_ids):
    ids = [int(valor) for valor in user_ids if str(valor).isdigit()]
    if not ids:
        return []
    return conn.execute("""
        SELECT id, username, email, rol
        FROM usuarios
        WHERE id = ANY(%s)
          AND email IS NOT NULL
          AND trim(email) <> ''
        ORDER BY username
    """, (ids,)).fetchall()


def sincronizar_circulares_urba(conn, anio, usuario=None):
    url = urba_circulares_url_anio(anio)
    req = UrlRequest(url, headers={"User-Agent": "SIG-RMR/1.0 (+https://sig.rudamachorugby.com)"})
    try:
        with urlopen(req, timeout=20) as resp:
            html_text = resp.read().decode("utf-8", errors="ignore")
    except HTTPError as error:
        raise RuntimeError(f"URBA devolvio {error.code} al consultar {url}.")
    except URLError as error:
        raise RuntimeError(f"No se pudo conectar con URBA: {error.reason}.")
    except TimeoutError:
        raise RuntimeError("URBA no respondio dentro del tiempo de espera. Probá sincronizar nuevamente en unos minutos.")
    except OSError as error:
        raise RuntimeError(f"No se pudo consultar URBA: {error}.")

    parser = UrbaCircularesParser(url)
    parser.feed(html_text)
    items = parser.items
    nuevas = []
    conn.execute("UPDATE urba_circulares SET nueva = 0 WHERE anio = %s", (anio,))
    for orden, item in enumerate(items, start=1):
        existente = conn.execute("""
            SELECT id
            FROM urba_circulares
            WHERE anio = %s
              AND url = %s
            LIMIT 1
        """, (anio, item["url"])).fetchone()
        if existente:
            conn.execute("""
                UPDATE urba_circulares
                SET titulo = %s,
                    origen_url = %s,
                    orden_fuente = %s,
                    actualizada_en = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (item["titulo"], url, orden, existente["id"]))
            continue
        fila = conn.execute("""
            INSERT INTO urba_circulares (
                anio, titulo, url, origen_url, orden_fuente, nueva,
                detectada_en, actualizada_en
            )
            VALUES (%s, %s, %s, %s, %s, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id, anio, titulo, url
        """, (anio, item["titulo"], item["url"], url, orden)).fetchone()
        nuevas.append(fila)

    guardar_app_setting(conn, "urba_circulares_sync_en", ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), usuario)
    guardar_app_setting(conn, "urba_circulares_sync_por", usuario or "", usuario)
    return {
        "anio": anio,
        "url": url,
        "total": len(items),
        "nuevas": nuevas,
    }


def enviar_notificacion_circulares_urba(conn, nuevas, usuario=None):
    config = obtener_config_circulares_urba(conn)
    destinatarios = usuarios_notificar_circulares_urba(conn, config["notify_user_ids"])
    if not nuevas or not destinatarios:
        return []
    asunto = f"Nuevas circulares URBA {nuevas[0]['anio']}"
    lineas = [
        "Se detectaron nuevas circulares en URBA:",
        "",
    ]
    for item in nuevas[:20]:
        lineas.append(f"- {item['titulo']}: {item['url']}")
    if len(nuevas) > 20:
        lineas.append("")
        lineas.append(f"Y {len(nuevas) - 20} circular(es) mas.")
    lineas.append("")
    lineas.append("Consulta completa en SIG > URBA.")
    cuerpo = "\n".join(lineas)
    resultados = []
    notificados_ids = []
    for destinatario in destinatarios:
        enviado, motivo = enviar_email(destinatario["email"], asunto, cuerpo)
        resultados.append((destinatario, enviado, motivo))
        if enviado:
            notificados_ids.append(destinatario["id"])
    if notificados_ids:
        ids_circulares = [item["id"] for item in nuevas]
        conn.execute("""
            UPDATE urba_circulares
            SET notificada_en = CURRENT_TIMESTAMP
            WHERE id = ANY(%s)
        """, (ids_circulares,))
    return resultados


@app.route("/gastos-compartidos")
def listar_gastos_compartidos():
    check = permiso_requerido("cuotas_ver")
    if check:
        return check

    estado_filtro = (request.args.get("estado") or "todos").strip().lower()
    categoria_filtro = (request.args.get("categoria") or "").strip()
    conn = get_connection()
    gastos = conn.execute("""
        SELECT
            g.*,
            COUNT(i.id) AS jugadores_total,
            COALESCE(SUM(i.importe), 0) AS total_asignado,
            COUNT(*) FILTER (WHERE i.estado = 'pagado') AS pagados,
            COUNT(*) FILTER (WHERE i.estado = 'exento') AS exentos,
            COUNT(*) FILTER (WHERE i.estado = 'pendiente') AS pendientes,
            STRING_AGG(DISTINCT COALESCE(NULLIF(j.categoria, ''), 'Sin categoria'), ', ') AS categorias,
            COUNT(*) FILTER (
                WHERE COALESCE(NULLIF(i.comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'rechazado')
                  AND i.estado = 'pendiente'
            ) AS comprobantes_revision
        FROM gastos_compartidos g
        LEFT JOIN gasto_compartido_items i ON i.gasto_id = g.id
        LEFT JOIN jugadores j ON j.id = i.jugador_id
        GROUP BY g.id
        ORDER BY COALESCE(g.fecha_evento::timestamp, g.fecha_vencimiento::timestamp, g.creado_en) DESC, g.id DESC
    """).fetchall()
    categorias_disponibles = sorted({
        categoria.strip()
        for gasto in gastos
        for categoria in (gasto.get("categorias") or "").split(",")
        if categoria.strip()
    })
    gastos_filtrados = []
    for gasto in gastos:
        if estado_filtro == "activos" and (gasto.get("estado") or "").lower() != "activo":
            continue
        if estado_filtro == "con_pendientes" and not (gasto.get("pendientes") or 0):
            continue
        if estado_filtro == "con_revision" and not (gasto.get("comprobantes_revision") or 0):
            continue
        if categoria_filtro:
            categorias_gasto = {(valor or "").strip() for valor in (gasto.get("categorias") or "").split(",")}
            if categoria_filtro not in categorias_gasto:
                continue
        gastos_filtrados.append(gasto)
    resumen = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE estado = 'Activo') AS activos,
            COALESCE(SUM(monto_total), 0) AS monto_total,
            COALESCE((
                SELECT SUM(importe)
                FROM gasto_compartido_items
                WHERE estado = 'pagado'
            ), 0) AS cobrado,
            COALESCE((
                SELECT COUNT(*)
                FROM gasto_compartido_items
                WHERE estado = 'pendiente'
                  AND COALESCE(NULLIF(comprobante_estado, ''), 'sin_comprobante') IN ('pendiente', 'rechazado')
            ), 0) AS comprobantes_revision
        FROM gastos_compartidos
    """).fetchone()
    conn.close()
    return render_template(
        "gastos_compartidos.html",
        gastos=gastos_filtrados,
        resumen=resumen,
        estado_filtro=estado_filtro,
        categoria_filtro=categoria_filtro,
        categorias_disponibles=categorias_disponibles,
    )


@app.route("/gastos-compartidos/nuevo", methods=["GET", "POST"])
def nuevo_gasto_compartido():
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    conn = get_connection()
    jugadores_activos = conn.execute("""
        SELECT id, apellido, nombre, categoria
        FROM jugadores
        WHERE estado = 'Activo'
        ORDER BY categoria NULLS LAST, apellido, nombre
    """).fetchall()
    eventos = opciones_eventos_gastos_compartidos(conn)
    categorias_jugadores = sorted({(jugador["categoria"] or "").strip() for jugador in jugadores_activos if (jugador["categoria"] or "").strip()})
    form_defaults = {
        "fuente_jugadores": request.args.get("fuente_jugadores", "manual").strip() or "manual",
        "modo_importe": request.args.get("modo_importe", "por_jugador").strip() or "por_jugador",
        "calendario_evento_id": request.args.get("calendario_evento_id", "").strip(),
        "titulo": request.args.get("titulo", "").strip(),
    }
    if form_defaults["calendario_evento_id"].isdigit() and not form_defaults["titulo"]:
        evento_prefill = next((evento for evento in eventos if str(evento["id"]) == form_defaults["calendario_evento_id"]), None)
        if evento_prefill:
            form_defaults["titulo"] = f"Tercer tiempo - {evento_prefill['label']}"

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        concepto = request.form.get("concepto", "").strip()
        calendario_evento_id_raw = request.form.get("calendario_evento_id", "").strip()
        calendario_evento_id = int(calendario_evento_id_raw) if calendario_evento_id_raw.isdigit() else None
        fuente = request.form.get("fuente_jugadores", "manual").strip()
        jugadores_ids = request.form.getlist("jugadores_ids")
        fecha_vencimiento = request.form.get("fecha_vencimiento", "").strip()
        modo_importe = request.form.get("modo_importe", "por_jugador").strip()
        monto_raw = request.form.get("monto", "").strip()
        observaciones = request.form.get("observaciones", "").strip()

        if not titulo or not monto_raw:
            conn.close()
            flash("Titulo e importe son obligatorios.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
            )

        try:
            monto = float(monto_raw)
        except ValueError:
            conn.close()
            flash("El importe debe ser numerico.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
            )

        seleccionados = jugadores_gasto_desde_fuente(conn, fuente, calendario_evento_id, jugadores_ids)
        if not seleccionados:
            conn.close()
            flash("Tenes que definir al menos un jugador que deba pagar.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
            )

        if modo_importe not in {"por_jugador", "total"}:
            modo_importe = "por_jugador"

        if modo_importe == "por_jugador":
            importes = [round(monto, 2) for _ in seleccionados]
            monto_total = round(sum(importes), 2)
            monto_por_jugador = round(monto, 2)
        else:
            importes = repartir_importe_gasto(round(monto, 2), len(seleccionados))
            monto_total = round(monto, 2)
            monto_por_jugador = round(monto_total / len(seleccionados), 2) if seleccionados else 0

        evento = None
        if calendario_evento_id:
            evento = conn.execute("""
                SELECT id, fecha, titulo
                FROM calendario_eventos
                WHERE id = %s
            """, (calendario_evento_id,)).fetchone()

        gasto = conn.execute("""
            INSERT INTO gastos_compartidos (
                titulo, concepto, calendario_evento_id, fecha_evento, fecha_vencimiento,
                modo_importe, monto_total, monto_por_jugador, estado, observaciones, creado_por
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Activo', %s, %s)
            RETURNING id
        """, (
            titulo,
            concepto or None,
            calendario_evento_id,
            evento["fecha"] if evento else None,
            fecha_vencimiento or None,
            modo_importe,
            monto_total,
            monto_por_jugador,
            observaciones or None,
            session.get("username"),
        )).fetchone()

        for jugador_item, importe_item in zip(seleccionados, importes):
            conn.execute("""
                INSERT INTO gasto_compartido_items (gasto_id, jugador_id, importe, estado)
                VALUES (%s, %s, %s, 'pendiente')
            """, (gasto["id"], jugador_item["id"], importe_item))

        conn.commit()
        conn.close()
        flash("Gasto compartido creado correctamente.", "ok")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto["id"]))

    conn.close()
    return render_template(
        "gasto_compartido_form.html",
        jugadores=jugadores_activos,
        eventos=eventos,
        categorias_jugadores=categorias_jugadores,
        form=form_defaults,
        seleccion_jugadores=form_defaults.get("jugadores_ids", []),
        modo_formulario="nuevo",
        gasto=None,
    )


@app.route("/gastos-compartidos/<int:gasto_id>/editar", methods=["GET", "POST"])
def editar_gasto_compartido(gasto_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    conn = get_connection()
    gasto = conn.execute("""
        SELECT *
        FROM gastos_compartidos
        WHERE id = %s
    """, (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if gasto_compartido_esta_cerrado(gasto):
        conn.close()
        flash("Ese gasto compartido ya esta cerrado y no se puede editar.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    jugadores_activos = conn.execute("""
        SELECT id, apellido, nombre, categoria
        FROM jugadores
        WHERE estado = 'Activo'
        ORDER BY categoria NULLS LAST, apellido, nombre
    """).fetchall()
    eventos = opciones_eventos_gastos_compartidos(conn)
    categorias_jugadores = sorted({(jugador["categoria"] or "").strip() for jugador in jugadores_activos if (jugador["categoria"] or "").strip()})
    items_existentes = conn.execute("""
        SELECT *
        FROM gasto_compartido_items
        WHERE gasto_id = %s
        ORDER BY id
    """, (gasto_id,)).fetchall()

    seleccion_existente = [str(item["jugador_id"]) for item in items_existentes]
    form_defaults = {
        "titulo": gasto.get("titulo") or "",
        "concepto": gasto.get("concepto") or "",
        "calendario_evento_id": str(gasto.get("calendario_evento_id") or ""),
        "fecha_vencimiento": gasto.get("fecha_vencimiento") or "",
        "fuente_jugadores": gasto.get("fuente_jugadores") or "manual",
        "modo_importe": gasto.get("modo_importe") or "por_jugador",
        "monto": str(gasto.get("monto_total") if (gasto.get("modo_importe") or "por_jugador") == "total" else gasto.get("monto_por_jugador") or ""),
        "observaciones": gasto.get("observaciones") or "",
    }

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        concepto = request.form.get("concepto", "").strip()
        calendario_evento_id_raw = request.form.get("calendario_evento_id", "").strip()
        calendario_evento_id = int(calendario_evento_id_raw) if calendario_evento_id_raw.isdigit() else None
        fuente = request.form.get("fuente_jugadores", "manual").strip()
        jugadores_ids = request.form.getlist("jugadores_ids")
        fecha_vencimiento = request.form.get("fecha_vencimiento", "").strip()
        modo_importe = request.form.get("modo_importe", "por_jugador").strip()
        monto_raw = request.form.get("monto", "").strip()
        observaciones = request.form.get("observaciones", "").strip()

        if not titulo or not monto_raw:
            conn.close()
            flash("Titulo e importe son obligatorios.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
                modo_formulario="editar",
                gasto=gasto,
            )

        try:
            monto = float(monto_raw)
        except ValueError:
            conn.close()
            flash("El importe debe ser numerico.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
                modo_formulario="editar",
                gasto=gasto,
            )

        seleccionados = jugadores_gasto_desde_fuente(conn, fuente, calendario_evento_id, jugadores_ids)
        if not seleccionados:
            conn.close()
            flash("Tenes que definir al menos un jugador que deba pagar.", "error")
            return render_template(
                "gasto_compartido_form.html",
                jugadores=jugadores_activos,
                eventos=eventos,
                categorias_jugadores=categorias_jugadores,
                form=request.form,
                seleccion_jugadores=request.form.getlist("jugadores_ids"),
                modo_formulario="editar",
                gasto=gasto,
            )

        if modo_importe not in {"por_jugador", "total"}:
            modo_importe = "por_jugador"

        if modo_importe == "por_jugador":
            importes = [round(monto, 2) for _ in seleccionados]
            monto_total = round(sum(importes), 2)
            monto_por_jugador = round(monto, 2)
        else:
            importes = repartir_importe_gasto(round(monto, 2), len(seleccionados))
            monto_total = round(monto, 2)
            monto_por_jugador = round(monto_total / len(seleccionados), 2) if seleccionados else 0

        evento = None
        if calendario_evento_id:
            evento = conn.execute("""
                SELECT id, fecha, titulo
                FROM calendario_eventos
                WHERE id = %s
            """, (calendario_evento_id,)).fetchone()

        conn.execute("""
            UPDATE gastos_compartidos
            SET titulo = %s,
                concepto = %s,
                calendario_evento_id = %s,
                fecha_evento = %s,
                fecha_vencimiento = %s,
                fuente_jugadores = %s,
                modo_importe = %s,
                monto_total = %s,
                monto_por_jugador = %s,
                observaciones = %s,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (
            titulo,
            concepto or None,
            calendario_evento_id,
            evento["fecha"] if evento else None,
            fecha_vencimiento or None,
            fuente,
            modo_importe,
            monto_total,
            monto_por_jugador,
            observaciones or None,
            gasto_id,
        ))

        seleccion_map = {jugador["id"]: importe for jugador, importe in zip(seleccionados, importes)}
        existentes_map = {item["jugador_id"]: item for item in items_existentes}
        protegidos = 0

        for jugador_id, item in existentes_map.items():
            if jugador_id in seleccion_map:
                if item.get("estado") != "pagado":
                    conn.execute("""
                        UPDATE gasto_compartido_items
                        SET importe = %s,
                            actualizado_en = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (seleccion_map[jugador_id], item["id"]))
                del seleccion_map[jugador_id]
                continue

            if item.get("estado") == "pagado" or item.get("comprobante_drive_file_id"):
                protegidos += 1
                continue

            conn.execute("DELETE FROM gasto_compartido_items WHERE id = %s", (item["id"],))

        for jugador_id, importe_item in seleccion_map.items():
            conn.execute("""
                INSERT INTO gasto_compartido_items (gasto_id, jugador_id, importe, estado)
                VALUES (%s, %s, %s, 'pendiente')
            """, (gasto_id, jugador_id, importe_item))

        conn.commit()
        conn.close()
        if protegidos:
            flash("Gasto actualizado. Se conservaron jugadores con pago o comprobante ya cargado.", "warning")
        else:
            flash("Gasto compartido actualizado correctamente.", "ok")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    conn.close()
    return render_template(
        "gasto_compartido_form.html",
        jugadores=jugadores_activos,
        eventos=eventos,
        categorias_jugadores=categorias_jugadores,
        form=form_defaults,
        seleccion_jugadores=seleccion_existente,
        modo_formulario="editar",
        gasto=gasto,
    )


@app.route("/gastos-compartidos/<int:gasto_id>")
def ver_gasto_compartido(gasto_id):
    check = permiso_requerido("cuotas_ver")
    if check:
        return check

    conn = get_connection()
    gasto = conn.execute("""
        SELECT g.*, ce.titulo AS evento_titulo
        FROM gastos_compartidos g
        LEFT JOIN calendario_eventos ce ON ce.id = g.calendario_evento_id
        WHERE g.id = %s
    """, (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))

    items = conn.execute("""
        SELECT
            i.*,
            j.apellido,
            j.nombre,
            j.categoria,
            j.email,
            j.email_tutor
        FROM gasto_compartido_items i
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.gasto_id = %s
        ORDER BY
            CASE i.estado
                WHEN 'pendiente' THEN 0
                WHEN 'pagado' THEN 1
                WHEN 'exento' THEN 2
                ELSE 3
            END,
            j.apellido,
            j.nombre
    """, (gasto_id,)).fetchall()
    resumen = conn.execute("""
        SELECT
            COUNT(*) AS total_jugadores,
            COUNT(*) FILTER (WHERE estado = 'pagado') AS pagados,
            COUNT(*) FILTER (WHERE estado = 'pendiente') AS pendientes,
            COUNT(*) FILTER (WHERE estado = 'exento') AS exentos,
            COALESCE(SUM(importe), 0) AS total_asignado,
            COALESCE(SUM(importe) FILTER (WHERE estado = 'pagado'), 0) AS total_cobrado,
            COALESCE(SUM(importe) FILTER (WHERE estado = 'pendiente'), 0) AS total_pendiente
        FROM gasto_compartido_items
        WHERE gasto_id = %s
    """, (gasto_id,)).fetchone()
    conn.close()
    return render_template("gasto_compartido_detalle.html", gasto=gasto, items=items, resumen=resumen)


@app.route("/gastos-compartidos/<int:gasto_id>/exportar")
def exportar_gasto_compartido(gasto_id):
    check = permiso_requerido("cuotas_ver")
    if check:
        return check
    conn = get_connection()
    gasto = conn.execute("SELECT * FROM gastos_compartidos WHERE id = %s", (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    items = conn.execute("""
        SELECT
            i.*,
            j.apellido,
            j.nombre,
            j.categoria,
            j.telefono,
            j.email
        FROM gasto_compartido_items i
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.gasto_id = %s
        ORDER BY j.apellido, j.nombre
    """, (gasto_id,)).fetchall()
    conn.close()

    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(exist_ok=True)
    nombre_archivo = secure_filename(f"gasto_compartido_{gasto_id}_{gasto.get('titulo') or 'detalle'}") or f"gasto_compartido_{gasto_id}"
    ruta = export_dir / f"{nombre_archivo}.csv"
    with ruta.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Apellido", "Nombre", "Categoria", "Importe", "Estado", "Fecha pago", "Telefono", "Email", "Comprobante"])
        for item in items:
            writer.writerow([
                item.get("apellido") or "",
                item.get("nombre") or "",
                item.get("categoria") or "",
                item.get("importe") or 0,
                item.get("estado") or "",
                item.get("fecha_pago") or "",
                item.get("telefono") or "",
                item.get("email") or "",
                item.get("comprobante_nombre") or "",
            ])
    return send_file(ruta, as_attachment=True, download_name=f"{nombre_archivo}.csv")


@app.route("/gastos-compartidos/items/<int:item_id>/email", methods=["POST"])
def enviar_recordatorio_gasto_compartido(item_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check
    conn = get_connection()
    item = conn.execute("""
        SELECT
            i.*,
            g.titulo,
            g.concepto,
            g.fecha_vencimiento,
            g.id AS gasto_id,
            g.estado AS gasto_estado,
            j.nombre,
            j.apellido,
            j.email,
            j.email_tutor
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.id = %s
    """, (item_id,)).fetchone()
    conn.close()
    if item is None:
        flash("Pago compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if gasto_compartido_esta_cerrado({"estado": item.get("gasto_estado")}):
        flash("El gasto compartido ya esta cerrado y no admite nuevos recordatorios.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))
    asunto = f"Gasto compartido pendiente - {item.get('titulo') or 'Ruda Macho Rugby Club'}"
    cuerpo = construir_texto_gasto_compartido(item)
    enviado, destinatario, motivo = enviar_email_jugador(item, asunto, cuerpo)
    flash("Recordatorio enviado por email." if enviado else mensaje_fallo_email(motivo, destinatario), "ok" if enviado else "error")
    return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))


@app.route("/gastos-compartidos/<int:gasto_id>/email-pendientes", methods=["POST"])
def enviar_recordatorios_gasto_compartido_lote(gasto_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check
    conn = get_connection()
    items = conn.execute("""
        SELECT
            i.*,
            g.titulo,
            g.concepto,
            g.fecha_vencimiento,
            g.id AS gasto_id,
            g.estado AS gasto_estado,
            j.nombre,
            j.apellido,
            j.email,
            j.email_tutor
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.gasto_id = %s
          AND i.estado = 'pendiente'
    """, (gasto_id,)).fetchall()
    conn.close()
    if items and gasto_compartido_esta_cerrado({"estado": items[0].get("gasto_estado")}):
        flash("Ese gasto compartido ya esta cerrado y no admite nuevos recordatorios.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))
    resultados = []
    for item in items:
        asunto = f"Gasto compartido pendiente - {item.get('titulo') or 'Ruda Macho Rugby Club'}"
        cuerpo = construir_texto_gasto_compartido(item)
        resultados.append(enviar_email_jugador(item, asunto, cuerpo))
    mensaje, categoria = resumir_envio_masivo_email(resultados, "recordatorio(s) de gasto compartido")
    flash(mensaje, categoria)
    return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))


@app.route("/gastos-compartidos/<int:gasto_id>/cerrar", methods=["POST"])
def cerrar_gasto_compartido(gasto_id):
    check = permiso_requerido("cuotas_gestionar", "caja_gestionar")
    if check:
        return check
    conn = get_connection()
    gasto = conn.execute("SELECT * FROM gastos_compartidos WHERE id = %s", (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if gasto_compartido_esta_cerrado(gasto):
        conn.close()
        flash("Ese gasto compartido ya estaba cerrado.", "warning")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    resumen = conn.execute("""
        SELECT
            COALESCE(SUM(importe) FILTER (WHERE estado = 'pagado'), 0) AS total_cobrado,
            COUNT(*) FILTER (WHERE estado = 'pendiente') AS pendientes
        FROM gasto_compartido_items
        WHERE gasto_id = %s
    """, (gasto_id,)).fetchone()
    total_cobrado = round(float(resumen.get("total_cobrado") or 0), 2)
    pendientes = int(resumen.get("pendientes") or 0)
    fecha_cierre = ahora_sig().strftime("%Y-%m-%d")
    if mes_esta_cerrado(fecha_cierre[:7]):
        conn.close()
        flash("No se puede cerrar el gasto porque el mes actual de caja ya esta cerrado.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    movimiento_id = None
    if total_cobrado > 0:
        movimiento = conn.execute("""
            INSERT INTO movimientos (tipo, concepto, monto, fecha, referencia)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
        """, (
            "ingreso",
            f"Cierre gasto compartido: {gasto.get('titulo') or f'#{gasto_id}'}",
            total_cobrado,
            fecha_cierre,
            f"Gasto compartido #{gasto_id}",
        )).fetchone()
        movimiento_id = movimiento["id"]

    conn.execute("""
        UPDATE gastos_compartidos
        SET estado = 'Cerrado',
            cerrado_en = CURRENT_TIMESTAMP,
            cerrado_por = %s,
            cierre_movimiento_id = %s,
            cierre_monto = %s,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (session.get("username"), movimiento_id, total_cobrado, gasto_id))
    conn.commit()
    conn.close()
    deuda_pendiente = (
        f" {pendientes} deuda(s) impaga(s) seguiran visibles en los perfiles y portales."
        if pendientes
        else ""
    )
    if movimiento_id:
        flash(
            "Gasto compartido cerrado y cobro registrado en caja como un unico ingreso."
            + deuda_pendiente,
            "ok",
        )
    else:
        flash(
            "Gasto compartido cerrado. No habia cobros pagados para registrar en caja."
            + deuda_pendiente,
            "ok",
        )
    return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))


@app.route("/gastos-compartidos/items/<int:item_id>/pago-posterior", methods=["POST"])
def registrar_pago_posterior_gasto_compartido(item_id):
    check = permiso_requerido("cuotas_gestionar", "caja_gestionar")
    if check:
        return check

    conn = get_connection()
    item = conn.execute("""
        SELECT
            i.*,
            g.id AS gasto_id,
            g.titulo,
            g.estado AS gasto_estado,
            j.nombre,
            j.apellido
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.id = %s
    """, (item_id,)).fetchone()

    if item is None:
        conn.close()
        flash("Pago compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if not gasto_compartido_esta_cerrado({"estado": item.get("gasto_estado")}):
        conn.close()
        flash("El pago posterior se usa solo para gastos compartidos cerrados.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))
    if item.get("estado") != "pendiente":
        conn.close()
        flash("Ese saldo ya no esta pendiente.", "warning")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))

    fecha_pago = validar_fecha_movimiento(request.form.get("fecha_pago")) or ahora_sig().strftime("%Y-%m-%d")
    if mes_esta_cerrado(fecha_pago[:7]):
        conn.close()
        flash("No se puede registrar el pago en un mes de caja cerrado.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))

    titulo_gasto = item.get("titulo") or f"#{item['gasto_id']}"
    movimiento = conn.execute("""
        INSERT INTO movimientos (tipo, concepto, monto, fecha, referencia)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id
    """, (
        "ingreso",
        f"Pago posterior gasto compartido: {titulo_gasto}",
        round(float(item.get("importe") or 0), 2),
        fecha_pago,
        f"Gasto compartido #{item['gasto_id']} item #{item_id}",
    )).fetchone()
    conn.execute("""
        UPDATE gasto_compartido_items
        SET estado = 'pagado',
            fecha_pago = %s,
            comprobante_estado = CASE
                WHEN comprobante_drive_file_id IS NOT NULL THEN 'aceptado'
                ELSE comprobante_estado
            END
        WHERE id = %s
    """, (fecha_pago, item_id))
    conn.commit()
    conn.close()

    registrar_auditoria("registrar_pago_posterior", "gasto_compartido_item", str(item_id), {
        "gasto_id": item["gasto_id"],
        "jugador_id": item["jugador_id"],
        "importe": item["importe"],
        "movimiento_id": movimiento["id"],
        "fecha_pago": fecha_pago,
    })
    flash("Pago posterior registrado en caja y deuda cancelada.", "ok")
    return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))


@app.route("/gastos-compartidos/<int:gasto_id>/reabrir", methods=["POST"])
def reabrir_gasto_compartido(gasto_id):
    if session.get("rol") != "admin":
        flash("Solo admins pueden reabrir un gasto compartido cerrado.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    conn = get_connection()
    gasto = conn.execute("SELECT * FROM gastos_compartidos WHERE id = %s", (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if not gasto_compartido_esta_cerrado(gasto):
        conn.close()
        flash("Ese gasto compartido ya esta abierto.", "warning")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))

    conn.execute("""
        UPDATE gastos_compartidos
        SET estado = 'Activo',
            cerrado_en = NULL,
            cerrado_por = NULL,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (gasto_id,))
    conn.commit()
    conn.close()
    flash("Gasto compartido reabierto. El ingreso ya registrado en caja se conserva.", "ok")
    return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))


@app.route("/gastos-compartidos/<int:gasto_id>/eliminar", methods=["POST"])
def eliminar_gasto_compartido(gasto_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check
    conn = get_connection()
    gasto = conn.execute("SELECT * FROM gastos_compartidos WHERE id = %s", (gasto_id,)).fetchone()
    if gasto is None:
        conn.close()
        flash("Gasto compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if gasto_compartido_esta_cerrado(gasto):
        conn.close()
        flash("No se puede eliminar un gasto compartido cerrado.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=gasto_id))
    conn.execute("DELETE FROM gasto_compartido_items WHERE gasto_id = %s", (gasto_id,))
    conn.execute("DELETE FROM gastos_compartidos WHERE id = %s", (gasto_id,))
    conn.commit()
    conn.close()
    flash("Gasto compartido eliminado.", "ok")
    return redirect(url_for("listar_gastos_compartidos"))


@app.route("/gastos-compartidos/items/<int:item_id>/comprobante")
def descargar_comprobante_gasto_compartido(item_id):
    check = permiso_requerido("cuotas_ver")
    if check:
        return check
    conn = get_connection()
    item = conn.execute("""
        SELECT i.*, g.id AS gasto_id
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        WHERE i.id = %s
    """, (item_id,)).fetchone()
    conn.close()
    if item is None:
        flash("Pago compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if not item.get("comprobante_drive_file_id"):
        flash("Este pago no tiene comprobante adjunto.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))
    try:
        archivo = descargar_drive_file(item["comprobante_drive_file_id"])
    except Exception as error:
        app.logger.exception("No se pudo descargar comprobante de gasto compartido %s.", item_id)
        flash(mensaje_error_drive(error, accion="descargar el comprobante"), "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))
    return send_file(
        archivo,
        mimetype=item.get("comprobante_mime_type") or "application/octet-stream",
        as_attachment=False,
        download_name=item.get("comprobante_nombre") or f"gasto_compartido_{item_id}",
    )


@app.route("/gastos-compartidos/items/<int:item_id>/estado", methods=["POST"])
def actualizar_item_gasto_compartido(item_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    accion = request.form.get("accion", "").strip()
    observaciones = request.form.get("observaciones", "").strip()
    conn = get_connection()
    item = conn.execute("""
        SELECT i.*, g.id AS gasto_id, g.estado AS gasto_estado
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        WHERE i.id = %s
    """, (item_id,)).fetchone()
    if item is None:
        conn.close()
        flash("Pago compartido no encontrado.", "error")
        return redirect(url_for("listar_gastos_compartidos"))
    if gasto_compartido_esta_cerrado({"estado": item.get("gasto_estado")}):
        conn.close()
        flash("El gasto compartido esta cerrado y ya no admite cambios.", "error")
        return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))

    ahora = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
    if accion == "aprobar":
        conn.execute("""
            UPDATE gasto_compartido_items
            SET estado = 'pagado',
                fecha_pago = CURRENT_DATE,
                comprobante_estado = 'aceptado',
                comprobante_revisado_en = %s,
                comprobante_revisado_por = %s,
                comprobante_observaciones = %s,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (ahora, session.get("username"), observaciones or None, item_id))
        flash("Comprobante aceptado.", "ok")
    elif accion == "rechazar":
        conn.execute("""
            UPDATE gasto_compartido_items
            SET estado = 'pendiente',
                fecha_pago = NULL,
                comprobante_estado = 'rechazado',
                comprobante_revisado_en = %s,
                comprobante_revisado_por = %s,
                comprobante_observaciones = %s,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (ahora, session.get("username"), observaciones or None, item_id))
        flash("Comprobante rechazado.", "warning")
    elif accion == "exento":
        conn.execute("""
            UPDATE gasto_compartido_items
            SET estado = 'exento',
                fecha_pago = NULL,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (item_id,))
        flash("Jugador marcado como exento.", "ok")
    elif accion == "marcar_pagado":
        conn.execute("""
            UPDATE gasto_compartido_items
            SET estado = 'pagado',
                fecha_pago = CURRENT_DATE,
                comprobante_estado = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN 'aceptado'
                    ELSE comprobante_estado
                END,
                comprobante_revisado_en = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN %s
                    ELSE comprobante_revisado_en
                END,
                comprobante_revisado_por = CASE
                    WHEN comprobante_drive_file_id IS NOT NULL THEN %s
                    ELSE comprobante_revisado_por
                END,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (ahora, session.get("username"), item_id))
        flash("Pago marcado como cobrado.", "ok")
    else:
        conn.execute("""
            UPDATE gasto_compartido_items
            SET estado = 'pendiente',
                fecha_pago = NULL,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (item_id,))
        flash("Pago restaurado a pendiente.", "ok")

    conn.commit()
    conn.close()
    return redirect(url_for("ver_gasto_compartido", gasto_id=item["gasto_id"]))


@app.route("/portal/<token>/gastos/<int:item_id>/comprobante", methods=["POST"])
def portal_subir_comprobante_gasto_compartido(token, item_id):
    conn = get_connection()
    item = conn.execute("""
        SELECT
            i.*,
            g.titulo,
            g.fecha_vencimiento,
            g.estado AS gasto_estado,
            j.id AS jugador_id,
            j.nombre,
            j.apellido
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.id = %s
          AND j.portal_token = %s
          AND COALESCE(j.portal_activo, 0) = 1
    """, (item_id, token)).fetchone()
    if item is None:
        conn.close()
        flash("No encontramos ese gasto en tu portal.", "error")
        return redirect(url_for("portal_jugador", token=token))
    if gasto_compartido_esta_cerrado({"estado": item.get("gasto_estado")}):
        conn.close()
        flash("Ese gasto compartido ya esta cerrado y no admite nuevos comprobantes.", "error")
        return redirect(url_for("portal_jugador", token=token))

    comprobante_pago = request.files.get("comprobante_pago")
    try:
        comprobante_info = subir_comprobante_gasto_compartido_a_drive(comprobante_pago, item)
    except (RuntimeError, ValueError) as error:
        conn.close()
        flash(str(error), "error")
        return redirect(url_for("portal_jugador", token=token))
    except Exception as error:
        conn.close()
        app.logger.exception("No se pudo subir comprobante de gasto compartido %s desde portal.", item_id)
        flash(mensaje_error_drive(error, accion="subir el comprobante"), "error")
        return redirect(url_for("portal_jugador", token=token))

    comprobante_fecha = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""
        UPDATE gasto_compartido_items
        SET comprobante_drive_file_id = %s,
            comprobante_nombre = %s,
            comprobante_mime_type = %s,
            comprobante_tamano = %s,
            comprobante_fecha = %s,
            comprobante_usuario = 'portal',
            comprobante_web_url = %s,
            comprobante_estado = 'pendiente',
            comprobante_revisado_en = NULL,
            comprobante_revisado_por = NULL,
            comprobante_observaciones = NULL,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (
        comprobante_info["file_id"],
        comprobante_info["nombre"],
        comprobante_info["mime_type"],
        comprobante_info["tamano"],
        comprobante_fecha,
        comprobante_info["web_url"],
        item_id,
    ))
    conn.commit()
    conn.close()
    registrar_auditoria(
        "subir_comprobante_gasto",
        "portal_jugador",
        str(item["jugador_id"]),
        {
            **detalle_actor_portal(item),
            "gasto_id": item["gasto_id"],
            "gasto": item.get("titulo") or "",
            "item_id": item_id,
            "archivo": comprobante_info["nombre"],
            "subido_en": comprobante_fecha,
        },
        username=username_portal_jugador(item),
        rol="portal",
    )
    flash("Comprobante enviado para revision.", "ok")
    return redirect(url_for("portal_jugador", token=token))


@app.route("/portal/<token>/gastos/<int:item_id>/comprobante/ver")
def portal_ver_comprobante_gasto_compartido(token, item_id):
    conn = get_connection()
    item = conn.execute("""
        SELECT i.*
        FROM gasto_compartido_items i
        JOIN jugadores j ON j.id = i.jugador_id
        WHERE i.id = %s
          AND j.portal_token = %s
          AND COALESCE(j.portal_activo, 0) = 1
    """, (item_id, token)).fetchone()
    conn.close()
    if item is None:
        flash("No encontramos ese comprobante en tu portal.", "error")
        return redirect(url_for("portal_jugador", token=token))
    if not item.get("comprobante_drive_file_id"):
        flash("Ese gasto no tiene comprobante cargado.", "error")
        return redirect(url_for("portal_jugador", token=token))
    try:
        archivo = descargar_drive_file(item["comprobante_drive_file_id"])
    except Exception as error:
        app.logger.exception("No se pudo descargar comprobante de gasto compartido %s desde portal.", item_id)
        flash(mensaje_error_drive(error, accion="descargar el comprobante"), "error")
        return redirect(url_for("portal_jugador", token=token))
    return send_file(
        archivo,
        mimetype=item.get("comprobante_mime_type") or "application/octet-stream",
        as_attachment=False,
        download_name=item.get("comprobante_nombre") or f"gasto_compartido_{item_id}",
    )


@app.route("/portal", methods=["GET", "POST"])
def portal_buscar():
    identificador = ""
    if request.method == "POST":
        identificador = normalizar_identificador_portal(request.form.get("identificador", ""))
        if not identificador:
            flash("Ingresa tu DNI.", "error")
            return render_template("portal_buscar.html", identificador=identificador)

        identificador_digitos = "".join(ch for ch in identificador if ch.isdigit())
        if not identificador_digitos:
            flash("Ingresa un DNI valido.", "error")
            return render_template("portal_buscar.html", identificador=identificador)

        conn = get_connection()
        jugadores = conn.execute("""
            SELECT id, portal_token
            FROM jugadores
            WHERE COALESCE(portal_activo, 0) = 1
              AND portal_token IS NOT NULL
              AND REGEXP_REPLACE(COALESCE(dni, ''), '[^0-9]', '', 'g') = %s
            ORDER BY id ASC
            LIMIT 2
        """, (identificador_digitos,)).fetchall()
        conn.close()

        if len(jugadores) == 1:
            return redirect(url_for("portal_jugador", token=jugadores[0]["portal_token"]))

        if len(jugadores) > 1:
            flash("Encontramos mas de un portal con ese DNI. Consulta con administracion.", "error")
        else:
            flash("No encontramos un portal activo con ese DNI. Revisalo o consulta con administracion.", "error")

    return render_template("portal_buscar.html", identificador=identificador)


@app.route("/portal/<token>")
def portal_jugador(token):
    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()

    if jugador is None:
        conn.close()
        abort(404)

    cuotas = conn.execute("""
        SELECT id, periodo, importe, pagado, fecha_vencimiento, fecha_pago,
               metodo_pago, becada, beca_porcentaje, descuento_beca,
               plan_pago_monto, plan_pago_detalle, anulada, anulacion_motivo,
               comprobante_drive_file_id, comprobante_nombre, comprobante_fecha,
               comprobante_mime_type,
               comprobante_estado, comprobante_observaciones
        FROM cuotas
        WHERE jugador_id = %s
        ORDER BY periodo DESC, id DESC
        LIMIT 24
    """, (jugador["id"],)).fetchall()

    deuda_cuotas = conn.execute("""
        SELECT COALESCE(SUM(importe), 0) AS total
        FROM cuotas
        WHERE jugador_id = %s
          AND pagado = 0
          AND COALESCE(anulada, 0) = 0
          AND COALESCE(importe, 0) > 0
    """, (jugador["id"],)).fetchone()["total"]

    ficha = conn.execute("""
        SELECT presentada, fecha_vencimiento, apto_fisico
        FROM fichas_medicas
        WHERE jugador_id = %s
    """, (jugador["id"],)).fetchone()

    documentos = conn.execute("""
        SELECT tipo, nombre, fecha_vencimiento, url
        FROM documentos_jugadores
        WHERE jugador_id = %s
        ORDER BY fecha_vencimiento DESC NULLS LAST, id DESC
    """, (jugador["id"],)).fetchall()

    historial_asistencia = conn.execute("""
        SELECT
            e.fecha,
            e.tipo,
            e.descripcion,
            a.estado_asistencia,
            a.presente,
            a.observaciones
        FROM asistencias a
        JOIN eventos_asistencia e ON e.id = a.evento_id
        WHERE a.jugador_id = %s
        ORDER BY e.fecha DESC, e.id DESC
        LIMIT 12
    """, (jugador["id"],)).fetchall()

    gastos_compartidos = conn.execute("""
        SELECT
            i.*,
            g.titulo,
            g.concepto,
            g.fecha_evento,
            g.fecha_vencimiento,
            g.estado AS gasto_estado
        FROM gasto_compartido_items i
        JOIN gastos_compartidos g ON g.id = i.gasto_id
        WHERE i.jugador_id = %s
          AND (
              g.estado = 'Activo'
              OR i.estado = 'pendiente'
          )
        ORDER BY
            CASE i.estado
                WHEN 'pendiente' THEN 0
                WHEN 'pagado' THEN 1
                WHEN 'exento' THEN 2
                ELSE 3
            END,
            COALESCE(g.fecha_vencimiento, g.fecha_evento, g.creado_en::text) DESC,
            i.id DESC
        LIMIT 20
    """, (jugador["id"],)).fetchall()

    lesiones_activas_portal = conn.execute("""
        SELECT
            id,
            fecha_lesion,
            tipo_lesion,
            zona_cuerpo,
            estado,
            etapa_recuperacion,
            proximo_control,
            fecha_retorno_estimada,
            tratamiento_hasta,
            observaciones
        FROM lesiones
        WHERE jugador_id = %s
          AND (
              estado = 'Activa'
              OR estado ILIKE 'En recuperaci%%'
          )
        ORDER BY fecha_lesion DESC, id DESC
        LIMIT 6
    """, (jugador["id"],)).fetchall()

    planes_pago = conn.execute("""
        SELECT *
        FROM planes_pago
        WHERE jugador_id = %s
        ORDER BY
            CASE WHEN estado = 'Activo' THEN 0 ELSE 1 END,
            fecha_inicio DESC,
            id DESC
        LIMIT 12
    """, (jugador["id"],)).fetchall()

    tests_recientes = conn.execute("""
        SELECT
            r.*,
            t.nombre AS test_nombre,
            t.unidad
        FROM test_resultados r
        JOIN test_tipos t ON t.id = r.test_id
        WHERE r.jugador_id = %s
        ORDER BY r.fecha DESC, r.id DESC
        LIMIT 24
    """, (jugador["id"],)).fetchall()

    portal_tests = conn.execute("""
        SELECT DISTINCT
            t.id,
            t.nombre,
            t.unidad,
            t.mayor_es_mejor
        FROM test_tipos t
        JOIN test_resultados r ON r.test_id = t.id
        WHERE r.jugador_id = %s
        ORDER BY t.nombre
    """, (jugador["id"],)).fetchall()

    portal_test_id_raw = request.args.get("test_id", "").strip()
    portal_test_id = int(portal_test_id_raw) if portal_test_id_raw.isdigit() else None
    if not portal_test_id and portal_tests:
        portal_test_id = portal_tests[0]["id"]

    portal_test_desde = validar_fecha_movimiento(request.args.get("test_desde", "").strip())
    portal_test_hasta = validar_fecha_movimiento(request.args.get("test_hasta", "").strip())
    portal_test_actual = next((test for test in portal_tests if test["id"] == portal_test_id), None)
    portal_test_resultados = []
    if portal_test_actual:
        filtros = ["r.jugador_id = %s", "r.test_id = %s"]
        params = [jugador["id"], portal_test_id]
        if portal_test_desde:
            filtros.append("r.fecha >= %s")
            params.append(portal_test_desde)
        if portal_test_hasta:
            filtros.append("r.fecha <= %s")
            params.append(portal_test_hasta)

        where = " AND ".join(filtros)
        portal_test_resultados = conn.execute(f"""
            SELECT
                r.*,
                j.apellido,
                j.nombre,
                j.categoria
            FROM test_resultados r
            JOIN jugadores j ON j.id = r.jugador_id
            WHERE {where}
            ORDER BY r.fecha, r.id
        """, params).fetchall()

    portal_test_grafico = construir_grafico_tests(portal_test_resultados)
    portal_test_comparativo = construir_comparativo_tests(portal_test_resultados, portal_test_actual)
    eventos_deportivos = obtener_eventos_deportivos_portal(jugador)
    confirmaciones_portal = {}
    evento_ids_confirmables = [evento["asistencia_evento_id"] for evento in eventos_deportivos if evento.get("asistencia_evento_id")]
    if evento_ids_confirmables:
        confirmaciones_portal = obtener_confirmaciones_portal(conn, evento_ids_confirmables, jugador["id"])
    cuenta_corriente = obtener_cuenta_corriente_jugador(conn, jugador["id"], limite=20)
    comunicaciones_portal = obtener_comunicaciones_portal_dia(conn, jugador)
    conn.close()

    documentos_por_vencer = 0
    hoy = ahora_sig().date()
    for documento in documentos:
        fecha_vencimiento = validar_fecha_movimiento(documento.get("fecha_vencimiento"))
        if not fecha_vencimiento:
            continue
        try:
            fecha = datetime.strptime(fecha_vencimiento, "%Y-%m-%d").date()
        except ValueError:
            continue
        if fecha <= hoy + timedelta(days=30):
            documentos_por_vencer += 1

    gastos_pendientes = [item for item in gastos_compartidos if item.get("estado") == "pendiente"]
    gastos_pagados = [item for item in gastos_compartidos if item.get("estado") == "pagado"]
    gasto_pendiente_total = round(sum(float(item.get("importe") or 0) for item in gastos_pendientes), 2)
    deuda = round(float(deuda_cuotas or 0) + gasto_pendiente_total, 2)

    portal_alertas = []
    if deuda > 0:
        portal_alertas.append({
            "nivel": "danger",
            "titulo": "Tenes deuda pendiente",
            "detalle": f"Actualmente tenes {formato_moneda(deuda)} pendientes de pago.",
        })
    if gastos_pendientes:
        portal_alertas.append({
            "nivel": "warning",
            "titulo": "Gastos compartidos pendientes",
            "detalle": f"Tenes {len(gastos_pendientes)} gasto(s) compartido(s) para revisar o pagar.",
        })
    ficha_portal = estado_ficha_portal(ficha)
    if ficha_portal["nivel"] != "success":
        portal_alertas.append({
            "nivel": ficha_portal["nivel"],
            "titulo": "Estado de ficha médica",
            "detalle": ficha_portal["label"],
        })
    if documentos_por_vencer:
        portal_alertas.append({
            "nivel": "warning",
            "titulo": "Documentacion por revisar",
            "detalle": f"Tenes {documentos_por_vencer} documento(s) con vencimiento cercano.",
        })
    if lesiones_activas_portal:
        portal_alertas.append({
            "nivel": "warning",
            "titulo": "Seguimiento de salud activo",
            "detalle": f"Hay {len(lesiones_activas_portal)} lesion(es) activa(s) o en recuperacion.",
        })

    for lesion in lesiones_activas_portal:
        lesion["semaforo"] = semaforo_lesion(lesion)
    for evento in eventos_deportivos:
        confirmacion = None
        if evento.get("asistencia_evento_id"):
            confirmacion = confirmaciones_portal.get((evento["asistencia_evento_id"], jugador["id"]))
        evento["confirmacion_portal"] = confirmacion

    calendario_ics_url = url_for("portal_calendario_ics", token=token, _external=True)
    calendario_webcal_url = calendario_ics_url.replace("https://", "webcal://", 1).replace("http://", "webcal://", 1)
    calendario_google_url = "https://calendar.google.com/calendar/r?cid=" + quote(calendario_webcal_url, safe="")
    calendario_android_url = (
        calendario_ics_url.replace("https://", "intent://", 1).replace("http://", "intent://", 1)
        + "#Intent;scheme=webcal;S.browser_fallback_url="
        + quote(calendario_ics_url, safe="")
        + ";end"
    )

    return render_template(
        "portal_jugador.html",
        jugador=jugador,
        cuotas=cuotas,
        deuda=deuda,
        deuda_cuotas=deuda_cuotas,
        ficha=ficha,
        documentos=documentos,
        historial_asistencia=historial_asistencia,
        lesiones_activas_portal=lesiones_activas_portal,
        portal_alertas=portal_alertas,
        ficha_portal=ficha_portal,
        planes_pago=planes_pago,
        gastos_compartidos=gastos_compartidos,
        gastos_pendientes=gastos_pendientes,
        gastos_pagados=gastos_pagados,
        gasto_pendiente_total=gasto_pendiente_total,
        tests_recientes=tests_recientes,
        resumen_tests=resumir_resultados_tests(tests_recientes),
        portal_tests=portal_tests,
        portal_test_id=portal_test_id,
        portal_test_desde=portal_test_desde or "",
        portal_test_hasta=portal_test_hasta or "",
        portal_test_actual=portal_test_actual,
        portal_test_grafico=portal_test_grafico,
        portal_test_comparativo=portal_test_comparativo,
        portal_test_resultados=portal_test_resultados,
        eventos_deportivos=eventos_deportivos,
        calendario_ics_url=calendario_ics_url,
        calendario_webcal_url=calendario_webcal_url,
        calendario_google_url=calendario_google_url,
        calendario_android_url=calendario_android_url,
        cuenta_corriente=cuenta_corriente,
        comunicaciones_portal=comunicaciones_portal,
        token=token,
    )


@app.route("/portal/<token>/eventos/<int:evento_id>/confirmar", methods=["POST"])
def portal_confirmar_asistencia(token, evento_id):
    estado = request.form.get("estado", "").strip()
    if estado not in PORTAL_ASISTENCIA_ESTADOS:
        flash("La confirmacion no es valida.", "error")
        return redirect(url_for("portal_jugador", token=token))

    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()
    if jugador is None:
        conn.close()
        abort(404)

    evento = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE id = %s
          AND COALESCE(publicar_portal, 0) = 1
    """, (evento_id,)).fetchone()
    if evento is None or not evento.get("asistencia_evento_id"):
        conn.close()
        flash("Ese evento no admite confirmacion desde el portal.", "error")
        return redirect(url_for("portal_jugador", token=token))

    if not categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria")):
        conn.close()
        abort(404)

    if es_evento_partido(evento):
        confirmado_en = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
        guardar_confirmacion_portal_sin_bienestar(conn, evento, jugador, estado)
        conn.commit()
        conn.close()
        registrar_auditoria(
            "confirmar_asistencia",
            "portal_jugador",
            str(jugador["id"]),
            {
                **detalle_actor_portal(jugador),
                "evento_id": evento["id"],
                "asistencia_evento_id": evento["asistencia_evento_id"],
                "evento": evento.get("titulo") or "",
                "fecha_evento": formato_fecha_hora_evento(evento),
                "estado": estado,
                "confirmado_en": confirmado_en,
                "bienestar": False,
            },
            username=username_portal_jugador(jugador),
            rol="portal",
        )
        flash("Confirmacion guardada.", "ok")
        return redirect(url_for("portal_jugador", token=token))

    conn.close()
    return redirect(url_for("portal_bienestar_asistencia", token=token, evento_id=evento_id, estado=estado))


@app.route("/portal/<token>/eventos/<int:evento_id>/bienestar", methods=["GET", "POST"])
def portal_bienestar_asistencia(token, evento_id):
    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()
    if jugador is None:
        conn.close()
        abort(404)

    evento = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE id = %s
          AND COALESCE(publicar_portal, 0) = 1
    """, (evento_id,)).fetchone()
    if evento is None or not evento.get("asistencia_evento_id"):
        conn.close()
        flash("Ese evento no admite confirmacion desde el portal.", "error")
        return redirect(url_for("portal_jugador", token=token))

    if not categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria")):
        conn.close()
        abort(404)

    actual = conn.execute("""
        SELECT *
        FROM portal_asistencia_confirmaciones
        WHERE evento_id = %s AND jugador_id = %s
    """, (evento["asistencia_evento_id"], jugador["id"])).fetchone()
    actual = dict(actual) if actual else {}
    try:
        actual["dolor_zonas_lista"] = json.loads(actual.get("dolor_zonas") or "[]")
    except (TypeError, ValueError):
        actual["dolor_zonas_lista"] = []

    estado = (request.form.get("estado") if request.method == "POST" else request.args.get("estado") or actual.get("estado") or "confirmado").strip()
    if estado not in PORTAL_ASISTENCIA_ESTADOS:
        estado = "confirmado"

    if es_evento_partido(evento):
        if request.method == "POST" or request.args.get("estado"):
            confirmado_en = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
            guardar_confirmacion_portal_sin_bienestar(conn, evento, jugador, estado)
            conn.commit()
            registrar_auditoria(
                "confirmar_asistencia",
                "portal_jugador",
                str(jugador["id"]),
                {
                    **detalle_actor_portal(jugador),
                    "evento_id": evento["id"],
                    "asistencia_evento_id": evento["asistencia_evento_id"],
                    "evento": evento.get("titulo") or "",
                    "fecha_evento": formato_fecha_hora_evento(evento),
                    "estado": estado,
                    "confirmado_en": confirmado_en,
                    "bienestar": False,
                },
                username=username_portal_jugador(jugador),
                rol="portal",
            )
            flash("Confirmacion guardada.", "ok")
        else:
            flash("Los partidos no requieren cuestionario de bienestar.", "ok")
        conn.close()
        return redirect(url_for("portal_jugador", token=token))

    if request.method == "POST":
        campos = {
            "sueno_calidad": request.form.get("sueno_calidad", "").strip(),
            "horas_sueno": request.form.get("horas_sueno", "").strip(),
            "doms": request.form.get("doms", "").strip(),
            "fatiga": request.form.get("fatiga", "").strip(),
            "estres": request.form.get("estres", "").strip(),
            "animo": request.form.get("animo", "").strip(),
            "motivacion": request.form.get("motivacion", "").strip(),
            "recuperacion": request.form.get("recuperacion", "").strip(),
        }
        if not all(campos.values()):
            conn.close()
            flash("Completa todas las respuestas obligatorias del bienestar.", "error")
            return render_template(
                "portal_bienestar.html",
                token=token,
                jugador=jugador,
                evento=evento,
                estado=estado,
                actual={**actual, **campos, "dolor_zonas_lista": request.form.getlist("dolor_zonas"), "dolor_otro": request.form.get("dolor_otro", "").strip(), "comentarios": request.form.get("comentarios", "").strip()},
                horas_opciones=BIENESTAR_HORAS_OPCIONES,
                dolor_zonas=BIENESTAR_DOLOR_ZONAS,
            )

        dolor_zonas = request.form.getlist("dolor_zonas") or ["No"]
        dolor_otro = request.form.get("dolor_otro", "").strip()
        comentarios = request.form.get("comentarios", "").strip()

        conn.execute("""
            INSERT INTO portal_asistencia_confirmaciones (
                evento_id, jugador_id, estado, sueno_calidad, horas_sueno, doms, fatiga, estres, animo,
                motivacion, recuperacion, dolor_zonas, dolor_otro, comentarios, creado_en, actualizado_en
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (evento_id, jugador_id)
            DO UPDATE SET
                estado = excluded.estado,
                sueno_calidad = excluded.sueno_calidad,
                horas_sueno = excluded.horas_sueno,
                doms = excluded.doms,
                fatiga = excluded.fatiga,
                estres = excluded.estres,
                animo = excluded.animo,
                motivacion = excluded.motivacion,
                recuperacion = excluded.recuperacion,
                dolor_zonas = excluded.dolor_zonas,
                dolor_otro = excluded.dolor_otro,
                comentarios = excluded.comentarios,
                actualizado_en = CURRENT_TIMESTAMP
        """, (
            evento["asistencia_evento_id"], jugador["id"], estado,
            int(campos["sueno_calidad"]), campos["horas_sueno"], int(campos["doms"]), int(campos["fatiga"]),
            int(campos["estres"]), int(campos["animo"]), int(campos["motivacion"]), int(campos["recuperacion"]),
            json.dumps(dolor_zonas, ensure_ascii=False), dolor_otro or None, comentarios or None,
        ))
        confirmado_en = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
        conn.commit()
        conn.close()

        registrar_auditoria(
            "confirmar_asistencia",
            "portal_jugador",
            str(jugador["id"]),
            {
                **detalle_actor_portal(jugador),
                "evento_id": evento["id"],
                "asistencia_evento_id": evento["asistencia_evento_id"],
                "evento": evento.get("titulo") or "",
                "fecha_evento": formato_fecha_hora_evento(evento),
                "estado": estado,
                "confirmado_en": confirmado_en,
                "bienestar": True,
            },
            username=username_portal_jugador(jugador),
            rol="portal",
        )
        flash("Confirmacion y bienestar guardados.", "ok")
        return redirect(url_for("portal_jugador", token=token))

    conn.close()
    return render_template(
        "portal_bienestar.html",
        token=token,
        jugador=jugador,
        evento=evento,
        estado=estado,
        actual=actual,
        horas_opciones=BIENESTAR_HORAS_OPCIONES,
        dolor_zonas=BIENESTAR_DOLOR_ZONAS,
    )


@app.route("/portal/<token>/calendario.ics")
def portal_calendario_ics(token):
    jugador, eventos = obtener_eventos_deportivos_ics(token)
    if jugador is None:
        abort(404)
    feed_url = url_for("portal_calendario_ics", token=token, _external=True)
    contenido = generar_ics_calendario(jugador, eventos, feed_url)
    return Response(
        contenido,
        mimetype="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": "inline; filename=calendario-sig.ics",
            "Cache-Control": "public, max-age=900",
        },
    )


@app.route("/portal/<token>/contacto", methods=["POST"])
def portal_actualizar_contacto(token):
    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()
    if jugador is None:
        conn.close()
        abort(404)

    data = {
        "nombre": request.form.get("nombre", "").strip(),
        "apellido": request.form.get("apellido", "").strip(),
        "dni": request.form.get("dni", "").strip(),
        "fecha_nacimiento": request.form.get("fecha_nacimiento", "").strip(),
        "telefono": request.form.get("telefono", "").strip(),
        "email": request.form.get("email", "").strip(),
        "direccion": request.form.get("direccion", "").strip(),
        "obra_social": request.form.get("obra_social", "").strip(),
        "numero_afiliado_obra_social": request.form.get("numero_afiliado_obra_social", "").strip(),
        "contacto_tutor": request.form.get("contacto_tutor", "").strip(),
        "parentesco_tutor": request.form.get("parentesco_tutor", "").strip(),
        "telefono_tutor": request.form.get("telefono_tutor", "").strip(),
        "email_tutor": request.form.get("email_tutor", "").strip(),
    }

    if not data["nombre"] or not data["apellido"]:
        conn.close()
        flash("Nombre y apellido son obligatorios.", "error")
        return redirect(url_for("portal_jugador", token=token))

    if data["dni"]:
        existente = conn.execute("""
            SELECT id
            FROM jugadores
            WHERE dni = %s AND id <> %s
        """, (data["dni"], jugador["id"])).fetchone()
        if existente:
            conn.close()
            flash("Ya existe otro jugador con ese DNI. Revisalo con administracion.", "error")
            return redirect(url_for("portal_jugador", token=token))

    campos_modificados = [
        campo
        for campo, valor in data.items()
        if str(jugador[campo] or "") != str(valor or "")
    ]
    labels_campos = {
        "nombre": "Nombre",
        "apellido": "Apellido",
        "dni": "DNI",
        "fecha_nacimiento": "Fecha de nacimiento",
        "telefono": "Telefono",
        "email": "Email",
        "direccion": "Direccion",
        "obra_social": "Obra social",
        "numero_afiliado_obra_social": "Numero de afiliado de obra social",
        "contacto_tutor": "Contacto familiar",
        "parentesco_tutor": "Parentesco",
        "telefono_tutor": "Telefono familiar",
        "email_tutor": "Email familiar",
    }
    cambios_detalle = {
        campo: {
            "label": labels_campos.get(campo, campo),
            "antes": jugador[campo] or "",
            "despues": data[campo] or "",
        }
        for campo in campos_modificados
    }

    conn.execute("""
        UPDATE jugadores
        SET nombre = %s,
            apellido = %s,
            dni = %s,
            fecha_nacimiento = %s,
            telefono = %s,
            email = %s,
            direccion = %s,
            obra_social = %s,
            numero_afiliado_obra_social = %s,
            contacto_tutor = %s,
            parentesco_tutor = %s,
            telefono_tutor = %s,
            email_tutor = %s,
            portal_actualizado_en = %s
        WHERE id = %s
    """, (
        data["nombre"],
        data["apellido"],
        data["dni"],
        data["fecha_nacimiento"],
        data["telefono"],
        data["email"],
        data["direccion"],
        data["obra_social"],
        data["numero_afiliado_obra_social"],
        data["contacto_tutor"],
        data["parentesco_tutor"],
        data["telefono_tutor"],
        data["email_tutor"],
        ahora_sig().strftime("%Y-%m-%d %H:%M:%S"),
        jugador["id"],
    ))
    if campos_modificados:
        campos_texto = ", ".join(labels_campos.get(campo, campo) for campo in campos_modificados)
        conn.execute("""
            INSERT INTO jugador_bitacora (jugador_id, tipo, nota, creado_por)
            VALUES (%s, 'general', %s, 'portal')
        """, (
            jugador["id"],
            f"El jugador actualizo datos personales desde el portal: {campos_texto}.",
        ))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "actualizar_contacto",
        "portal_jugador",
        str(jugador["id"]),
        {
            **detalle_actor_portal(jugador),
            "campos": campos_modificados or list(data.keys()),
            "cambios": cambios_detalle,
        },
        username=username_portal_jugador(jugador),
        rol="portal",
    )
    flash("Datos personales actualizados.", "ok")
    return redirect(url_for("portal_jugador", token=token))


@app.route("/portal/<token>/cuotas/<int:cuota_id>/comprobante", methods=["POST"])
def portal_subir_comprobante(token, cuota_id):
    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.*,
            j.nombre,
            j.apellido
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
          AND j.portal_token = %s
          AND COALESCE(j.portal_activo, 0) = 1
    """, (cuota_id, token)).fetchone()
    if cuota is None:
        conn.close()
        abort(404)

    if cuota["pagado"]:
        conn.close()
        flash("La cuota ya figura pagada.", "error")
        return redirect(url_for("portal_jugador", token=token))

    if cuota.get("anulada"):
        conn.close()
        flash("La cuota esta anulada por un plan de pago.", "error")
        return redirect(url_for("portal_jugador", token=token))

    comprobante_pago = request.files.get("comprobante_pago")
    referencia_pago = request.form.get("referencia_pago", "").strip()
    try:
        comprobante_info = subir_comprobante_a_drive(comprobante_pago, cuota)
    except (RuntimeError, ValueError) as error:
        conn.close()
        flash(str(error), "error")
        return redirect(url_for("portal_jugador", token=token))
    except Exception as error:
        app.logger.exception("No se pudo subir comprobante desde portal para cuota %s.", cuota_id)
        conn.close()
        flash(mensaje_error_drive(error), "error")
        return redirect(url_for("portal_jugador", token=token))

    comprobante_fecha = ahora_sig().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""
        UPDATE cuotas
        SET referencia_pago = COALESCE(NULLIF(%s, ''), referencia_pago),
            comprobante_drive_file_id = %s,
            comprobante_nombre = %s,
            comprobante_mime_type = %s,
            comprobante_tamano = %s,
            comprobante_fecha = %s,
            comprobante_usuario = 'portal',
            comprobante_web_url = %s,
            comprobante_estado = 'pendiente',
            comprobante_revisado_en = NULL,
            comprobante_revisado_por = NULL,
            comprobante_observaciones = NULL
        WHERE id = %s
    """, (
        referencia_pago,
        comprobante_info["file_id"],
        comprobante_info["nombre"],
        comprobante_info["mime_type"],
        comprobante_info["tamano"],
        comprobante_fecha,
        comprobante_info["web_url"],
        cuota_id,
    ))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "subir_comprobante",
        "portal_jugador",
        str(cuota["jugador_id"]),
        {
            **detalle_actor_portal(cuota),
            "cuota_id": cuota_id,
            "periodo": cuota.get("periodo") or "",
            "archivo": comprobante_info["nombre"],
            "subido_en": comprobante_fecha,
        },
        username=username_portal_jugador(cuota),
        rol="portal",
    )
    flash("Comprobante recibido. La cuota queda pendiente de validacion interna.", "ok")
    return redirect(url_for("portal_jugador", token=token))


@app.route("/portal/<token>/cuotas/<int:cuota_id>/comprobante/ver")
def portal_ver_comprobante(token, cuota_id):
    conn = get_connection()
    cuota = conn.execute("""
        SELECT c.*
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
          AND j.portal_token = %s
          AND COALESCE(j.portal_activo, 0) = 1
    """, (cuota_id, token)).fetchone()
    conn.close()

    if cuota is None:
        abort(404)

    if not cuota["comprobante_drive_file_id"]:
        flash("La cuota no tiene comprobante adjunto.", "error")
        return redirect(url_for("portal_jugador", token=token))

    try:
        archivo = descargar_drive_file(cuota["comprobante_drive_file_id"])
    except RuntimeError as error:
        flash(str(error), "error")
        return redirect(url_for("portal_jugador", token=token))
    except Exception as error:
        app.logger.exception("No se pudo descargar comprobante desde portal para cuota %s.", cuota_id)
        flash(mensaje_error_drive(error, accion="descargar el comprobante"), "error")
        return redirect(url_for("portal_jugador", token=token))

    return send_file(
        archivo,
        mimetype=cuota["comprobante_mime_type"] or "application/octet-stream",
        as_attachment=False,
        download_name=cuota["comprobante_nombre"] or f"comprobante_cuota_{cuota_id}",
    )


@app.route("/portal/<token>/cuotas/<int:cuota_id>/recibo")
def portal_descargar_recibo(token, cuota_id):
    conn = get_connection()
    cuota = conn.execute("""
        SELECT c.*
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
          AND j.portal_token = %s
          AND COALESCE(j.portal_activo, 0) = 1
    """, (cuota_id, token)).fetchone()
    conn.close()

    if cuota is None:
        abort(404)
    if not cuota["pagado"]:
        flash("El recibo esta disponible cuando la cuota figura pagada.", "error")
        return redirect(url_for("portal_jugador", token=token))

    archivo = BASE_DIR / "recibos" / f"recibo_cuota_{cuota_id}.pdf"
    if not archivo.exists():
        archivo = generar_recibo_pdf(cuota_id)
    if archivo is None or not archivo.exists():
        flash("No se pudo generar el recibo.", "error")
        return redirect(url_for("portal_jugador", token=token))

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"recibo_cuota_{cuota_id}.pdf",
    )


@app.route("/portal/<token>/constancia")
def portal_descargar_constancia(token):
    conn = get_connection()
    jugador = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE portal_token = %s
          AND COALESCE(portal_activo, 0) = 1
    """, (token,)).fetchone()
    conn.close()

    if jugador is None:
        abort(404)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    ancho, alto = A4
    y = alto - 30 * mm

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(25 * mm, y, "Constancia de jugador activo")
    y -= 14 * mm

    pdf.setFont("Helvetica", 11)
    pdf.drawString(25 * mm, y, f"Club: Ruda Macho Rugby Club")
    y -= 8 * mm
    pdf.drawString(25 * mm, y, f"Fecha de emision: {ahora_sig().strftime('%Y-%m-%d')}")
    y -= 14 * mm

    nombre_completo = f"{jugador.get('apellido') or ''}, {jugador.get('nombre') or ''}".strip(", ")
    pdf.drawString(25 * mm, y, f"Jugador/a: {nombre_completo or '-'}")
    y -= 8 * mm
    pdf.drawString(25 * mm, y, f"DNI: {jugador.get('dni') or '-'}")
    y -= 8 * mm
    pdf.drawString(25 * mm, y, f"Categoria: {jugador.get('categoria') or '-'}")
    y -= 8 * mm
    pdf.drawString(25 * mm, y, f"Numero de socio del club: {jugador.get('numero_socio') or '-'}")
    y -= 16 * mm

    pdf.drawString(
        25 * mm,
        y,
        "Se deja constancia de que la persona indicada posee portal activo en el sistema del club.",
    )
    y -= 24 * mm

    pdf.line(25 * mm, y, 90 * mm, y)
    pdf.drawString(25 * mm, y - 7 * mm, "Tesoreria / Administracion")
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawString(25 * mm, 20 * mm, "Ruda Macho Rugby Club - Sistema Integral de Gestion")
    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="constancia_jugador.pdf",
    )


@app.route("/calendario")
def ver_calendario():
    check = permiso_requerido("calendario_ver")
    if check:
        return check

    mes = normalizar_mes(request.args.get("mes"), ahora_sig().strftime("%Y-%m"))
    calendario = obtener_calendario(mes)
    mes_actual = datetime.strptime(mes, "%Y-%m")

    if mes_actual.month == 1:
        mes_anterior = mes_actual.replace(year=mes_actual.year - 1, month=12)
    else:
        mes_anterior = mes_actual.replace(month=mes_actual.month - 1)

    if mes_actual.month == 12:
        mes_siguiente = mes_actual.replace(year=mes_actual.year + 1, month=1)
    else:
        mes_siguiente = mes_actual.replace(month=mes_actual.month + 1)

    return render_template(
        "calendario.html",
        calendario=calendario,
        mes=mes,
        mes_anterior=mes_anterior.strftime("%Y-%m"),
        mes_siguiente=mes_siguiente.strftime("%Y-%m"),
    )


@app.route("/calendario/nuevo", methods=["GET", "POST"])
def nuevo_evento_calendario():
    check = permiso_requerido("calendario_gestionar", "asistencia_gestionar")
    if check:
        return check
    origen = request.form.get("origen") or request.args.get("origen", "")
    if origen != "asistencia":
        origen = ""
    if not tiene_permiso("calendario_gestionar") and origen != "asistencia":
        flash("No tenes permiso para acceder a esa seccion.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        data = {
            "fecha": request.form.get("fecha", "").strip(),
            "hora_inicio": normalizar_hora_evento(request.form.get("hora_inicio", "")),
            "duracion_minutos": normalizar_duracion_evento(request.form.get("duracion_minutos", "90")),
            "tipo": request.form.get("tipo", "").strip(),
            "titulo": request.form.get("titulo", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "ubicacion": request.form.get("ubicacion", "").strip(),
            "categoria": request.form.get("categoria", "").strip(),
            "convocatoria_texto": request.form.get("convocatoria_texto", "").strip(),
            "convocatoria_cierre": request.form.get("convocatoria_cierre", "").strip(),
            "minuta_post_evento": request.form.get("minuta_post_evento", "").strip(),
            "publicar_portal": 1 if request.form.get("publicar_portal") == "on" else 0,
            "crear_asistencia": 1 if request.form.get("crear_asistencia") == "on" else 0,
            "origen": origen,
        }

        if calendario_evento_requiere_asistencia(data["tipo"]):
            data["crear_asistencia"] = 1 if request.form.get("crear_asistencia", "on") == "on" else 0
        if origen == "asistencia":
            data["crear_asistencia"] = 1

        if not data["fecha"] or not data["tipo"] or not data["titulo"]:
            flash("Fecha, tipo y titulo son obligatorios.", "error")
            return render_template("calendario_evento_form.html", evento=data)

        try:
            datetime.strptime(data["fecha"], "%Y-%m-%d")
        except ValueError:
            flash("La fecha del evento no es valida.", "error")
            return render_template("calendario_evento_form.html", evento=data)

        if request.form.get("hora_inicio") and not data["hora_inicio"]:
            flash("La hora debe tener formato HH:MM.", "error")
            return render_template("calendario_evento_form.html", evento=data)

        conn = get_connection()
        asistencia_evento_id = None
        if data["crear_asistencia"]:
            asistencia_evento_id = crear_evento_asistencia_desde_calendario(conn, data)

        evento_id = conn.execute("""
            INSERT INTO calendario_eventos (
                fecha, tipo, titulo, descripcion, ubicacion, categoria,
                hora_inicio, duracion_minutos, publicar_portal, asistencia_evento_id,
                convocatoria_texto, convocatoria_cierre, minuta_post_evento
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data["fecha"],
            data["tipo"],
            data["titulo"],
            data["descripcion"],
            data["ubicacion"],
            data["categoria"],
            data["hora_inicio"] or None,
            data["duracion_minutos"],
            data["publicar_portal"],
            asistencia_evento_id,
            data["convocatoria_texto"] or None,
            data["convocatoria_cierre"] or None,
            data["minuta_post_evento"] or None,
        )).fetchone()["id"]
        conn.commit()
        conn.close()

        registrar_auditoria("crear", "calendario_evento", str(evento_id), {
            "tipo": data["tipo"],
            "fecha": data["fecha"],
            "asistencia_evento_id": asistencia_evento_id,
        })

        if asistencia_evento_id:
            flash("Evento agregado al calendario y listo para tomar asistencia.", "ok")
        else:
            flash("Evento agregado al calendario.", "ok")
        if origen == "asistencia":
            return redirect(url_for("listar_eventos_asistencia"))
        return redirect(url_for("ver_calendario", mes=data["fecha"][:7]))

    return render_template(
        "calendario_evento_form.html",
        evento={"publicar_portal": 1, "crear_asistencia": 1, "duracion_minutos": 90, "origen": origen},
    )


@app.route("/calendario/<int:evento_id>/editar", methods=["GET", "POST"])
def editar_evento_calendario(evento_id):
    check = permiso_requerido("calendario_gestionar", "asistencia_gestionar")
    if check:
        return check
    origen = request.form.get("origen") or request.args.get("origen", "")
    if origen != "asistencia":
        origen = ""
    if not tiene_permiso("calendario_gestionar") and origen != "asistencia":
        flash("No tenes permiso para acceder a esa seccion.", "error")
        return redirect(url_for("index"))

    conn = get_connection()
    evento = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE id = %s
    """, (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("ver_calendario"))
    if not tiene_permiso("calendario_gestionar") and not evento.get("asistencia_evento_id"):
        conn.close()
        flash("No tenes permiso para acceder a esa seccion.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        data = {
            "fecha": request.form.get("fecha", "").strip(),
            "hora_inicio": normalizar_hora_evento(request.form.get("hora_inicio", "")),
            "duracion_minutos": normalizar_duracion_evento(request.form.get("duracion_minutos", "90")),
            "tipo": request.form.get("tipo", "").strip(),
            "titulo": request.form.get("titulo", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "ubicacion": request.form.get("ubicacion", "").strip(),
            "categoria": request.form.get("categoria", "").strip(),
            "convocatoria_texto": request.form.get("convocatoria_texto", "").strip(),
            "convocatoria_cierre": request.form.get("convocatoria_cierre", "").strip(),
            "minuta_post_evento": request.form.get("minuta_post_evento", "").strip(),
            "publicar_portal": 1 if request.form.get("publicar_portal") == "on" else 0,
            "crear_asistencia": 1 if request.form.get("crear_asistencia") == "on" else 0,
            "origen": origen,
        }

        if calendario_evento_requiere_asistencia(data["tipo"]):
            data["crear_asistencia"] = 1 if request.form.get("crear_asistencia", "on") == "on" else 0
        if origen == "asistencia":
            data["crear_asistencia"] = 1

        if not data["fecha"] or not data["tipo"] or not data["titulo"]:
            conn.close()
            flash("Fecha, tipo y titulo son obligatorios.", "error")
            data["id"] = evento_id
            return render_template("calendario_evento_form.html", evento=data)

        try:
            datetime.strptime(data["fecha"], "%Y-%m-%d")
        except ValueError:
            conn.close()
            flash("La fecha del evento no es valida.", "error")
            data["id"] = evento_id
            return render_template("calendario_evento_form.html", evento=data)

        if request.form.get("hora_inicio") and not data["hora_inicio"]:
            conn.close()
            flash("La hora debe tener formato HH:MM.", "error")
            data["id"] = evento_id
            return render_template("calendario_evento_form.html", evento=data)

        asistencia_evento_id = evento.get("asistencia_evento_id")
        if data["crear_asistencia"]:
            if asistencia_evento_id:
                conn.execute("""
                    UPDATE eventos_asistencia
                    SET fecha = %s,
                        tipo = %s,
                        descripcion = %s
                    WHERE id = %s
                """, (
                    data["fecha"],
                    data["tipo"],
                    data["titulo"] if data["titulo"] != data["tipo"] else (data["descripcion"] or ""),
                    asistencia_evento_id,
                ))
            else:
                asistencia_evento_id = crear_evento_asistencia_desde_calendario(conn, data)
        elif asistencia_evento_id:
            conn.execute("DELETE FROM portal_asistencia_confirmaciones WHERE evento_id = %s", (asistencia_evento_id,))
            conn.execute("DELETE FROM asistencias WHERE evento_id = %s", (asistencia_evento_id,))
            conn.execute("DELETE FROM aspirante_asistencias WHERE evento_id = %s", (asistencia_evento_id,))
            conn.execute("DELETE FROM eventos_asistencia WHERE id = %s", (asistencia_evento_id,))
            asistencia_evento_id = None

        conn.execute("""
            UPDATE calendario_eventos
            SET fecha = %s,
                tipo = %s,
                titulo = %s,
                descripcion = %s,
                ubicacion = %s,
                categoria = %s,
                hora_inicio = %s,
                duracion_minutos = %s,
                publicar_portal = %s,
                asistencia_evento_id = %s,
                convocatoria_texto = %s,
                convocatoria_cierre = %s,
                minuta_post_evento = %s
            WHERE id = %s
        """, (
            data["fecha"],
            data["tipo"],
            data["titulo"],
            data["descripcion"],
            data["ubicacion"],
            data["categoria"],
            data["hora_inicio"] or None,
            data["duracion_minutos"],
            data["publicar_portal"],
            asistencia_evento_id,
            data["convocatoria_texto"] or None,
            data["convocatoria_cierre"] or None,
            data["minuta_post_evento"] or None,
            evento_id,
        ))
        conn.commit()
        conn.close()

        registrar_auditoria("editar", "calendario_evento", str(evento_id), {
            "tipo": data["tipo"],
            "fecha": data["fecha"],
            "asistencia_evento_id": asistencia_evento_id,
        })

        flash("Evento actualizado.", "ok")
        if origen == "asistencia":
            return redirect(url_for("listar_eventos_asistencia"))
        return redirect(url_for("ver_calendario", mes=data["fecha"][:7]))

    conn.close()
    evento = dict(evento)
    evento["crear_asistencia"] = 1 if evento.get("asistencia_evento_id") else 0
    evento["origen"] = origen
    return render_template("calendario_evento_form.html", evento=evento)


@app.route("/calendario/<int:evento_id>/eliminar", methods=["POST"])
def eliminar_evento_calendario(evento_id):
    check = permiso_requerido("calendario_gestionar")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE id = %s
    """, (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("ver_calendario"))

    asistencia_evento_id = evento.get("asistencia_evento_id")
    if asistencia_evento_id:
        conn.execute("DELETE FROM portal_asistencia_confirmaciones WHERE evento_id = %s", (asistencia_evento_id,))
        conn.execute("DELETE FROM asistencias WHERE evento_id = %s", (asistencia_evento_id,))
        conn.execute("DELETE FROM aspirante_asistencias WHERE evento_id = %s", (asistencia_evento_id,))
        conn.execute("DELETE FROM eventos_asistencia WHERE id = %s", (asistencia_evento_id,))

    conn.execute("DELETE FROM calendario_eventos WHERE id = %s", (evento_id,))
    conn.commit()
    conn.close()

    registrar_auditoria("eliminar", "calendario_evento", str(evento_id), {
        "fecha": evento.get("fecha"),
        "tipo": evento.get("tipo"),
        "titulo": evento.get("titulo"),
        "asistencia_evento_id": asistencia_evento_id,
    })

    flash("Evento eliminado.", "ok")
    return redirect(url_for("ver_calendario", mes=(evento["fecha"] or ahora_sig().strftime("%Y-%m-%d"))[:7]))


@app.route("/calendario/<int:evento_id>/recordatorio", methods=["POST"])
def enviar_recordatorio_evento_calendario(evento_id):
    check = permiso_requerido("comunicaciones_ver", "calendario_gestionar")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("""
        SELECT *
        FROM calendario_eventos
        WHERE id = %s
    """, (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("ver_calendario"))

    jugadores = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE estado = 'Activo'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY apellido, nombre
    """).fetchall()
    conn.close()

    enviados = 0
    for jugador in jugadores:
        if not categoria_evento_aplica(evento.get("categoria"), jugador.get("categoria")):
            continue
        portal_url = None
        if jugador.get("portal_token") and jugador.get("portal_activo") and evento.get("asistencia_evento_id"):
            portal_url = url_for("portal_jugador", token=jugador["portal_token"], _external=True)
        cuerpo = construir_texto_recordatorio_evento(jugador, evento, portal_url=portal_url)
        enviado, _, _ = enviar_email_jugador(jugador, f"Recordatorio: {evento['titulo']}", cuerpo)
        enviados += 1 if enviado else 0

    registrar_auditoria("enviar_recordatorio", "calendario_evento", str(evento_id), {"cantidad": enviados})
    flash(f"Se enviaron {enviados} recordatorios del evento.", "ok" if enviados else "error")
    return redirect(url_for("ver_calendario", mes=(evento["fecha"] or ahora_sig().strftime("%Y-%m-%d"))[:7]))


@app.route("/alertas")
def ver_alertas():
    check = permiso_requerido("alertas_finanzas", "alertas_salud")
    if check:
        return check

    puede_ver_finanzas = tiene_permiso("alertas_finanzas")
    puede_ver_salud = tiene_permiso("alertas_salud")
    alertas = obtener_alertas()
    alertas = filtrar_alertas_por_permisos(
        alertas,
        puede_ver_finanzas=puede_ver_finanzas,
        puede_ver_salud=puede_ver_salud,
    )
    return render_template(
        "alertas.html",
        alertas=alertas,
        puede_ver_finanzas=puede_ver_finanzas,
        puede_ver_salud=puede_ver_salud,
    )


@app.route("/salud")
def panel_salud():
    check = permiso_requerido("salud_ver")
    if check:
        return check

    return render_template("salud_panel.html", panel=obtener_panel_salud())


@app.route("/reportes")
def ver_reportes():
    check = permiso_requerido("reportes_ver")
    if check:
        return check

    filtros = filtros_reportes()
    reportes = obtener_reportes(filtros["desde"], filtros["hasta"])

    return render_template(
        "reportes.html",
        filtros=filtros,
        reportes=reportes,
    )


@app.route("/urba/circulares")
def listar_circulares_urba():
    anios = anios_circulares_urba()
    anio = request.args.get("anio", str(ahora_sig().year)).strip()
    anio = int(anio) if anio.isdigit() and int(anio) in anios else ahora_sig().year
    conn = get_connection()
    circulares = conn.execute("""
        SELECT *
        FROM urba_circulares
        WHERE anio = %s
        ORDER BY COALESCE(orden_fuente, 9999), id
    """, (anio,)).fetchall()
    usuarios = []
    config = obtener_config_circulares_urba(conn)
    if session.get("rol") == "admin":
        usuarios = conn.execute("""
            SELECT id, username, email, rol
            FROM usuarios
            WHERE email IS NOT NULL
              AND trim(email) <> ''
            ORDER BY username
        """).fetchall()
    resumen = conn.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE nueva = 1) AS nuevas,
            COUNT(DISTINCT anio) AS anios_sincronizados
        FROM urba_circulares
    """).fetchone()
    conn.close()
    return render_template(
        "circulares_urba.html",
        anio=anio,
        anios=anios,
        circulares=circulares,
        config=config,
        usuarios_notificables=usuarios,
        resumen=resumen,
    )


@app.route("/urba/circulares/sync", methods=["POST"])
def sincronizar_circulares_urba_view():
    if session.get("rol") != "admin":
        flash("Solo admins pueden sincronizar circulares URBA.", "error")
        return redirect(url_for("listar_circulares_urba"))
    anio_raw = (request.form.get("anio") or "").strip()
    anios = anios_circulares_urba()
    anio = int(anio_raw) if anio_raw.isdigit() and int(anio_raw) in anios else ahora_sig().year
    conn = get_connection()
    try:
        resultado = sincronizar_circulares_urba(conn, anio, session.get("username"))
        notificaciones = enviar_notificacion_circulares_urba(conn, resultado["nuevas"], session.get("username"))
        conn.commit()
    except RuntimeError as error:
        conn.rollback()
        conn.close()
        flash(str(error), "error")
        return redirect(url_for("listar_circulares_urba", anio=anio))
    conn.close()
    extras = []
    if resultado["nuevas"]:
        extras.append(f"{len(resultado['nuevas'])} nueva(s)")
    if notificaciones:
        enviados = sum(1 for _, ok, _ in notificaciones if ok)
        extras.append(f"{enviados} email(s) enviados")
    detalle = f" ({', '.join(extras)})" if extras else ""
    flash(f"Sincronizacion URBA {anio} completada{detalle}.", "ok")
    return redirect(url_for("listar_circulares_urba", anio=anio))


@app.route("/urba/circulares/notificaciones", methods=["POST"])
def guardar_notificaciones_circulares_urba():
    if session.get("rol") != "admin":
        flash("Solo admins pueden configurar notificaciones de circulares URBA.", "error")
        return redirect(url_for("listar_circulares_urba"))
    seleccion = [int(valor) for valor in request.form.getlist("notify_user_ids") if str(valor).isdigit()]
    conn = get_connection()
    guardar_app_setting(
        conn,
        "urba_circulares_notify_user_ids",
        json.dumps(seleccion, ensure_ascii=False),
        session.get("username"),
    )
    conn.commit()
    conn.close()
    flash("Usuarios de notificacion automatica actualizados.", "ok")
    return redirect(url_for("listar_circulares_urba", anio=request.form.get("anio") or ahora_sig().year))


def estilizar_hoja_reporte(ws, header_row=1):
    encabezado = PatternFill("solid", fgColor="1F2937")
    thin = Side(style="thin", color="D1D5DB")

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = encabezado
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def agregar_hoja_reporte(wb, titulo, encabezados, filas):
    ws = wb.create_sheet(titulo)
    ws.append(encabezados)
    for fila in filas:
        append_fila_reporte(ws, fila)
    estilizar_hoja_reporte(ws)
    return ws


def normalizar_valor_excel_reporte(valor):
    if isinstance(valor, datetime) and valor.tzinfo is not None and valor.utcoffset() is not None:
        return valor.astimezone(APP_TZ).strftime("%Y-%m-%d %H:%M:%S")
    return valor


def append_fila_reporte(ws, fila):
    ws.append([normalizar_valor_excel_reporte(valor) for valor in fila])


def aplicar_formato_columnas(ws, formatos):
    for columna, formato in formatos.items():
        for cell in ws[columna][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = formato


@app.route("/reportes/exportar")
def exportar_reportes():
    check = permiso_requerido("reportes_ver")
    if check:
        return check

    filtros = filtros_reportes()
    reportes = obtener_reportes(filtros["desde"], filtros["hasta"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Reporte", "Valor"])
    append_fila_reporte(ws, ["Periodo", f"{filtros['desde']} a {filtros['hasta']}"])
    append_fila_reporte(ws, ["Ingresos", reportes["resumen"]["ingresos"]])
    append_fila_reporte(ws, ["Egresos", reportes["resumen"]["egresos"]])
    append_fila_reporte(ws, ["Resultado", reportes["resumen"]["resultado"]])
    append_fila_reporte(ws, ["Cuotas cobradas", reportes["resumen"]["cuotas_cobradas"]])
    append_fila_reporte(ws, ["Cuotas pagadas", reportes["resumen"]["cuotas_pagadas"]])
    append_fila_reporte(ws, ["Deuda total pendiente", reportes["resumen"]["deuda"]])
    append_fila_reporte(ws, ["Deuda vencida", reportes["resumen"]["deuda_vencida"]])
    append_fila_reporte(ws, ["Total bonificado por becas", reportes["resumen"]["total_bonificado_becas"]])
    append_fila_reporte(ws, ["Cuotas becadas", reportes["resumen"]["cuotas_becadas"]])
    append_fila_reporte(ws, ["Becas totales", reportes["resumen"]["becas_totales"]])
    append_fila_reporte(ws, ["Becas parciales", reportes["resumen"]["becas_parciales"]])
    append_fila_reporte(ws, ["Jugadores activos", reportes["resumen"]["jugadores_activos"]])
    append_fila_reporte(ws, ["Asistencia promedio", f"{reportes['resumen']['asistencia_porcentaje']}%"])
    estilizar_hoja_reporte(ws)
    for row in (3, 4, 5, 6, 8, 9):
        ws.cell(row=row, column=2).number_format = '$ #,##0'
    for row in (7, 10):
        ws.cell(row=row, column=2).number_format = '#,##0'

    ws_mensual = agregar_hoja_reporte(
        wb,
        "Mensual",
        [
            "Mes", "Ingresos", "Egresos", "Resultado", "Movimientos",
            "Cuotas emitidas", "Cuotas pagadas", "Cuotas pendientes",
            "Total emitido", "Total cobrado", "Cuotas becadas", "Becas totales",
            "Becas parciales", "Total bonificado"
        ],
        [
            [
                fila["mes"],
                fila["ingresos"],
                fila["egresos"],
                fila["resultado"],
                fila["movimientos"],
                fila["cuotas_emitidas"],
                fila["cuotas_pagadas"],
                fila["cuotas_pendientes"],
                fila["total_emitido"],
                fila["total_cobrado"],
                fila["cuotas_becadas"],
                fila["becas_totales"],
                fila["becas_parciales"],
                fila["total_bonificado"],
            ]
            for fila in reportes["mensual"]
        ],
    )
    aplicar_formato_columnas(ws_mensual, {
        "B": '$ #,##0',
        "C": '$ #,##0',
        "D": '$ #,##0',
        "E": '#,##0',
        "F": '#,##0',
        "G": '#,##0',
        "H": '#,##0',
        "I": '$ #,##0',
        "J": '$ #,##0',
        "K": '#,##0',
        "L": '#,##0',
        "M": '#,##0',
        "N": '$ #,##0',
    })

    ws_becas = agregar_hoja_reporte(
        wb,
        "Becas",
        [
            "Apellido", "Nombre", "Categoria", "Porcentaje", "Desde", "Hasta",
            "Cuotas becadas", "Total bonificado", "Motivo"
        ],
        [
            [
                fila["apellido"],
                fila["nombre"],
                fila["categoria"],
                fila["beca_porcentaje"],
                fila["beca_desde"],
                fila["beca_hasta"],
                fila["cuotas_becadas"],
                fila["total_bonificado"],
                fila["beca_motivo"],
            ]
            for fila in reportes["becas_jugadores"]
        ],
    )
    aplicar_formato_columnas(ws_becas, {
        "D": '0.00',
        "G": '#,##0',
        "H": '$ #,##0',
    })

    ws_deuda = agregar_hoja_reporte(
        wb,
        "Deuda categoria",
        ["Categoria", "Jugadores", "Cuotas pendientes", "Deuda", "Deuda vencida"],
        [
            [
                fila["categoria"],
                fila["jugadores"],
                fila["cuotas_pendientes"],
                fila["deuda"],
                fila["deuda_vencida"],
            ]
            for fila in reportes["deuda_por_categoria"]
        ],
    )
    aplicar_formato_columnas(ws_deuda, {
        "B": '#,##0',
        "C": '#,##0',
        "D": '$ #,##0',
        "E": '$ #,##0',
    })

    ws_egresos = agregar_hoja_reporte(
        wb,
        "Egresos concepto",
        ["Concepto", "Cantidad", "Total"],
        [
            [fila["concepto"], fila["cantidad"], fila["total"]]
            for fila in reportes["egresos_por_concepto"]
        ],
    )
    aplicar_formato_columnas(ws_egresos, {
        "B": '#,##0',
        "C": '$ #,##0',
    })

    ws_morosos = agregar_hoja_reporte(
        wb,
        "Morosos",
        [
            "Apellido", "Nombre", "Categoria", "Telefono", "Email",
            "Cuotas pendientes", "Cuotas vencidas", "Deuda", "Primer vencimiento"
        ],
        [
            [
                fila["apellido"],
                fila["nombre"],
                fila["categoria"],
                fila["telefono"],
                fila["email"],
                fila["cuotas_pendientes"],
                fila["cuotas_vencidas"],
                fila["deuda"],
                fila["primer_vencimiento"],
            ]
            for fila in reportes["morosos_recurrentes"]
        ],
    )
    aplicar_formato_columnas(ws_morosos, {
        "F": '#,##0',
        "G": '#,##0',
        "H": '$ #,##0',
    })

    ws_asistencia = agregar_hoja_reporte(
        wb,
        "Asistencia",
        ["Categoria", "Eventos", "Registros", "Presentes", "Ausentes", "Porcentaje"],
        [
            [
                fila["categoria"],
                fila["eventos"],
                fila["registros"],
                fila["presentes"],
                fila["ausentes"],
                fila["porcentaje"],
            ]
            for fila in reportes["asistencia_por_categoria"]
        ],
    )
    aplicar_formato_columnas(ws_asistencia, {
        "B": '#,##0',
        "C": '#,##0',
        "D": '#,##0',
        "E": '#,##0',
        "F": '0.0',
    })

    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(exist_ok=True)

    archivo = export_dir / f"reportes_{filtros['desde']}_a_{filtros['hasta']}.xlsx"
    wb.save(archivo)

    registrar_auditoria(
        "exportar_ok",
        "reportes",
        f"{filtros['desde']}:{filtros['hasta']}",
        {"formato": "xlsx"},
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"reportes_{filtros['desde']}_a_{filtros['hasta']}.xlsx"
    )


@app.route("/exportar/datos")
def exportar_datos_integral():
    check = permiso_requerido("reportes_ver")
    if check:
        return check

    conn = get_connection()

    jugadores = conn.execute("""
        SELECT
            id, apellido, nombre, dni, categoria, estado, fecha_ingreso,
            telefono, email, contacto_tutor, telefono_tutor, email_tutor,
            beca_activa, beca_porcentaje, beca_desde, beca_hasta, beca_motivo
        FROM jugadores
        ORDER BY apellido, nombre
    """).fetchall()

    deudores = conn.execute("""
        SELECT
            j.id,
            j.apellido,
            j.nombre,
            j.dni,
            j.categoria,
            j.telefono,
            j.email,
            COUNT(c.id) AS cuotas_pendientes,
            SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1 ELSE 0
                END
            ) AS cuotas_vencidas,
            COALESCE(SUM(c.importe), 0) AS deuda,
            MIN(c.fecha_vencimiento) AS primer_vencimiento
        FROM jugadores j
        JOIN cuotas c ON c.jugador_id = j.id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY j.id, j.apellido, j.nombre, j.dni, j.categoria, j.telefono, j.email
        HAVING COALESCE(SUM(c.importe), 0) > 0
        ORDER BY deuda DESC, j.apellido, j.nombre
    """).fetchall()

    asistencia = conn.execute("""
        SELECT
            j.apellido,
            j.nombre,
            j.categoria,
            COUNT(a.id) AS registros,
            SUM(CASE WHEN COALESCE(a.presente, 0) = 1 THEN 1 ELSE 0 END) AS presentes,
            SUM(CASE WHEN COALESCE(a.presente, 0) = 0 THEN 1 ELSE 0 END) AS ausentes,
            ROUND(
                CASE WHEN COUNT(a.id) = 0 THEN 0
                ELSE (SUM(CASE WHEN COALESCE(a.presente, 0) = 1 THEN 1 ELSE 0 END)::numeric / COUNT(a.id)) * 100
                END,
                1
            ) AS porcentaje
        FROM jugadores j
        LEFT JOIN asistencias a ON a.jugador_id = j.id
        GROUP BY j.id, j.apellido, j.nombre, j.categoria
        ORDER BY j.apellido, j.nombre
    """).fetchall()

    becas = conn.execute("""
        SELECT
            j.apellido,
            j.nombre,
            j.categoria,
            j.beca_activa,
            j.beca_porcentaje,
            j.beca_desde,
            j.beca_hasta,
            j.beca_motivo,
            COUNT(c.id) FILTER (WHERE COALESCE(c.becada, 0) = 1) AS cuotas_becadas,
            COALESCE(SUM(c.descuento_beca), 0) AS total_bonificado
        FROM jugadores j
        LEFT JOIN cuotas c ON c.jugador_id = j.id
        WHERE COALESCE(j.beca_activa, 0) = 1
           OR COALESCE(c.becada, 0) = 1
        GROUP BY
            j.id, j.apellido, j.nombre, j.categoria, j.beca_activa,
            j.beca_porcentaje, j.beca_desde, j.beca_hasta, j.beca_motivo
        ORDER BY j.apellido, j.nombre
    """).fetchall()

    fichas = conn.execute("""
        SELECT
            j.apellido,
            j.nombre,
            j.categoria,
            f.presentada,
            f.apto_fisico,
            f.fecha_vencimiento,
            f.contacto_emergencia,
            f.telefono_emergencia,
            f.documento_nombre,
            f.ocr_fecha,
            f.observaciones
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        ORDER BY j.apellido, j.nombre
    """).fetchall()

    auditoria = []
    if tiene_permiso("auditoria_ver"):
        auditoria = conn.execute("""
            SELECT fecha, username, rol, accion, entidad, entidad_id, ip
            FROM auditoria
            ORDER BY fecha DESC, id DESC
            LIMIT 2000
        """).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Jugadores"
    ws.append([
        "ID", "Apellido", "Nombre", "DNI", "Categoria", "Estado", "Ingreso",
        "Telefono", "Email", "Contacto tutor", "Telefono tutor", "Email tutor",
        "Beca activa", "Beca %", "Beca desde", "Beca hasta", "Motivo beca"
    ])
    for jugador in jugadores:
        append_fila_reporte(ws, [
            jugador["id"],
            jugador["apellido"],
            jugador["nombre"],
            jugador["dni"],
            jugador["categoria"],
            jugador["estado"],
            jugador["fecha_ingreso"],
            jugador["telefono"],
            jugador["email"],
            jugador["contacto_tutor"],
            jugador["telefono_tutor"],
            jugador["email_tutor"],
            "Si" if jugador["beca_activa"] else "No",
            jugador["beca_porcentaje"],
            jugador["beca_desde"],
            jugador["beca_hasta"],
            jugador["beca_motivo"],
        ])
    estilizar_hoja_reporte(ws)

    ws_deudores = agregar_hoja_reporte(
        wb,
        "Deudores",
        [
            "ID", "Apellido", "Nombre", "DNI", "Categoria", "Telefono", "Email",
            "Cuotas pendientes", "Cuotas vencidas", "Deuda", "Primer vencimiento"
        ],
        [
            [
                fila["id"],
                fila["apellido"],
                fila["nombre"],
                fila["dni"],
                fila["categoria"],
                fila["telefono"],
                fila["email"],
                fila["cuotas_pendientes"],
                fila["cuotas_vencidas"],
                fila["deuda"],
                fila["primer_vencimiento"],
            ]
            for fila in deudores
        ],
    )
    aplicar_formato_columnas(ws_deudores, {"H": '#,##0', "I": '#,##0', "J": '$ #,##0'})

    ws_asistencia = agregar_hoja_reporte(
        wb,
        "Asistencia",
        ["Apellido", "Nombre", "Categoria", "Registros", "Presentes", "Ausentes", "Porcentaje"],
        [
            [
                fila["apellido"],
                fila["nombre"],
                fila["categoria"],
                fila["registros"],
                fila["presentes"],
                fila["ausentes"],
                fila["porcentaje"],
            ]
            for fila in asistencia
        ],
    )
    aplicar_formato_columnas(ws_asistencia, {"D": '#,##0', "E": '#,##0', "F": '#,##0', "G": '0.0'})

    ws_becas = agregar_hoja_reporte(
        wb,
        "Becas",
        [
            "Apellido", "Nombre", "Categoria", "Activa", "Porcentaje", "Desde",
            "Hasta", "Motivo", "Cuotas becadas", "Total bonificado"
        ],
        [
            [
                fila["apellido"],
                fila["nombre"],
                fila["categoria"],
                "Si" if fila["beca_activa"] else "No",
                fila["beca_porcentaje"],
                fila["beca_desde"],
                fila["beca_hasta"],
                fila["beca_motivo"],
                fila["cuotas_becadas"],
                fila["total_bonificado"],
            ]
            for fila in becas
        ],
    )
    aplicar_formato_columnas(ws_becas, {"E": '0.00', "I": '#,##0', "J": '$ #,##0'})

    agregar_hoja_reporte(
        wb,
        "Fichas médicas",
        [
            "Apellido", "Nombre", "Categoria", "Presentada", "Apto fisico",
            "Vencimiento", "Contacto emergencia", "Telefono emergencia",
            "Documento", "OCR fecha", "Observaciones"
        ],
        [
            [
                fila["apellido"],
                fila["nombre"],
                fila["categoria"],
                "Si" if fila["presentada"] else "No",
                "Si" if fila["apto_fisico"] else "No",
                fila["fecha_vencimiento"],
                fila["contacto_emergencia"],
                fila["telefono_emergencia"],
                fila["documento_nombre"],
                fila["ocr_fecha"],
                fila["observaciones"],
            ]
            for fila in fichas
        ],
    )

    if auditoria:
        agregar_hoja_reporte(
            wb,
            "Auditoria",
            ["Fecha", "Usuario", "Rol", "Accion", "Entidad", "Entidad ID", "IP"],
            [
                [
                    fila["fecha"],
                    fila["username"],
                    fila["rol"],
                    fila["accion"],
                    fila["entidad"],
                    fila["entidad_id"],
                    fila["ip"],
                ]
                for fila in auditoria
            ],
        )

    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(exist_ok=True)
    fecha = ahora_sig().strftime("%Y%m%d_%H%M")
    archivo = export_dir / f"sig_export_integral_{fecha}.xlsx"
    wb.save(archivo)

    registrar_auditoria(
        "exportar_ok",
        "datos_integrales",
        None,
        {
            "formato": "xlsx",
            "jugadores": len(jugadores),
            "deudores": len(deudores),
            "incluye_auditoria": bool(auditoria),
        },
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"sig_export_integral_{fecha}.xlsx",
    )


def template_morosos_default():
    return (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situacion o avisarnos si ya realizaste el pago. Gracias."
    )


def construir_texto_recordatorio_cuota(cuota):
    nombre = nombre_jugador_corto(cuota)
    estado = "vencio" if (cuota.get("dias_vencida") or 0) > 0 else "vence"
    fecha = cuota.get("fecha_vencimiento") or "-"
    return (
        f"Hola {nombre}, te escribimos de Ruda Macho Rugby Club.\n\n"
        f"La cuota {cuota.get('periodo') or '-'} por {formato_moneda(cuota.get('importe') or 0)} {estado} el {fecha}.\n"
        "Si ya realizaste el pago, podes responder este mensaje o cargar el comprobante desde tu portal.\n\n"
        "Gracias."
    )


def construir_texto_recordatorio_ficha(ficha):
    nombre = nombre_jugador_corto(ficha)
    if ficha.get("estado_documento") == "vencida":
        estado = f"vencio el {ficha.get('fecha_vencimiento') or '-'}"
    elif ficha.get("estado_documento") == "por_vencer":
        estado = f"vence el {ficha.get('fecha_vencimiento') or '-'}"
    else:
        estado = "figura pendiente de carga"
    return (
        f"Hola {nombre}, te escribimos de Ruda Macho Rugby Club.\n\n"
        f"La ficha m?dica {estado}. Cuando puedas, acercanos la actualizaci?n o cargala por los canales habituales.\n\n"
        "Gracias."
    )


@app.route("/comunicaciones")
def ver_comunicaciones():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    template_default = (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situación o avisarnos si ya realizaste el pago. Gracias."
    )
    template_default = template_morosos_default()
    template = request.args.get("mensaje", template_default).strip() or template_default
    morosos = obtener_morosos_para_comunicacion()

    comunicaciones = []
    for jugador in morosos:
        mensaje = mensaje_moroso(template, jugador)
        telefono = valor_texto_contacto(jugador["telefono"]) or valor_texto_contacto(jugador["telefono_tutor"])
        telefono_whatsapp = normalizar_telefono_whatsapp(telefono)
        if telefono_whatsapp:
            whatsapp_url = f"https://wa.me/{telefono_whatsapp}?text={quote(mensaje)}"
        else:
            whatsapp_url = f"https://wa.me/?text={quote(mensaje)}"

        comunicaciones.append({
            "jugador": jugador,
            "mensaje": mensaje,
            "email": email_jugador_preferido(jugador),
            "telefono_whatsapp": telefono_whatsapp,
            "whatsapp_url": whatsapp_url,
        })

    return render_template(
        "comunicaciones.html",
        template=template,
        comunicaciones=comunicaciones,
    )


@app.route("/comunicaciones/<int:jugador_id>/email", methods=["POST"])
def enviar_email_comunicacion_moroso(jugador_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    template_default = (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situaci\u00f3n o avisarnos si ya realizaste el pago. Gracias."
    )
    template_default = template_morosos_default()
    template = request.form.get("mensaje", template_default).strip() or template_default
    jugadores = obtener_morosos_para_comunicacion()
    jugador = next((item for item in jugadores if item["id"] == jugador_id), None)
    if jugador is None:
        flash("Jugador no encontrado en el listado de deuda.", "error")
        return redirect(url_for("ver_comunicaciones", mensaje=template))

    mensaje = mensaje_moroso(template, jugador)
    asunto = f"Estado de cuotas - {jugador['apellido']}, {jugador['nombre']}"
    enviado, destinatario, motivo = enviar_email_jugador(jugador, asunto, mensaje)
    if enviado:
        registrar_auditoria("enviar_recordatorio", "moroso", str(jugador_id), {"destinatario": destinatario, "tipo": "comunicacion_moroso"})
        flash("Email enviado.", "ok")
    else:
        flash(mensaje_fallo_email(motivo, destinatario), "error")
    return redirect(url_for("ver_comunicaciones", mensaje=template))


@app.route("/comunicaciones/email-lote", methods=["POST"])
def enviar_email_comunicacion_morosos_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    template_default = (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situaci\u00f3n o avisarnos si ya realizaste el pago. Gracias."
    )
    template_default = template_morosos_default()
    template = request.form.get("mensaje", template_default).strip() or template_default
    jugadores = obtener_morosos_para_comunicacion()
    resultados = []
    for jugador in jugadores:
        mensaje = mensaje_moroso(template, jugador)
        asunto = f"Estado de cuotas - {jugador['apellido']}, {jugador['nombre']}"
        resultados.append(enviar_email_jugador(jugador, asunto, mensaje))

    enviados = sum(1 for ok, _, _ in resultados if ok)
    registrar_auditoria("enviar_recordatorio", "morosos", None, {"cantidad": enviados, "tipo": "comunicacion_morosos"})
    mensaje_resultado, nivel = resumir_envio_masivo_email(resultados, "emails de comunicaci\u00f3n")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_comunicaciones", mensaje=template))


@app.route("/comunicaciones/<int:jugador_id>/whatsapp", methods=["POST"])
def enviar_whatsapp_comunicacion_moroso(jugador_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    template_default = (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situacion o avisarnos si ya realizaste el pago. Gracias."
    )
    template_default = template_morosos_default()
    template = request.form.get("mensaje", template_default).strip() or template_default
    jugadores = obtener_morosos_para_comunicacion()
    jugador = next((item for item in jugadores if item["id"] == jugador_id), None)
    if jugador is None:
        flash("Jugador no encontrado en el listado de deuda.", "error")
        return redirect(url_for("ver_comunicaciones", mensaje=template))

    enviado, destinatario, motivo, detalle = enviar_whatsapp_recordatorio_cuota_template(
        jugador,
        tipo="comunicacion_moroso",
        entidad="moroso",
        entidad_id=str(jugador_id),
    )
    if enviado:
        registrar_auditoria("enviar_recordatorio", "moroso", str(jugador_id), {"destinatario": destinatario, "tipo": "comunicacion_moroso_whatsapp"})
        flash("WhatsApp enviado.", "ok")
    else:
        flash(mensaje_fallo_whatsapp(motivo, destinatario, detalle), "error")
    return redirect(url_for("ver_comunicaciones", mensaje=template))


@app.route("/comunicaciones/whatsapp-lote", methods=["POST"])
def enviar_whatsapp_comunicacion_morosos_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    template_default = (
        "Hola {nombre}, te escribimos de Ruda Macho Rugby Club. "
        "Registramos {cuotas_pendientes} cuota(s) pendiente(s) por {deuda}. "
        "Te pedimos regularizar la situacion o avisarnos si ya realizaste el pago. Gracias."
    )
    template_default = template_morosos_default()
    template = request.form.get("mensaje", template_default).strip() or template_default
    jugadores = obtener_morosos_para_comunicacion()
    resultados = []
    for jugador in jugadores:
        resultados.append(enviar_whatsapp_recordatorio_cuota_template(
            jugador,
            tipo="comunicacion_moroso",
            entidad="moroso",
            entidad_id=str(jugador["id"]),
        ))

    enviados = sum(1 for ok, _, _, _ in resultados if ok)
    registrar_auditoria("enviar_recordatorio", "morosos", None, {"cantidad": enviados, "tipo": "comunicacion_morosos_whatsapp"})
    mensaje_resultado, nivel = resumir_envio_masivo_whatsapp(resultados, "WhatsApps de comunicacion")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_comunicaciones", mensaje=template))


@app.route("/comunicaciones/app", methods=["GET", "POST"])
def enviar_notificacion_app_manual():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()
    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        mensaje = request.form.get("mensaje", "").strip()
        destino = request.form.get("destino", "portal_todos").strip()
        categoria = request.form.get("categoria", "").strip()
        jugador_id_raw = request.form.get("jugador_id", "").strip()
        jugador_id = int(jugador_id_raw) if jugador_id_raw.isdigit() else None
        url = normalizar_url_push(request.form.get("url"), fallback=url_for("index"))
        mostrar_portal = 1 if request.form.get("mostrar_portal") == "on" else 0
        visible_hasta = validar_fecha_movimiento(request.form.get("visible_hasta", "").strip()) or ahora_sig().strftime("%Y-%m-%d")

        if not titulo or not mensaje:
            conn.close()
            flash("Titulo y mensaje son obligatorios.", "error")
            return redirect(url_for("enviar_notificacion_app_manual"))
        if destino == "categoria" and not categoria:
            conn.close()
            flash("Elegí una categoria para enviar la notificacion.", "error")
            return redirect(url_for("enviar_notificacion_app_manual"))
        if destino == "jugador" and not jugador_id:
            conn.close()
            flash("Elegí un jugador para enviar la notificacion.", "error")
            return redirect(url_for("enviar_notificacion_app_manual"))

        destinatarios = obtener_destinatarios_push_manual(conn, destino, categoria=categoria, jugador_id=jugador_id)
        payload = {
            "title": titulo,
            "body": mensaje,
            "url": url,
            "icon": pwa_icon_url("192"),
        }
        enviados = 0
        errores = []
        for destinatario in destinatarios:
            try:
                subscription = json.loads(destinatario["subscription_json"] or "{}")
            except (TypeError, ValueError):
                subscription = {}
            ok, error = enviar_push_subscription(subscription, payload)
            if ok:
                enviados += 1
            else:
                errores.append(error or "Error desconocido")
                if error and ("410" in error or "404" in error):
                    desactivar_suscripcion_push(conn, destinatario["endpoint"])

        envio = conn.execute("""
            INSERT INTO pwa_push_envios (
                titulo, mensaje, destino, categoria, jugador_id, url,
                enviados, errores, detalle, creado_por, mostrar_portal, visible_hasta
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            titulo,
            mensaje,
            destino,
            categoria or None,
            jugador_id,
            url,
            enviados,
            len(errores),
            json.dumps({"errores": errores[:20], "destinatarios": len(destinatarios)}, ensure_ascii=False),
            session.get("username"),
            mostrar_portal,
            visible_hasta,
        )).fetchone()
        conn.commit()
        conn.close()

        registrar_auditoria("enviar", "pwa_push_manual", str(envio["id"]), {
            "destino": destino,
            "categoria": categoria,
            "jugador_id": jugador_id,
            "destinatarios": len(destinatarios),
            "enviados": enviados,
            "errores": len(errores),
            "mostrar_portal": bool(mostrar_portal),
            "visible_hasta": visible_hasta,
        })
        if not destinatarios:
            flash("No hay dispositivos suscriptos para ese destino.", "warning")
        elif errores:
            flash(f"Notificacion enviada a {enviados} dispositivo(s), con {len(errores)} error(es).", "warning")
        else:
            flash(f"Notificacion enviada a {enviados} dispositivo(s).", "ok")
        return redirect(url_for("enviar_notificacion_app_manual"))

    categorias = conn.execute("""
        SELECT DISTINCT COALESCE(NULLIF(categoria, ''), 'Sin categoria') AS categoria
        FROM jugadores
        ORDER BY categoria
    """).fetchall()
    jugadores = conn.execute("""
        SELECT id, apellido, nombre, categoria
        FROM jugadores
        WHERE COALESCE(estado, 'Activo') <> 'Baja'
        ORDER BY apellido, nombre
        LIMIT 500
    """).fetchall()
    historial = conn.execute("""
        SELECT e.*, j.apellido, j.nombre
        FROM pwa_push_envios e
        LEFT JOIN jugadores j ON j.id = e.jugador_id
        ORDER BY e.creado_en DESC, e.id DESC
        LIMIT 30
    """).fetchall()
    resumen_push = resumen_suscripciones_push(conn)
    conn.close()

    return render_template(
        "notificaciones_app.html",
        categorias=categorias,
        jugadores=jugadores,
        historial=historial,
        resumen_push=resumen_push,
        push_configurado=bool(PWA_VAPID_PUBLIC_KEY and PWA_VAPID_PRIVATE_KEY and webpush),
    )


@app.route("/notificaciones")
def ver_notificaciones():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    datos = obtener_notificaciones_operativas()

    cuotas_vencidas = []
    for cuota in datos["cuotas_vencidas"]:
        mensaje = (
            f"Hola {cuota['nombre']}, registramos pendiente la cuota {cuota['periodo']} "
            f"por {formato_moneda(cuota['importe'])}, vencida el {cuota['fecha_vencimiento']}. "
            "Por favor avisanos si ya fue abonada. Gracias."
        )
        cuotas_vencidas.append({
            "item": cuota,
            "mensaje": mensaje,
            "whatsapp_url": whatsapp_mensaje(cuota["telefono_tutor"] or cuota["telefono"], mensaje),
        })

    cuotas_por_vencer = []
    for cuota in datos["cuotas_por_vencer"]:
        mensaje = (
            f"Hola {cuota['nombre']}, te recordamos que la cuota {cuota['periodo']} "
            f"por {formato_moneda(cuota['importe'])} vence el {cuota['fecha_vencimiento']}. Gracias."
        )
        cuotas_por_vencer.append({
            "item": cuota,
            "mensaje": mensaje,
            "whatsapp_url": whatsapp_mensaje(cuota["telefono_tutor"] or cuota["telefono"], mensaje),
        })

    fichas = []
    for ficha in datos["fichas"]:
        if ficha["estado_documento"] == "vencida":
            texto_estado = f"esta vencida desde el {ficha['fecha_vencimiento']}"
        elif ficha["estado_documento"] == "por_vencer":
            texto_estado = f"vence el {ficha['fecha_vencimiento']}"
        else:
            texto_estado = "figura pendiente de carga"
        mensaje = (
            f"Hola {ficha['nombre']}, la ficha m?dica {texto_estado}. "
            "Cuando puedas, acercanos la actualizacion. Gracias."
        )
        fichas.append({
            "item": ficha,
            "mensaje": mensaje,
            "whatsapp_url": whatsapp_mensaje(ficha["telefono_tutor"] or ficha["telefono"], mensaje),
        })

    asistencia_baja = []
    for jugador in datos["asistencia_baja"]:
        registros = jugador["registros"] or 0
        presentes = jugador["presentes"] or 0
        porcentaje = round((presentes / registros) * 100, 1) if registros else 0
        mensaje = (
            f"Hola {jugador['nombre']}, notamos baja asistencia en los \u00faltimos entrenamientos "
            f"({porcentaje}%). Queremos saber si est\u00e1 todo bien y c\u00f3mo podemos acompa\u00f1ar."
        )
        asistencia_baja.append({
            "item": jugador,
            "porcentaje": porcentaje,
            "mensaje": mensaje,
            "whatsapp_url": whatsapp_mensaje(jugador["telefono_tutor"] or jugador["telefono"], mensaje),
        })

    return render_template(
        "notificaciones.html",
        cuotas_vencidas=cuotas_vencidas,
        cuotas_por_vencer=cuotas_por_vencer,
        fichas=fichas,
        asistencia_baja=asistencia_baja,
        comprobantes_pendientes=datos["comprobantes_pendientes"],
        whatsapp_conversaciones=datos["whatsapp_conversaciones"],
        secretaria_documentos=datos["secretaria_documentos"],
        ahijadxs_objetivo=datos["ahijadxs_objetivo"],
        cambios_portal=datos["cambios_portal"],
    )


@app.route("/notificaciones/descartar", methods=["POST"])
def descartar_notificacion():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    tipo = request.form.get("tipo", "")
    entidad_id = request.form.get("entidad_id", "")
    if descartar_notificacion_usuario(tipo, entidad_id):
        registrar_auditoria(
            "descartar",
            "notificacion",
            clave_notificacion(tipo, entidad_id),
            {"tipo": tipo, "entidad_id": entidad_id},
        )
        flash("Notificación descartada.", "ok")
    else:
        flash("No se pudo descartar la notificación.", "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/descartar-lote", methods=["POST"])
def descartar_notificaciones_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    modo = request.form.get("modo", "seleccionadas")
    campo = "all_items" if modo == "todas" else "items"
    items = [
        parsear_notificacion_form_value(valor)
        for valor in request.form.getlist(campo)
    ]
    cantidad = descartar_notificaciones_usuario(items)
    if cantidad:
        registrar_auditoria(
            "descartar_lote",
            "notificacion",
            None,
            {"cantidad": cantidad, "modo": modo},
        )
        flash(f"{cantidad} notificaciones descartadas.", "ok")
    else:
        flash("No seleccionaste notificaciones para descartar.", "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/comunicacion/whatsapp")
def ver_whatsapp_inbox():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conversaciones = listar_whatsapp_conversaciones()
    jugadores_disponibles = listar_jugadores_basicos_para_whatsapp()
    webhook_eventos = listar_whatsapp_webhook_eventos()
    telefono = normalizar_telefono_whatsapp(request.args.get("telefono", ""))
    if not telefono and conversaciones:
        telefono = conversaciones[0]["telefono"]

    mensajes = obtener_whatsapp_conversacion(telefono) if telefono else []
    if telefono and mensajes:
        conn = get_connection()
        conn.execute("""
            UPDATE whatsapp_mensajes
            SET leido = 1
            WHERE direccion = 'in'
              AND COALESCE(leido, 0) = 0
              AND (telefono = %s OR wa_id = %s)
        """, (telefono, telefono))
        conn.commit()
        conn.close()
        mensajes = obtener_whatsapp_conversacion(telefono)

    conversacion_actual = next((item for item in conversaciones if item["telefono"] == telefono), None)
    return render_template(
        "whatsapp_inbox.html",
        conversaciones=conversaciones,
        conversacion_actual=conversacion_actual,
        mensajes=mensajes,
        telefono_actual=telefono,
        jugadores_disponibles=jugadores_disponibles,
        webhook_eventos=webhook_eventos,
        respuestas_rapidas=WHATSAPP_RESPUESTAS_RAPIDAS,
    )


@app.route("/comunicacion/whatsapp/estado")
def estado_whatsapp_inbox():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    return Response(
        json.dumps(obtener_estado_whatsapp_inbox(), ensure_ascii=False),
        status=200,
        mimetype="application/json",
    )


@app.route("/manifest.webmanifest")
def pwa_manifest():
    manifest = {
        "name": "SIG - Sistema Integral de Gestion",
        "short_name": "SIG",
        "description": "Portal y administracion del club.",
        "start_url": url_for("index"),
        "scope": "/",
        "display": "standalone",
        "background_color": "#0f172a",
        "theme_color": "#0f172a",
        "orientation": "portrait-primary",
        "icons": [
            {"src": pwa_icon_url("192"), "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": pwa_icon_url("512"), "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
        "shortcuts": [
            {
                "name": "Portal jugador",
                "short_name": "Portal",
                "url": url_for("portal_buscar"),
                "description": "Abrir portal del jugador.",
            },
            {
                "name": "Notificaciones",
                "short_name": "Avisos",
                "url": url_for("ver_notificaciones"),
                "description": "Ver notificaciones operativas.",
            },
        ],
    }
    return Response(json.dumps(manifest, ensure_ascii=False), mimetype="application/manifest+json")


@app.route("/service-worker.js")
def pwa_service_worker():
    response = send_file(BASE_DIR / "static" / "service-worker.js", mimetype="application/javascript")
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/pwa/config")
def pwa_config():
    return jsonify({
        "vapidPublicKey": PWA_VAPID_PUBLIC_KEY,
        "pushEnabled": bool(PWA_VAPID_PUBLIC_KEY and PWA_VAPID_PRIVATE_KEY and webpush),
    })


@app.route("/pwa/push/subscribe", methods=["POST"])
def pwa_push_subscribe():
    data = request.get_json(silent=True) or {}
    subscription = data.get("subscription") or {}
    portal_token = data.get("portal_token") or ""

    conn = get_connection()
    actor = actor_push_actual(conn, portal_token=portal_token)
    if not actor:
        conn.close()
        abort(403)
    try:
        guardar_suscripcion_push(conn, subscription, actor, request.headers.get("User-Agent", ""))
    except ValueError as error:
        conn.rollback()
        conn.close()
        return jsonify({"ok": False, "error": str(error)}), 400
    conn.commit()
    conn.close()
    registrar_auditoria("suscribir", "pwa_push", None, {"actor": actor["tipo"]})
    return jsonify({"ok": True})


@app.route("/pwa/push/unsubscribe", methods=["POST"])
def pwa_push_unsubscribe():
    data = request.get_json(silent=True) or {}
    endpoint = data.get("endpoint") or ""
    conn = get_connection()
    desactivar_suscripcion_push(conn, endpoint)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/pwa/push/test", methods=["POST"])
def pwa_push_test():
    data = request.get_json(silent=True) or {}
    portal_token = data.get("portal_token") or ""
    conn = get_connection()
    actor = actor_push_actual(conn, portal_token=portal_token)
    conn.close()
    if not actor:
        abort(403)

    payload = {
        "title": "SIG",
        "body": "Notificaciones activadas correctamente.",
        "url": url_for("portal_jugador", token=portal_token) if actor["tipo"] == "portal" and portal_token else url_for("index"),
        "icon": pwa_icon_url("192"),
    }
    if actor["tipo"] == "portal":
        resultado = enviar_push_por_actor("portal", payload, jugador_id=actor["jugador_id"])
    else:
        resultado = enviar_push_por_actor("usuario", payload, usuario_id=actor["usuario_id"])
    return jsonify({"ok": not resultado["errores"], **resultado})


@app.route("/presencia/heartbeat", methods=["POST"])
def presencia_heartbeat():
    if "user_id" not in session:
        return Response("unauthorized", status=401, mimetype="text/plain")

    username = normalizar_username(session.get("username", ""))
    if not username:
        return Response("bad request", status=400, mimetype="text/plain")

    registrar_presencia_usuario(username)
    return Response(
        json.dumps({"ok": True}, ensure_ascii=False),
        status=200,
        mimetype="application/json",
    )


@app.route("/comunicacion/whatsapp/responder", methods=["POST"])
def responder_whatsapp_inbox():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    telefono = normalizar_telefono_whatsapp(request.form.get("telefono", ""))
    mensaje = (request.form.get("mensaje", "") or "").strip()
    if not telefono:
        flash("No hay un teléfono válido para responder.", "error")
        return redirect(url_for("ver_whatsapp_inbox"))
    if not mensaje:
        flash("Escribí un mensaje para responder por WhatsApp.", "error")
        return redirect(url_for("ver_whatsapp_inbox", telefono=telefono))

    jugador = buscar_jugador_por_whatsapp(telefono) or {}
    enviado, destinatario, motivo, detalle = enviar_whatsapp_meta(
        telefono,
        mensaje,
        tipo="inbox_respuesta",
        entidad="whatsapp_inbox",
        entidad_id=telefono,
        jugador_id=jugador.get("id"),
    )
    if enviado:
        registrar_auditoria(
            "responder",
            "whatsapp_inbox",
            telefono,
            {"destinatario": destinatario, "jugador_id": jugador.get("id")},
        )
        flash("Respuesta enviada por WhatsApp.", "ok")
    else:
        flash(mensaje_fallo_whatsapp(motivo, destinatario, detalle), "error")
    return redirect(url_for("ver_whatsapp_inbox", telefono=telefono))


@app.route("/comunicacion/whatsapp/vincular", methods=["POST"])
def vincular_whatsapp_inbox():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    telefono = normalizar_telefono_whatsapp(request.form.get("telefono", ""))
    jugador_id_raw = (request.form.get("jugador_id", "") or "").strip()
    if not telefono:
        flash("No hay un teléfono válido para vincular.", "error")
        return redirect(url_for("ver_whatsapp_inbox"))

    jugador_id = int(jugador_id_raw) if jugador_id_raw.isdigit() else None
    conn = get_connection()
    conn.execute("""
        UPDATE whatsapp_mensajes
        SET jugador_id = %s
        WHERE telefono = %s OR wa_id = %s
    """, (jugador_id, telefono, telefono))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "vincular",
        "whatsapp_inbox",
        telefono,
        {"jugador_id": jugador_id},
    )
    if jugador_id:
        flash("Conversación vinculada al jugador.", "ok")
    else:
        flash("Conversación desvinculada del jugador.", "ok")
    return redirect(url_for("ver_whatsapp_inbox", telefono=telefono))


@app.route("/notificaciones/cuotas/<int:cuota_id>/email", methods=["POST"])
def enviar_recordatorio_cuota(cuota_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.jugador_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            j.nombre,
            j.apellido,
            j.email,
            j.email_tutor,
            j.portal_token,
            j.portal_activo
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
    """, (cuota_id,)).fetchone()
    conn.close()
    if cuota is None:
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("ver_notificaciones"))

    cuerpo = construir_texto_recordatorio_cuota(cuota)
    enviado, destinatario, motivo = enviar_email_jugador(cuota, f"Recordatorio de cuota {cuota['periodo']}", cuerpo)
    if enviado:
        registrar_auditoria("enviar_recordatorio", "cuota", str(cuota_id), {"destinatario": destinatario, "tipo": "cuota"})
        flash("Recordatorio de cuota enviado.", "ok")
    else:
        flash(mensaje_fallo_email(motivo, destinatario), "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/cuotas/<int:cuota_id>/whatsapp", methods=["POST"])
def enviar_recordatorio_cuota_whatsapp(cuota_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()
    cuota = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.periodo,
            c.importe,
            c.fecha_vencimiento,
            CASE
                WHEN c.fecha_vencimiento IS NOT NULL
                 AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                 AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                THEN GREATEST((CURRENT_DATE - c.fecha_vencimiento::date), 0)
                ELSE 0
            END AS dias_vencida,
            j.id AS jugador_id,
            j.nombre,
            j.apellido,
            j.telefono,
            j.telefono_tutor
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
    """, (cuota_id,)).fetchone()
    conn.close()
    if cuota is None:
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("ver_notificaciones"))

    enviado, destinatario, motivo, detalle = enviar_whatsapp_recordatorio_cuota_template(
        cuota,
        tipo="recordatorio_cuota",
        entidad="cuota",
        entidad_id=str(cuota_id),
    )
    if enviado:
        registrar_auditoria("enviar_recordatorio", "cuota", str(cuota_id), {"destinatario": destinatario, "tipo": "cuota_whatsapp"})
        flash("Recordatorio de cuota enviado por WhatsApp.", "ok")
    else:
        flash(mensaje_fallo_whatsapp(motivo, destinatario, detalle), "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/fichas/<int:jugador_id>/email", methods=["POST"])
def enviar_recordatorio_ficha(jugador_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()
    ficha = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.nombre,
            j.apellido,
            j.email,
            j.email_tutor,
            j.portal_token,
            j.portal_activo,
            f.fecha_vencimiento,
            CASE
                WHEN f.id IS NULL OR f.fecha_vencimiento IS NULL OR NULLIF(f.fecha_vencimiento::text, '') IS NULL THEN 'faltante'
                WHEN f.fecha_vencimiento::text !~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN 'faltante'
                WHEN f.fecha_vencimiento::date < CURRENT_DATE THEN 'vencida'
                WHEN f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days' THEN 'por_vencer'
                ELSE 'vigente'
            END AS estado_documento
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.id = %s
    """, (jugador_id,)).fetchone()
    conn.close()
    if ficha is None:
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("ver_notificaciones"))

    cuerpo = construir_texto_recordatorio_ficha(ficha)
    enviado, destinatario, motivo = enviar_email_jugador(ficha, "Recordatorio de ficha m\u00e9dica", cuerpo)
    if enviado:
        registrar_auditoria("enviar_recordatorio", "ficha_medica", str(jugador_id), {"destinatario": destinatario, "tipo": "ficha_medica"})
        flash("Recordatorio de ficha m\u00e9dica enviado.", "ok")
    else:
        flash(mensaje_fallo_email(motivo, destinatario), "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/fichas/<int:jugador_id>/whatsapp", methods=["POST"])
def enviar_recordatorio_ficha_whatsapp(jugador_id):
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()
    ficha = conn.execute("""
        SELECT
            j.id AS jugador_id,
            j.nombre,
            j.apellido,
            j.telefono,
            j.telefono_tutor,
            f.fecha_vencimiento,
            CASE
                WHEN f.id IS NULL OR f.fecha_vencimiento IS NULL OR NULLIF(f.fecha_vencimiento::text, '') IS NULL THEN 'faltante'
                WHEN f.fecha_vencimiento::text !~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$' THEN 'faltante'
                WHEN f.fecha_vencimiento::date < CURRENT_DATE THEN 'vencida'
                WHEN f.fecha_vencimiento::date <= CURRENT_DATE + INTERVAL '30 days' THEN 'por_vencer'
                ELSE 'vigente'
            END AS estado_documento
        FROM jugadores j
        LEFT JOIN fichas_medicas f ON f.jugador_id = j.id
        WHERE j.id = %s
    """, (jugador_id,)).fetchone()
    conn.close()
    if ficha is None:
        flash("Jugador no encontrado.", "error")
        return redirect(url_for("ver_notificaciones"))

    cuerpo = construir_texto_recordatorio_ficha(ficha)
    enviado, destinatario, motivo, detalle = enviar_whatsapp_jugador(
        ficha,
        cuerpo,
        tipo="recordatorio_ficha",
        entidad="ficha_medica",
        entidad_id=str(jugador_id),
    )
    if enviado:
        registrar_auditoria("enviar_recordatorio", "ficha_medica", str(jugador_id), {"destinatario": destinatario, "tipo": "ficha_medica_whatsapp"})
        flash("Recordatorio de ficha médica enviado por WhatsApp.", "ok")
    else:
        flash(mensaje_fallo_whatsapp(motivo, destinatario, detalle), "error")
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/cuotas/email-lote", methods=["POST"])
def enviar_recordatorios_cuotas_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    modo = request.form.get("modo", "vencidas").strip()
    datos = obtener_notificaciones_operativas()
    cuotas = datos["cuotas_vencidas"] if modo == "vencidas" else datos["cuotas_por_vencer"]
    resultados = []
    for cuota in cuotas:
        cuerpo = construir_texto_recordatorio_cuota(cuota)
        resultados.append(enviar_email_jugador(cuota, f"Recordatorio de cuota {cuota['periodo']}", cuerpo))
    mensaje_resultado, nivel = resumir_envio_masivo_email(resultados, "recordatorios por email")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/cuotas/whatsapp-lote", methods=["POST"])
def enviar_recordatorios_cuotas_whatsapp_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    modo = request.form.get("modo", "vencidas").strip()
    datos = obtener_notificaciones_operativas()
    cuotas = datos["cuotas_vencidas"] if modo == "vencidas" else datos["cuotas_por_vencer"]
    resultados = []
    for cuota in cuotas:
        resultados.append(enviar_whatsapp_recordatorio_cuota_template(
            cuota,
            tipo="recordatorio_cuota",
            entidad="cuota",
            entidad_id=str(cuota["cuota_id"]),
        ))
    mensaje_resultado, nivel = resumir_envio_masivo_whatsapp(resultados, "recordatorios por WhatsApp")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/fichas/email-lote", methods=["POST"])
def enviar_recordatorios_fichas_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    datos = obtener_notificaciones_operativas()
    resultados = []
    for ficha in datos["fichas"]:
        cuerpo = construir_texto_recordatorio_ficha(ficha)
        resultados.append(enviar_email_jugador(ficha, "Recordatorio de ficha m\u00e9dica", cuerpo))
    mensaje_resultado, nivel = resumir_envio_masivo_email(resultados, "recordatorios de ficha m\u00e9dica")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_notificaciones"))


@app.route("/notificaciones/fichas/whatsapp-lote", methods=["POST"])
def enviar_recordatorios_fichas_whatsapp_lote():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    datos = obtener_notificaciones_operativas()
    resultados = []
    for ficha in datos["fichas"]:
        cuerpo = construir_texto_recordatorio_ficha(ficha)
        resultados.append(enviar_whatsapp_jugador(
            ficha,
            cuerpo,
            tipo="recordatorio_ficha",
            entidad="ficha_medica",
            entidad_id=str(ficha["jugador_id"]),
        ))
    mensaje_resultado, nivel = resumir_envio_masivo_whatsapp(resultados, "recordatorios de ficha médica por WhatsApp")
    flash(mensaje_resultado, nivel)
    return redirect(url_for("ver_notificaciones"))

@app.route("/exportar/morosos")
def exportar_morosos():
    check = permiso_requerido("comunicaciones_ver")
    if check:
        return check

    conn = get_connection()

    morosos = conn.execute("""
        SELECT
            j.apellido,
            j.nombre,
            j.dni,
            j.categoria,
            j.telefono,
            j.email,
            COALESCE(SUM(c.importe), 0) AS deuda,
            SUM(
                CASE
                    WHEN c.fecha_vencimiento IS NOT NULL
                     AND NULLIF(c.fecha_vencimiento::text, '') IS NOT NULL
                     AND c.fecha_vencimiento::text ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                     AND c.fecha_vencimiento::date < CURRENT_DATE
                    THEN 1
                    ELSE 0
                END
            ) AS cuotas_vencidas
        FROM jugadores j
        JOIN cuotas c ON j.id = c.jugador_id
        WHERE c.pagado = 0
          AND COALESCE(c.importe, 0) > 0
        GROUP BY j.id, j.apellido, j.nombre, j.dni, j.categoria, j.telefono, j.email
        HAVING COALESCE(SUM(c.importe), 0) > 0
        ORDER BY deuda DESC, j.apellido, j.nombre
    """).fetchall()

    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow([
        "Apellido",
        "Nombre",
        "DNI",
        "Categoria",
        "Telefono",
        "Email",
        "Deuda",
        "Cuotas vencidas"
    ])

    for jugador in morosos:
        writer.writerow([
            jugador["apellido"],
            jugador["nombre"],
            jugador["dni"],
            jugador["categoria"],
            jugador["telefono"],
            jugador["email"],
            jugador["deuda"],
            jugador["cuotas_vencidas"]
        ])

    contenido = output.getvalue()
    output.close()

    registrar_auditoria(
        "exportar_ok",
        "morosos",
        None,
        {"formato": "csv", "cantidad_registros": len(morosos)},
    )

    return Response(
        contenido,
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=morosos.csv"
        }
    )

@app.route("/backup")
def backup_db():
    check = permiso_requerido("backup_ver")
    if check:
        return check

    registrar_auditoria(
        "backup_info",
        "backup",
        None,
        {"mensaje": "Consulta de instrucciones de backup Cloud SQL"},
    )
    flash("En la versión con Cloud SQL, el backup se gestiona desde Google Cloud SQL. Usá los backups automáticos o exportaciones desde la consola de Google Cloud.", "ok")
    return render_template(
        "sistema_admin.html",
        estado=obtener_estado_sistema_admin(),
        solo_backup=True,
    )


@app.route("/admin/sistema/backup", methods=["POST"])
def crear_backup_cloud_sql():
    check = permiso_requerido("backup_gestionar")
    if check:
        return check

    try:
        resultado = solicitar_backup_cloud_sql()
    except Exception as error:
        app.logger.exception("No se pudo solicitar backup manual de Cloud SQL.")
        registrar_auditoria(
            "backup_manual_error",
            "backup",
            None,
            {"error": str(error)},
        )
        flash(f"No se pudo solicitar el backup: {error}", "error")
        return redirect(url_for("panel_sistema_admin" if session.get("rol") == "admin" else "backup_db"))

    registrar_auditoria(
        "backup_manual_solicitado",
        "backup",
        str(resultado.get("targetId") or resultado.get("name") or ""),
        {
            "operation": resultado.get("name"),
            "status": resultado.get("status"),
            "target_id": resultado.get("targetId"),
        },
    )
    flash("Backup manual solicitado. Puede tardar unos minutos en aparecer como completado.", "ok")
    return redirect(url_for("panel_sistema_admin" if session.get("rol") == "admin" else "backup_db"))


@app.route("/admin/sistema")
def panel_sistema_admin():
    check = rol_requerido("admin")
    if check:
        return check

    return render_template(
        "sistema_admin.html",
        estado=obtener_estado_sistema_admin(),
        solo_backup=False,
    )


@app.route("/admin/sistema/automatizaciones", methods=["POST"])
def configurar_automatizaciones():
    check = rol_requerido("admin")
    if check:
        return check
    conn = get_connection()
    guardar_app_setting(conn, "automation_reminders_enabled", "true" if request.form.get("recordatorios_activos") else "false", session.get("username"))
    guardar_app_setting(conn, "automation_reminders_days_before", str(int_setting(request.form.get("dias_antes"), 3, 0, 30)), session.get("username"))
    guardar_app_setting(conn, "automation_invoices_enabled", "true" if request.form.get("facturas_activas") else "false", session.get("username"))
    conn.commit()
    conn.close()
    registrar_auditoria("configurar", "automatizaciones", None, {
        "recordatorios_activos": bool(request.form.get("recordatorios_activos")),
        "dias_antes": int_setting(request.form.get("dias_antes"), 3, 0, 30),
        "facturas_activas": bool(request.form.get("facturas_activas")),
    })
    flash("Automatizaciones actualizadas.", "ok")
    return redirect(url_for("panel_sistema_admin"))


@app.route("/admin/sistema/automatizaciones/ejecutar", methods=["POST"])
def ejecutar_automatizaciones_admin():
    check = rol_requerido("admin")
    if check:
        return check
    resultado = ejecutar_automatizaciones(session.get("username") or "admin")
    flash(
        f"Ejecucion completada: {resultado['recordatorios_enviados']} recordatorio(s) enviado(s), "
        f"{len(resultado['errores'])} error(es).",
        "warning" if resultado["errores"] else "ok",
    )
    return redirect(url_for("panel_sistema_admin"))


@app.route("/tasks/automatizaciones", methods=["POST"])
def ejecutar_automatizaciones_programadas():
    token_configurado = os.environ.get("AUTOMATION_TOKEN", "").strip()
    token_recibido = request.headers.get("X-Automation-Token", "").strip()
    if not token_configurado or not secrets.compare_digest(token_configurado, token_recibido):
        abort(403)
    return jsonify(ejecutar_automatizaciones("scheduler"))


@app.route("/admin/sistema/facturas-email", methods=["GET", "POST"])
def configurar_facturas_email():
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()
    if request.method == "POST":
        auditoria_cuentas = []
        for indice in (1, 2):
            defaults = factura_email_defaults(indice)
            suffix = "" if indice == 1 else "2"
            host = request.form.get(f"host{suffix}", "").strip()
            port = int_setting(request.form.get(f"port{suffix}"), defaults["port"], 1, 65535)
            user = request.form.get(f"user{suffix}", "").strip()
            folder = request.form.get(f"folder{suffix}", "").strip() or "INBOX"
            use_ssl = request.form.get(f"use_ssl{suffix}") == "on"
            search_days = int_setting(request.form.get(f"search_days{suffix}"), defaults["search_days"], 1, 365)
            max_messages = int_setting(request.form.get(f"max_messages{suffix}"), defaults["max_messages"], 1, 500)
            secret_name = request.form.get(f"secret_name{suffix}", "").strip() or defaults["secret_name"]
            password = request.form.get(f"password{suffix}", "")

            if indice == 1 and (not host or not user):
                conn.close()
                flash("Host y usuario IMAP de la cuenta 1 son obligatorios.", "error")
                return redirect(url_for("configurar_facturas_email"))
            if indice == 2 and not any([host, user, password]):
                for nombre in (
                    "imap_host", "imap_port", "imap_user", "imap_folder", "imap_use_ssl",
                    "search_days", "max_messages", "secret_name",
                    "secret_actualizado_en", "secret_actualizado_por",
                ):
                    guardar_app_setting(conn, factura_email_setting_key(nombre, indice), "", session.get("username"))
                continue
            if indice == 2 and (not host or not user):
                conn.close()
                flash("Para usar la cuenta 2 completá host y usuario.", "error")
                return redirect(url_for("configurar_facturas_email"))

            try:
                if password:
                    secret_path = guardar_secret_manager(secret_name, password)
                    guardar_app_setting(conn, factura_email_setting_key("secret_name", indice), secret_path, session.get("username"))
                    guardar_app_setting(conn, factura_email_setting_key("secret_actualizado_en", indice), ahora_sig().strftime("%Y-%m-%d %H:%M:%S"), session.get("username"))
                    guardar_app_setting(conn, factura_email_setting_key("secret_actualizado_por", indice), session.get("username") or "", session.get("username"))
                else:
                    guardar_app_setting(conn, factura_email_setting_key("secret_name", indice), secret_name, session.get("username"))
            except Exception as error:
                conn.rollback()
                conn.close()
                app.logger.exception("No se pudo guardar secreto IMAP de facturas.")
                flash(f"No se pudo guardar el secreto de la cuenta {indice}: {error}", "error")
                return redirect(url_for("configurar_facturas_email"))

            guardar_app_setting(conn, factura_email_setting_key("imap_host", indice), host, session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("imap_port", indice), str(port), session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("imap_user", indice), user, session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("imap_folder", indice), folder, session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("imap_use_ssl", indice), "true" if use_ssl else "false", session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("search_days", indice), str(search_days), session.get("username"))
            guardar_app_setting(conn, factura_email_setting_key("max_messages", indice), str(max_messages), session.get("username"))
            auditoria_cuentas.append({
                "indice": indice,
                "host": host,
                "user": user,
                "folder": folder,
                "secret_name": secret_name,
                "password_actualizada": bool(password),
            })
        conn.commit()
        conn.close()

        registrar_auditoria("configurar", "facturas_email", None, {
            "cuentas": auditoria_cuentas,
        })
        flash("Configuracion de facturas por email guardada.", "ok")
        return redirect(url_for("configurar_facturas_email"))

    configs = obtener_factura_email_configs(conn)
    conn.close()
    return render_template("facturas_email_config.html", config=configs[0], config2=configs[1], configs=configs)


@app.route("/admin/sugerencias-recomendaciones")
def listar_sugerencias_recomendaciones():
    check = permiso_requerido("sugerencias_ver")
    if check:
        return check

    filtros = {
        "q": request.args.get("q", "").strip(),
        "tipo": request.args.get("tipo", "").strip().lower(),
        "categoria": request.args.get("categoria", "").strip().lower(),
        "email_estado": request.args.get("email_estado", "").strip().lower(),
        "seguimiento_estado": request.args.get("seguimiento_estado", "").strip().lower(),
    }
    categorias_validas = {item["clave"] for item in SUGERENCIA_RECOMENDACION_CATEGORIAS}
    estados_validos = set(SUGERENCIA_EMAIL_ESTADOS)
    estados_seguimiento_validos = set(SUGERENCIA_SEGUIMIENTO_ESTADOS)

    condiciones = ["tipo IN ('sugerencia', 'recomendacion')"]
    params = []
    if filtros["q"]:
        like = f"%{filtros['q']}%"
        condiciones.append("(mensaje ILIKE %s OR nombre ILIKE %s OR contacto ILIKE %s)")
        params.extend([like, like, like])
    if filtros["tipo"] in {"sugerencia", "recomendacion"}:
        condiciones.append("tipo = %s")
        params.append(filtros["tipo"])
    else:
        filtros["tipo"] = ""
    if filtros["categoria"] in categorias_validas:
        condiciones.append("categoria = %s")
        params.append(filtros["categoria"])
    else:
        filtros["categoria"] = ""
    if filtros["email_estado"] in estados_validos:
        condiciones.append("email_estado = %s")
        params.append(filtros["email_estado"])
    else:
        filtros["email_estado"] = ""
    if filtros["seguimiento_estado"] in estados_seguimiento_validos:
        condiciones.append("seguimiento_estado = %s")
        params.append(filtros["seguimiento_estado"])
    else:
        filtros["seguimiento_estado"] = ""

    where_sql = "WHERE " + " AND ".join(condiciones) if condiciones else ""
    conn = get_connection()
    registros = conn.execute(f"""
        SELECT *
        FROM sugerencias_denuncias
        {where_sql}
        ORDER BY creado_en DESC, id DESC
        LIMIT 300
    """, params).fetchall()
    resumen = conn.execute(f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN tipo = 'sugerencia' THEN 1 ELSE 0 END) AS sugerencias,
            SUM(CASE WHEN tipo = 'recomendacion' THEN 1 ELSE 0 END) AS recomendaciones,
            SUM(CASE WHEN email_estado NOT IN ('enviado', 'parcial') OR email_estado IS NULL THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN seguimiento_estado = 'nuevo' OR seguimiento_estado IS NULL THEN 1 ELSE 0 END) AS nuevos,
            SUM(CASE WHEN seguimiento_estado = 'en_revision' THEN 1 ELSE 0 END) AS en_revision
        FROM sugerencias_denuncias
        {where_sql}
    """, params).fetchone()
    resumen_total = conn.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN tipo = 'sugerencia' THEN 1 ELSE 0 END) AS sugerencias,
            SUM(CASE WHEN tipo = 'recomendacion' THEN 1 ELSE 0 END) AS recomendaciones,
            SUM(CASE WHEN email_estado NOT IN ('enviado', 'parcial') OR email_estado IS NULL THEN 1 ELSE 0 END) AS pendientes,
            SUM(CASE WHEN seguimiento_estado = 'nuevo' OR seguimiento_estado IS NULL THEN 1 ELSE 0 END) AS nuevos,
            SUM(CASE WHEN seguimiento_estado = 'en_revision' THEN 1 ELSE 0 END) AS en_revision
        FROM sugerencias_denuncias
        WHERE tipo IN ('sugerencia', 'recomendacion')
    """).fetchone()
    conn.close()

    categoria_labels = {item["clave"]: item["nombre"] for item in SUGERENCIA_RECOMENDACION_CATEGORIAS}
    registros = [dict(registro) for registro in registros]
    for registro in registros:
        registro["destinatarios_lista"] = leer_lista_config(registro.get("destinatarios"))
        registro["email_info"] = info_email_estado_sugerencia(registro.get("email_estado"))
        registro["seguimiento_info"] = info_seguimiento_estado_sugerencia(registro.get("seguimiento_estado"))
        registro["categoria_label"] = categoria_labels.get(registro.get("categoria"), registro.get("categoria") or "General")
        registro["puede_gestionar"] = puede_gestionar_tipo_sugerencia(registro.get("tipo"))

    return render_template(
        "sugerencias_recomendaciones_admin.html",
        registros=registros,
        resumen=resumen,
        resumen_total=resumen_total,
        filtros=filtros,
        categorias=SUGERENCIA_RECOMENDACION_CATEGORIAS,
        estados=SUGERENCIA_EMAIL_ESTADOS,
        estados_seguimiento=SUGERENCIA_SEGUIMIENTO_ESTADOS,
        puede_configurar_sugerencias=tiene_permiso("sugerencias_configurar"),
        puede_gestionar_sugerencias=tiene_permiso("sugerencias_gestionar"),
    )


@app.route("/admin/sugerencias-recomendaciones/<int:registro_id>/seguimiento", methods=["POST"])
def actualizar_sugerencia_recomendacion(registro_id):
    check = permiso_requerido("sugerencias_gestionar")
    if check:
        return check

    conn = get_connection()
    registro = conn.execute("SELECT * FROM sugerencias_denuncias WHERE id = %s", (registro_id,)).fetchone()
    if not registro:
        conn.close()
        flash("Registro no encontrado.", "error")
        return redirect(url_for("listar_sugerencias_recomendaciones"))
    if not puede_ver_tipo_sugerencia(registro["tipo"]):
        conn.close()
        flash("No tenes permiso para gestionar este registro.", "error")
        return redirect(url_for("listar_sugerencias_recomendaciones"))

    seguimiento_estado = normalizar_estado_seguimiento_sugerencia(request.form.get("seguimiento_estado"))
    notas_internas = request.form.get("notas_internas", "").strip()
    conn.execute("""
        UPDATE sugerencias_denuncias
        SET seguimiento_estado = %s,
            notas_internas = %s,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (seguimiento_estado, notas_internas or None, session.get("username"), registro_id))
    conn.commit()
    conn.close()

    registrar_auditoria("actualizar", "sugerencia_recomendacion", str(registro_id), {
        "seguimiento_estado": seguimiento_estado,
        "notas_internas": bool(notas_internas),
    })
    flash("Seguimiento actualizado.", "ok")
    return redirect(destino_interno(request.form.get("next"), fallback="listar_sugerencias_recomendaciones"))


@app.route("/admin/sugerencias-recomendaciones/<int:registro_id>/reenviar", methods=["POST"])
def reenviar_sugerencia_recomendacion(registro_id):
    check = permiso_requerido("sugerencias_gestionar")
    if check:
        return check

    conn = get_connection()
    registro = conn.execute("SELECT * FROM sugerencias_denuncias WHERE id = %s", (registro_id,)).fetchone()
    if not registro:
        conn.close()
        flash("Registro no encontrado.", "error")
        return redirect(url_for("listar_sugerencias_recomendaciones"))
    if not puede_ver_tipo_sugerencia(registro["tipo"]):
        conn.close()
        flash("No tenes permiso para reenviar este registro.", "error")
        return redirect(url_for("listar_sugerencias_recomendaciones"))

    data = {
        "tipo": registro["tipo"],
        "categoria": registro["categoria"],
        "anonima": bool(registro["anonima"]),
        "nombre": registro["nombre"] or "",
        "contacto": registro["contacto"] or "",
        "mensaje": registro["mensaje"] or "",
    }
    destinatarios = obtener_destinatarios_sugerencias(conn)
    email_estado, enviados = enviar_notificacion_sugerencia_recomendacion(data, registro_id, destinatarios)
    conn.execute("""
        UPDATE sugerencias_denuncias
        SET email_estado = %s,
            destinatarios = %s,
            notificacion_reintentos = COALESCE(notificacion_reintentos, 0) + 1,
            notificado_en = CASE WHEN %s IN ('enviado', 'parcial') THEN CURRENT_TIMESTAMP ELSE notificado_en END,
            actualizado_en = CURRENT_TIMESTAMP,
            actualizado_por = %s
        WHERE id = %s
    """, (
        email_estado,
        json.dumps(destinatarios, ensure_ascii=False),
        email_estado,
        session.get("username"),
        registro_id,
    ))
    conn.commit()
    conn.close()

    registrar_auditoria("reenviar", "sugerencia_recomendacion", str(registro_id), {
        "email_estado": email_estado,
        "emails_enviados": enviados,
        "destinatarios": len(destinatarios),
    })
    if email_estado in {"enviado", "parcial"}:
        flash("Notificacion reenviada.", "ok")
    else:
        flash("No se pudo reenviar la notificacion. El registro sigue disponible en la bandeja.", "warning")
    return redirect(destino_interno(request.form.get("next"), fallback="listar_sugerencias_recomendaciones"))


@app.route("/admin/sugerencias-recomendaciones/config", methods=["GET", "POST"])
def configurar_sugerencias_recomendaciones():
    check = permiso_requerido("sugerencias_configurar")
    if check:
        return check

    conn = get_connection()

    if request.method == "POST":
        directiva_texto = request.form.get("directiva_emails", "")
        directiva_emails, directiva_invalidos = parsear_emails_config(directiva_texto)

        form_config = {
            "directiva_emails": directiva_emails,
            "actualizado_en": None,
            "actualizado_por": None,
        }
        if directiva_invalidos:
            conn.close()
            flash("Hay emails con formato invalido: " + ", ".join(directiva_invalidos), "error")
            return render_template("sugerencias_recomendaciones_config.html", config=form_config)

        if not directiva_emails:
            conn.close()
            flash("Configura al menos un email de Comision Directiva.", "error")
            return render_template("sugerencias_recomendaciones_config.html", config=form_config)

        guardar_app_setting(conn, SUGERENCIAS_DIRECTIVA_EMAILS_KEY, serializar_lista_config(directiva_emails), session.get("username"))
        conn.commit()
        conn.close()

        registrar_auditoria("configurar", "sugerencia_recomendacion_config", None, {
            "directiva_emails": len(directiva_emails),
        })
        flash("Configuracion de sugerencias y recomendaciones guardada.", "ok")
        return redirect(url_for("configurar_sugerencias_recomendaciones"))

    config = obtener_sugerencias_config(conn)
    conn.close()
    return render_template("sugerencias_recomendaciones_config.html", config=config)


@app.route("/admin/sugerencias-denuncias")
def listar_sugerencias_denuncias_legacy():
    return redirect(url_for("listar_sugerencias_recomendaciones"), code=301)


@app.route("/admin/sugerencias-denuncias/config")
def configurar_sugerencias_denuncias_legacy():
    return redirect(url_for("configurar_sugerencias_recomendaciones"), code=301)


@app.route("/admin/mantenimiento", methods=["GET", "POST"])
def configurar_mantenimiento():
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()

    if request.method == "POST":
        activo = request.form.get("activo") == "on"
        mensaje = request.form.get("mensaje", "").strip() or MAINTENANCE_DEFAULT_MESSAGE

        guardar_app_setting(conn, "maintenance_mode", "true" if activo else "false", session.get("username"))
        guardar_app_setting(conn, "maintenance_message", mensaje, session.get("username"))
        conn.commit()

        g.mantenimiento = obtener_config_mantenimiento(conn)
        conn.close()

        if activo:
            flash("Modo mantenimiento activado. Solo admin puede usar el sistema.", "ok")
        else:
            flash("Modo mantenimiento desactivado. El sistema vuelve a estar disponible.", "ok")
        return redirect(url_for("configurar_mantenimiento"))

    mantenimiento = obtener_config_mantenimiento(conn)
    conn.close()

    return render_template(
        "mantenimiento_admin.html",
        mantenimiento=mantenimiento,
        mensaje_default=MAINTENANCE_DEFAULT_MESSAGE,
    )


@app.route("/auditoria")
def ver_auditoria():
    check = permiso_requerido("auditoria_ver")
    if check:
        return check

    filtros = {
        "q": request.args.get("q", "").strip(),
        "accion": request.args.get("accion", "").strip(),
        "entidad": request.args.get("entidad", "").strip(),
        "usuario": request.args.get("usuario", "").strip(),
        "desde": request.args.get("desde", "").strip(),
        "hasta": request.args.get("hasta", "").strip(),
    }

    condiciones = []
    params = []

    if filtros["q"]:
        patron = f"%{filtros['q']}%"
        condiciones.append("""
            (
                accion ILIKE %s OR entidad ILIKE %s OR entidad_id ILIKE %s
                OR detalle ILIKE %s OR ip ILIKE %s OR username ILIKE %s
            )
        """)
        params.extend([patron, patron, patron, patron, patron, patron])

    if filtros["accion"]:
        condiciones.append("accion = %s")
        params.append(filtros["accion"])

    if filtros["entidad"]:
        condiciones.append("entidad = %s")
        params.append(filtros["entidad"])

    if filtros["usuario"]:
        condiciones.append("username ILIKE %s")
        params.append(f"%{filtros['usuario']}%")

    if validar_fecha_movimiento(filtros["desde"]):
        condiciones.append("fecha::date >= %s::date")
        params.append(filtros["desde"])
    else:
        filtros["desde"] = ""

    if validar_fecha_movimiento(filtros["hasta"]):
        condiciones.append("fecha::date <= %s::date")
        params.append(filtros["hasta"])
    else:
        filtros["hasta"] = ""

    where_sql = ""
    if condiciones:
        where_sql = "WHERE " + " AND ".join(condiciones)

    conn = get_connection()
    registros = conn.execute(f"""
        SELECT *
        FROM auditoria
        {where_sql}
        ORDER BY fecha DESC, id DESC
        LIMIT 300
    """, tuple(params)).fetchall()
    registros = enriquecer_actores_auditoria(conn, registros)

    acciones = conn.execute("""
        SELECT DISTINCT accion
        FROM auditoria
        ORDER BY accion
    """).fetchall()

    entidades = conn.execute("""
        SELECT DISTINCT entidad
        FROM auditoria
        WHERE entidad IS NOT NULL
        ORDER BY entidad
    """).fetchall()

    conn.close()

    return render_template(
        "auditoria.html",
        registros=registros,
        filtros=filtros,
        acciones=[row["accion"] for row in acciones],
        entidades=[row["entidad"] for row in entidades],
    )


@app.route("/auditoria/exportar")
def exportar_auditoria():
    check = permiso_requerido("auditoria_ver")
    if check:
        return check

    filtros = {
        "q": request.args.get("q", "").strip(),
        "accion": request.args.get("accion", "").strip(),
        "entidad": request.args.get("entidad", "").strip(),
        "usuario": request.args.get("usuario", "").strip(),
        "desde": request.args.get("desde", "").strip(),
        "hasta": request.args.get("hasta", "").strip(),
    }

    condiciones = []
    params = []
    if filtros["q"]:
        patron = f"%{filtros['q']}%"
        condiciones.append("""
            (
                accion ILIKE %s OR entidad ILIKE %s OR entidad_id ILIKE %s
                OR detalle ILIKE %s OR ip ILIKE %s OR username ILIKE %s
            )
        """)
        params.extend([patron, patron, patron, patron, patron, patron])
    if filtros["accion"]:
        condiciones.append("accion = %s")
        params.append(filtros["accion"])
    if filtros["entidad"]:
        condiciones.append("entidad = %s")
        params.append(filtros["entidad"])
    if filtros["usuario"]:
        condiciones.append("username ILIKE %s")
        params.append(f"%{filtros['usuario']}%")
    if validar_fecha_movimiento(filtros["desde"]):
        condiciones.append("fecha::date >= %s::date")
        params.append(filtros["desde"])
    if validar_fecha_movimiento(filtros["hasta"]):
        condiciones.append("fecha::date <= %s::date")
        params.append(filtros["hasta"])

    where_sql = "WHERE " + " AND ".join(condiciones) if condiciones else ""
    conn = get_connection()
    registros = conn.execute(f"""
        SELECT *
        FROM auditoria
        {where_sql}
        ORDER BY fecha DESC, id DESC
        LIMIT 5000
    """, tuple(params)).fetchall()
    registros = enriquecer_actores_auditoria(conn, registros)
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Fecha", "Usuario", "Rol", "Accion", "Entidad", "Entidad ID", "IP", "Detalle"])
    for registro in registros:
        writer.writerow([
            formato_fecha_hora(registro["fecha"]),
            registro["actor_display"] or registro["username"],
            registro["rol"],
            registro["accion"],
            registro["entidad"],
            registro["entidad_id"],
            registro["ip"],
            registro["detalle"],
        ])

    contenido = output.getvalue()
    output.close()
    registrar_auditoria("exportar_ok", "auditoria", None, {"cantidad_registros": len(registros)})

    return Response(
        contenido,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=auditoria.csv"},
    )


@app.route("/seguridad")
def ver_seguridad():
    check = permiso_requerido("seguridad_ver")
    if check:
        return check

    conn = get_connection()
    login_resumen = conn.execute("""
        SELECT
            COUNT(*) AS intentos,
            SUM(CASE WHEN success = 1 THEN 1 ELSE 0 END) AS exitosos,
            SUM(CASE WHEN success = 0 THEN 1 ELSE 0 END) AS fallidos
        FROM login_attempts
        WHERE fecha >= CURRENT_TIMESTAMP - INTERVAL '7 days'
    """).fetchone()

    intentos_recientes = conn.execute("""
        SELECT *
        FROM login_attempts
        ORDER BY fecha DESC, id DESC
        LIMIT 80
    """).fetchall()

    ips_fallidas = conn.execute("""
        SELECT ip, username, COUNT(*) AS fallidos, MAX(fecha) AS ultimo
        FROM login_attempts
        WHERE success = 0
          AND fecha >= CURRENT_TIMESTAMP - INTERVAL '24 hours'
        GROUP BY ip, username
        HAVING COUNT(*) >= 2
        ORDER BY fallidos DESC, ultimo DESC
        LIMIT 20
    """).fetchall()

    acciones_sensibles = conn.execute("""
        SELECT *
        FROM auditoria
        WHERE accion ILIKE %s
           OR accion ILIKE %s
           OR accion ILIKE %s
           OR accion ILIKE %s
           OR accion ILIKE %s
           OR accion ILIKE %s
           OR accion ILIKE %s
        ORDER BY fecha DESC, id DESC
        LIMIT 80
    """, (
        "%eliminar%",
        "%password%",
        "%login%",
        "%portal%",
        "%mantenimiento%",
        "%rol%",
        "%usuario%",
    )).fetchall()
    conn.close()

    return render_template(
        "seguridad.html",
        login_resumen=login_resumen,
        intentos_recientes=intentos_recientes,
        ips_fallidas=ips_fallidas,
        acciones_sensibles=acciones_sensibles,
        max_login_attempts=MAX_LOGIN_ATTEMPTS,
        login_window=LOGIN_ATTEMPT_WINDOW_MINUTES,
    )


def obtener_roles(conn):
    return conn.execute("""
        SELECT id, nombre, descripcion, sistema
        FROM roles
        ORDER BY sistema DESC, nombre ASC
    """).fetchall()


def permisos_desde_formulario():
    return normalizar_permisos(request.form.getlist("permisos"))


@app.route("/roles")
def listar_roles():
    check = permiso_requerido("roles_gestionar")
    if check:
        return check

    conn = get_connection()
    roles = conn.execute("""
        SELECT
            r.id,
            r.nombre,
            r.descripcion,
            r.permisos,
            r.sistema,
            COUNT(u.id) AS usuarios
        FROM roles r
        LEFT JOIN usuarios u ON u.rol = r.nombre
        GROUP BY r.id, r.nombre, r.descripcion, r.permisos, r.sistema
        ORDER BY r.sistema DESC, r.nombre ASC
    """).fetchall()
    conn.close()

    roles = [dict(rol) for rol in roles]
    for rol in roles:
        rol["cantidad_permisos"] = len(deserializar_permisos(rol["permisos"], rol["nombre"]))

    return render_template("roles.html", roles=roles)


@app.route("/roles/nuevo", methods=["GET", "POST"])
def nuevo_rol():
    check = permiso_requerido("roles_gestionar")
    if check:
        return check

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip().lower()
        descripcion = request.form.get("descripcion", "").strip()
        permisos = permisos_desde_formulario()

        if not nombre:
            flash("El nombre del rol es obligatorio.", "error")
            return render_template(
                "rol_form.html",
                rol={"nombre": nombre, "descripcion": descripcion},
                grupos=grupos_permisos(),
                permisos_seleccionados=permisos,
                modo="nuevo",
            )

        if not re.match(r"^[a-z0-9_ -]+$", nombre):
            flash("El rol solo puede usar letras, números, espacios, guiones o guiones bajos.", "error")
            return render_template(
                "rol_form.html",
                rol={"nombre": nombre, "descripcion": descripcion},
                grupos=grupos_permisos(),
                permisos_seleccionados=permisos,
                modo="nuevo",
            )

        conn = get_connection()
        existente = conn.execute("""
            SELECT id
            FROM roles
            WHERE nombre = %s
        """, (nombre,)).fetchone()

        if existente:
            conn.close()
            flash("Ya existe un rol con ese nombre.", "error")
            return render_template(
                "rol_form.html",
                rol={"nombre": nombre, "descripcion": descripcion},
                grupos=grupos_permisos(),
                permisos_seleccionados=permisos,
                modo="nuevo",
            )

        conn.execute("""
            INSERT INTO roles (nombre, descripcion, permisos, sistema)
            VALUES (%s, %s, %s, 0)
        """, (nombre, descripcion, serializar_permisos(permisos)))
        conn.commit()
        conn.close()

        flash("Rol creado correctamente.", "ok")
        return redirect(url_for("listar_roles"))

    return render_template(
        "rol_form.html",
        rol=None,
        grupos=grupos_permisos(),
        permisos_seleccionados=[],
        modo="nuevo",
    )


@app.route("/roles/<int:rol_id>/editar", methods=["GET", "POST"])
def editar_rol(rol_id):
    check = permiso_requerido("roles_gestionar")
    if check:
        return check

    conn = get_connection()
    rol = conn.execute("""
        SELECT *
        FROM roles
        WHERE id = %s
    """, (rol_id,)).fetchone()

    if rol is None:
        conn.close()
        flash("Rol no encontrado.", "error")
        return redirect(url_for("listar_roles"))

    if request.method == "POST":
        descripcion = request.form.get("descripcion", "").strip()
        permisos = permisos_desde_formulario()
        conn.execute("""
            UPDATE roles
            SET descripcion = %s,
                permisos = %s
            WHERE id = %s
        """, (descripcion, serializar_permisos(permisos), rol_id))
        conn.commit()
        conn.close()
        flash("Rol actualizado correctamente. Los usuarios con este rol deberán volver a iniciar sesión para tomar los nuevos permisos.", "ok")
        return redirect(url_for("listar_roles"))

    permisos_actuales = deserializar_permisos(rol["permisos"], rol["nombre"])
    conn.close()
    return render_template(
        "rol_form.html",
        rol=rol,
        grupos=grupos_permisos(),
        permisos_seleccionados=permisos_actuales,
        modo="editar",
    )


@app.route("/roles/<int:rol_id>/eliminar", methods=["POST"])
def eliminar_rol(rol_id):
    check = permiso_requerido("roles_gestionar")
    if check:
        return check

    conn = get_connection()
    rol = conn.execute("""
        SELECT *
        FROM roles
        WHERE id = %s
    """, (rol_id,)).fetchone()

    if rol is None:
        conn.close()
        flash("Rol no encontrado.", "error")
        return redirect(url_for("listar_roles"))

    if rol["sistema"]:
        conn.close()
        flash("No se pueden eliminar los roles base del sistema.", "error")
        return redirect(url_for("listar_roles"))

    usuarios = conn.execute("""
        SELECT COUNT(*) AS total
        FROM usuarios
        WHERE rol = %s
    """, (rol["nombre"],)).fetchone()["total"]

    if usuarios:
        conn.close()
        flash("No se puede eliminar un rol asignado a usuarios.", "error")
        return redirect(url_for("listar_roles"))

    conn.execute("DELETE FROM roles WHERE id = %s", (rol_id,))
    conn.commit()
    conn.close()
    flash("Rol eliminado correctamente.", "ok")
    return redirect(url_for("listar_roles"))


@app.route("/usuarios/nuevo", methods=["GET", "POST"])
def nuevo_usuario():
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()
    roles = obtener_roles(conn)

    if request.method == "POST":
        username = normalizar_username(request.form.get("username", ""))
        email = normalizar_email(request.form.get("email", ""))
        password = request.form.get("password", "")
        rol = request.form.get("rol", "tesorero")

        rol_existe = conn.execute("""
            SELECT id
            FROM roles
            WHERE nombre = %s
        """, (rol,)).fetchone()

        if not rol_existe:
            conn.close()
            flash("Rol invalido.", "error")
            return render_template("usuario_form.html", username=username, email=email, rol=rol, roles=roles)

        if not username or not password:
            conn.close()
            flash("Usuario y clave son obligatorios.", "error")
            return render_template("usuario_form.html", username=username, email=email, rol=rol, roles=roles)

        password_hash = generate_password_hash(password)

        existente = conn.execute("""
            SELECT id FROM usuarios
            WHERE lower(username) = %s
               OR (email IS NOT NULL AND email <> '' AND lower(email) = %s)
        """, (username, email)).fetchone()

        if existente:
            conn.close()
            flash("Ese usuario o email ya existe.", "error")
            return render_template("usuario_form.html", username=username, email=email, rol=rol, roles=roles)

        conn.execute("""
            INSERT INTO usuarios (username, email, password, rol, debe_cambiar_password, onboarding_visto)
            VALUES (%s, %s, %s, %s, 1, 0)
        """, (username, email or None, password_hash, rol))

        conn.commit()
        conn.close()

        flash(f"Usuario creado correctamente con rol {rol}.", "ok")
        return redirect(url_for("listar_usuarios"))

    conn.close()
    return render_template("usuario_form.html", roles=roles)


@app.route("/usuarios")
def listar_usuarios():
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()

    usuarios = conn.execute("""
        SELECT id, username, email, rol, debe_cambiar_password, onboarding_visto, ultimo_login
        FROM usuarios
        ORDER BY username
    """).fetchall()

    conn.close()

    return render_template("usuarios.html", usuarios=usuarios)


@app.route("/usuarios/<int:usuario_id>/editar", methods=["GET", "POST"])
def editar_usuario(usuario_id):
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()
    usuario = conn.execute("""
        SELECT id, username, email, rol
        FROM usuarios
        WHERE id = %s
    """, (usuario_id,)).fetchone()

    if usuario is None:
        conn.close()
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("listar_usuarios"))

    roles = obtener_roles(conn)

    if request.method == "POST":
        username = normalizar_username(request.form.get("username", ""))
        email = normalizar_email(request.form.get("email", ""))
        rol = request.form.get("rol", "").strip()
        usuario_form = {
            **usuario,
            "username": username,
            "email": email,
            "rol": rol,
        }

        if not username:
            conn.close()
            flash("El usuario es obligatorio.", "error")
            return render_template("usuario_form.html", usuario=usuario_form, rol=rol, roles=roles, modo="editar")

        rol_existe = conn.execute("""
            SELECT id
            FROM roles
            WHERE nombre = %s
        """, (rol,)).fetchone()

        if not rol_existe:
            conn.close()
            flash("Rol invalido.", "error")
            return render_template("usuario_form.html", usuario=usuario_form, rol=rol, roles=roles, modo="editar")

        email_param = email or None
        if email_param:
            duplicado = conn.execute("""
                SELECT id
                FROM usuarios
                WHERE id <> %s
                  AND (
                      lower(username) = %s
                      OR lower(email) = %s
                  )
                LIMIT 1
            """, (usuario_id, username, email_param)).fetchone()
        else:
            duplicado = conn.execute("""
                SELECT id
                FROM usuarios
                WHERE id <> %s
                  AND lower(username) = %s
                LIMIT 1
            """, (usuario_id, username)).fetchone()

        if duplicado:
            conn.close()
            flash("Ese usuario o email ya existe.", "error")
            return render_template("usuario_form.html", usuario=usuario_form, rol=rol, roles=roles, modo="editar")

        conn.execute("""
            UPDATE usuarios
            SET username = %s,
                email = %s,
                rol = %s
            WHERE id = %s
        """, (username, email_param, rol, usuario_id))
        permisos_actualizados = cargar_permisos_rol(conn, rol)
        conn.commit()
        conn.close()

        if session.get("user_id") == usuario_id:
            session["username"] = username
            session["rol"] = rol
            session["permisos"] = permisos_actualizados
            flash("Tu usuario fue actualizado.", "ok")
        else:
            flash("Usuario actualizado correctamente. Si estaba conectado, debera volver a iniciar sesion para tomar los cambios.", "ok")
        return redirect(url_for("listar_usuarios"))

    conn.close()
    return render_template("usuario_form.html", usuario=usuario, rol=usuario["rol"], roles=roles, modo="editar")


@app.route("/usuarios/<int:usuario_id>/password", methods=["GET", "POST"])
def resetear_password_usuario(usuario_id):
    check = rol_requerido("admin")
    if check:
        return check

    if session.get("user_id") == usuario_id:
        return redirect(url_for("cambiar_mi_password"))

    conn = get_connection()
    usuario = conn.execute("""
        SELECT id, username, email, rol
        FROM usuarios
        WHERE id = %s
    """, (usuario_id,)).fetchone()

    if usuario is None:
        conn.close()
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("listar_usuarios"))

    if request.method == "POST":
        password_nueva = request.form.get("password_nueva", "")
        password_confirmacion = request.form.get("password_confirmacion", "")

        error = validar_password_nueva(password_nueva, password_confirmacion)
        if error:
            conn.close()
            flash(error, "error")
            return render_template("password_form.html", usuario=usuario, modo="admin")

        conn.execute("""
            UPDATE usuarios
            SET password = %s,
                debe_cambiar_password = 1
            WHERE id = %s
        """, (generate_password_hash(password_nueva), usuario_id))
        conn.commit()
        conn.close()

        flash(f"Clave de {usuario['username']} actualizada. Debera cambiarla al ingresar.", "ok")
        return redirect(url_for("listar_usuarios"))

    conn.close()
    return render_template("password_form.html", usuario=usuario, modo="admin")

@app.route("/usuarios/<int:usuario_id>/eliminar", methods=["POST"])
def eliminar_usuario(usuario_id):
    check = rol_requerido("admin")
    if check:
        return check

    if session.get("user_id") == usuario_id:
        flash("No podés eliminar tu propio usuario mientras estás logueado.", "error")
        return redirect(url_for("listar_usuarios"))

    conn = get_connection()
    conn.execute("DELETE FROM password_reset_tokens WHERE usuario_id = %s", (usuario_id,))
    conn.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
    conn.commit()
    conn.close()

    flash("Usuario eliminado correctamente.", "ok")
    return redirect(url_for("listar_usuarios"))

@app.route("/cuotas/<int:cuota_id>/eliminar", methods=["POST"])
def eliminar_cuota(cuota_id):
    check = permiso_requerido("cuotas_gestionar")
    if check:
        return check

    conn = get_connection()

    cuota = conn.execute("""
        SELECT jugador_id, anulada, plan_pago_id
        FROM cuotas
        WHERE id = %s
    """, (cuota_id,)).fetchone()

    if cuota is None:
        conn.close()
        flash("Cuota no encontrada.", "error")
        return redirect(url_for("listar_jugadores"))

    if cuota.get("anulada") and cuota.get("plan_pago_id"):
        conn.close()
        flash("La cuota esta incluida en un plan de pago. Elimina el plan para restaurarla o quitarla.", "error")
        return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

    conn.execute("DELETE FROM cuotas WHERE id = %s", (cuota_id,))
    conn.commit()
    conn.close()

    flash("Cuota eliminada correctamente.", "ok")
    return redirect(url_for("ver_cuotas", jugador_id=cuota["jugador_id"]))

def generar_recibo_pdf(cuota_id):
    conn = get_connection()

    datos = conn.execute("""
        SELECT
            c.id AS cuota_id,
            c.periodo,
            c.importe,
            c.fecha_pago,
            c.fecha_vencimiento,
            c.numero_recibo,
            c.metodo_pago,
            c.referencia_pago,
            c.importe_original,
            c.descuento_beca,
            c.beca_porcentaje,
            c.becada,
            j.nombre,
            j.apellido,
            j.dni,
            j.categoria
        FROM cuotas c
        JOIN jugadores j ON j.id = c.jugador_id
        WHERE c.id = %s
    """, (cuota_id,)).fetchone()

    conn.close()

    if datos is None:
        return None

    recibos_dir = BASE_DIR / "recibos"
    recibos_dir.mkdir(exist_ok=True)
    archivo = recibos_dir / f"recibo_cuota_{cuota_id}.pdf"

    pdf = canvas.Canvas(str(archivo), pagesize=A4)
    width, height = A4
    margen_izquierdo = 22 * mm
    margen_derecho = 188 * mm

    pdf.setFillColor(colors.HexColor("#f7fafb"))
    pdf.rect(0, 0, width, height, stroke=0, fill=1)

    header_x = 18 * mm
    header_y = height - 52 * mm
    header_w = 174 * mm
    header_h = 34 * mm

    pdf.setFillColor(colors.white)
    pdf.roundRect(header_x, header_y, header_w, header_h, 7 * mm, stroke=0, fill=1)
    pdf.setStrokeColor(colors.HexColor("#d6a443"))
    pdf.setLineWidth(1.2)
    pdf.roundRect(header_x, header_y, header_w, header_h, 7 * mm, stroke=1, fill=0)

    logo_path = BASE_DIR / "static" / "img" / "logo.png"
    firma_path = BASE_DIR / "static" / "img" / "firma-tesoreria.png"

    if logo_path.exists():
        pdf.drawImage(
            ImageReader(str(logo_path)),
            margen_izquierdo + 1 * mm,
            header_y + 6 * mm,
            width=20 * mm,
            height=20 * mm,
            preserveAspectRatio=True,
            mask="auto"
        )
        texto_x = margen_izquierdo + 28 * mm
    else:
        texto_x = margen_izquierdo

    titulo_y = header_y + header_h - 10 * mm
    subtitulo_y = titulo_y - 6.5 * mm
    meta_y = titulo_y
    meta_sub_y = subtitulo_y

    pdf.setFillColor(colors.HexColor("#10231e"))
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(texto_x, titulo_y, "Ruda Macho Rugby Club")

    pdf.setFillColor(colors.HexColor("#475569"))
    pdf.setFont("Helvetica", 10)
    pdf.drawString(texto_x, subtitulo_y, "Recibo interno no v\u00e1lido como factura")

    pdf.setFillColor(colors.HexColor("#10231e"))
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawRightString(margen_derecho - 2 * mm, meta_y, f"Nro. {datos['numero_recibo'] or datos['cuota_id']}")

    pdf.setFillColor(colors.HexColor("#475569"))
    pdf.setFont("Helvetica", 10)
    pdf.drawRightString(margen_derecho - 2 * mm, meta_sub_y, f"Emitido: {ahora_sig().strftime('%d/%m/%Y')}")

    pdf.setStrokeColor(colors.HexColor("#d6dee8"))
    pdf.setLineWidth(1)
    pdf.line(margen_izquierdo, header_y - 5 * mm, margen_derecho, header_y - 5 * mm)

    pdf.setFillColor(colors.HexColor("#10231e"))
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(margen_izquierdo, header_y - 16 * mm, "RECIBO DE CUOTA")

    filas = [
        ("Jugador", f"{datos['apellido']}, {datos['nombre']}"),
        ("DNI", str(datos['dni'] or "-")),
        ("Categor\u00eda", str(datos['categoria'] or "-")),
        ("Per\u00edodo abonado", str(datos['periodo'])),
        ("Importe abonado", f"${int(float(datos['importe'])):,}".replace(",", ".")),
    ]
    if datos["becada"]:
        original = f"${int(float(datos['importe_original'] or datos['importe'])):,}".replace(",", ".")
        descuento = f"${int(float(datos['descuento_beca'] or 0)):,}".replace(",", ".")
        porcentaje_beca_pdf = float(datos["beca_porcentaje"] or 0)
        filas.append(("Beca aplicada", f"{porcentaje_beca_pdf:g}% - Original {original} - Descuento {descuento}"))
    filas.extend([
        ("Fecha de pago", str(datos['fecha_pago'] or "-")),
        ("Vencimiento original", str(datos['fecha_vencimiento'] or "-")),
        ("M\u00e9todo de pago", str(datos['metodo_pago'] or "-")),
        ("Referencia", str(datos['referencia_pago'] or "-")),
    ])

    fila_alto = 8 * mm
    card_top = header_y - 24 * mm
    card_height = (len(filas) * fila_alto) + (18 * mm)
    card_bottom = card_top - card_height

    pdf.setFillColor(colors.white)
    pdf.roundRect(18 * mm, card_bottom, 174 * mm, card_height, 6 * mm, stroke=0, fill=1)
    pdf.setStrokeColor(colors.HexColor("#d6dee8"))
    pdf.roundRect(18 * mm, card_bottom, 174 * mm, card_height, 6 * mm, stroke=1, fill=0)

    y = card_top - 12 * mm
    for label_texto, valor_texto in filas:
        pdf.setFillColor(colors.HexColor("#64748b"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(margen_izquierdo, y, label_texto)
        pdf.setFillColor(colors.HexColor("#17212f"))
        pdf.setFont("Helvetica", 12)
        pdf.drawString(62 * mm, y, valor_texto)
        y -= fila_alto

    firma_y = card_bottom - 18 * mm
    pdf.setStrokeColor(colors.HexColor("#94a3b8"))
    pdf.line(margen_izquierdo, firma_y, 94 * mm, firma_y)
    if firma_path.exists():
        pdf.drawImage(
            ImageReader(str(firma_path)),
            margen_izquierdo,
            firma_y - 14 * mm,
            width=42 * mm,
            height=14 * mm,
            preserveAspectRatio=True,
            mask="auto",
        )
    pdf.setFillColor(colors.HexColor("#17212f"))
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(margen_izquierdo, firma_y - 21 * mm, TESORERO_FIRMA_NOMBRE)
    pdf.setFont("Helvetica", 10)
    pdf.setFillColor(colors.HexColor("#64748b"))
    pdf.drawString(margen_izquierdo, firma_y - 27 * mm, f"{TESORERO_FIRMA_CARGO} - Ruda Macho Rugby Club")

    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawString(margen_izquierdo, max(12 * mm, firma_y - 40 * mm), "Ruda Macho Rugby Club - Sistema Integral de Gesti\u00f3n")

    pdf.save()
    return archivo

@app.route("/cuotas/<int:cuota_id>/recibo")
def descargar_recibo(cuota_id):
    check = permiso_requerido("cuotas_ver")
    if check:
        return check

    archivo = generar_recibo_pdf(cuota_id)

    if archivo is None or not archivo.exists():
        flash("No se pudo generar el recibo.", "error")
        return redirect(url_for("listar_jugadores"))

    registrar_auditoria(
        "descargar_ok",
        "recibo",
        str(cuota_id),
        {"archivo": archivo.name},
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"recibo_cuota_{cuota_id}.pdf"
    )

@app.route("/movimientos/<int:movimiento_id>/comprobante")
def descargar_comprobante_movimiento(movimiento_id):
    check = permiso_requerido("caja_ver", "caja_gestionar")
    if check:
        return check

    conn = get_connection()
    movimiento = conn.execute("""
        SELECT *
        FROM movimientos
        WHERE id = %s
    """, (movimiento_id,)).fetchone()
    conn.close()

    if movimiento is None:
        flash("Movimiento no encontrado.", "error")
        return redirect(url_for("ver_caja"))
    if not movimiento.get("comprobante_drive_file_id"):
        flash("El movimiento no tiene comprobante adjunto.", "error")
        return redirect(url_for("ver_caja", mes=movimiento["fecha"][:7]))

    try:
        archivo = descargar_drive_file(movimiento["comprobante_drive_file_id"])
    except RuntimeError as error:
        flash(str(error), "error")
        return redirect(url_for("ver_caja", mes=movimiento["fecha"][:7]))
    except Exception as error:
        app.logger.exception("No se pudo descargar comprobante de movimiento %s.", movimiento_id)
        flash(mensaje_error_drive(error, carpeta="Caja", accion="descargar el comprobante"), "error")
        return redirect(url_for("ver_caja", mes=movimiento["fecha"][:7]))

    registrar_auditoria(
        "descargar_ok",
        "comprobante_movimiento",
        str(movimiento_id),
        {"archivo": movimiento.get("comprobante_nombre"), "drive_file_id": movimiento.get("comprobante_drive_file_id")},
    )

    return send_file(
        archivo,
        mimetype=movimiento.get("comprobante_mime_type") or "application/octet-stream",
        as_attachment=False,
        download_name=movimiento.get("comprobante_nombre") or f"comprobante_movimiento_{movimiento_id}",
    )


@app.route("/caja/exportar")
def exportar_caja():
    check = permiso_requerido("caja_ver")
    if check:
        return check

    mes = request.args.get("mes")
    mes_actual = mes or ahora_sig().strftime("%Y-%m")

    conn = get_connection()

    movimientos = conn.execute("""
        SELECT *
        FROM movimientos
        WHERE substring(fecha from 1 for 7) = %s
          AND COALESCE(anulado, 0) = 0
        ORDER BY fecha ASC, id ASC
    """, (mes_actual,)).fetchall()

    ingresos_mes = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'ingreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    egresos_mes = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'egreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes_actual,)).fetchone()["total"]

    conn.close()

    resultado_mes = ingresos_mes - egresos_mes

    wb = Workbook()
    ws = wb.active
    ws.title = "Caja"

    ws["A1"] = "Ruda Macho Rugby Club"
    ws["A2"] = f"Balance de caja - {mes_actual}"

    ws["A4"] = "Ingresos del mes"
    ws["B4"] = ingresos_mes
    ws["A5"] = "Egresos del mes"
    ws["B5"] = egresos_mes
    ws["A6"] = "Resultado del mes"
    ws["B6"] = resultado_mes

    encabezados = ["Fecha", "Tipo", "Concepto", "Referencia", "Operacion", "Comprobante", "Monto"]
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=8, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F2937")
        celda.alignment = Alignment(horizontal="center")

    fila = 9
    for m in movimientos:
        ws.cell(row=fila, column=1).value = m["fecha"]
        ws.cell(row=fila, column=2).value = m["tipo"]
        ws.cell(row=fila, column=3).value = m["concepto"]
        ws.cell(row=fila, column=4).value = m["referencia"]
        ws.cell(row=fila, column=5).value = m.get("comprobante_operacion")
        ws.cell(row=fila, column=6).value = m.get("comprobante_nombre")
        ws.cell(row=fila, column=7).value = m["monto"]

        if m["tipo"] == "ingreso":
            ws.cell(row=fila, column=2).font = Font(color="166534", bold=True)
            ws.cell(row=fila, column=7).font = Font(color="166534", bold=True)
        else:
            ws.cell(row=fila, column=2).font = Font(color="DC2626", bold=True)
            ws.cell(row=fila, column=7).font = Font(color="DC2626", bold=True)

        fila += 1

    for row in range(4, 7):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '$ #,##0'

    for row in range(9, fila):
        ws.cell(row=row, column=7).number_format = '$ #,##0'

    thin = Side(style="thin", color="D1D5DB")
    for row in ws.iter_rows(min_row=8, max_row=max(fila - 1, 8), min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(size=12)

    widths = {
        "A": 14,
        "B": 14,
        "C": 40,
        "D": 32,
        "E": 24,
        "F": 32,
        "G": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A9"

    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(exist_ok=True)

    archivo = export_dir / f"caja_{mes_actual}.xlsx"
    wb.save(archivo)

    registrar_auditoria(
        "exportar_ok",
        "caja",
        mes_actual,
        {"formato": "xlsx", "cantidad_registros": len(movimientos)},
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=f"caja_{mes_actual}.xlsx"
    )

@app.route("/movimientos/<int:movimiento_id>/editar", methods=["GET", "POST"])
def editar_movimiento(movimiento_id):
    check = permiso_requerido("caja_gestionar")
    if check:
        return check

    conn = get_connection()

    movimiento = conn.execute("""
        SELECT *
        FROM movimientos
        WHERE id = %s
    """, (movimiento_id,)).fetchone()

    if movimiento is None:
        conn.close()
        flash("Movimiento no encontrado.", "error")
        return redirect(url_for("ver_caja"))

    if movimiento["anulado"]:
        conn.close()
        flash("No se puede editar un movimiento anulado.", "error")
        return redirect(url_for("ver_caja", mes=movimiento["fecha"][:7]))

    if request.method == "POST":
        tipo = request.form.get("tipo")
        concepto = request.form.get("concepto")
        monto = request.form.get("monto", "").strip()
        fecha = validar_fecha_movimiento(request.form.get("fecha", "").strip())
        referencia = request.form.get("referencia", "").strip()
        comprobante_pago = request.files.get("comprobante_pago")

        if not fecha:
            flash("La fecha del movimiento no es valida.", "error")
            movimiento_form = dict(movimiento)
            movimiento_form.update({"tipo": tipo, "concepto": concepto, "monto": monto, "fecha": request.form.get("fecha", "").strip(), "referencia": referencia})
            conn.close()
            return render_template("movimiento_form_editar.html", movimiento=movimiento_form)

        mes_destino = fecha[:7]
        if mes_esta_cerrado(mes_destino):
            flash("No se puede mover un movimiento a un mes cerrado.", "error")
            movimiento_form = dict(movimiento)
            movimiento_form.update({"tipo": tipo, "concepto": concepto, "monto": monto, "fecha": fecha, "referencia": referencia})
            conn.close()
            return render_template("movimiento_form_editar.html", movimiento=movimiento_form)

        comprobante_info = None
        numero_operacion = ""
        monto_ocr = ""
        ocr_texto = ""
        if comprobante_pago and comprobante_pago.filename:
            try:
                comprobante_info, numero_operacion, monto_ocr, ocr_texto = procesar_comprobante_movimiento(
                    comprobante_pago,
                    {"id": movimiento_id, "tipo": tipo, "concepto": concepto, "fecha": fecha},
                    existing_file_id=movimiento.get("comprobante_drive_file_id"),
                )
            except (RuntimeError, ValueError) as error:
                flash(str(error), "error")
                movimiento_form = dict(movimiento)
                movimiento_form.update({"tipo": tipo, "concepto": concepto, "monto": monto, "fecha": fecha, "referencia": referencia})
                conn.close()
                return render_template("movimiento_form_editar.html", movimiento=movimiento_form)
            except Exception as error:
                app.logger.exception("No se pudo procesar comprobante de movimiento de caja %s.", movimiento_id)
                flash(mensaje_error_drive(error, carpeta="Caja", accion="subir o leer el comprobante"), "error")
                movimiento_form = dict(movimiento)
                movimiento_form.update({"tipo": tipo, "concepto": concepto, "monto": monto, "fecha": fecha, "referencia": referencia})
                conn.close()
                return render_template("movimiento_form_editar.html", movimiento=movimiento_form)

        if not referencia and numero_operacion:
            referencia = numero_operacion
        if not monto and monto_ocr:
            monto = monto_ocr
        if not monto:
            flash("Debe indicar un monto o adjuntar un comprobante donde pueda leerse el monto.", "error")
            movimiento_form = dict(movimiento)
            movimiento_form.update({"tipo": tipo, "concepto": concepto, "monto": monto, "fecha": fecha, "referencia": referencia})
            conn.close()
            return render_template("movimiento_form_editar.html", movimiento=movimiento_form)

        conn.execute("""
            UPDATE movimientos
            SET tipo = %s,
                concepto = %s,
                monto = %s,
                fecha = %s,
                referencia = %s,
                comprobante_drive_file_id = COALESCE(%s, comprobante_drive_file_id),
                comprobante_nombre = COALESCE(%s, comprobante_nombre),
                comprobante_mime_type = COALESCE(%s, comprobante_mime_type),
                comprobante_tamano = COALESCE(%s, comprobante_tamano),
                comprobante_fecha = COALESCE(%s, comprobante_fecha),
                comprobante_usuario = COALESCE(%s, comprobante_usuario),
                comprobante_web_url = COALESCE(%s, comprobante_web_url),
                comprobante_operacion = COALESCE(%s, comprobante_operacion),
                comprobante_ocr_texto = COALESCE(%s, comprobante_ocr_texto)
            WHERE id = %s
        """, (
            tipo, concepto, monto, fecha, referencia,
            comprobante_info["file_id"] if comprobante_info else None,
            comprobante_info["nombre"] if comprobante_info else None,
            comprobante_info["mime_type"] if comprobante_info else None,
            comprobante_info["tamano"] if comprobante_info else None,
            ahora_sig().strftime("%Y-%m-%d %H:%M:%S") if comprobante_info else None,
            session.get("username") if comprobante_info else None,
            comprobante_info["web_url"] if comprobante_info else None,
            numero_operacion or None,
            ocr_texto or None,
            movimiento_id,
        ))
        conn.commit()
        conn.close()

        flash("Movimiento actualizado.", "ok")
        return redirect(url_for("ver_caja", mes=mes_destino))

    mes_movimiento = movimiento["fecha"][:7]

    if mes_esta_cerrado(mes_movimiento):
        conn.close()
        flash("No se puede editar un movimiento de un mes cerrado.", "error")
        return redirect(url_for("ver_caja", mes=mes_movimiento))

    conn.close()
    return render_template("movimiento_form_editar.html", movimiento=movimiento)

@app.route("/movimientos/<int:movimiento_id>/eliminar", methods=["POST"])
def eliminar_movimiento(movimiento_id):
    check = permiso_requerido("caja_gestionar")
    if check:
        return check

    conn = get_connection()

    movimiento = conn.execute("""
        SELECT *
        FROM movimientos
        WHERE id = %s
    """, (movimiento_id,)).fetchone()

    if movimiento is None:
        conn.close()
        flash("Movimiento no encontrado.", "error")
        return redirect(url_for("ver_caja"))

    if movimiento["anulado"]:
        conn.close()
        flash("Ese movimiento ya estaba anulado.", "error")
        return redirect(url_for("ver_caja", mes=movimiento["fecha"][:7]))

    mes_movimiento = movimiento["fecha"][:7]
    if mes_esta_cerrado(mes_movimiento):
        conn.close()
        flash("No se puede anular un movimiento de un mes cerrado.", "error")
        return redirect(url_for("ver_caja", mes=mes_movimiento))

    motivo = request.form.get("motivo_anulacion", "").strip()
    if not motivo:
        conn.close()
        flash("Debe indicar un motivo para anular el movimiento.", "error")
        return redirect(url_for("ver_caja", mes=mes_movimiento))

    conn.execute("""
        UPDATE movimientos
        SET anulado = 1,
            fecha_anulacion = CURRENT_TIMESTAMP,
            usuario_anulacion = %s,
            motivo_anulacion = %s
        WHERE id = %s
    """, (session.get("username"), motivo, movimiento_id))
    conn.commit()
    conn.close()

    registrar_auditoria(
        "anular_detalle",
        "movimiento_caja",
        str(movimiento_id),
        {
            "fecha": movimiento["fecha"],
            "tipo": movimiento["tipo"],
            "concepto": movimiento["concepto"],
            "monto": movimiento["monto"],
            "referencia": movimiento["referencia"],
            "motivo_anulacion": motivo,
            "usuario_anulacion": session.get("username"),
        },
    )

    flash("Movimiento anulado correctamente.", "ok")
    return redirect(url_for("ver_caja", mes=mes_movimiento))

def mes_esta_cerrado(mes):
    conn = get_connection()
    cierre = conn.execute("""
        SELECT id
        FROM cierres_mensuales
        WHERE mes = %s
    """, (mes,)).fetchone()
    conn.close()
    return cierre is not None

@app.route("/caja/cerrar", methods=["POST"])
def cerrar_mes():
    check = permiso_requerido("caja_gestionar")
    if check:
        return check

    mes = request.form.get("mes", "").strip()

    if not mes:
        flash("Debe indicar un mes para cerrar.", "error")
        return redirect(url_for("ver_caja"))

    conn = get_connection()

    cierre_existente = conn.execute("""
        SELECT id
        FROM cierres_mensuales
        WHERE mes = %s
    """, (mes,)).fetchone()

    if cierre_existente:
        conn.close()
        flash("Ese mes ya está cerrado.", "error")
        return redirect(url_for("ver_caja", mes=mes))

    ingresos = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'ingreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes,)).fetchone()["total"]

    egresos = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total
        FROM movimientos
        WHERE tipo = 'egreso'
          AND COALESCE(anulado, 0) = 0
          AND substring(fecha from 1 for 7) = %s
    """, (mes,)).fetchone()["total"]

    resultado = ingresos - egresos

    conn.execute("""
        INSERT INTO cierres_mensuales (
            mes, ingresos, egresos, resultado, fecha_cierre, usuario
        )
        VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, %s)
    """, (
        mes,
        ingresos,
        egresos,
        resultado,
        session.get("username")
    ))

    conn.commit()
    conn.close()

    flash(f"Mes {mes} cerrado correctamente.", "ok")
    return redirect(url_for("ver_caja", mes=mes))


@app.route("/tests")
def listar_tests():
    check = permiso_requerido("tests_ver")
    if check:
        return check

    conn = get_connection()

    tests = conn.execute("""
        SELECT
            t.*,
            COUNT(r.id) AS mediciones,
            COUNT(DISTINCT r.jugador_id) AS jugadores_medidos,
            MAX(r.fecha) AS ultima_fecha
        FROM test_tipos t
        LEFT JOIN test_resultados r ON r.test_id = t.id
        GROUP BY t.id
        ORDER BY t.activo DESC, t.nombre
    """).fetchall()

    recientes = conn.execute("""
        SELECT
            r.*,
            t.nombre AS test_nombre,
            t.unidad,
            j.apellido,
            j.nombre,
            j.categoria
        FROM test_resultados r
        JOIN test_tipos t ON t.id = r.test_id
        JOIN jugadores j ON j.id = r.jugador_id
        ORDER BY r.fecha DESC, r.id DESC
        LIMIT 20
    """).fetchall()

    conn.close()

    return render_template("tests.html", tests=tests, recientes=recientes)


@app.route("/tests/nuevo", methods=["GET", "POST"])
def nuevo_test_tipo():
    check = permiso_requerido("tests_gestionar")
    if check:
        return check

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        unidad = request.form.get("unidad", "").strip()
        puntaje_min = parsear_puntaje_test(request.form.get("puntaje_min"))
        puntaje_max = parsear_puntaje_test(request.form.get("puntaje_max"))
        mayor_es_mejor = 1 if request.form.get("mayor_es_mejor") else 0
        activo = 1 if request.form.get("activo") else 0

        if not nombre:
            flash("El nombre del test es obligatorio.", "error")
            return render_template("test_form.html", test=request.form)

        if puntaje_min is not None and puntaje_max is not None and puntaje_min > puntaje_max:
            flash("El puntaje minimo no puede ser mayor al maximo.", "error")
            return render_template("test_form.html", test=request.form)

        conn = get_connection()
        try:
            conn.execute("""
                INSERT INTO test_tipos (
                    nombre, descripcion, unidad, puntaje_min, puntaje_max,
                    mayor_es_mejor, activo, creado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                nombre,
                descripcion,
                unidad,
                puntaje_min,
                puntaje_max,
                mayor_es_mejor,
                activo,
                session.get("username"),
            ))
            conn.commit()
        except Exception:
            conn.rollback()
            conn.close()
            flash("No se pudo crear el test. Revisá que el nombre no esté repetido.", "error")
            return render_template("test_form.html", test=request.form)

        conn.close()
        flash("Test creado correctamente.", "ok")
        return redirect(url_for("listar_tests"))

    return render_template("test_form.html", test={"mayor_es_mejor": 1, "activo": 1})


@app.route("/tests/<int:test_id>/cargar", methods=["GET", "POST"])
def cargar_test_resultados(test_id):
    check = permiso_requerido("tests_gestionar")
    if check:
        return check

    conn = get_connection()

    test = conn.execute("""
        SELECT *
        FROM test_tipos
        WHERE id = %s
    """, (test_id,)).fetchone()

    if test is None:
        conn.close()
        flash("Test no encontrado.", "error")
        return redirect(url_for("listar_tests"))

    jugadores = conn.execute("""
        SELECT id, apellido, nombre, dni, categoria, estado
        FROM jugadores
        WHERE COALESCE(estado, 'Activo') <> 'Baja'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY categoria, apellido, nombre
    """).fetchall()

    if request.method == "POST":
        fecha = normalizar_fecha_test(request.form.get("fecha"))
        cargados = 0
        omitidos = 0

        for jugador in jugadores:
            puntaje = parsear_puntaje_test(request.form.get(f"puntaje_{jugador['id']}"))
            observaciones = request.form.get(f"obs_{jugador['id']}", "").strip()

            if puntaje is None:
                omitidos += 1
                continue

            conn.execute("""
                INSERT INTO test_resultados (
                    test_id, jugador_id, fecha, puntaje, observaciones, creado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                test_id,
                jugador["id"],
                fecha,
                puntaje,
                observaciones,
                session.get("username"),
            ))
            cargados += 1

        conn.commit()
        conn.close()

        if cargados:
            flash(f"Se cargaron {cargados} mediciones.", "ok")
            if omitidos:
                flash(f"{omitidos} jugadores quedaron sin puntaje en esta carga.", "info")
            return redirect(url_for("graficos_tests", test_id=test_id))

        flash("No se cargaron puntajes. Completá al menos una medición.", "error")
        return redirect(url_for("cargar_test_resultados", test_id=test_id))

    conn.close()
    return render_template(
        "test_carga.html",
        test=test,
        jugadores=jugadores,
        fecha_hoy=ahora_sig().strftime("%Y-%m-%d"),
    )


@app.route("/tests/importar", methods=["GET", "POST"])
def importar_test_resultados():
    check = permiso_requerido("tests_gestionar")
    if check:
        return check

    conn = get_connection()
    tests = obtener_test_tipos(conn, solo_activos=True)
    batches_recientes = obtener_test_importaciones_batch_recientes(conn)

    if request.method == "POST":
        archivo = request.files.get("archivo")
        test_id_form = request.form.get("test_id", "").strip()
        test_id_fijo = int(test_id_form) if test_id_form.isdigit() else None

        if not archivo or not archivo.filename:
            flash("Seleccioná un archivo Excel para importar.", "error")
            conn.close()
            return render_template("test_importar.html", tests=tests, batches_recientes=batches_recientes)

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
        except Exception:
            conn.close()
            flash("No se pudo leer el Excel. Verificá el archivo.", "error")
            return render_template("test_importar.html", tests=tests, batches_recientes=batches_recientes)

        if not filas:
            conn.close()
            flash("El archivo no tiene filas para importar.", "error")
            return render_template("test_importar.html", tests=tests, batches_recientes=batches_recientes)

        headers = [normalizar_header_excel(valor) for valor in filas[0]]
        tests_por_nombre = {
            normalizar_header_excel(test["nombre"]): test["id"]
            for test in tests
        }

        jugadores = obtener_jugadores_selector(conn)
        batch_id = f"{ahora_sig().strftime('%Y%m%d%H%M%S')}_{secrets.token_urlsafe(6)}"
        pendientes = 0

        for numero_fila, row in enumerate(filas[1:], start=2):
            datos = {
                headers[index]: row[index] if index < len(row) else None
                for index in range(len(headers))
            }
            nombre_completo = str(
                datos.get("jugador") or
                datos.get("nombre_completo") or
                datos.get("jugador_nombre") or
                ""
            ).strip()
            nombre = str(datos.get("nombre") or "").strip()
            apellido = str(datos.get("apellido") or "").strip()
            jugador_sugerido, confianza, motivo = sugerir_jugador_por_nombre_test(datos, jugadores)
            puntaje = parsear_puntaje_test(
                datos.get("puntaje") or datos.get("score") or datos.get("valor") or datos.get("resultado")
            )
            fecha = normalizar_fecha_test(datos.get("fecha"), usar_hoy_si_vacia=False)
            observaciones = str(datos.get("observaciones") or datos.get("obs") or "").strip()

            test_id = test_id_fijo
            nombre_test = datos.get("test") or datos.get("test_nombre") or datos.get("nombre_test")
            if not test_id:
                test_id = tests_por_nombre.get(normalizar_header_excel(nombre_test))

            errores_fila = []
            if not jugador_sugerido:
                errores_fila.append("Revisar jugador")
            if not test_id:
                errores_fila.append("Revisar test")
            if not fecha:
                errores_fila.append("Revisar fecha")
            if puntaje is None:
                errores_fila.append("Revisar puntaje")

            error = ", ".join(errores_fila) if errores_fila else None
            conn.execute("""
                INSERT INTO test_importaciones_batch (
                    batch_id, estado, fila, test_id, test_nombre,
                    jugador_sugerido_id, confianza, motivo,
                    nombre_excel, apellido_excel, nombre_completo_excel,
                    fecha, puntaje, observaciones, error, creado_por
                )
                VALUES (%s, 'pendiente', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                batch_id,
                numero_fila,
                test_id,
                str(nombre_test or "").strip() or None,
                jugador_sugerido["id"] if jugador_sugerido else None,
                confianza,
                motivo,
                nombre or None,
                apellido or None,
                nombre_completo or None,
                fecha,
                puntaje,
                observaciones or None,
                error,
                session.get("username"),
            ))
            pendientes += 1

        conn.commit()
        conn.close()

        if pendientes:
            flash(f"Se prepararon {pendientes} mediciones para revisar antes de confirmar.", "ok")
            return redirect(url_for("revisar_test_importacion", batch_id=batch_id))

        flash("El Excel no tenia filas para revisar.", "error")
        return redirect(url_for("importar_test_resultados"))
    conn.close()
    return render_template("test_importar.html", tests=tests, batches_recientes=batches_recientes)


@app.route("/tests/importar/<batch_id>/revisar", methods=["GET", "POST"])
def revisar_test_importacion(batch_id):
    check = permiso_requerido("tests_gestionar")
    if check:
        return check

    conn = get_connection()
    tests = obtener_test_tipos(conn, solo_activos=True)
    jugadores = obtener_jugadores_selector(conn)

    if request.method == "POST":
        item_ids = request.form.getlist("item_ids")
        procesadas = 0
        omitidas = 0
        errores = 0

        for item_id in item_ids:
            if request.form.get(f"procesar_{item_id}") != "on":
                omitidas += 1
                continue

            jugador_id = request.form.get(f"jugador_id_{item_id}", "").strip()
            test_id = request.form.get(f"test_id_{item_id}", "").strip()
            fecha = normalizar_fecha_test(request.form.get(f"fecha_{item_id}"), usar_hoy_si_vacia=False)
            puntaje = parsear_puntaje_test(request.form.get(f"puntaje_{item_id}"))
            observaciones = request.form.get(f"observaciones_{item_id}", "").strip()

            if not jugador_id or not jugador_id.isdigit():
                errores += 1
                flash(f"La fila #{item_id} no tiene jugador asignado.", "error")
                continue
            if not test_id or not test_id.isdigit():
                errores += 1
                flash(f"La fila #{item_id} no tiene test asignado.", "error")
                continue
            if not fecha:
                errores += 1
                flash(f"La fila #{item_id} no tiene fecha valida.", "error")
                continue
            if puntaje is None:
                errores += 1
                flash(f"La fila #{item_id} no tiene puntaje valido.", "error")
                continue

            item = conn.execute("""
                SELECT *
                FROM test_importaciones_batch
                WHERE id = %s AND batch_id = %s AND estado = 'pendiente'
                FOR UPDATE
            """, (item_id, batch_id)).fetchone()

            jugador = conn.execute("""
                SELECT id
                FROM jugadores
                WHERE id = %s
            """, (jugador_id,)).fetchone()

            test = conn.execute("""
                SELECT id
                FROM test_tipos
                WHERE id = %s
            """, (test_id,)).fetchone()

            if not item or not jugador or not test:
                errores += 1
                flash(f"No se encontro la fila pendiente #{item_id}, el jugador o el test.", "error")
                continue

            conn.execute("""
                INSERT INTO test_resultados (
                    test_id, jugador_id, fecha, puntaje, observaciones, creado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                int(test_id),
                int(jugador_id),
                fecha,
                puntaje,
                observaciones or item["observaciones"],
                session.get("username"),
            ))

            conn.execute("""
                UPDATE test_importaciones_batch
                SET estado = 'procesado',
                    test_id = %s,
                    jugador_id = %s,
                    fecha = %s,
                    puntaje = %s,
                    observaciones = %s,
                    procesado_en = CURRENT_TIMESTAMP,
                    procesado_por = %s,
                    error = NULL
                WHERE id = %s
            """, (
                int(test_id),
                int(jugador_id),
                fecha,
                puntaje,
                observaciones or item["observaciones"],
                session.get("username"),
                item["id"],
            ))
            procesadas += 1

        conn.commit()
        conn.close()

        if procesadas:
            flash(f"Se confirmaron {procesadas} mediciones.", "ok")
        if omitidas:
            flash(f"{omitidas} mediciones quedaron pendientes.", "warning")
        if errores:
            flash(f"{errores} mediciones requieren revision.", "error")
        return redirect(url_for("revisar_test_importacion", batch_id=batch_id))

    items = conn.execute("""
        SELECT
            b.*,
            js.apellido AS sugerido_apellido,
            js.nombre AS sugerido_nombre,
            js.dni AS sugerido_dni,
            ja.apellido AS asignado_apellido,
            ja.nombre AS asignado_nombre,
            t.nombre AS test_nombre_actual
        FROM test_importaciones_batch b
        LEFT JOIN jugadores js ON js.id = b.jugador_sugerido_id
        LEFT JOIN jugadores ja ON ja.id = b.jugador_id
        LEFT JOIN test_tipos t ON t.id = b.test_id
        WHERE b.batch_id = %s
        ORDER BY b.id
    """, (batch_id,)).fetchall()

    conn.close()

    if not items:
        flash("No se encontro la tanda de importacion.", "error")
        return redirect(url_for("importar_test_resultados"))

    return render_template(
        "test_importar_revision.html",
        batch_id=batch_id,
        items=items,
        tests=tests,
        jugadores=jugadores,
    )


def armar_filtros_tests_desde_request(incluir_test=True):
    test_id_raw = request.args.get("test_id", "").strip()
    test_id = int(test_id_raw) if test_id_raw.isdigit() else None
    categoria = request.args.get("categoria", "").strip()
    desde = validar_fecha_movimiento(request.args.get("desde", "").strip())
    hasta = validar_fecha_movimiento(request.args.get("hasta", "").strip())
    jugadores_seleccionados = [
        int(valor)
        for valor in request.args.getlist("jugadores")
        if str(valor).isdigit()
    ]

    filtros = []
    params = []
    if incluir_test and test_id:
        filtros.append("r.test_id = %s")
        params.append(test_id)
    if categoria:
        filtros.append("j.categoria = %s")
        params.append(categoria)
    if desde:
        filtros.append("r.fecha >= %s")
        params.append(desde)
    if hasta:
        filtros.append("r.fecha <= %s")
        params.append(hasta)
    if jugadores_seleccionados:
        filtros.append("j.id = ANY(%s)")
        params.append(jugadores_seleccionados)

    return {
        "test_id": test_id,
        "categoria": categoria,
        "desde": desde,
        "hasta": hasta,
        "jugadores": jugadores_seleccionados,
        "sql": " AND ".join(filtros),
        "params": params,
    }


@app.route("/tests/exportar")
def exportar_tests():
    check = permiso_requerido("tests_ver")
    if check:
        return check

    filtros = armar_filtros_tests_desde_request()
    where_sql = f"WHERE {filtros['sql']}" if filtros["sql"] else ""

    conn = get_connection()
    resultados = conn.execute(f"""
        SELECT
            r.id,
            r.fecha,
            r.puntaje,
            r.observaciones,
            r.creado_en,
            r.creado_por,
            t.id AS test_id,
            t.nombre AS test_nombre,
            t.unidad,
            t.mayor_es_mejor,
            j.id AS jugador_id,
            j.apellido,
            j.nombre,
            j.dni,
            j.categoria,
            j.estado
        FROM test_resultados r
        JOIN test_tipos t ON t.id = r.test_id
        JOIN jugadores j ON j.id = r.jugador_id
        {where_sql}
        ORDER BY t.nombre, r.fecha DESC, j.apellido, j.nombre, r.id DESC
    """, filtros["params"]).fetchall()

    test_actual = None
    if filtros["test_id"]:
        test_actual = conn.execute("""
            SELECT *
            FROM test_tipos
            WHERE id = %s
        """, (filtros["test_id"],)).fetchone()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append([
        "Fecha",
        "Test",
        "Unidad",
        "Puntaje",
        "Apellido",
        "Nombre",
        "DNI",
        "Categoria",
        "Estado jugador",
        "Observaciones",
        "Cargado por",
        "Cargado en",
        "Resultado ID",
        "Jugador ID",
        "Test ID",
    ])

    resumen = {}
    for item in resultados:
        append_fila_reporte(ws, [
            item["fecha"],
            item["test_nombre"],
            item["unidad"] or "",
            item["puntaje"],
            item["apellido"],
            item["nombre"],
            item["dni"],
            item["categoria"] or "",
            item["estado"] or "",
            item["observaciones"] or "",
            item["creado_por"] or "",
            str(item["creado_en"] or ""),
            item["id"],
            item["jugador_id"],
            item["test_id"],
        ])

        clave = item["test_id"]
        if clave not in resumen:
            resumen[clave] = {
                "test": item["test_nombre"],
                "unidad": item["unidad"] or "",
                "mayor_es_mejor": item["mayor_es_mejor"],
                "puntajes": [],
                "jugadores": set(),
                "fechas": [],
            }
        resumen[clave]["puntajes"].append(float(item["puntaje"]))
        resumen[clave]["jugadores"].add(item["jugador_id"])
        resumen[clave]["fechas"].append(item["fecha"])

    estilizar_hoja_reporte(ws)
    aplicar_formato_columnas(ws, {"D": "0.00"})

    filas_resumen = []
    for item in sorted(resumen.values(), key=lambda valor: valor["test"]):
        puntajes = item["puntajes"]
        fechas = sorted(item["fechas"])
        filas_resumen.append([
            item["test"],
            item["unidad"],
            len(puntajes),
            len(item["jugadores"]),
            fechas[0] if fechas else "",
            fechas[-1] if fechas else "",
            round(sum(puntajes) / len(puntajes), 2) if puntajes else "",
            min(puntajes) if puntajes else "",
            max(puntajes) if puntajes else "",
            "Mayor valor es mejor" if item["mayor_es_mejor"] else "Menor valor es mejor",
        ])
    ws_resumen = agregar_hoja_reporte(wb, "Resumen tests", [
        "Test",
        "Unidad",
        "Mediciones",
        "Jugadores medidos",
        "Primera fecha",
        "Ultima fecha",
        "Promedio",
        "Minimo",
        "Maximo",
        "Criterio",
    ], filas_resumen)
    aplicar_formato_columnas(ws_resumen, {"G": "0.00", "H": "0.00", "I": "0.00"})

    if test_actual:
        comparativo = construir_comparativo_tests(resultados, test_actual)
        encabezados = ["Jugador", "Categoria"] + comparativo["fechas"] + ["Ultimo cambio", "Estado"]
        filas_comparativo = []
        for fila in comparativo["filas"]:
            filas_comparativo.append([
                fila["nombre"],
                fila["categoria"],
                *[fila["valores"].get(fecha, "") for fecha in comparativo["fechas"]],
                fila["delta"] if fila["delta"] is not None else "",
                fila["estado_label"],
            ])
        ws_comparativo = agregar_hoja_reporte(wb, "Comparativo", encabezados, filas_comparativo)
        for col_idx in range(3, 3 + len(comparativo["fechas"])):
            col_letter = get_column_letter(col_idx)
            aplicar_formato_columnas(ws_comparativo, {col_letter: "0.00"})

    filtros_texto = [
        ["Test", test_actual["nombre"] if test_actual else "Todos"],
        ["Categoria", filtros["categoria"] or "Todas"],
        ["Desde", filtros["desde"] or "Sin filtro"],
        ["Hasta", filtros["hasta"] or "Sin filtro"],
        ["Jugadores seleccionados", len(filtros["jugadores"]) if filtros["jugadores"] else "Todos"],
        ["Registros exportados", len(resultados)],
        ["Generado", ahora_sig().strftime("%Y-%m-%d %H:%M")],
        ["Usuario", session.get("username") or ""],
    ]
    agregar_hoja_reporte(wb, "Filtros", ["Filtro", "Valor"], filtros_texto)

    filename = f"tests_deportivos_{ahora_sig().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join("exports", filename)
    os.makedirs("exports", exist_ok=True)
    wb.save(path)

    registrar_auditoria(
        "exportar_ok",
        "test_deportivo",
        str(filtros["test_id"]) if filtros["test_id"] else None,
        {
            "formato": "xlsx",
            "cantidad_registros": len(resultados),
            "categoria": filtros["categoria"],
            "desde": filtros["desde"],
            "hasta": filtros["hasta"],
            "jugadores": len(filtros["jugadores"]),
        },
    )

    return send_file(path, as_attachment=True, download_name=filename)


@app.route("/tests/graficos")
def graficos_tests():
    check = permiso_requerido("tests_ver")
    if check:
        return check

    conn = get_connection()

    tests = obtener_test_tipos(conn, solo_activos=True)
    categorias = conn.execute("""
        SELECT DISTINCT categoria
        FROM jugadores
        WHERE categoria IS NOT NULL AND TRIM(categoria) <> ''
        ORDER BY categoria
    """).fetchall()
    jugadores = conn.execute("""
        SELECT id, apellido, nombre, categoria
        FROM jugadores
        WHERE COALESCE(estado, 'Activo') <> 'Baja'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY categoria, apellido, nombre
    """).fetchall()

    test_id_raw = request.args.get("test_id", "").strip()
    test_id = int(test_id_raw) if test_id_raw.isdigit() else None
    if not test_id and tests:
        test_id = tests[0]["id"]

    categoria = request.args.get("categoria", "").strip()
    desde = validar_fecha_movimiento(request.args.get("desde", "").strip())
    hasta = validar_fecha_movimiento(request.args.get("hasta", "").strip())
    jugadores_seleccionados = [
        int(valor)
        for valor in request.args.getlist("jugadores")
        if str(valor).isdigit()
    ]

    resultados = []
    if test_id:
        filtros = ["r.test_id = %s"]
        params = [test_id]

        if categoria:
            filtros.append("j.categoria = %s")
            params.append(categoria)
        if desde:
            filtros.append("r.fecha >= %s")
            params.append(desde)
        if hasta:
            filtros.append("r.fecha <= %s")
            params.append(hasta)
        if jugadores_seleccionados:
            filtros.append("j.id = ANY(%s)")
            params.append(jugadores_seleccionados)

        where = " AND ".join(filtros)
        resultados = conn.execute(f"""
            SELECT
                r.*,
                j.apellido,
                j.nombre,
                j.categoria
            FROM test_resultados r
            JOIN jugadores j ON j.id = r.jugador_id
            WHERE {where}
            ORDER BY j.apellido, j.nombre, r.fecha
        """, params).fetchall()

    test_actual = next((test for test in tests if test["id"] == test_id), None)
    grafico = construir_grafico_tests(resultados)
    comparativo = construir_comparativo_tests(resultados, test_actual)

    conn.close()
    return render_template(
        "tests_graficos.html",
        tests=tests,
        test_actual=test_actual,
        categorias=categorias,
        jugadores=jugadores,
        jugadores_seleccionados=jugadores_seleccionados,
        categoria=categoria,
        desde=desde or "",
        hasta=hasta or "",
        grafico=grafico,
        comparativo=comparativo,
        resultados=resultados,
        test_id=test_id,
    )


@app.route("/asistencia")
def listar_eventos_asistencia():
    check = permiso_requerido("asistencia_ver")
    if check:
        return check

    conn = get_connection()

    eventos = conn.execute("""
        SELECT
            e.*,
            ce.id AS calendario_evento_id
        FROM eventos_asistencia e
        LEFT JOIN calendario_eventos ce ON ce.asistencia_evento_id = e.id
        ORDER BY
            CASE
                WHEN e.fecha >= CURRENT_DATE::text THEN 0
                ELSE 1
            END,
            CASE
                WHEN e.fecha >= CURRENT_DATE::text THEN e.fecha
            END ASC,
            CASE
                WHEN e.fecha < CURRENT_DATE::text THEN e.fecha
            END DESC,
            e.id DESC
    """).fetchall()

    conn.close()

    return render_template("asistencia_eventos.html", eventos=eventos)


@app.route("/asistencia/nuevo", methods=["GET", "POST"])
def nuevo_evento_asistencia():
    check = permiso_requerido("asistencia_gestionar")
    if check:
        return check

    if request.method == "GET":
        return redirect(url_for("nuevo_evento_calendario", origen="asistencia"))

    if request.method == "POST":
        fecha = request.form.get("fecha", "").strip()
        tipo = request.form.get("tipo", "").strip()
        descripcion = request.form.get("descripcion", "").strip()

        if not fecha or not tipo:
            flash("Fecha y tipo son obligatorios.", "error")
            return render_template("asistencia_evento_form.html")

        conn = get_connection()
        conn.execute("""
            INSERT INTO eventos_asistencia (fecha, tipo, descripcion)
            VALUES (%s, %s, %s)
        """, (fecha, tipo, descripcion))
        conn.commit()
        conn.close()

        flash("Evento de asistencia creado.", "ok")
        return redirect(url_for("listar_eventos_asistencia"))

    return render_template("asistencia_evento_form.html", evento={})


@app.route("/asistencia/<int:evento_id>/editar", methods=["GET", "POST"])
def editar_evento_asistencia(evento_id):
    check = permiso_requerido("asistencia_gestionar")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("""
        SELECT *
        FROM eventos_asistencia
        WHERE id = %s
    """, (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))

    calendario_evento = conn.execute("""
        SELECT id
        FROM calendario_eventos
        WHERE asistencia_evento_id = %s
    """, (evento_id,)).fetchone()
    if calendario_evento is not None:
        conn.close()
        return redirect(url_for("editar_evento_calendario", evento_id=calendario_evento["id"], origen="asistencia"))

    if request.method == "POST":
        fecha = request.form.get("fecha", "").strip()
        tipo = request.form.get("tipo", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        if not fecha or not tipo:
            conn.close()
            flash("Fecha y tipo son obligatorios.", "error")
            return render_template("asistencia_evento_form.html", evento=request.form)
        if evento.get("cerrado"):
            conn.close()
            flash("El evento está cerrado. Solo un admin puede reabrirlo para modificar asistencia o datos.", "error")
            return redirect(url_for("tomar_asistencia", evento_id=evento_id))

        conn.execute("""
            UPDATE eventos_asistencia
            SET fecha = %s,
                tipo = %s,
                descripcion = %s
            WHERE id = %s
        """, (fecha, tipo, descripcion or None, evento_id))
        conn.commit()
        conn.close()
        registrar_auditoria("editar", "asistencia_evento", str(evento_id), {
            "fecha": fecha,
            "tipo": tipo,
        })
        flash("Evento de asistencia actualizado.", "ok")
        return redirect(url_for("listar_eventos_asistencia"))

    conn.close()
    return render_template("asistencia_evento_form.html", evento=evento)


@app.route("/asistencia/<int:evento_id>/cerrar", methods=["POST"])
def cerrar_evento_asistencia(evento_id):
    check = permiso_requerido("asistencia_gestionar")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("SELECT * FROM eventos_asistencia WHERE id = %s", (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))
    if evento.get("cerrado"):
        conn.close()
        flash("El evento ya estaba cerrado.", "warning")
        return redirect(url_for("tomar_asistencia", evento_id=evento_id))

    conn.execute("""
        UPDATE eventos_asistencia
        SET cerrado = 1,
            cerrado_en = CURRENT_TIMESTAMP,
            cerrado_por = %s
        WHERE id = %s
    """, (session.get("username"), evento_id))
    conn.commit()
    conn.close()
    flash("Evento cerrado. La asistencia quedó bloqueada.", "ok")
    return redirect(url_for("tomar_asistencia", evento_id=evento_id))


@app.route("/asistencia/<int:evento_id>/reabrir", methods=["POST"])
def reabrir_evento_asistencia(evento_id):
    check = rol_requerido("admin")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("SELECT * FROM eventos_asistencia WHERE id = %s", (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))

    conn.execute("""
        UPDATE eventos_asistencia
        SET cerrado = 0,
            cerrado_en = NULL,
            cerrado_por = NULL
        WHERE id = %s
    """, (evento_id,))
    conn.commit()
    conn.close()
    flash("Evento reabierto. Ya se puede volver a modificar la asistencia.", "ok")
    return redirect(url_for("tomar_asistencia", evento_id=evento_id))


@app.route("/asistencia/<int:evento_id>/eliminar", methods=["POST"])
def eliminar_evento_asistencia(evento_id):
    check = permiso_requerido("asistencia_gestionar")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("SELECT * FROM eventos_asistencia WHERE id = %s", (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))

    conn.execute("DELETE FROM portal_asistencia_confirmaciones WHERE evento_id = %s", (evento_id,))
    conn.execute("DELETE FROM asistencias WHERE evento_id = %s", (evento_id,))
    conn.execute("DELETE FROM aspirante_asistencias WHERE evento_id = %s", (evento_id,))
    conn.execute("DELETE FROM eventos_asistencia WHERE id = %s", (evento_id,))
    conn.commit()
    conn.close()
    flash("Evento de asistencia eliminado.", "ok")
    return redirect(url_for("listar_eventos_asistencia"))


@app.route("/asistencia/<int:evento_id>/exportar")
def exportar_evento_asistencia(evento_id):
    check = permiso_requerido("asistencia_ver")
    if check:
        return check

    conn = get_connection()
    evento = conn.execute("SELECT * FROM eventos_asistencia WHERE id = %s", (evento_id,)).fetchone()
    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))

    jugadores = conn.execute("""
        SELECT
            'Jugador' AS tipo_persona,
            j.id AS persona_id,
            j.apellido,
            j.nombre,
            j.dni,
            j.categoria,
            a.estado_asistencia,
            a.presente,
            a.observaciones,
            p.estado AS confirmacion_portal,
            p.sueno_calidad,
            p.horas_sueno,
            p.doms,
            p.fatiga,
            p.estres,
            p.animo,
            p.motivacion,
            p.recuperacion,
            p.dolor_zonas,
            p.dolor_otro,
            p.comentarios AS comentarios_portal,
            p.actualizado_en AS portal_actualizado_en
        FROM jugadores j
        LEFT JOIN asistencias a
          ON a.jugador_id = j.id
         AND a.evento_id = %s
        LEFT JOIN portal_asistencia_confirmaciones p
          ON p.jugador_id = j.id
         AND p.evento_id = %s
        WHERE j.estado = 'Activo'
          AND COALESCE(j.tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY j.apellido, j.nombre
    """, (evento_id, evento_id)).fetchall()

    aspirantes = conn.execute("""
        SELECT
            'Ahijadx' AS tipo_persona,
            a2.id AS persona_id,
            a2.apellido,
            a2.nombre,
            a2.dni,
            NULL::TEXT AS categoria,
            aa.estado_asistencia,
            aa.presente,
            aa.observaciones,
            NULL::TEXT AS confirmacion_portal,
            NULL::INTEGER AS sueno_calidad,
            NULL::TEXT AS horas_sueno,
            NULL::INTEGER AS doms,
            NULL::INTEGER AS fatiga,
            NULL::INTEGER AS estres,
            NULL::INTEGER AS animo,
            NULL::INTEGER AS motivacion,
            NULL::INTEGER AS recuperacion,
            NULL::TEXT AS dolor_zonas,
            NULL::TEXT AS dolor_otro,
            NULL::TEXT AS comentarios_portal,
            NULL::TIMESTAMPTZ AS portal_actualizado_en
        FROM aspirantes a2
        LEFT JOIN aspirante_asistencias aa
          ON aa.aspirante_id = a2.id
         AND aa.evento_id = %s
        WHERE a2.estado = 'Aspirante'
        ORDER BY a2.apellido, a2.nombre
    """, (evento_id,)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"
    ws.append([
        "Fecha",
        "Tipo evento",
        "Descripcion",
        "Tipo persona",
        "ID persona",
        "Apellido",
        "Nombre",
        "DNI",
        "Categoria",
        "Estado asistencia",
        "Presente",
        "Observaciones asistencia",
        "Confirmacion portal",
        "Sueno calidad",
        "Horas sueno",
        "DOMS",
        "Fatiga",
        "Estres",
        "Animo",
        "Motivacion",
        "Recuperacion",
        "Molestias",
        "Dolor otro",
        "Comentarios portal",
        "Portal actualizado en",
    ])

    filas = list(jugadores) + list(aspirantes)
    for fila in filas:
        try:
            dolor_zonas = ", ".join(json.loads(fila.get("dolor_zonas") or "[]"))
        except (TypeError, ValueError):
            dolor_zonas = fila.get("dolor_zonas") or ""

        estado_asistencia = fila.get("estado_asistencia")
        presente = fila.get("presente")
        append_fila_reporte(ws, [
            evento.get("fecha") or "",
            evento.get("tipo") or "",
            evento.get("descripcion") or "",
            fila["tipo_persona"],
            fila["persona_id"],
            fila["apellido"],
            fila["nombre"],
            fila.get("dni") or "",
            fila.get("categoria") or "",
            estado_asistencia or "Sin registrar",
            "Si" if presente else ("No" if presente == 0 else ""),
            fila.get("observaciones") or "",
            fila.get("confirmacion_portal") or "",
            fila.get("sueno_calidad") or "",
            fila.get("horas_sueno") or "",
            fila.get("doms") or "",
            fila.get("fatiga") or "",
            fila.get("estres") or "",
            fila.get("animo") or "",
            fila.get("motivacion") or "",
            fila.get("recuperacion") or "",
            dolor_zonas,
            fila.get("dolor_otro") or "",
            fila.get("comentarios_portal") or "",
            fila.get("portal_actualizado_en") or "",
        ])
    estilizar_hoja_reporte(ws)

    resumen_filas = [
        ["Evento ID", evento_id],
        ["Fecha", evento.get("fecha") or ""],
        ["Tipo", evento.get("tipo") or ""],
        ["Descripcion", evento.get("descripcion") or ""],
        ["Jugadores", len(jugadores)],
        ["Ahijadxs", len(aspirantes)],
        ["Total participantes", len(filas)],
        ["Con asistencia registrada", sum(1 for fila in filas if fila.get("estado_asistencia"))],
        ["Confirmaciones portal", sum(1 for fila in filas if fila.get("confirmacion_portal"))],
        ["Generado", ahora_sig().strftime("%Y-%m-%d %H:%M")],
        ["Usuario", session.get("username") or ""],
    ]
    agregar_hoja_reporte(wb, "Resumen", ["Dato", "Valor"], resumen_filas)

    registrar_auditoria("exportar", "asistencia_evento", str(evento_id), {
        "formato": "xlsx",
        "filas": len(filas),
        "fecha": evento.get("fecha"),
        "tipo": evento.get("tipo"),
    })

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"asistencia_{evento_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/asistencia/<int:evento_id>", methods=["GET", "POST"])
def tomar_asistencia(evento_id):
    check = permiso_requerido("asistencia_ver")
    if check:
        return check

    conn = get_connection()

    evento = conn.execute("""
        SELECT *
        FROM eventos_asistencia
        WHERE id = %s
    """, (evento_id,)).fetchone()

    if evento is None:
        conn.close()
        flash("Evento no encontrado.", "error")
        return redirect(url_for("listar_eventos_asistencia"))

    jugadores = conn.execute("""
        SELECT *
        FROM jugadores
        WHERE estado = 'Activo'
          AND COALESCE(tipo_miembro, 'Jugador') = 'Jugador'
        ORDER BY apellido, nombre
    """).fetchall()

    aspirantes = conn.execute("""
        SELECT *
        FROM aspirantes
        WHERE estado = 'Aspirante'
        ORDER BY apellido, nombre
    """).fetchall()

    if request.method == "POST":
        check = permiso_requerido("asistencia_gestionar")
        if check:
            conn.close()
            return check
        if evento.get("cerrado"):
            conn.close()
            flash("El evento está cerrado. Solo un admin puede reabrirlo.", "error")
            return redirect(url_for("tomar_asistencia", evento_id=evento_id))

        for jugador in jugadores:
            estado_asistencia = request.form.get(
                f"estado_jugador_{jugador['id']}",
                "ausente",
            )
            if estado_asistencia not in {"ausente", "a_tiempo", "tarde"}:
                estado_asistencia = "ausente"

            presente = 1 if estado_asistencia in {"a_tiempo", "tarde"} else 0
            observaciones = request.form.get(f"obs_jugador_{jugador['id']}", "").strip()

            conn.execute("""
                INSERT INTO asistencias (
                    evento_id, jugador_id, presente, estado_asistencia, observaciones
                )
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT(evento_id, jugador_id)
                DO UPDATE SET
                    presente = excluded.presente,
                    estado_asistencia = excluded.estado_asistencia,
                    observaciones = excluded.observaciones
            """, (
                evento_id,
                jugador["id"],
                presente,
                estado_asistencia,
                observaciones,
            ))

        for aspirante in aspirantes:
            estado_asistencia = request.form.get(
                f"estado_aspirante_{aspirante['id']}",
                "ausente",
            )
            if estado_asistencia not in {"ausente", "a_tiempo", "tarde"}:
                estado_asistencia = "ausente"

            presente = 1 if estado_asistencia in {"a_tiempo", "tarde"} else 0
            observaciones = request.form.get(f"obs_aspirante_{aspirante['id']}", "").strip()

            conn.execute("""
                INSERT INTO aspirante_asistencias (
                    evento_id, aspirante_id, presente, estado_asistencia, observaciones
                )
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT(evento_id, aspirante_id)
                DO UPDATE SET
                    presente = excluded.presente,
                    estado_asistencia = excluded.estado_asistencia,
                    observaciones = excluded.observaciones
            """, (
                evento_id,
                aspirante["id"],
                presente,
                estado_asistencia,
                observaciones,
            ))

        conn.commit()
        conn.close()

        registrar_auditoria("guardar", "asistencia_evento", str(evento_id), {
            "jugadores": len(jugadores),
            "aspirantes": len(aspirantes),
        })

        flash("Asistencia guardada.", "ok")
        return redirect(url_for("listar_eventos_asistencia"))

    asistencias = conn.execute("""
        SELECT *
        FROM asistencias
        WHERE evento_id = %s
    """, (evento_id,)).fetchall()

    aspirante_asistencias = conn.execute("""
        SELECT *
        FROM aspirante_asistencias
        WHERE evento_id = %s
    """, (evento_id,)).fetchall()

    confirmaciones_portal_rows = conn.execute("""
        SELECT *
        FROM portal_asistencia_confirmaciones
        WHERE evento_id = %s
    """, (evento_id,)).fetchall()

    conn.close()

    confirmaciones_portal = []
    for row in confirmaciones_portal_rows:
        item = dict(row)
        try:
            item["dolor_zonas_lista"] = json.loads(item.get("dolor_zonas") or "[]")
        except (TypeError, ValueError):
            item["dolor_zonas_lista"] = []
        item["bienestar_resumen"] = resumen_bienestar_confirmacion(item)
        item["bienestar_completo"] = item["bienestar_resumen"] is not None
        confirmaciones_portal.append(item)

    asistencias_por_jugador = {
        a["jugador_id"]: a for a in asistencias
    }
    asistencias_por_aspirante = {
        a["aspirante_id"]: a for a in aspirante_asistencias
    }
    confirmaciones_por_jugador = {
        c["jugador_id"]: c for c in confirmaciones_portal
    }

    participantes = []
    for jugador in jugadores:
        item = dict(jugador)
        item["tipo"] = "jugador"
        item["tipo_label"] = "Jugador"
        item["form_key"] = f"jugador_{jugador['id']}"
        item["asistencia"] = asistencias_por_jugador.get(jugador["id"])
        item["confirmacion_portal"] = confirmaciones_por_jugador.get(jugador["id"])
        participantes.append(item)

    for aspirante in aspirantes:
        item = dict(aspirante)
        item["tipo"] = "aspirante"
        item["tipo_label"] = "Ahijadx"
        item["form_key"] = f"aspirante_{aspirante['id']}"
        item["asistencia"] = asistencias_por_aspirante.get(aspirante["id"])
        item["confirmacion_portal"] = None
        participantes.append(item)

    bienestar_completado = [c for c in confirmaciones_portal if c.get("bienestar_completo")]
    promedio_general = None
    if bienestar_completado:
        suma_promedios = 0
        cantidad_promedios = 0
        dolores_totales = {}
        for item in bienestar_completado:
            resumen = item.get("bienestar_resumen") or {}
            promedio = resumen.get("promedio")
            if promedio is not None:
                suma_promedios += promedio
                cantidad_promedios += 1
            for dolor in resumen.get("dolores") or []:
                dolores_totales[dolor] = dolores_totales.get(dolor, 0) + 1
        if cantidad_promedios:
            promedio_general = round(suma_promedios / cantidad_promedios, 1)
        dolores_frecuentes = [
            nombre for nombre, _cantidad in sorted(
                dolores_totales.items(),
                key=lambda item: (-item[1], item[0].lower()),
            )[:3]
        ]
    else:
        dolores_frecuentes = []

    bienestar_resumen_evento = {
        "confirmaciones": len(confirmaciones_portal),
        "completados": len(bienestar_completado),
        "rojos": sum(
            1 for item in bienestar_completado
            if (item.get("bienestar_resumen") or {}).get("nivel") == "danger"
        ),
        "amarillos": sum(
            1 for item in bienestar_completado
            if (item.get("bienestar_resumen") or {}).get("nivel") == "warning"
        ),
        "verdes": sum(
            1 for item in bienestar_completado
            if (item.get("bienestar_resumen") or {}).get("nivel") == "success"
        ),
        "promedio_general": promedio_general,
        "dolores_frecuentes": dolores_frecuentes,
    }
    resumen_categoria = {}
    for participante in participantes:
        categoria = (participante.get("categoria") or "Sin categoria").strip() or "Sin categoria"
        bucket = resumen_categoria.setdefault(categoria, {"total": 0, "confirmados": 0, "alertas": 0})
        bucket["total"] += 1
        if participante.get("confirmacion_portal"):
            bucket["confirmados"] += 1
            bienestar = participante["confirmacion_portal"].get("bienestar_resumen") or {}
            if bienestar.get("nivel") == "danger":
                bucket["alertas"] += 1
    resumen_categoria = [{"categoria": categoria, **datos} for categoria, datos in sorted(resumen_categoria.items())]

    return render_template(
        "tomar_asistencia.html",
        evento=evento,
        participantes=participantes,
        puede_reabrir=session.get("rol") == "admin",
        requiere_bienestar=not es_evento_partido(evento),
        bienestar_resumen_evento=bienestar_resumen_evento,
        bienestar_por_categoria=resumen_categoria,
    )

if os.environ.get("INIT_DB", "true").lower() in {"1", "true", "yes", "on"}:
    init_db()

if __name__ == "__main__":
    app.run(debug=True)
